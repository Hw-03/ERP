"""
Sync legacy metadata and current stock from ERP Excel workbooks into backend/erp.db.

Usage:
    python backend/sync_excel_stock.py
    cd backend && python sync_excel_stock.py
"""

from __future__ import annotations

import csv
import os
import re
import sys
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
CSV_PATH = DATA_DIR / "ERP_Master_DB.csv"
SQLITE_PATH = BACKEND_DIR / "erp.db"

sys.path.insert(0, str(BACKEND_DIR))
os.environ["DATABASE_URL"] = f"sqlite:///{SQLITE_PATH.as_posix()}"

from app.database import SessionLocal
from app.models import Inventory, Item


CATEGORY_TO_FILE_TYPE: dict[str, str] = {
    "RM": "원자재",
    "TA": "조립자재",
    "TF": "조립자재",
    "HA": "발생부자재",
    "HF": "발생부자재",
    "VA": "발생부자재",
    "VF": "발생부자재",
    "AA": "조립자재",
    "AF": "조립자재",
    "FG": "완제품",
}

CATEGORY_TO_PART: dict[str, str] = {
    "RM": "자재창고",
    "TA": "튜닝파트",
    "TF": "튜닝파트",
    "HA": "고압파트",
    "HF": "고압파트",
    "VA": "진공파트",
    "VF": "진공파트",
    "AA": "조립출하",
    "AF": "조립출하",
    "FG": "출하",
}


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def to_decimal(value: object | None) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except Exception:
        return None


def to_number(value: object | None) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return 0.0


def load_master_rows() -> list[dict[str, str]]:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def apply_metadata(db, rows: list[dict[str, str]]) -> int:
    updated = 0
    for row in rows:
        item_code = (row.get("item_id") or "").strip()
        if not item_code:
            continue

        raw_category_code = (row.get("category_code") or "").strip().upper()
        changed = (
            db.query(Item)
            .filter(Item.item_code == item_code)
            .update(
                {
                    Item.barcode: item_code,
                    Item.legacy_file_type: CATEGORY_TO_FILE_TYPE.get(raw_category_code, "미분류"),
                    Item.legacy_part: CATEGORY_TO_PART.get(raw_category_code, "자재창고"),
                    Item.legacy_item_type: normalize_text(row.get("part_type")) or None,
                    Item.legacy_model: normalize_text(row.get("model_ref")) or "공용",
                    Item.supplier: normalize_text(row.get("supplier")) or None,
                    Item.min_stock: to_decimal(row.get("safety_stock")),
                },
                synchronize_session=False,
            )
        )
        updated += changed

    return updated


def queue_rows(rows: list[dict[str, str]]):
    by_name_spec: dict[tuple[str, str], list[str]] = defaultdict(list)
    by_name: dict[str, list[str]] = defaultdict(list)

    for row in rows:
        item_code = (row.get("item_id") or "").strip()
        if not item_code:
            continue
        name = normalize_text(row.get("original_name_a") or row.get("std_name"))
        spec = normalize_text(row.get("std_spec"))
        by_name_spec[(name, spec)].append(item_code)
        by_name[name].append(item_code)

    return by_name_spec, by_name


def consume_match(
    name: str,
    spec: str,
    by_name_spec: dict[tuple[str, str], list[str]],
    by_name: dict[str, list[str]],
    used: set[str],
) -> str | None:
    for queue in (by_name_spec.get((name, spec), []), by_name.get(name, [])):
        while queue and queue[0] in used:
            queue.pop(0)
        if queue:
            item_code = queue.pop(0)
            used.add(item_code)
            return item_code
    return None


def sum_row_cells(ws, row_index: int, start_col: int, end_col: int) -> float:
    total = 0.0
    for values in ws.iter_rows(min_row=row_index, max_row=row_index, min_col=start_col, max_col=end_col, values_only=True):
        for cell in values:
            total += to_number(cell)
    return total


def load_stock_updates(rows: list[dict[str, str]]) -> list[tuple[str, float, str]]:
    updates: list[tuple[str, float, str]] = []
    by_name_spec, by_name = queue_rows(rows)
    used: set[str] = set()

    f704 = next((path for path in DATA_DIR.iterdir() if path.suffix.lower() == ".xlsx" and path.name.startswith("F704")), None)
    if f704 is not None:
        workbook = load_workbook(f704, read_only=True, data_only=True)
        sheet = workbook["26.03월"] if "26.03월" in workbook.sheetnames else workbook[workbook.sheetnames[-1]]
        for values in sheet.iter_rows(min_row=4, values_only=True):
            row = list(values)
            name = normalize_text(row[3] if len(row) > 3 else None)
            spec = normalize_text(row[4] if len(row) > 4 else None)
            if not name:
                continue
            quantity = (
                to_number(row[10] if len(row) > 10 else None)
                + to_number(row[21] if len(row) > 21 else None)
                - to_number(row[30] if len(row) > 30 else None)
                - to_number(row[51] if len(row) > 51 else None)
            )
            item_code = consume_match(name, spec, by_name_spec, by_name, used)
            if item_code:
                updates.append((item_code, quantity, f704.name))

    for path in DATA_DIR.iterdir():
        if path.suffix.lower() != ".xlsx" or path.name.startswith("F704"):
            continue

        workbook = load_workbook(path, read_only=True, data_only=True)
        main_sheet = workbook[workbook.sheetnames[0]]
        incoming_sheet = workbook[workbook.sheetnames[1]]
        outgoing_sheet = workbook[workbook.sheetnames[2]]

        if main_sheet.title == "조립 자재":
            start_row = 3
            name_col = 4
            prev_col = 6
        else:
            start_row = 3
            name_col = 5
            prev_col = 7

        for row_index, values in enumerate(main_sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            row = list(values)
            name = normalize_text(row[name_col - 1] if len(row) >= name_col else None)
            if not name or name in {"품명", "항목"}:
                continue

            prev_stock = to_number(row[prev_col - 1] if len(row) >= prev_col else None)
            incoming = sum_row_cells(incoming_sheet, row_index, 6, 36)
            outgoing = sum_row_cells(outgoing_sheet, row_index, 6, 36)
            quantity = prev_stock + incoming - outgoing

            item_code = consume_match(name, "", by_name_spec, by_name, used)
            if item_code:
                updates.append((item_code, quantity, path.name))

    return updates


def sync() -> None:
    rows = load_master_rows()
    stock_updates = load_stock_updates(rows)

    db = SessionLocal()
    try:
        metadata_updated = apply_metadata(db, rows)

        stock_counter: Counter[str] = Counter()
        for item_code, quantity, source in stock_updates:
            item_id = db.query(Item.item_id).filter(Item.item_code == item_code).scalar()
            if not item_id:
                continue
            changed = (
                db.query(Inventory)
                .filter(Inventory.item_id == item_id)
                .update({Inventory.quantity: Decimal(str(quantity))}, synchronize_session=False)
            )
            if changed:
                stock_counter[source] += 1

        db.commit()

        print(f"Metadata updated: {metadata_updated}")
        print(f"Stock updates applied: {len(stock_updates)}")
        for source, count in stock_counter.items():
            print(f"  {source}: {count}")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    sync()
