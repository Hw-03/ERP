"""Excel export helpers using openpyxl."""

from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def apply_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def make_xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
