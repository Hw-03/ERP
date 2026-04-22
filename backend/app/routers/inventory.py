"""Inventory router for summary, receipts, shipments, transfers, defective, and transactions."""

import uuid
import csv
from decimal import Decimal
from io import StringIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    CategoryEnum,
    DepartmentEnum,
    Inventory,
    InventoryLocation,
    Item,
    LocationStatusEnum,
    ShipPackage,
    TransactionLog,
    TransactionTypeEnum,
)
from app.schemas import (
    CategorySummary,
    InventoryAdjust,
    InventoryLocationResponse,
    InventoryReceive,
    InventoryResponse,
    InventoryShip,
    InventorySummaryResponse,
    MarkDefectiveRequest,
    PackageShipRequest,
    SupplierReturnRequest,
    DeptTransferRequest,
    TransactionLogResponse,
    TransactionLogUpdate,
    TransferRequest,
)
from app.services import inventory as inventory_svc


def _list_locations(db: Session, item_id: uuid.UUID) -> List[InventoryLocationResponse]:
    rows = (
        db.query(InventoryLocation)
        .filter(InventoryLocation.item_id == item_id, InventoryLocation.quantity > 0)
        .all()
    )
    return [
        InventoryLocationResponse(
            department=row.department,
            status=row.status,
            quantity=row.quantity or Decimal("0"),
        )
        for row in rows
    ]


def _to_response(db: Session, inv: Inventory) -> InventoryResponse:
    """Build InventoryResponse with bucket breakdown."""
    pending = inv.pending_quantity or Decimal("0")
    wh = inv.warehouse_qty or Decimal("0")
    prod = inventory_svc.production_total(db, inv.item_id)
    defect = inventory_svc.defective_total(db, inv.item_id)
    total = wh + prod + defect
    return InventoryResponse(
        inventory_id=inv.inventory_id,
        item_id=inv.item_id,
        quantity=total,
        warehouse_qty=wh,
        production_total=prod,
        defective_total=defect,
        pending_quantity=pending,
        available_quantity=wh + prod - pending,
        last_reserver_name=inv.last_reserver_name,
        location=inv.location,
        updated_at=inv.updated_at,
        locations=_list_locations(db, inv.item_id),
    )


router = APIRouter()

CATEGORY_LABELS = {
    CategoryEnum.RM: "원자재",
    CategoryEnum.TA: "튜브 반제품",
    CategoryEnum.TF: "튜브 완제품",
    CategoryEnum.HA: "고압 반제품",
    CategoryEnum.HF: "고압 완제품",
    CategoryEnum.VA: "진공 반제품",
    CategoryEnum.VF: "진공 완제품",
    CategoryEnum.BA: "조립 반제품",
    CategoryEnum.BF: "조립 완제품",
    CategoryEnum.FG: "완제품",
    CategoryEnum.UK: "미분류 품목",
}

CATEGORY_ORDER = [
    CategoryEnum.RM,
    CategoryEnum.TA,
    CategoryEnum.TF,
    CategoryEnum.HA,
    CategoryEnum.HF,
    CategoryEnum.VA,
    CategoryEnum.VF,
    CategoryEnum.BA,
    CategoryEnum.BF,
    CategoryEnum.FG,
    CategoryEnum.UK,
]


@router.get("/summary", response_model=InventorySummaryResponse)
def get_inventory_summary(db: Session = Depends(get_db)):
    rows = (
        db.query(
            Item.category,
            func.count(Item.item_id).label("item_count"),
            func.coalesce(func.sum(Inventory.quantity), 0).label("total_quantity"),
            func.coalesce(func.sum(Inventory.warehouse_qty), 0).label("warehouse_sum"),
        )
        .outerjoin(Inventory, Item.item_id == Inventory.item_id)
        .group_by(Item.category)
        .all()
    )

    # category별 production / defective 합 (InventoryLocation 조인)
    loc_rows = (
        db.query(
            Item.category,
            InventoryLocation.status,
            func.coalesce(func.sum(InventoryLocation.quantity), 0).label("loc_sum"),
        )
        .join(InventoryLocation, InventoryLocation.item_id == Item.item_id)
        .group_by(Item.category, InventoryLocation.status)
        .all()
    )
    prod_map: dict = {}
    defect_map: dict = {}
    for cat, st, val in loc_rows:
        v = Decimal(str(val or 0))
        if st == LocationStatusEnum.PRODUCTION:
            prod_map[cat] = prod_map.get(cat, Decimal("0")) + v
        elif st == LocationStatusEnum.DEFECTIVE:
            defect_map[cat] = defect_map.get(cat, Decimal("0")) + v

    summary_map = {
        row.category: {
            "item_count": row.item_count,
            "total_quantity": Decimal(str(row.total_quantity)),
            "warehouse_sum": Decimal(str(row.warehouse_sum)),
        }
        for row in rows
    }

    total_items = sum(value["item_count"] for value in summary_map.values())
    total_quantity = sum(value["total_quantity"] for value in summary_map.values())
    uk_count = summary_map.get(CategoryEnum.UK, {}).get("item_count", 0)

    categories = []
    for category in CATEGORY_ORDER:
        data = summary_map.get(
            category,
            {"item_count": 0, "total_quantity": Decimal("0"), "warehouse_sum": Decimal("0")},
        )
        categories.append(
            CategorySummary(
                category=category,
                category_label=CATEGORY_LABELS[category],
                item_count=data["item_count"],
                total_quantity=data["total_quantity"],
                warehouse_qty_sum=data["warehouse_sum"],
                production_qty_sum=prod_map.get(category, Decimal("0")),
                defective_qty_sum=defect_map.get(category, Decimal("0")),
            )
        )

    return InventorySummaryResponse(
        categories=categories,
        total_items=total_items,
        total_quantity=total_quantity,
        uk_item_count=uk_count,
    )


@router.post("/receive", response_model=InventoryResponse, status_code=status.HTTP_201_CREATED)
def receive_inventory(payload: InventoryReceive, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")

    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    inventory_svc.receive_confirmed(db, payload.item_id, payload.quantity, bucket="warehouse")
    if payload.location:
        inventory.location = payload.location

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.RECEIVE,
            quantity_change=payload.quantity,
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.notes,
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/ship", response_model=InventoryResponse, status_code=status.HTTP_200_OK)
def ship_inventory(payload: InventoryShip, db: Session = Depends(get_db)):
    """출고: 출하부 PRODUCTION 에서만 차감."""
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")

    inventory = db.query(Inventory).filter(Inventory.item_id == payload.item_id).first()
    if not inventory:
        raise HTTPException(status_code=404, detail="출고할 재고가 존재하지 않습니다.")

    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.consume_from_department(
            db, payload.item_id, payload.quantity, DepartmentEnum.SHIPPING
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"{exc} 다른 부서에서 출하부로 먼저 이동해 주세요.",
        )

    if payload.location is not None:
        inventory.location = payload.location

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.SHIP,
            quantity_change=-payload.quantity,
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.notes,
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/ship-package", status_code=status.HTTP_200_OK)
def ship_package(payload: PackageShipRequest, db: Session = Depends(get_db)):
    """패키지 출고: 모든 구성품을 출하부 PRODUCTION 에서 차감."""
    package = db.query(ShipPackage).filter(ShipPackage.package_id == payload.package_id).first()
    if not package:
        raise HTTPException(status_code=404, detail="출하 패키지를 찾을 수 없습니다.")

    if not package.items:
        raise HTTPException(status_code=400, detail="패키지에 등록된 품목이 없습니다.")

    shortages: list[str] = []
    for package_item in package.items:
        loc = (
            db.query(InventoryLocation)
            .filter(
                InventoryLocation.item_id == package_item.item_id,
                InventoryLocation.department == DepartmentEnum.SHIPPING,
                InventoryLocation.status == LocationStatusEnum.PRODUCTION,
            )
            .first()
        )
        current = (loc.quantity if loc else None) or Decimal("0")
        required_qty = package_item.quantity * payload.quantity
        if current < required_qty:
            shortages.append(
                f"[{package_item.item.erp_code}] {package_item.item.item_name}: "
                f"필요 {required_qty}, 출하부 보유 {current}"
            )

    if shortages:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "출하부 재고가 부족합니다. 다른 부서에서 출하부로 먼저 이동하세요.",
                "shortages": shortages,
            },
        )

    shipped_items = []
    for package_item in package.items:
        required_qty = package_item.quantity * payload.quantity
        inventory = db.query(Inventory).filter(
            Inventory.item_id == package_item.item_id
        ).with_for_update().first()
        before_qty = inventory.quantity if inventory else Decimal("0")
        inventory_svc.consume_from_department(
            db, package_item.item_id, required_qty, DepartmentEnum.SHIPPING
        )

        db.add(
            TransactionLog(
                item_id=package_item.item_id,
                transaction_type=TransactionTypeEnum.SHIP,
                quantity_change=-required_qty,
                quantity_before=before_qty,
                quantity_after=inventory.quantity if inventory else Decimal("0"),
                reference_no=payload.reference_no,
                produced_by=payload.produced_by,
                notes=payload.notes or f"[출하 패키지] {package.name} x {payload.quantity}",
            )
        )
        shipped_items.append(
            {
                "item_id": str(package_item.item_id),
                "erp_code": package_item.item.erp_code,
                "item_name": package_item.item.item_name,
                "quantity": float(required_qty),
                "stock_after": float(inventory.quantity if inventory else 0),
            }
        )

    db.commit()
    return {
        "message": f"{package.name} 패키지 {payload.quantity}건 출고 완료",
        "package_name": package.name,
        "quantity": float(payload.quantity),
        "items": shipped_items,
    }


@router.post("/adjust", response_model=InventoryResponse, status_code=status.HTTP_200_OK)
def adjust_inventory(payload: InventoryAdjust, db: Session = Depends(get_db)):
    """재고 조정: warehouse_qty를 직접 설정. (기존 의미 유지)

    payload.quantity 는 조정 후 창고 수량. production/defective 는 건드리지 않음."""
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")

    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    wh_before = inventory.warehouse_qty or Decimal("0")
    delta = payload.quantity - wh_before

    inventory.warehouse_qty = payload.quantity
    inventory_svc._sync_total(db, payload.item_id)
    if payload.location is not None:
        inventory.location = payload.location

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.ADJUST,
            quantity_change=delta,
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.reason,
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


# ---------------------------------------------------------------------------
# 신규: 이동 / 불량 / 공급업체 반품
# ---------------------------------------------------------------------------


@router.post("/transfer-to-production", response_model=InventoryResponse)
def transfer_to_production(payload: TransferRequest, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")
    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.transfer_to_production(
            db, payload.item_id, payload.quantity, payload.department
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_TO_PROD,
            quantity_change=Decimal("0"),
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.notes or f"창고 → {payload.department.value} 이동 ({payload.quantity})",
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/transfer-to-warehouse", response_model=InventoryResponse)
def transfer_to_warehouse(payload: TransferRequest, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")
    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.transfer_to_warehouse(
            db, payload.item_id, payload.quantity, payload.department
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_TO_WH,
            quantity_change=Decimal("0"),
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.notes or f"{payload.department.value} → 창고 복귀 ({payload.quantity})",
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/transfer-between-depts", response_model=InventoryResponse)
def transfer_between_depts(payload: DeptTransferRequest, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")
    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.transfer_between_departments(
            db, payload.item_id, payload.quantity,
            payload.from_department, payload.to_department,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_DEPT,
            quantity_change=Decimal("0"),
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            notes=payload.notes or f"{payload.from_department.value} → {payload.to_department.value} 이동 ({payload.quantity})",
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/mark-defective", response_model=InventoryResponse)
def mark_defective(payload: MarkDefectiveRequest, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")
    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.mark_defective(
            db, payload.item_id, payload.quantity,
            source=payload.source,
            target_dept=payload.target_department,
            source_dept=payload.source_department,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    note = (
        f"불량 등록 [{payload.source}"
        + (f"/{payload.source_department.value}" if payload.source_department else "")
        + f"] → {payload.target_department.value} 격리 ({payload.quantity})"
        + (f" — {payload.reason}" if payload.reason else "")
    )
    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.MARK_DEFECTIVE,
            quantity_change=Decimal("0"),
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=None,
            produced_by=payload.operator,
            notes=note,
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.post("/return-to-supplier", response_model=InventoryResponse)
def return_to_supplier(payload: SupplierReturnRequest, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == payload.item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다.")
    inventory = inventory_svc.get_or_create_inventory(db, payload.item_id)
    qty_before = inventory.quantity or Decimal("0")
    try:
        inventory_svc.return_to_supplier(
            db, payload.item_id, payload.quantity, payload.from_department
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))

    db.add(
        TransactionLog(
            item_id=payload.item_id,
            transaction_type=TransactionTypeEnum.SUPPLIER_RETURN,
            quantity_change=-payload.quantity,
            quantity_before=qty_before,
            quantity_after=inventory.quantity,
            reference_no=payload.reference_no,
            produced_by=payload.operator,
            notes=payload.notes or f"공급업체 반품 ({payload.from_department.value} 불량 {payload.quantity})",
        )
    )
    db.commit()
    db.refresh(inventory)
    return _to_response(db, inventory)


@router.get("/locations/{item_id}", response_model=List[InventoryLocationResponse])
def get_item_locations(item_id: uuid.UUID, db: Session = Depends(get_db)):
    """품목의 부서×상태 분포 조회 (수량 0인 행 포함, 모든 분포 노출)."""
    rows = (
        db.query(InventoryLocation)
        .filter(InventoryLocation.item_id == item_id)
        .all()
    )
    return [
        InventoryLocationResponse(
            department=row.department,
            status=row.status,
            quantity=row.quantity or Decimal("0"),
        )
        for row in rows
    ]


@router.get("", response_model=List[InventoryResponse])
def list_inventory(
    category: Optional[CategoryEnum] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    query = db.query(Inventory).join(Item, Inventory.item_id == Item.item_id)
    if category:
        query = query.filter(Item.category == category)

    rows = query.order_by(Item.erp_code).offset(skip).limit(limit).all()
    return [_to_response(db, inv) for inv in rows]


@router.get("/transactions", response_model=List[TransactionLogResponse])
def list_transactions(
    item_id: Optional[uuid.UUID] = Query(None),
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    reference_no: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    query = db.query(TransactionLog, Item).join(Item, TransactionLog.item_id == Item.item_id)

    if item_id:
        query = query.filter(TransactionLog.item_id == item_id)
    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    if reference_no:
        query = query.filter(TransactionLog.reference_no == reference_no)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Item.item_name.ilike(pattern),
                Item.erp_code.ilike(pattern),
                TransactionLog.reference_no.ilike(pattern),
                TransactionLog.notes.ilike(pattern),
                TransactionLog.produced_by.ilike(pattern),
            )
        )

    rows = query.order_by(TransactionLog.created_at.desc()).offset(skip).limit(limit).all()

    return [
        TransactionLogResponse(
            log_id=log.log_id,
            item_id=log.item_id,
            erp_code=item.erp_code,
            item_name=item.item_name,
            item_category=item.category,
            item_unit=item.unit,
            transaction_type=log.transaction_type,
            quantity_change=log.quantity_change,
            quantity_before=log.quantity_before,
            quantity_after=log.quantity_after,
            reference_no=log.reference_no,
            produced_by=log.produced_by,
            notes=log.notes,
            created_at=log.created_at,
        )
        for log, item in rows
    ]


@router.get("/transactions/export.csv")
def export_transactions_csv(
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(TransactionLog, Item).join(Item, TransactionLog.item_id == Item.item_id)

    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Item.item_name.ilike(pattern),
                Item.erp_code.ilike(pattern),
                TransactionLog.reference_no.ilike(pattern),
                TransactionLog.notes.ilike(pattern),
                TransactionLog.produced_by.ilike(pattern),
            )
        )

    rows = query.order_by(TransactionLog.created_at.desc()).all()

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "created_at",
            "transaction_type",
            "erp_code",
            "item_name",
            "category",
            "quantity_change",
            "quantity_before",
            "quantity_after",
            "reference_no",
            "produced_by",
            "notes",
        ]
    )
    for log, item in rows:
        writer.writerow(
            [
                log.created_at.isoformat(),
                log.transaction_type.value,
                item.erp_code or "",
                item.item_name,
                item.category.value,
                float(log.quantity_change),
                "" if log.quantity_before is None else float(log.quantity_before),
                "" if log.quantity_after is None else float(log.quantity_after),
                log.reference_no or "",
                log.produced_by or "",
                log.notes or "",
            ]
        )

    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="transactions-export.csv"'},
    )


_TX_ROW_COLOR = {
    "RECEIVE":   "D4EDDA",
    "PRODUCE":   "CCE5FF",
    "SHIP":      "F8D7DA",
    "ADJUST":    "FFF3CD",
    "BACKFLUSH": "FFE5B4",
    "TRANSFER_TO_PROD": "E0E7FF",
    "TRANSFER_TO_WH":   "E0E7FF",
    "TRANSFER_DEPT":    "E0E7FF",
    "MARK_DEFECTIVE":   "FBD9D7",
    "SUPPLIER_RETURN":  "F4C7C3",
}


@router.get("/transactions/export.xlsx")
def export_transactions_xlsx(
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    from datetime import date as _date
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from app.utils.excel import apply_header, auto_width, make_xlsx_response

    query = db.query(TransactionLog, Item).join(Item, TransactionLog.item_id == Item.item_id)
    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Item.item_name.ilike(pattern),
                Item.erp_code.ilike(pattern),
                TransactionLog.reference_no.ilike(pattern),
                TransactionLog.notes.ilike(pattern),
                TransactionLog.produced_by.ilike(pattern),
            )
        )

    rows = query.order_by(TransactionLog.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "거래 이력"

    tx_label = {
        "RECEIVE": "입고", "PRODUCE": "생산입고", "SHIP": "출고",
        "ADJUST": "재고조정", "BACKFLUSH": "자동차감",
        "TRANSFER_TO_PROD": "창고→부서", "TRANSFER_TO_WH": "부서→창고",
        "TRANSFER_DEPT": "부서간 이동",
        "MARK_DEFECTIVE": "불량 등록", "SUPPLIER_RETURN": "공급업체 반품",
    }

    columns = [
        "일시", "유형", "ERP코드", "품목명", "카테고리",
        "수량변화", "이전재고", "이후재고", "참조번호", "담당자", "메모",
    ]
    apply_header(ws, columns)

    positive_font = Font(color="1A7C3C", bold=True)
    negative_font = Font(color="CC0000", bold=True)

    for log, item in rows:
        tx_val = log.transaction_type.value
        row_data = [
            log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
            tx_label.get(tx_val, tx_val),
            item.erp_code or "",
            item.item_name,
            item.category.value,
            float(log.quantity_change),
            float(log.quantity_before) if log.quantity_before is not None else "",
            float(log.quantity_after) if log.quantity_after is not None else "",
            log.reference_no or "",
            log.produced_by or "",
            log.notes or "",
        ]
        ws.append(row_data)

        row_idx = ws.max_row
        hex_color = _TX_ROW_COLOR.get(tx_val, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws[row_idx]:
            cell.fill = row_fill

        qty_change_cell = ws.cell(row=row_idx, column=6)
        if isinstance(qty_change_cell.value, float):
            qty_change_cell.font = positive_font if qty_change_cell.value >= 0 else negative_font

    auto_width(ws)
    filename = f"transactions-{_date.today().strftime('%Y%m%d')}.xlsx"
    return make_xlsx_response(wb, filename)


@router.put("/transactions/{log_id}", response_model=TransactionLogResponse)
def update_transaction_notes(
    log_id: uuid.UUID,
    payload: TransactionLogUpdate,
    db: Session = Depends(get_db),
):
    """Update the notes field of a transaction log entry."""
    log = db.query(TransactionLog).filter(TransactionLog.log_id == log_id).first()
    if not log:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Transaction not found")
    item = db.query(Item).filter(Item.item_id == log.item_id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    log.notes = payload.notes
    db.commit()
    db.refresh(log)

    return TransactionLogResponse(
        log_id=log.log_id,
        item_id=log.item_id,
        erp_code=item.erp_code,
        item_name=item.item_name,
        item_category=item.category,
        item_unit=item.unit,
        transaction_type=log.transaction_type,
        quantity_change=log.quantity_change,
        quantity_before=log.quantity_before,
        quantity_after=log.quantity_after,
        reference_no=log.reference_no,
        produced_by=log.produced_by,
        notes=log.notes,
        created_at=log.created_at,
    )
