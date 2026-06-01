"""Items router for item master CRUD operations."""

from datetime import UTC, datetime
import csv
from io import StringIO
import uuid
from typing import Annotated, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from sqlalchemy import func

from app.database import get_db
from app.dependencies.admin import require_admin_pin
from app.models import BOM, Inventory, InventoryLocation, Item, LocationStatusEnum
from app.routers._errors import ErrorCode, http_error
from app.schemas import (
    BomCompletionUpdate,
    InventoryLocationResponse,
    ItemCreate,
    ItemReorderPayload,
    ItemResponse,
    ItemUpdate,
    ItemWithInventory,
)
from app.utils.mes_code import (
    mes_code_to_model_slots,
    make_mes_code,
    next_serial_no,
    slots_to_model_symbol,
)
from app.models import ProductSymbol
from app.services import audit
from app.services import inventory as inventory_svc
from app.services import stock_math
from app.services._tx import commit_and_refresh
from app._evt import emit as _evt_emit
from app.services.export_helpers import csv_streaming_response
from app.services.reorder import reorder_by_display_order

router = APIRouter()


def _build_item_query(db: Session):
    return db.query(Item, Inventory).outerjoin(Inventory, Item.item_id == Inventory.item_id)


def _to_item_with_inventory(
    db: Session,
    item: Item,
    inventory: Optional[Inventory],
    *,
    figures: Optional[stock_math.StockFigures] = None,
    locations: Optional[list[InventoryLocationResponse]] = None,
    model_slots: Optional[list[int]] = None,
) -> ItemWithInventory:
    """ItemWithInventory DTO 조립.

    성능 모드 (Phase C bulk prefetch 용): 호출측이 figures / locations / model_slots 를
    미리 채워 넣으면 DB 쿼리를 추가로 발생시키지 않는다. 인자를 생략하면 기존처럼
    단건 쿼리를 수행한다 (단건 상세 조회용).
    """
    from decimal import Decimal as _D

    fig = figures if figures is not None else stock_math.compute_for(db, item.item_id)

    if locations is None:
        loc_rows = (
            db.query(InventoryLocation)
            .filter(InventoryLocation.item_id == item.item_id, InventoryLocation.quantity > 0)
            .all()
        )
        locations = [
            InventoryLocationResponse(
                department=row.department,
                status=row.status,
                quantity=row.quantity or _D("0"),
            )
            for row in loc_rows
        ]

    if model_slots is None:
        # 회사 규약: mes_code prefix(첫 '-' 앞 글자열) 가 모델을 결정.
        # 예: "8-AR-0307" → SOLO(slot 3), "78-PR-0042" → COCOON+SOLO(slot 2,3).
        # 별도 item_models 테이블 없이 코드만 보면 됨.
        model_slots = mes_code_to_model_slots(item.mes_code)

    return ItemWithInventory(
        item_id=item.item_id,
        item_name=item.item_name,
        unit=item.unit,
        legacy_part=item.legacy_part,
        legacy_item_type=item.legacy_item_type,
        supplier=item.supplier,
        min_stock=item.min_stock,
        mes_code=item.mes_code,
        model_symbol=item.model_symbol,
        model_slots=model_slots,
        process_type_code=item.process_type_code,
        serial_no=item.serial_no,
        bom_completed_at=item.bom_completed_at,
        deleted_at=item.deleted_at,
        created_at=item.created_at,
        updated_at=item.updated_at,
        quantity=fig.total,
        warehouse_qty=fig.warehouse_qty,
        production_total=fig.production_total,
        defective_total=fig.defective_total,
        pending_quantity=fig.pending,
        available_quantity=fig.available,
        last_reserver_name=inventory.last_reserver_name if inventory else None,
        location=inventory.location if inventory else None,
        locations=locations,
    )


@router.post("", response_model=ItemResponse, status_code=status.HTTP_201_CREATED)
def create_item(payload: ItemCreate, request: Request, db: Session = Depends(get_db)):
    pt = payload.process_type_code or None
    model_slots = payload.model_slots or []
    model_sym = slots_to_model_symbol(model_slots) if model_slots else ""

    # (item_name, process_type_code) 동일한 활성 품목이 이미 있으면 409.
    # SQLite 라 DB 레벨 UniqueConstraint 가 없는 상태에서 앱 레벨 안전망.
    existing_dup = (
        db.query(Item)
        .filter(
            Item.item_name == payload.item_name,
            Item.process_type_code == pt,
            Item.deleted_at.is_(None),
        )
        .first()
    )
    if existing_dup is not None:
        raise http_error(
            status.HTTP_409_CONFLICT,
            ErrorCode.CONFLICT,
            f"같은 카테고리에 이미 '{payload.item_name}' 품목이 있습니다.",
        )

    serial = None
    mes_code = None
    if pt and model_sym:
        serial = next_serial_no(model_sym, pt, db)
        mes_code = make_mes_code(model_sym, pt, serial)

    # 신규 항목은 목록 맨 끝으로. sort_order 미설정 시 NULL 이 되어 SQLite 가 맨앞에 정렬해버림.
    next_sort = (db.query(func.max(Item.sort_order)).scalar() or 0) + 1

    item = Item(
        item_name=payload.item_name,
        unit=payload.unit,
        legacy_part=payload.legacy_part,
        legacy_item_type=payload.legacy_item_type,
        supplier=payload.supplier,
        min_stock=payload.min_stock,
        process_type_code=pt,
        model_symbol=model_sym or None,
        serial_no=serial,
        mes_code=mes_code,
        sort_order=next_sort,
    )
    db.add(item)
    db.flush()

    init_qty = payload.initial_quantity if payload.initial_quantity is not None else 0
    inventory = Inventory(item_id=item.item_id, quantity=init_qty, warehouse_qty=init_qty)
    db.add(inventory)

    audit.record(
        db,
        request=request,
        action="item.create",
        target_type="item",
        target_id=str(item.item_id),
        payload_summary=f"{item.item_name} ({item.mes_code or 'no-code'}, init {init_qty})",
    )

    commit_and_refresh(db, item)
    _evt_emit(
        "item_create",
        request=request,
        item=item.mes_code or "-",
        name=item.item_name,
        init_qty=str(init_qty),
    )
    return item


@router.get("", response_model=List[ItemWithInventory])
def list_items(
    process_type_code: Optional[str] = Query(None, description="process_type_code 필터 (TR/HR/.../PF 18개)"),
    search: Optional[str] = Query(None, description="품목명, 품목코드, 위치 검색"),
    legacy_part: Optional[str] = Query(None, description="레거시 파트 필터"),
    department: Optional[str] = Query(None, description="부서 필터 (창고|조립|고압|진공|튜닝|튜브|출하|…)"),
    legacy_item_type: Optional[str] = Query(None, description="레거시 품목 유형 필터"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    query = _build_item_query(db)

    if process_type_code:
        query = query.filter(Item.process_type_code == process_type_code)

    if legacy_part:
        query = query.filter(Item.legacy_part == legacy_part)

    if department:
        if department == "창고":
            # 창고 재고가 1 이상인 품목
            query = query.filter(Inventory.warehouse_qty > 0)
        else:
            dept_item_ids = (
                db.query(InventoryLocation.item_id)
                .filter(
                    InventoryLocation.department == department,
                    InventoryLocation.quantity > 0,
                )
                .subquery()
            )
            query = query.filter(Item.item_id.in_(dept_item_ids))

    if legacy_item_type:
        query = query.filter(Item.legacy_item_type == legacy_item_type)

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Item.item_name.ilike(pattern),
                Item.mes_code.ilike(pattern),
                Inventory.location.ilike(pattern),
            )
        )

    rows = query.order_by(
        Item.deleted_at.is_(None).desc(),
        Item.sort_order,
        Item.mes_code,
    ).offset(skip).limit(limit).all()
    if not rows:
        return []

    # bulk prefetch — N+1 제거. 기존에는 item 1건당 4 쿼리씩 나갔음.
    item_ids = [it.item_id for it, _ in rows]
    figures_map = stock_math.bulk_compute(db, item_ids)

    from decimal import Decimal as _D

    loc_rows = (
        db.query(InventoryLocation)
        .filter(InventoryLocation.item_id.in_(item_ids), InventoryLocation.quantity > 0)
        .all()
    )
    locations_by_item: dict = {}
    for row in loc_rows:
        locations_by_item.setdefault(row.item_id, []).append(
            InventoryLocationResponse(
                department=row.department,
                status=row.status,
                quantity=row.quantity or _D("0"),
            )
        )

    return [
        _to_item_with_inventory(
            db,
            item,
            inv,
            figures=figures_map.get(item.item_id),
            locations=locations_by_item.get(item.item_id, []),
            model_slots=mes_code_to_model_slots(item.mes_code),
        )
        for item, inv in rows
    ]


@router.patch("/reorder", summary="품목 표시 순서 재배치")
def reorder_items(
    payload: ItemReorderPayload,
    _admin: Annotated[None, Depends(require_admin_pin)],
    db: Session = Depends(get_db),
):
    """드래그 reorder 결과를 영구 저장. 모델 reorder 와 동일 패턴.

    - PIN 검증 후 payload.items 의 (item_id, display_order) 쌍을 Item.sort_order 로 일괄 갱신.
    - 존재하지 않는 item_id 는 조용히 스킵 (부분 갱신 허용).
    """
    reorder_by_display_order(
        db, Item, "item_id",
        [(item.item_id, item.display_order) for item in payload.items],
        order_field="sort_order",
    )
    db.commit()
    return {"ok": True}


@router.get("/export.csv")
def export_items_csv(db: Session = Depends(get_db)):
    rows = _build_item_query(db).order_by(Item.mes_code).all()

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["mes_code", "item_name", "process_type_code", "unit", "quantity", "location", "updated_at"])

    for item, inventory in rows:
        writer.writerow(
            [
                item.mes_code or "",
                item.item_name,
                item.process_type_code or "",
                item.unit,
                float(inventory.quantity) if inventory else 0,
                inventory.location if inventory else "",
                item.updated_at.isoformat(),
            ]
        )

    return csv_streaming_response(buffer, "items-export.csv")


# process_type_code 의 prefix(부서 계열) 1글자 → 행 색상 (xlsx export 용).
# README 기준 6개 부서: T(튜브)/H(고압)/V(진공)/N(튜닝)/A(조립)/P(출하).
_PROCESS_PREFIX_ROW_COLOR = {
    "T": "D6F5F5",  # 튜브
    "H": "FFF8D6",  # 고압
    "V": "EAD6FF",  # 진공
    "N": "FFE8D6",  # 튜닝
    "A": "D6E8FF",  # 조립
    "P": "D6F5E0",  # 출하
}


def _row_color_for(process_type_code: Optional[str]) -> str:
    if not process_type_code:
        return "EBEBEB"
    return _PROCESS_PREFIX_ROW_COLOR.get(process_type_code[0], "FFFFFF")


@router.get("/export.xlsx")
def export_items_xlsx(
    process_type_code: Optional[str] = Query(None, description="process_type_code 필터"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    from datetime import date as _date
    from decimal import Decimal as _D
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from app.utils.excel import apply_header, auto_width, make_xlsx_response

    query = _build_item_query(db)
    if process_type_code:
        query = query.filter(Item.process_type_code == process_type_code)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Item.item_name.ilike(pattern),
                Item.mes_code.ilike(pattern),
                Inventory.location.ilike(pattern),
            )
        )
    rows = query.order_by(Item.process_type_code, Item.mes_code).all()

    # 가용수량 계산: stock_math 를 한 번 bulk 로 불러 경로별 재구현을 피한다.
    figures_map = stock_math.bulk_compute(db, [it.item_id for it, _ in rows])

    wb = Workbook()
    ws = wb.active
    ws.title = "품목 마스터"

    columns = [
        "품목 코드", "품목명", "공정코드", "단위",
        "재고수량", "가용수량", "예약수량", "위치", "공급업체", "안전재고", "수정일",
    ]
    apply_header(ws, columns)

    red_font = Font(color="CC0000", bold=True)

    for item, inventory in rows:
        fig = figures_map.get(item.item_id) or stock_math.StockFigures()
        qty = float(fig.total)
        pending = float(fig.pending)
        available = float(fig.available)
        min_stock = float(item.min_stock) if item.min_stock else None

        row_data = [
            item.mes_code or "",
            item.item_name,
            item.process_type_code or "",
            item.unit,
            qty,
            available,
            pending,
            inventory.location if inventory else "",
            item.supplier or "",
            min_stock or "",
            item.updated_at.strftime("%Y-%m-%d %H:%M") if item.updated_at else "",
        ]
        ws.append(row_data)

        row_idx = ws.max_row
        hex_color = _row_color_for(item.process_type_code)
        row_fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws[row_idx]:
            cell.fill = row_fill

        # 재고수량 < 안전재고이면 빨간 글씨
        qty_cell = ws.cell(row=row_idx, column=6)
        if min_stock and qty < min_stock:
            qty_cell.font = red_font

    auto_width(ws)
    filename = f"items-{_date.today().strftime('%Y%m%d')}.xlsx"
    return make_xlsx_response(wb, filename)


@router.get("/{item_id}", response_model=ItemWithInventory)
def get_item(item_id: uuid.UUID, db: Session = Depends(get_db)):
    row = _build_item_query(db).filter(Item.item_id == item_id).first()
    if not row:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")

    item, inventory = row
    # 단건이지만 list 경로와 동일하게 bulk_compute 사용 — compute_for 단건 호출 경로 정리.
    figures_map = stock_math.bulk_compute(db, [item.item_id])
    figures = figures_map.get(item.item_id)
    return _to_item_with_inventory(db, item, inventory, figures=figures)


@router.put("/{item_id}", response_model=ItemWithInventory)
def update_item(item_id: uuid.UUID, payload: ItemUpdate, request: Request, db: Session = Depends(get_db)):
    item = db.query(Item).filter(Item.item_id == item_id).first()
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")

    changed: list[str] = []
    # process_type/model/option 변경 감지 — mes_code 자동 재계산용.
    # 일반 필드(item_name·unit 등)는 코드와 무관하므로 그대로 setattr.
    for field in (
        "item_name", "unit",
        "legacy_part", "legacy_item_type",
        "supplier", "min_stock",
    ):
        new_val = getattr(payload, field)
        if new_val is not None and getattr(item, field) != new_val:
            setattr(item, field, new_val)
            changed.append(field)

    # 모델·카테고리 변경은 mes_code 재계산 트리거.
    # 의도: 사용자는 모델·카테고리만 바꾸고 mes_code 는 자동 부여.
    #   - 모델만 변경 → serial 유지, prefix 만 새로.
    #   - 카테고리 변경 → 새 카테고리의 next_serial_no 부여 (번호 풀이 다르므로).
    new_pt = payload.process_type_code if payload.process_type_code is not None else item.process_type_code
    if payload.model_slots is not None:
        new_model_sym = slots_to_model_symbol(payload.model_slots) or None
    else:
        new_model_sym = item.model_symbol

    pt_changed = payload.process_type_code is not None and payload.process_type_code != item.process_type_code
    model_changed = payload.model_slots is not None and new_model_sym != item.model_symbol

    if pt_changed or model_changed:
        # serial 결정: 카테고리 변경 시 새 카테고리에서 next_serial_no, 그 외엔 기존 유지.
        if pt_changed:
            if new_model_sym and new_pt:
                item.serial_no = next_serial_no(new_model_sym, new_pt, db)
            item.process_type_code = new_pt
            changed.append("process_type_code")
        if model_changed:
            item.model_symbol = new_model_sym
            changed.append("model_slots")

        # mes_code 재조립. model/category 가 비어있으면 None 으로 둠 (등록 직후 미완 상태와 동일).
        if new_model_sym and new_pt and item.serial_no:
            item.mes_code = make_mes_code(new_model_sym, new_pt, item.serial_no)
            # 안전망 — 중복 검사 (이론상 발생 안 해야 함).
            dup = db.query(Item).filter(
                Item.mes_code == item.mes_code,
                Item.item_id != item.item_id,
            ).first()
            if dup is not None:
                raise http_error(
                    409,
                    ErrorCode.CONFLICT,
                    f"'{item.mes_code}' 코드가 이미 사용 중입니다. 데이터 점검이 필요합니다.",
                )
            changed.append("mes_code")
        else:
            item.mes_code = None

    item.updated_at = datetime.now(UTC).replace(tzinfo=None)

    if changed:
        audit.record(
            db,
            request=request,
            action="item.update",
            target_type="item",
            target_id=str(item.item_id),
            payload_summary=f"{item.item_name}: {', '.join(changed)}",
        )

    commit_and_refresh(db, item)
    if changed:
        _evt_emit(
            "item_update",
            request=request,
            item=item.mes_code or "-",
            changed=",".join(changed),
        )
    # 응답에 inventory 동봉 — 좌측 list API 와 동일한 ItemWithInventory 형태로 보내
    # 저장 직후 우측 카드의 재고/창고 표시가 빈칸이 되는 잔여 UI 버그 방지.
    inventory = db.query(Inventory).filter(Inventory.item_id == item.item_id).first()
    return _to_item_with_inventory(db, item, inventory)


@router.patch("/{item_id}/bom-completion", response_model=ItemResponse)
def update_bom_completion(
    item_id: uuid.UUID,
    payload: BomCompletionUpdate,
    request: Request,
    db: Session = Depends(get_db),
):
    """BOM 완료 상태 토글 — 사용자가 명시적으로 누를 때만 set/clear."""
    item = db.query(Item).filter(Item.item_id == item_id).first()
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")

    item.bom_completed_at = datetime.now(UTC).replace(tzinfo=None) if payload.completed else None
    item.updated_at = datetime.now(UTC).replace(tzinfo=None)

    audit.record(
        db,
        request=request,
        action="item.bom_completion",
        target_type="item",
        target_id=str(item.item_id),
        payload_summary=f"{item.item_name}: {'완료' if payload.completed else '완료 해제'}",
    )

    commit_and_refresh(db, item)
    return item


@router.patch("/{item_id}/soft-delete", response_model=ItemResponse)
def soft_delete_item(item_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """품목 소프트 삭제 — deleted_at 세팅 + BOM 연결 제거. 입출고 내역은 보존."""
    item = db.query(Item).filter(Item.item_id == item_id).first()
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")
    if item.deleted_at is not None:
        raise http_error(409, ErrorCode.CONFLICT, "이미 삭제된 품목입니다.")

    db.query(BOM).filter(
        (BOM.parent_item_id == item_id) | (BOM.child_item_id == item_id)
    ).delete(synchronize_session=False)

    item.deleted_at = datetime.now(UTC).replace(tzinfo=None)
    item.updated_at = datetime.now(UTC).replace(tzinfo=None)

    audit.record(
        db,
        request=request,
        action="item.delete",
        target_type="item",
        target_id=str(item.item_id),
        payload_summary=f"{item.item_name} ({item.mes_code or 'no-code'})",
    )

    commit_and_refresh(db, item)
    _evt_emit(
        "item_delete",
        request=request,
        item=item.mes_code or "-",
        name=item.item_name,
    )
    return item


@router.patch("/{item_id}/restore", response_model=ItemResponse)
def restore_item(item_id: uuid.UUID, request: Request, db: Session = Depends(get_db)):
    """삭제된 품목 복구 — deleted_at 초기화."""
    item = db.query(Item).filter(Item.item_id == item_id).first()
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")
    if item.deleted_at is None:
        raise http_error(409, ErrorCode.CONFLICT, "삭제되지 않은 품목입니다.")

    item.deleted_at = None
    item.updated_at = datetime.now(UTC).replace(tzinfo=None)

    audit.record(
        db,
        request=request,
        action="item.restore",
        target_type="item",
        target_id=str(item.item_id),
        payload_summary=f"{item.item_name} ({item.mes_code or 'no-code'})",
    )

    commit_and_refresh(db, item)
    _evt_emit(
        "item_restore",
        request=request,
        item=item.mes_code or "-",
        name=item.item_name,
    )
    return item
