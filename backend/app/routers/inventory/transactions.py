"""거래 이력: /transactions, /transactions/export.csv|.xlsx, 메타/수량 수정 + 수정 이력."""

from __future__ import annotations

import base64
import binascii
import csv
import json
import uuid
from datetime import UTC, date, datetime, time
from decimal import Decimal
from io import StringIO
from typing import List, Optional

from fastapi import APIRouter, Depends, Query, Request, status
from pydantic import BaseModel, Field
from sqlalchemy import case, extract, func, or_, select
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import (
    Employee,
    IoBatch,
    Inventory,
    Item,
    TransactionEditLog,
    TransactionLog,
    TransactionTypeEnum,
)
from app.routers._errors import ErrorCode, http_error
from app.schemas import (
    TransactionDisplayGroupPageResponse,
    TransactionDisplayGroupResponse,
    TransactionEditLogResponse,
    TransactionLogResponse,
    TransactionMetaEditRequest,
    TransactionQuantityCorrectionRequest,
    TransactionQuantityCorrectionResponse,
)
from app.services import transaction_actions as transaction_actions_svc
from app.services.export_helpers import csv_streaming_response
from app.services.pin_auth import verify_pin
from app.utils.search import build_normalized_search_filter
from app._actor import set_actor
from app.routers.inventory._tx_filters import (
    _SUMMARY_WAREHOUSE_TYPES,
    _SUMMARY_DEPT_TYPES,
    _SUMMARY_ADJUST_TYPES,
    _department_label_expr,
    _process_step_filter,
    _model_filter,
    _department_filter,
    _operation_filter,
    _apply_common_filters,
    _history_visibility_filter,
    _history_request_date_expr,
    _batch_name_map,
    _stock_request_info_map,
    _to_log_response,
)
from app.repositories import item_repository, inventory_repository


router = APIRouter()


# 단일 export 요청에서 허용하는 최대 행 수. 운영 PC 메모리 보호용 안전 상한.
EXPORT_MAX_ROWS = 50_000
DISPLAY_GROUP_PAGE_SIZE = 100
OPERATION_KEYS_DESCRIPTION = (
    "화면 작업 종류 키. 예: warehouse,process,defect,item_conversion,shipping"
)


# 수량 보정이 허용되는 거래 타입.
# ADJUST 제외: 절대값 지정 방식이라 delta 보정 정책 애매 (별도 정책 확정 필요).
# TRANSFER_*: 부서 버킷 정보가 TransactionLog에 없어 1차 미지원.
QUANTITY_CORRECTABLE = {
    TransactionTypeEnum.RECEIVE,
    TransactionTypeEnum.SHIP,
}

class TransactionSummaryResponse(BaseModel):
    """입출고 내역 화면 KPI — 조건 전체 카운트(페이지네이션과 무관)."""
    total: int
    warehouse_count: int
    dept_count: int
    adjust_count: int
    # 전 거래의 실제 부서 또는 창고별 카운트 {라벨: 건수}.
    department_counts: dict[str, int] = {}


class ReferenceSummaryResponse(BaseModel):
    """참조번호 기반 목록 묶음의 전체 요약."""

    reference_no: str
    shipping_phase: Optional[str] = None
    log_count: int
    item_count: int
    total_quantity: float
    unit: Optional[str] = None


def _display_group_cursor(group: TransactionDisplayGroupResponse) -> str:
    """대표 행의 정렬 기준과 안정 키를 포함한 불투명 커서."""
    sort_at, created_at, log_id = _display_group_sort_tuple(group)
    payload = {
        "type": group.type,
        "key": group.key,
        "sort_at": sort_at,
        "created_at": created_at,
        "log_id": log_id,
    }
    encoded = base64.urlsafe_b64encode(json.dumps(payload, separators=(",", ":")).encode()).decode()
    return encoded.rstrip("=")


def _decode_display_group_cursor(cursor: str) -> dict[str, str | None]:
    try:
        padded = cursor + "=" * (-len(cursor) % 4)
        value = json.loads(base64.urlsafe_b64decode(padded.encode()).decode())
        if not isinstance(value, dict) or not all(
            isinstance(value.get(field), str) for field in ("type", "key", "sort_at", "created_at", "log_id")
        ):
            raise ValueError
        return value
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError, binascii.Error):
        raise http_error(400, ErrorCode.BAD_REQUEST, "유효하지 않은 대표 행 페이지 커서입니다.")


def _display_group_sort_tuple(
    group: TransactionDisplayGroupResponse,
) -> tuple[str, str, str]:
    # 불량 lifecycle은 화면 의미상 부모→후속 로그 순서를 보존한다. 다만 대표 행의
    # 정렬·커서는 실제 목록에서 먼저 나타나는(가장 최신인) 후속 로그를 기준으로 삼아야
    # 그 사이에 생긴 다른 대표 행이 다음 페이지에서 누락되지 않는다.
    anchor = max(
        group.logs,
        key=lambda log: (
            (log.requested_at or log.created_at).isoformat(),
            log.created_at.isoformat(),
            str(log.log_id),
        ),
    ) if group.type == "defect_lifecycle" else group.logs[0]
    return (
        (anchor.requested_at or anchor.created_at).isoformat(),
        anchor.created_at.isoformat(),
        str(anchor.log_id),
    )


def _is_after_display_group_cursor(
    group: TransactionDisplayGroupResponse,
    cursor: dict[str, str | None],
) -> bool:
    """내림차순 대표 행 정렬에서 커서 뒤에 남아야 하는 그룹인지 판단한다."""
    return _display_group_sort_tuple(group) < (
        cursor["sort_at"] or "",
        cursor["created_at"] or "",
        cursor["log_id"] or "",
    )


def _reference_group_key(log: TransactionLogResponse) -> str:
    return f"{log.reference_no or ''}::{log.shipping_phase or ''}"


def _defect_actor(log: TransactionLogResponse) -> Optional[str]:
    return (log.requester_name or log.produced_by or "").strip() or None


def _defect_reason_key(log: TransactionLogResponse) -> Optional[str]:
    category = (log.reason_category or "").strip()
    memo = (log.reason_memo or "").strip()
    return f"{category}::{memo}" if category or memo else None


def _is_matching_defect_lifecycle(parent: TransactionLogResponse, child: TransactionLogResponse) -> bool:
    if child.transaction_type not in {
        TransactionTypeEnum.DEFECT_SCRAP,
        TransactionTypeEnum.SUPPLIER_RETURN,
        TransactionTypeEnum.DISASSEMBLE,
    }:
        return False
    if parent.item_id != child.item_id or abs(parent.quantity_change) != abs(child.quantity_change):
        return False
    parent_actor, child_actor = _defect_actor(parent), _defect_actor(child)
    if not parent_actor or parent_actor != child_actor:
        return False
    parent_department = (parent.department or "").strip()
    child_department = (child.department or "").strip()
    if not parent_department or parent_department != child_department:
        return False
    parent_reason, child_reason = _defect_reason_key(parent), _defect_reason_key(child)
    if not parent_reason or parent_reason != child_reason:
        return False
    elapsed = (child.created_at - parent.created_at).total_seconds()
    return 0 <= elapsed <= 60


def _find_defect_lifecycle_pairs(
    logs: list[TransactionLogResponse],
) -> list[tuple[TransactionLogResponse, TransactionLogResponse]]:
    chronological = sorted(logs, key=lambda log: log.created_at)
    used: set[uuid.UUID] = set()
    pairs: list[tuple[TransactionLogResponse, TransactionLogResponse]] = []
    for index, parent in enumerate(chronological):
        if parent.transaction_type != TransactionTypeEnum.MARK_DEFECTIVE or parent.log_id in used:
            continue
        child = next(
            (
                candidate
                for candidate in chronological[index + 1 :]
                if candidate.log_id not in used and _is_matching_defect_lifecycle(parent, candidate)
            ),
            None,
        )
        if child is not None:
            used.update({parent.log_id, child.log_id})
            pairs.append((parent, child))
    return pairs


def _build_display_groups(logs: list[TransactionLogResponse]) -> list[TransactionDisplayGroupResponse]:
    """프런트엔드 buildGroups 와 같은 우선순위로 화면 대표 행을 만든다."""
    op_batches: dict[uuid.UUID, list[TransactionLogResponse]] = {}
    reference_batches: dict[str, list[TransactionLogResponse]] = {}
    pairs = _find_defect_lifecycle_pairs(logs)
    pair_by_log_id: dict[uuid.UUID, tuple[TransactionLogResponse, TransactionLogResponse, uuid.UUID]] = {}
    log_positions = {log.log_id: index for index, log in enumerate(logs)}
    for parent, child in pairs:
        anchor_id = parent.log_id if log_positions[parent.log_id] <= log_positions[child.log_id] else child.log_id
        pair_by_log_id[parent.log_id] = (parent, child, anchor_id)
        pair_by_log_id[child.log_id] = (parent, child, anchor_id)
    for log in logs:
        if log.operation_batch_id:
            op_batches.setdefault(log.operation_batch_id, []).append(log)
        elif log.reference_no:
            reference_batches.setdefault(_reference_group_key(log), []).append(log)

    groups: list[TransactionDisplayGroupResponse] = []
    seen_operation_batches: set[uuid.UUID] = set()
    seen_reference_batches: set[str] = set()
    for log in logs:
        pair = pair_by_log_id.get(log.log_id)
        if pair:
            parent, child, anchor_id = pair
            if anchor_id == log.log_id:
                groups.append(TransactionDisplayGroupResponse(
                    type="defect_lifecycle",
                    key=f"defect-lifecycle:{parent.log_id}:{child.log_id}",
                    logs=[parent, child],
                ))
            continue
        if log.operation_batch_id:
            batch_id = log.operation_batch_id
            if batch_id in seen_operation_batches:
                continue
            seen_operation_batches.add(batch_id)
            batch_logs = op_batches[batch_id]
            groups.append(TransactionDisplayGroupResponse(
                type="solo" if len(batch_logs) == 1 else "op_batch",
                key=str(batch_id) if len(batch_logs) > 1 else f"solo:{batch_logs[0].log_id}",
                logs=batch_logs,
            ))
        elif log.reference_no:
            reference_key = _reference_group_key(log)
            if reference_key in seen_reference_batches:
                continue
            seen_reference_batches.add(reference_key)
            reference_logs = reference_batches[reference_key]
            groups.append(TransactionDisplayGroupResponse(
                type="solo" if len(reference_logs) == 1 else "batch",
                key=reference_key if len(reference_logs) > 1 else f"solo:{reference_logs[0].log_id}",
                logs=reference_logs,
            ))
        else:
            groups.append(TransactionDisplayGroupResponse(type="solo", key=f"solo:{log.log_id}", logs=[log]))
    return groups


def _require_export_range(start_date: Optional[date], end_date: Optional[date]) -> tuple[datetime, datetime]:
    """export 는 start_date / end_date 모두 필수. 없으면 400."""
    if start_date is None or end_date is None:
        raise http_error(
            status_code=400,
            code=ErrorCode.EXPORT_RANGE_REQUIRED,
            message="export 는 start_date 와 end_date 를 모두 지정해야 합니다.",
        )
    return datetime.combine(start_date, time.min), datetime.combine(end_date, time.max)


def _enforce_export_limit(count: int) -> None:
    if count > EXPORT_MAX_ROWS:
        raise http_error(
            status_code=422,
            code=ErrorCode.EXPORT_RANGE_TOO_LARGE,
            message=f"범위 내 행 수가 {count:,}건으로 상한({EXPORT_MAX_ROWS:,})을 초과합니다. 기간을 좁혀 주세요.",
            row_count=count,
            max_rows=EXPORT_MAX_ROWS,
        )


_TX_ROW_COLOR = {
    "RECEIVE":   "D4EDDA",
    "PRODUCE":   "CCE5FF",
    "SHIP":      "F8D7DA",
    "ADJUST":    "FFF3CD",
    "BACKFLUSH": "FFE5B4",
    "TRANSFER_TO_PROD": "E0E7FF",
    "TRANSFER_TO_WH":   "E0E7FF",
    "TRANSFER_DEPT":    "E0E7FF",
    "MARK_DEFECTIVE":   "FBD9D7",
    "SUPPLIER_RETURN":  "F4C7C3",
    "INTERNAL_USE":     "FCE8B2",
}


def _log_snapshot(log: TransactionLog) -> dict:
    """TransactionLog의 가변 필드 스냅샷 (JSON 직렬화 가능 형태)."""
    return {
        "transaction_type": log.transaction_type.value if log.transaction_type else None,
        "quantity_change": str(log.quantity_change) if log.quantity_change is not None else None,
        "reference_no": log.reference_no,
        "produced_by": log.produced_by,
        "notes": log.notes,
    }


def _verify_editor(
    db: Session,
    employee_id: uuid.UUID,
    pin: str,
    request: Optional[Request] = None,
) -> Employee:
    """수정자 직원 + PIN 검증. 작업자 식별용 — 실제 보안 인증이 아님."""
    employee = db.query(Employee).filter(Employee.employee_id == employee_id).first()
    if not employee:
        raise http_error(404, ErrorCode.NOT_FOUND, "수정자 직원을 찾을 수 없습니다.")
    if not bool(employee.is_active):
        raise http_error(403, ErrorCode.FORBIDDEN, "비활성 직원은 거래를 수정할 수 없습니다.")
    if not verify_pin(employee.pin_hash, pin):
        raise http_error(403, ErrorCode.FORBIDDEN, "PIN이 올바르지 않습니다.")
    set_actor(request, employee)
    return employee


@router.get("/transactions/monthly-counts", summary="연도별 월별 거래 카운트")
def monthly_counts(
    year: int = Query(..., ge=2020, le=2100),
    db: Session = Depends(get_db),
) -> dict[str, int]:
    """주어진 year의 월별 거래 건수 반환. 빈 month는 0.

    응답 예: { "2026-01": 142, "2026-02": 89, ..., "2026-12": 0 }
    archived_at 이 있는 레코드는 제외한다.
    """
    request_date_expr = _history_request_date_expr()
    rows = (
        db.query(
            extract("month", request_date_expr).label("month"),
            func.count(TransactionLog.log_id).label("count"),
        )
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
        .filter(
            extract("year", request_date_expr) == year,
            TransactionLog.archived_at.is_(None),
            _history_visibility_filter(),
        )
        .group_by(extract("month", request_date_expr))
        .all()
    )
    counts = {f"{year:04d}-{int(r.month):02d}": int(r.count) for r in rows}
    return {f"{year:04d}-{m:02d}": counts.get(f"{year:04d}-{m:02d}", 0) for m in range(1, 13)}


@router.get("/transactions", response_model=List[TransactionLogResponse])
def list_transactions(
    item_id: Optional[uuid.UUID] = Query(None),
    operation_batch_id: Optional[uuid.UUID] = Query(None),
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    transaction_types: Optional[str] = Query(None, description="쉼표 구분 복수 타입. 예: RECEIVE,SHIP"),
    operation_keys: Optional[str] = Query(None, description=OPERATION_KEYS_DESCRIPTION),
    reference_no: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(
        None, description="부서 라벨 필터 (쉼표 복수). 예: 창고,조립,고압"
    ),
    model: Optional[str] = Query(None, description="제품 모델명 필터 (쉼표 복수)"),
    process_step: Optional[str] = Query(
        None, description="공정 구분 필터 R/A/F (쉼표 복수)"
    ),
    date_from: Optional[date] = Query(None, description="포함 시작일 YYYY-MM-DD"),
    date_to: Optional[date] = Query(None, description="포함 종료일 YYYY-MM-DD"),
    include_archived: bool = Query(False, description="archived_at 이 있는 레코드 포함 여부"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=2000),
    db: Session = Depends(get_db),
):
    # edit_count 서브쿼리: 각 TransactionLog의 수정 이력 건수
    edit_count_sq = (
        select(func.count(TransactionEditLog.edit_id))
        .where(TransactionEditLog.original_log_id == TransactionLog.log_id)
        .correlate(TransactionLog)
        .scalar_subquery()
    )

    # Keep logs without an IoBatch while allowing requester-name search.
    query = (
        db.query(TransactionLog, Item, edit_count_sq.label("edit_count"))
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )

    if item_id:
        query = query.filter(TransactionLog.item_id == item_id)

    if operation_batch_id:
        query = query.filter(TransactionLog.operation_batch_id == operation_batch_id)

    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)

    if reference_no:
        query = query.filter(TransactionLog.reference_no == reference_no)

    query = _apply_common_filters(
        query,
        db,
        transaction_types=transaction_types,
        operation_keys=operation_keys,
        search=search,
        department=department,
        model=model,
        process_step=process_step,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived,
    )

    requested_at_order = _history_request_date_expr()
    rows = (
        query.order_by(
            requested_at_order.desc(),
            TransactionLog.created_at.desc(),
            TransactionLog.log_id.desc(),
        )
        .offset(skip)
        .limit(limit)
        .all()
    )

    # Fill requester/approver names from IoBatch or StockRequest reference_no.
    batch_ids = {log.operation_batch_id for log, _, _ in rows if log.operation_batch_id}
    batch_map = _batch_name_map(db, batch_ids)
    reference_nos = {log.reference_no for log, _, _ in rows if log.reference_no}
    sr_map = _stock_request_info_map(db, reference_nos)

    result = []
    for log, item, edit_count in rows:
        info = batch_map.get(log.operation_batch_id)
        if info is None and log.reference_no:
            info = sr_map.get(log.reference_no)
        result.append(
            _to_log_response(
                log,
                item,
                int(edit_count or 0),
                requester_name=info.requester_name if info else None,
                approver_name=info.approver_name if info else None,
                requested_at=info.requested_at if info else None,
                approved_at=info.approved_at if info else None,
            )
        )
    return result


@router.get(
    "/transactions/display-groups",
    response_model=TransactionDisplayGroupPageResponse,
)
def list_transaction_display_groups(
    item_id: Optional[uuid.UUID] = Query(None),
    operation_batch_id: Optional[uuid.UUID] = Query(None),
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    transaction_types: Optional[str] = Query(None),
    operation_keys: Optional[str] = Query(None, description=OPERATION_KEYS_DESCRIPTION),
    reference_no: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    model: Optional[str] = Query(None),
    process_step: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    include_archived: bool = Query(False),
    cursor: Optional[str] = Query(None),
    limit: int = Query(DISPLAY_GROUP_PAGE_SIZE, ge=1, le=DISPLAY_GROUP_PAGE_SIZE),
    db: Session = Depends(get_db),
) -> TransactionDisplayGroupPageResponse:
    """필터된 로그를 화면 대표 행으로 묶어, 완결된 그룹 단위로 페이지를 반환한다."""
    edit_count_sq = (
        select(func.count(TransactionEditLog.edit_id))
        .where(TransactionEditLog.original_log_id == TransactionLog.log_id)
        .correlate(TransactionLog)
        .scalar_subquery()
    )
    query = (
        db.query(TransactionLog, Item, edit_count_sq.label("edit_count"))
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )
    if item_id:
        query = query.filter(TransactionLog.item_id == item_id)
    if operation_batch_id:
        query = query.filter(TransactionLog.operation_batch_id == operation_batch_id)
    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    if reference_no:
        query = query.filter(TransactionLog.reference_no == reference_no)
    query = _apply_common_filters(
        query,
        db,
        transaction_types=transaction_types,
        operation_keys=operation_keys,
        search=search,
        department=department,
        model=model,
        process_step=process_step,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived,
    )
    requested_at_order = _history_request_date_expr()
    rows = query.order_by(
        requested_at_order.desc(),
        TransactionLog.created_at.desc(),
        TransactionLog.log_id.desc(),
    ).all()
    batch_ids = {log.operation_batch_id for log, _, _ in rows if log.operation_batch_id}
    batch_map = _batch_name_map(db, batch_ids)
    reference_nos = {log.reference_no for log, _, _ in rows if log.reference_no}
    stock_request_map = _stock_request_info_map(db, reference_nos)
    logs = []
    for log, item, edit_count in rows:
        info = batch_map.get(log.operation_batch_id)
        if info is None and log.reference_no:
            info = stock_request_map.get(log.reference_no)
        logs.append(
            _to_log_response(
                log,
                item,
                int(edit_count or 0),
                requester_name=info.requester_name if info else None,
                approver_name=info.approver_name if info else None,
                requested_at=info.requested_at if info else None,
                approved_at=info.approved_at if info else None,
            )
        )
    groups = _build_display_groups(logs)
    if cursor:
        cursor_value = _decode_display_group_cursor(cursor)
        groups = [group for group in groups if _is_after_display_group_cursor(group, cursor_value)]
    page_groups = groups[:limit]
    has_more = len(page_groups) < len(groups)
    return TransactionDisplayGroupPageResponse(
        groups=page_groups,
        next_cursor=_display_group_cursor(page_groups[-1]) if has_more and page_groups else None,
        has_more=has_more,
    )


@router.get(
    "/transactions/reference-summaries",
    response_model=List[ReferenceSummaryResponse],
)
def list_reference_summaries(
    transaction_types: Optional[str] = Query(None, description="쉼표 구분 복수 타입"),
    operation_keys: Optional[str] = Query(None, description=OPERATION_KEYS_DESCRIPTION),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None, description="부서 라벨 필터"),
    model: Optional[str] = Query(None, description="제품 모델명 필터"),
    process_step: Optional[str] = Query(None, description="공정 구분 필터"),
    date_from: Optional[date] = Query(None, description="포함 시작일 YYYY-MM-DD"),
    date_to: Optional[date] = Query(None, description="포함 종료일 YYYY-MM-DD"),
    include_archived: bool = Query(False),
    db: Session = Depends(get_db),
):
    """페이지네이션과 무관하게 참조번호 묶음의 전체 수량을 반환한다."""
    query = (
        db.query(TransactionLog)
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )
    query = _apply_common_filters(
        query,
        db,
        transaction_types=transaction_types,
        operation_keys=operation_keys,
        search=search,
        department=department,
        model=model,
        process_step=process_step,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived,
    ).filter(
        TransactionLog.operation_batch_id.is_(None),
        TransactionLog.reference_no.isnot(None),
        TransactionLog.reference_no != "",
    )

    quantity = func.abs(func.coalesce(TransactionLog.transfer_qty, TransactionLog.quantity_change))
    unit_count = func.count(func.distinct(Item.unit))
    rows = (
        query.with_entities(
            TransactionLog.reference_no.label("reference_no"),
            TransactionLog.shipping_phase.label("shipping_phase"),
            func.count(TransactionLog.log_id).label("log_count"),
            func.count(func.distinct(TransactionLog.item_id)).label("item_count"),
            func.coalesce(func.sum(quantity), 0).label("total_quantity"),
            case((unit_count == 1, func.min(Item.unit)), else_=None).label("unit"),
        )
        .group_by(TransactionLog.reference_no, TransactionLog.shipping_phase)
        .order_by(TransactionLog.reference_no, TransactionLog.shipping_phase)
        .all()
    )
    return [
        ReferenceSummaryResponse(
            reference_no=row.reference_no,
            shipping_phase=row.shipping_phase,
            log_count=int(row.log_count or 0),
            item_count=int(row.item_count or 0),
            total_quantity=float(row.total_quantity or 0),
            unit=row.unit,
        )
        for row in rows
    ]


@router.get("/transactions/summary", response_model=TransactionSummaryResponse)
def get_transactions_summary(
    transaction_types: Optional[str] = Query(None, description="쉼표 구분 복수 타입"),
    operation_keys: Optional[str] = Query(None, description=OPERATION_KEYS_DESCRIPTION),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(
        None, description="부서 라벨 필터 (쉼표 복수). 예: 창고,조립,고압"
    ),
    model: Optional[str] = Query(None, description="제품 모델명 필터 (쉼표 복수)"),
    process_step: Optional[str] = Query(
        None, description="공정 구분 필터 R/A/F (쉼표 복수)"
    ),
    date_from: Optional[date] = Query(None, description="포함 시작일 YYYY-MM-DD"),
    date_to: Optional[date] = Query(None, description="포함 종료일 YYYY-MM-DD"),
    include_archived: bool = Query(False),
    db: Session = Depends(get_db),
):
    """KPI 카드용 카운트 집계. list_transactions 와 동일한 필터를 받지만 row 가 아니라
    숫자만 반환 — 화면에 로드된 100건이 아니라 조건 전체를 보여주기 위함.
    department_counts 는 전 거래를 실제 부서명 또는 '창고'로 묶은 카운트.
    """
    # list_transactions 와 동일한 join 패턴 — search 가 IoBatch.requester_name 까지 닿게.
    query = (
        db.query(TransactionLog)
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )

    # list_transactions 와 동일한 공통 필터 빌더로 위임.
    query = _apply_common_filters(
        query,
        db,
        transaction_types=transaction_types,
        operation_keys=operation_keys,
        search=search,
        department=department,
        model=model,
        process_step=process_step,
        date_from=date_from,
        date_to=date_to,
        include_archived=include_archived,
    )

    # 한 번의 집계 쿼리로 4개 카운트.
    agg = query.with_entities(
        func.count(TransactionLog.log_id).label("total"),
        func.coalesce(
            func.sum(case((TransactionLog.transaction_type.in_(_SUMMARY_WAREHOUSE_TYPES), 1), else_=0)),
            0,
        ).label("warehouse_count"),
        func.coalesce(
            func.sum(case((TransactionLog.transaction_type.in_(_SUMMARY_DEPT_TYPES), 1), else_=0)),
            0,
        ).label("dept_count"),
        func.coalesce(
            func.sum(case((TransactionLog.transaction_type.in_(_SUMMARY_ADJUST_TYPES), 1), else_=0)),
            0,
        ).label("adjust_count"),
    ).one()

    # 전 거래 부서별 카운트. 제한 필터 없음 — '창고'도 집계됨.
    # 마이그레이션 전/비정상 미분류 NULL은 UI 분류값으로 만들지 않고 건너뛴다.
    dexpr = _department_label_expr()
    dept_rows = (
        query.with_entities(dexpr.label("dept"), func.count(TransactionLog.log_id).label("c"))
        .group_by(dexpr)
        .all()
    )
    department_counts = {r.dept: int(r.c) for r in dept_rows if r.dept is not None}

    return TransactionSummaryResponse(
        total=int(agg.total or 0),
        warehouse_count=int(agg.warehouse_count or 0),
        dept_count=int(agg.dept_count or 0),
        adjust_count=int(agg.adjust_count or 0),
        department_counts=department_counts,
    )


@router.get("/transactions/export.csv")
def export_transactions_csv(
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    search: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None, description="필수. 포함 시작일 YYYY-MM-DD"),
    end_date: Optional[date] = Query(None, description="필수. 포함 종료일 YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    start_dt, end_dt = _require_export_range(start_date, end_date)

    # 목록 조회와 동일하게 IoBatch outerjoin — search 가 requester_name 까지 닿고
    # 요청자/승인자 컬럼을 채운다(operation_batch_id NULL row 보존 위해 outerjoin).
    query = (
        db.query(TransactionLog, Item)
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )
    query = query.filter(
        TransactionLog.created_at >= start_dt,
        TransactionLog.created_at <= end_dt,
    )

    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    search_filter = build_normalized_search_filter(
        search,
        Item.item_name,
        Item.mes_code,
        TransactionLog.reference_no,
        TransactionLog.notes,
        TransactionLog.produced_by,
        IoBatch.requester_name,
    )
    if search_filter is not None:
        query = query.filter(search_filter)

    _enforce_export_limit(query.count())
    rows = query.order_by(TransactionLog.created_at.desc()).all()
    batch_map = _batch_name_map(
        db, {log.operation_batch_id for log, _ in rows if log.operation_batch_id}
    )
    sr_map = _stock_request_info_map(
        db, {log.reference_no for log, _ in rows if log.reference_no}
    )

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "created_at",
            "transaction_type",
            "mes_code",
            "item_name",
            "process_type_code",
            "quantity_change",
            "quantity_before",
            "quantity_after",
            "reference_no",
            "produced_by",
            "requester_name",
            "approver_name",
            "notes",
        ]
    )
    for log, item in rows:
        info = batch_map.get(log.operation_batch_id)
        if info is None and log.reference_no:
            info = sr_map.get(log.reference_no)
        requester = info.requester_name if info else None
        approver = info.approver_name if info else None
        writer.writerow(
            [
                log.created_at.isoformat(),
                log.transaction_type.value,
                item.mes_code or "",
                item.item_name,
                item.process_type_code or "",
                float(log.quantity_change),
                "" if log.quantity_before is None else float(log.quantity_before),
                "" if log.quantity_after is None else float(log.quantity_after),
                log.reference_no or "",
                log.produced_by or "",
                requester or "",
                approver or "",
                log.notes or "",
            ]
        )

    return csv_streaming_response(buffer, "transactions-export.csv")


@router.get("/transactions/export.xlsx")
def export_transactions_xlsx(
    transaction_type: Optional[TransactionTypeEnum] = Query(None),
    search: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None, description="필수. 포함 시작일 YYYY-MM-DD"),
    end_date: Optional[date] = Query(None, description="필수. 포함 종료일 YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from app.utils.excel import apply_header, auto_width, make_xlsx_response

    start_dt, end_dt = _require_export_range(start_date, end_date)

    # 목록 조회와 동일하게 IoBatch outerjoin — search 가 requester_name 까지 닿고
    # 요청자/승인자 컬럼을 채운다.
    query = (
        db.query(TransactionLog, Item)
        .join(Item, TransactionLog.item_id == Item.item_id)
        .outerjoin(IoBatch, TransactionLog.operation_batch_id == IoBatch.batch_id)
    )
    query = query.filter(
        TransactionLog.created_at >= start_dt,
        TransactionLog.created_at <= end_dt,
    )
    if transaction_type:
        query = query.filter(TransactionLog.transaction_type == transaction_type)
    search_filter = build_normalized_search_filter(
        search,
        Item.item_name,
        Item.mes_code,
        TransactionLog.reference_no,
        TransactionLog.notes,
        TransactionLog.produced_by,
        IoBatch.requester_name,
    )
    if search_filter is not None:
        query = query.filter(search_filter)

    _enforce_export_limit(query.count())
    rows = query.order_by(TransactionLog.created_at.desc()).all()
    batch_map = _batch_name_map(
        db, {log.operation_batch_id for log, _ in rows if log.operation_batch_id}
    )
    sr_map = _stock_request_info_map(
        db, {log.reference_no for log, _ in rows if log.reference_no}
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "거래 이력"

    tx_label = {
        "RECEIVE": "입고", "PRODUCE": "생산입고", "SHIP": "출고",
        "ADJUST": "재고조정", "BACKFLUSH": "자동차감",
        "TRANSFER_TO_PROD": "창고→부서", "TRANSFER_TO_WH": "부서→창고",
        "TRANSFER_DEPT": "부서간 이동",
        "MARK_DEFECTIVE": "불량 등록", "SUPPLIER_RETURN": "공급업체 반품",
        "INTERNAL_USE": "사내 사용",
    }

    columns = [
        "일시", "유형", "품목 코드", "품목명", "공정코드",
        "수량변화", "이전재고", "이후재고", "참조번호", "담당자", "요청자", "승인자", "메모",
    ]
    apply_header(ws, columns)

    positive_font = Font(color="1A7C3C", bold=True)
    negative_font = Font(color="CC0000", bold=True)

    for log, item in rows:
        tx_val = log.transaction_type.value
        info = batch_map.get(log.operation_batch_id)
        if info is None and log.reference_no:
            info = sr_map.get(log.reference_no)
        requester = info.requester_name if info else None
        approver = info.approver_name if info else None
        row_data = [
            log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
            (
                "AS 반출"
                if tx_val == "INTERNAL_USE" and log.department == "AS"
                else "연구소 반출"
                if tx_val == "INTERNAL_USE" and log.department == "연구"
                else tx_label.get(tx_val, tx_val)
            ),
            item.mes_code or "",
            item.item_name,
            item.process_type_code or "",
            float(log.quantity_change),
            float(log.quantity_before) if log.quantity_before is not None else "",
            float(log.quantity_after) if log.quantity_after is not None else "",
            log.reference_no or "",
            log.produced_by or "",
            requester or "",
            approver or "",
            log.notes or "",
        ]
        ws.append(row_data)

        row_idx = ws.max_row
        hex_color = _TX_ROW_COLOR.get(tx_val, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=hex_color)
        for cell in ws[row_idx]:
            cell.fill = row_fill

        qty_change_cell = ws.cell(row=row_idx, column=6)
        if isinstance(qty_change_cell.value, float):
            qty_change_cell.font = positive_font if qty_change_cell.value >= 0 else negative_font

    auto_width(ws)
    filename = f"transactions-{date.today().strftime('%Y%m%d')}.xlsx"
    return make_xlsx_response(wb, filename)


# 기존 PUT 엔드포인트는 3차에서 reason+PIN을 요구하는 메타 수정으로 일원화됨.
# 호환성을 위해 PUT 경로는 제거하고 /meta-edit 으로 대체.


@router.post("/transactions/{log_id}/meta-edit", response_model=TransactionLogResponse)
def meta_edit_transaction(
    log_id: uuid.UUID,
    payload: TransactionMetaEditRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """거래 메타데이터(notes/reference_no/produced_by) 수정. 재고에 영향 없음.

    원본 TransactionLog의 메타 필드는 직접 업데이트하지만, 변경 전/후 스냅샷을
    TransactionEditLog에 기록하여 감사 이력을 남긴다.
    """
    editor = _verify_editor(db, payload.edited_by_employee_id, payload.edited_by_pin, request)
    try:
        log, item = transaction_actions_svc.edit_transaction_metadata(
            db,
            log_id=log_id,
            editor=editor,
            reason=payload.reason,
            notes=payload.notes,
            reference_no=payload.reference_no,
            produced_by=payload.produced_by,
            request=request,
        )
    except (
        transaction_actions_svc.TransactionLogNotFound,
        transaction_actions_svc.TransactionItemNotFound,
    ) as exc:
        raise http_error(404, ErrorCode.NOT_FOUND, str(exc))
    except transaction_actions_svc.UnsupportedTransactionMetadata as exc:
        raise http_error(422, ErrorCode.BUSINESS_RULE, str(exc))
    db.refresh(log)

    edit_count = (
        db.query(func.count(TransactionEditLog.edit_id))
        .filter(TransactionEditLog.original_log_id == log.log_id)
        .scalar()
        or 0
    )
    return _to_log_response(log, item, int(edit_count))


@router.get(
    "/transactions/{log_id}/edits",
    response_model=List[TransactionEditLogResponse],
)
def list_transaction_edits(log_id: uuid.UUID, db: Session = Depends(get_db)):
    """특정 거래의 수정 이력 (최신순)."""
    log = db.query(TransactionLog).filter(TransactionLog.log_id == log_id).first()
    if not log:
        raise http_error(404, ErrorCode.NOT_FOUND, "거래를 찾을 수 없습니다.")

    edits = (
        db.query(TransactionEditLog)
        .filter(TransactionEditLog.original_log_id == log_id)
        .order_by(TransactionEditLog.created_at.desc())
        .all()
    )
    return edits


@router.post(
    "/transactions/{log_id}/quantity-correction",
    response_model=TransactionQuantityCorrectionResponse,
)
def quantity_correct_transaction(
    log_id: uuid.UUID,
    payload: TransactionQuantityCorrectionRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """RECEIVE/SHIP 수량 보정. 원본은 보존하고 차액만 ADJUST 거래로 보정한다.

    - delta = new_quantity_change - original.quantity_change
    - new_warehouse = inventory.warehouse_qty + delta (음수 방지 검증)
    - adjust_warehouse() 서비스 호출로 재고 동기화
    - ADJUST TransactionLog 생성 + TransactionEditLog 기록
    """
    log = db.query(TransactionLog).filter(TransactionLog.log_id == log_id).first()
    if not log:
        raise http_error(404, ErrorCode.NOT_FOUND, "거래를 찾을 수 없습니다.")
    item = item_repository.get(db, log.item_id)
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")

    if log.transaction_type not in QUANTITY_CORRECTABLE:
        raise http_error(
            422,
            ErrorCode.BUSINESS_RULE,
            f"수량 보정은 RECEIVE / SHIP 유형만 지원합니다 (현재: {log.transaction_type.value}).",
        )

    new_qty = payload.quantity_change

    # SHIP 부호 검증: SHIP은 음수여야 함
    if log.transaction_type == TransactionTypeEnum.SHIP and new_qty >= 0:
        raise http_error(
            422,
            ErrorCode.BUSINESS_RULE,
            "SHIP의 수량 변화량은 음수여야 합니다 (UI에서 양수 입력 시 자동 음수 변환 필요).",
        )
    if log.transaction_type == TransactionTypeEnum.RECEIVE and new_qty <= 0:
        raise http_error(
            422,
            ErrorCode.BUSINESS_RULE,
            "RECEIVE의 수량 변화량은 양수여야 합니다.",
        )

    # 동일 거래에 이미 수량 보정 이력이 있으면 추가 보정 차단 (정책 미확정)
    existing_correction = (
        db.query(TransactionEditLog.edit_id)
        .filter(
            TransactionEditLog.original_log_id == log.log_id,
            TransactionEditLog.correction_log_id.isnot(None),
        )
        .first()
    )
    if existing_correction is not None:
        raise http_error(
            422,
            ErrorCode.BUSINESS_RULE,
            "이미 수량 보정된 거래입니다. 추가 보정은 별도 정책 확정 후 가능합니다.",
        )

    editor = _verify_editor(db, payload.edited_by_employee_id, payload.edited_by_pin, request)

    delta = new_qty - log.quantity_change

    # 재고 검증: 보정 후 warehouse_qty >= max(0, pending_quantity)
    inv = inventory_repository.get(db, log.item_id)
    if not inv:
        raise http_error(404, ErrorCode.NOT_FOUND, "재고 레코드를 찾을 수 없습니다.")

    new_warehouse = inv.warehouse_qty + delta
    if new_warehouse < 0:
        raise http_error(
            422,
            ErrorCode.STOCK_SHORTAGE,
            f"재고 부족: 보정 후 창고 재고가 {float(new_warehouse)}로 음수가 됩니다.",
        )
    # pending_quantity가 None인 레거시 레코드 방어
    pending = inv.pending_quantity or Decimal("0")
    if new_warehouse < pending:
        raise http_error(
            422,
            ErrorCode.STOCK_SHORTAGE,
            "예약 수량보다 창고 재고가 낮아질 수 없습니다.",
        )

    before = _log_snapshot(log)
    correction_log = transaction_actions_svc.correct_transaction_quantity(
        db,
        log=log,
        editor=editor,
        new_warehouse=new_warehouse,
        delta=delta,
        reason=payload.reason,
        before=before,
        request=request,
    )
    db.refresh(log)
    db.refresh(correction_log)

    # edit_count 갱신
    edit_count_orig = (
        db.query(func.count(TransactionEditLog.edit_id))
        .filter(TransactionEditLog.original_log_id == log.log_id)
        .scalar()
        or 0
    )
    return TransactionQuantityCorrectionResponse(
        original=_to_log_response(log, item, int(edit_count_orig)),
        correction=_to_log_response(correction_log, item, 0),
    )


# ---------------------------------------------------------------------------
# 거래 취소
# ---------------------------------------------------------------------------

class TransactionCancelRequest(BaseModel):
    reason: str = Field(..., min_length=1, description="취소 사유 (필수)")
    employee_code: str = Field(..., min_length=1, max_length=30, description="처리자 사번")
    pin: str = Field(..., min_length=1, max_length=20)


@router.post("/transactions/{log_id}/cancel", response_model=TransactionLogResponse)
def cancel_transaction(
    log_id: uuid.UUID,
    payload: TransactionCancelRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """거래 취소 — 내역 유지 + 재고 자동 롤백 + '취소됨' 표시.

    권한: 요청자 본인(producer_employee_id) 또는 결재 권한자(warehouse_role / department_role != none).
    BOM 배치(PRODUCE+BACKFLUSH)는 operation_batch_id 단위로 일괄 취소.
    """
    log = db.query(TransactionLog).filter(TransactionLog.log_id == log_id).first()
    if not log:
        raise http_error(404, ErrorCode.NOT_FOUND, "거래를 찾을 수 없습니다.")

    if bool(getattr(log, "cancelled", False)):
        raise http_error(422, ErrorCode.BUSINESS_RULE, "이미 취소된 거래입니다.")

    item = item_repository.get(db, log.item_id)
    if not item:
        raise http_error(404, ErrorCode.NOT_FOUND, "품목을 찾을 수 없습니다.")

    canceller = db.query(Employee).filter(Employee.employee_code == payload.employee_code).first()
    if not canceller:
        raise http_error(404, ErrorCode.NOT_FOUND, "직원을 찾을 수 없습니다.")
    if not bool(canceller.is_active):
        raise http_error(403, ErrorCode.FORBIDDEN, "비활성 직원은 거래를 취소할 수 없습니다.")
    if not verify_pin(canceller.pin_hash, payload.pin):
        raise http_error(403, ErrorCode.FORBIDDEN, "PIN이 올바르지 않습니다.")

    # 권한 체크: 본인(요청자) 또는 결재 권한자
    # 요청자 식별 — 히스토리 화면의 '요청자' 표기와 동일한 우선순위로 판정한다:
    #   1) producer_employee_id
    #   2) operation_batch_id -> IoBatch.requester_employee_id
    # produced_by is only a display snapshot and is not trusted for authorization.
    requester_eid: Optional[str] = None
    if log.producer_employee_id is not None:
        requester_eid = str(log.producer_employee_id)
    elif log.operation_batch_id is not None:
        batch = (
            db.query(IoBatch)
            .filter(IoBatch.batch_id == log.operation_batch_id)
            .first()
        )
        if batch is not None and batch.requester_employee_id is not None:
            requester_eid = str(batch.requester_employee_id)
    is_self = requester_eid == str(canceller.employee_id) if requester_eid is not None else False
    is_approver = (
        (getattr(canceller, "warehouse_role", None) or "none").lower() != "none"
        or (getattr(canceller, "department_role", None) or "none").lower() != "none"
    )
    if not (is_self or is_approver):
        raise http_error(403, ErrorCode.FORBIDDEN, "본인 거래 또는 결재 권한자만 취소할 수 있습니다.")

    set_actor(request, canceller)

    try:
        transaction_actions_svc.cancel_transaction(
            db,
            log=log,
            canceller=canceller,
            reason=payload.reason,
            request=request,
        )
    except transaction_actions_svc.TransactionInventoryNotFound as exc:
        raise http_error(404, ErrorCode.NOT_FOUND, str(exc))
    except ValueError as exc:
        raise http_error(422, ErrorCode.BUSINESS_RULE, str(exc))
    db.refresh(log)

    edit_count = (
        db.query(func.count(TransactionEditLog.edit_id))
        .filter(TransactionEditLog.original_log_id == log.log_id)
        .scalar()
        or 0
    )
    return _to_log_response(log, item, int(edit_count))
