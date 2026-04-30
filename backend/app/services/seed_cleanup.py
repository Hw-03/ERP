"""seed_cleanup.py — 722 정리본 엑셀을 DB에 적재하는 호출 가능 서비스.

scripts/dev/import_inventory_cleanup.py 의 핵심 로직 추출.
settings./reset 엔드포인트가 이 함수를 호출한다.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from sqlalchemy.orm import Session

from app.models import Inventory, InventoryLocation, Item, LocationStatusEnum

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore[assignment]

REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_EXCEL_PATH = REPO_ROOT / "outputs" / "inventory_cleanup" / "생산부_재고_매칭작업_정리본.xlsx"

EXPECTED_ROWS = 722
EXPECTED_TOTAL_QTY = Decimal("108924")
DEFAULT_MIN_STOCK = Decimal("200")

DEPT_MAP: dict[str, str] = {
    "T": "튜브",
    "H": "고압",
    "V": "진공",
    "N": "튜닝",
    "A": "조립",
    "P": "출하",
}


def _parse_erp_code(raw: str) -> tuple[str, str, int, str | None]:
    parts = str(raw).strip().split("-")
    if len(parts) < 3:
        raise ValueError(f"ERP 코드 형식 오류: {raw!r}")
    model_symbol = parts[0]
    process_type_code = parts[1]
    try:
        serial_no = int(parts[2])
    except ValueError:
        raise ValueError(f"시리얼 번호 파싱 오류: {raw!r}")
    option_code = parts[3] if len(parts) >= 4 else None
    return model_symbol, process_type_code, serial_no, option_code


def _load_excel(excel_path: Path) -> list[dict]:
    if openpyxl is None:
        raise RuntimeError("openpyxl 미설치: pip install openpyxl")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    col_erp  = next((i for i, h in enumerate(headers) if "ERP" in h.upper() or "코드" in h), None)
    col_name = next((i for i, h in enumerate(headers) if "품명" in h or "name" in h.lower()), None)
    col_type = next((i for i, h in enumerate(headers) if "분류" in h), None)
    col_qty  = next((i for i, h in enumerate(headers) if "현재고" in h or "수량" in h), None)

    if any(v is None for v in [col_erp, col_name, col_qty]):
        raise ValueError(f"필수 헤더 없음. 감지된 헤더: {headers}")

    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        erp_code  = ws.cell(r, col_erp + 1).value
        item_name = ws.cell(r, col_name + 1).value
        legacy_item_type = ws.cell(r, col_type + 1).value if col_type is not None else None
        qty_raw   = ws.cell(r, col_qty + 1).value

        if not erp_code or not item_name:
            continue

        rows.append({
            "erp_code": str(erp_code).strip(),
            "item_name": str(item_name).strip(),
            "legacy_item_type": str(legacy_item_type).strip() if legacy_item_type else None,
            "quantity": Decimal(str(qty_raw or 0)),
        })
    return rows


def run_cleanup_import(
    db: Session,
    excel_path: Path | None = None,
    *,
    dry_run: bool = False,
) -> dict:
    """722 정리본 엑셀을 파싱해 Item + Inventory + InventoryLocation 적재.

    호출 전 반드시 items / inventory / inventory_locations 테이블이 비어 있어야 한다.
    dry_run=True 면 파싱·검증만 하고 DB는 변경하지 않는다.

    Returns:
        {"rows": int, "total_qty": Decimal, "ok": bool, "errors": list[str]}
    """
    path = excel_path or DEFAULT_EXCEL_PATH
    rows = _load_excel(path)
    errors: list[str] = []

    if len(rows) != EXPECTED_ROWS:
        errors.append(f"행수 불일치: expected={EXPECTED_ROWS}, got={len(rows)}")

    total_qty = sum(r["quantity"] for r in rows)
    if total_qty != EXPECTED_TOTAL_QTY:
        errors.append(f"재고 합계 불일치: expected={EXPECTED_TOTAL_QTY}, got={total_qty}")

    from app.models import ProcessType
    valid_codes = {pt.code for pt in db.query(ProcessType).all()}
    if not valid_codes:
        raise RuntimeError("process_types 테이블이 비어있음. bootstrap_db.py --seed 먼저 실행하세요.")

    parsed: list[dict] = []
    erp_seen: set[str] = set()
    for i, row in enumerate(rows):
        erp = row["erp_code"]
        if erp in erp_seen:
            raise ValueError(f"ERP 코드 중복: {erp}")
        erp_seen.add(erp)

        model_symbol, pt_code, serial_no, option_code = _parse_erp_code(erp)

        if pt_code not in valid_codes:
            raise ValueError(f"유효하지 않은 process_type_code: {pt_code!r} (ERP={erp})")

        dept = DEPT_MAP.get(pt_code[0])
        if dept is None:
            raise ValueError(f"부서 매핑 실패: process_type_code={pt_code!r} (ERP={erp})")

        parsed.append({
            "erp": erp,
            "item_name": row["item_name"],
            "legacy_item_type": row["legacy_item_type"],
            "model_symbol": model_symbol,
            "pt_code": pt_code,
            "serial_no": serial_no,
            "option_code": option_code,
            "dept": dept,
            "quantity": row["quantity"],
            "sort_order": i + 1,
        })

    if dry_run:
        return {"rows": len(parsed), "total_qty": total_qty, "ok": not errors, "errors": errors}

    items_to_add = [
        Item(
            item_code=p["erp"],
            erp_code=p["erp"],
            barcode=p["erp"],
            item_name=p["item_name"],
            unit="EA",
            model_symbol=p["model_symbol"],
            process_type_code=p["pt_code"],
            serial_no=p["serial_no"],
            option_code=p["option_code"],
            legacy_item_type=p["legacy_item_type"],
            sort_order=p["sort_order"],
            min_stock=DEFAULT_MIN_STOCK,
        )
        for p in parsed
    ]
    db.add_all(items_to_add)
    db.flush()

    inventories = []
    locations = []
    for item_obj, p in zip(items_to_add, parsed):
        qty = p["quantity"]
        inventories.append(Inventory(
            item_id=item_obj.item_id,
            quantity=qty,
            warehouse_qty=Decimal("0"),
            pending_quantity=Decimal("0"),
        ))
        if qty > 0:
            locations.append(InventoryLocation(
                item_id=item_obj.item_id,
                department=p["dept"],
                status=LocationStatusEnum.PRODUCTION,
                quantity=qty,
            ))

    db.add_all(inventories)
    db.add_all(locations)
    db.commit()

    return {"rows": len(parsed), "total_qty": total_qty, "ok": not errors, "errors": errors}
