from __future__ import annotations

from datetime import datetime
import os
from pathlib import Path
import hashlib
import sqlite3

from openpyxl import Workbook, load_workbook
import pytest


def test_compare_inventory_separates_warehouse_and_department_and_warns_unconfirmed() -> None:
    from scripts.ops.verify_inventory_snapshot import MesItem, SourceRow, compare_inventory

    item = MesItem(
        item_id="item-1",
        mes_code="6-AR-0001",
        item_name="ADX6000 BOTTOM BLOCK",
        process_type_code="AR",
        deleted_at=None,
        warehouse_quantity=8,
        department_quantities={"조립": 2},
    )
    rows = (
        SourceRow(
            source_kind="warehouse",
            source_group="warehouse",
            source_path=Path("warehouse.xlsx"),
            sheet_name="26.08월",
            row_number=4,
            original_name="BOTTOM BLOCK",
            mes_name="ADX6000 BOTTOM BLOCK",
            mes_code="6-AR-0001",
            confirmation="O",
            quantity=10,
        ),
        SourceRow(
            source_kind="department",
            source_group="assembly_shipping",
            source_path=Path("assembly.xlsx"),
            sheet_name="조립 자재",
            row_number=3,
            original_name="BOTTOM BLOCK",
            mes_name="ADX6000 BOTTOM BLOCK",
            mes_code="6-AR-0001",
            confirmation="",
            quantity=2,
        ),
    )

    result = compare_inventory(rows, (item,))
    by_bucket = {(row.bucket, row.mes_code): row for row in result.comparisons}

    assert by_bucket[("창고", "6-AR-0001")].excel_quantity == 10
    assert by_bucket[("창고", "6-AR-0001")].mes_quantity == 8
    assert by_bucket[("창고", "6-AR-0001")].difference == -2
    assert by_bucket[("창고", "6-AR-0001")].status == "수량 불일치"
    assert by_bucket[("조립", "6-AR-0001")].status == "일치"
    assert any(
        issue.issue_type == "담당자 미확인" and issue.row_number == 3
        for issue in result.issues
    )


def test_discover_department_workbooks_accepts_next_month_names(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import discover_department_workbooks

    names = {
        "high_vacuum_tuning": "2026.09_생산부 자재_고압,진공,튜닝파트.xlsx",
        "assembly_shipping": "2026.09_생산부 자재_조립,출하파트.xlsx",
        "tube": "2026.09_생산부 자재_튜브 파트.xlsx",
        "finished": "2026.09_출하_완제품.xlsx",
    }
    for name in names.values():
        (tmp_path / name).touch()

    discovered = discover_department_workbooks(tmp_path)

    assert {group: path.name for group, path in discovered.items()} == names


def test_discover_department_workbooks_rejects_multiple_candidates(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import (
        InventoryVerificationError,
        discover_department_workbooks,
    )

    for name in (
        "2026.09_생산부 자재_고압,진공,튜닝파트.xlsx",
        "복사본_2026.09_생산부 자재_고압,진공,튜닝파트.xlsx",
        "2026.09_생산부 자재_조립,출하파트.xlsx",
        "2026.09_생산부 자재_튜브 파트.xlsx",
        "2026.09_출하_완제품.xlsx",
    ):
        (tmp_path / name).touch()

    with pytest.raises(InventoryVerificationError, match="high_vacuum_tuning"):
        discover_department_workbooks(tmp_path)


def test_extract_rows_finds_moved_headers_in_first_ten_rows(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import extract_rows_from_workbook

    path = tmp_path / "assembly.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "조립 자재"
    sheet["B4"] = "품 목"
    sheet["D4"] = "현재고"
    sheet["F4"] = "MES 품목명"
    sheet["H4"] = "MES 코드"
    sheet["J4"] = "담당자 확인"
    sheet["B5"] = "기존 품명"
    sheet["D5"] = 7
    sheet["F5"] = "현재 MES 품명"
    sheet["H5"] = "6-AR-0001"
    sheet["J5"] = "O"
    workbook.save(path)
    workbook.close()

    rows = extract_rows_from_workbook(
        path,
        source_kind="department",
        source_group="assembly_shipping",
        sheet_name="조립 자재",
    )

    assert len(rows) == 1
    assert rows[0].row_number == 5
    assert rows[0].original_name == "기존 품명"
    assert rows[0].quantity == 7
    assert rows[0].mes_name == "현재 MES 품명"
    assert rows[0].mes_code == "6-AR-0001"
    assert rows[0].confirmation == "O"


def test_extract_rows_uses_rightmost_original_name_when_header_is_duplicated(
    tmp_path: Path,
) -> None:
    from scripts.ops.verify_inventory_snapshot import extract_rows_from_workbook

    path = tmp_path / "assembly.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "조립 자재"
    sheet["B2"] = "품 목"
    sheet["D2"] = "품 목"
    sheet["I2"] = "현재고"
    sheet["J2"] = "MES 품목명"
    sheet["L2"] = "MES 코드"
    sheet["M2"] = "담당자 확인"
    sheet["B3"] = "사출 및 공용 자재"
    sheet["D3"] = "LCD GUIDE"
    sheet["I3"] = 2
    sheet["J3"] = "DX3000 LCD GUIDE"
    sheet["L3"] = "3-AR-0003"
    sheet["M3"] = "O"
    workbook.save(path)
    workbook.close()

    rows = extract_rows_from_workbook(
        path,
        source_kind="department",
        source_group="assembly_shipping",
        sheet_name="조립 자재",
    )

    assert rows[0].original_name == "LCD GUIDE"


def test_finished_goods_prefers_total_over_current_stock(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import extract_rows_from_workbook

    path = tmp_path / "finished.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "완제품"
    sheet["B2"] = "품 목"
    sheet["I2"] = "현재고"
    sheet["L2"] = "총 합"
    sheet["V2"] = "MES 품목명"
    sheet["X2"] = "MES 코드"
    sheet["Y2"] = "담당자 확인"
    sheet["B3"] = "DX3000"
    sheet["I3"] = 3
    sheet["L3"] = 13
    sheet["V3"] = "DX3000 완제품"
    sheet["X3"] = "3-AF-0001"
    sheet["Y3"] = "O"
    workbook.save(path)
    workbook.close()

    rows = extract_rows_from_workbook(
        path,
        source_kind="department",
        source_group="finished",
        sheet_name="완제품",
    )

    assert rows[0].quantity == 13


def test_latest_warehouse_sheet_is_selected_by_month(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import extract_rows_from_workbook

    path = tmp_path / "warehouse.xlsx"
    workbook = Workbook()
    old = workbook.active
    old.title = "26.08월"
    latest = workbook.create_sheet("26.09월")
    for sheet, quantity in ((old, 3), (latest, 9)):
        sheet["B3"] = "품명"
        sheet["D3"] = "현재고"
        sheet["F3"] = "MES 품목명"
        sheet["H3"] = "MES 코드"
        sheet["J3"] = "담당자 확인"
        sheet["B4"] = "창고 품명"
        sheet["D4"] = quantity
        sheet["F4"] = "현재 MES 품명"
        sheet["H4"] = "6-AR-0001"
        sheet["J4"] = "O"
    workbook.save(path)
    workbook.close()

    rows = extract_rows_from_workbook(
        path,
        source_kind="warehouse",
        source_group="warehouse",
    )

    assert len(rows) == 1
    assert rows[0].sheet_name == "26.09월"
    assert rows[0].quantity == 9


def test_copy_and_load_mes_snapshot_is_read_only_and_sums_production_locations(
    tmp_path: Path,
) -> None:
    from scripts.ops.verify_inventory_snapshot import copy_sqlite_snapshot, load_mes_items

    source = tmp_path / "employee.db"
    snapshot = tmp_path / "snapshot.db"
    with sqlite3.connect(source) as connection:
        connection.executescript(
            """
            CREATE TABLE items (
                item_id TEXT PRIMARY KEY,
                mes_code TEXT NOT NULL,
                item_name TEXT NOT NULL,
                process_type_code TEXT NOT NULL,
                deleted_at TEXT
            );
            CREATE TABLE inventory (
                item_id TEXT PRIMARY KEY,
                warehouse_qty INTEGER NOT NULL
            );
            CREATE TABLE inventory_locations (
                item_id TEXT NOT NULL,
                department TEXT NOT NULL,
                status TEXT NOT NULL,
                quantity INTEGER NOT NULL
            );
            INSERT INTO items VALUES ('item-1', '6-AR-0001', '품목 1', 'AR', NULL);
            INSERT INTO inventory VALUES ('item-1', 11);
            INSERT INTO inventory_locations VALUES ('item-1', '조립', 'PRODUCTION', 2);
            INSERT INTO inventory_locations VALUES ('item-1', '조립', 'PRODUCTION', 3);
            INSERT INTO inventory_locations VALUES ('item-1', '조립', 'DEFECTIVE', 99);
            """
        )
    before_hash = hashlib.sha256(source.read_bytes()).hexdigest()

    copy_sqlite_snapshot(source, snapshot)
    items = load_mes_items(snapshot)

    assert hashlib.sha256(source.read_bytes()).hexdigest() == before_hash
    assert len(items) == 1
    assert items[0].warehouse_quantity == 11
    assert items[0].department_quantities == {"조립": 5}


def test_write_reports_creates_matching_excel_and_json_summaries(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import (
        ComparisonRow,
        RunMetadata,
        VerificationIssue,
        VerificationResult,
        write_reports,
    )

    result = VerificationResult(
        comparisons=(
            ComparisonRow(
                bucket="창고",
                mes_code="6-AR-0001",
                mes_name="품목 1",
                excel_quantity=10,
                mes_quantity=8,
                difference=-2,
                status="수량 불일치",
                confirmed=True,
                source_rows=("warehouse.xlsx!26.08월!4",),
            ),
            ComparisonRow(
                bucket="조립",
                mes_code="6-AR-0001",
                mes_name="품목 1",
                excel_quantity=2,
                mes_quantity=2,
                difference=0,
                status="일치",
                confirmed=False,
                source_rows=("assembly.xlsx!조립 자재!3",),
            ),
        ),
        issues=(
            VerificationIssue(
                severity="경고",
                issue_type="담당자 미확인",
                message="확인 필요",
                source_path="assembly.xlsx",
                sheet_name="조립 자재",
                row_number=3,
                mes_code="6-AR-0001",
            ),
        ),
    )
    metadata = RunMetadata(
        verified_at="2026-08-10T17:00:00+09:00",
        source_hashes={"warehouse.xlsx": "source-hash"},
        employee_db_path="employee.db",
        employee_db_hash="db-hash",
        warehouse_sheet="26.08월",
    )

    paths = write_reports(
        result,
        metadata,
        tmp_path,
        timestamp=datetime(2026, 8, 10, 17, 0, 0),
    )

    assert paths.xlsx_path.name == "inventory_verification_20260810_170000.xlsx"
    assert paths.json_path.name == "inventory_verification_20260810_170000.json"
    workbook = load_workbook(paths.xlsx_path, data_only=False)
    try:
        assert workbook.sheetnames == ["요약", "수량차이", "품목오류", "전체대조"]
        assert workbook["수량차이"]["A2"].value == "창고"
        assert workbook["수량차이"]["F2"].value == -2
        assert workbook["전체대조"]["G2"].fill.fgColor.rgb == "FFF4CCCC"
        assert workbook["전체대조"]["G3"].fill.fgColor.rgb == "FFC6EFCE"
        merged_ranges = {str(cell_range) for cell_range in workbook["요약"].merged_cells.ranges}
        assert "A22:B22" in merged_ranges
        assert "C22:F22" in merged_ranges
        assert "A23:B23" in merged_ranges
        assert "C23:F23" in merged_ranges
        assert workbook["요약"].row_dimensions[23].height == 45
    finally:
        workbook.close()
    payload = __import__("json").loads(paths.json_path.read_text(encoding="utf-8"))
    assert payload["summary"]["comparison_count"] == 2
    assert payload["summary"]["difference_count"] == 1
    assert payload["summary"]["issue_count"] == 1
    assert len(payload["comparisons"]) == 2


def test_assert_sources_unlocked_rejects_excel_owner_file(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import (
        InventoryVerificationError,
        assert_sources_unlocked,
    )

    workbook_path = tmp_path / "inventory.xlsx"
    workbook_path.touch()
    (tmp_path / "~$inventory.xlsx").touch()

    with pytest.raises(InventoryVerificationError, match="열려"):
        assert_sources_unlocked((workbook_path,))


@pytest.mark.skipif(os.name != "nt", reason="Microsoft Excel COM is Windows-only")
def test_calculate_workbook_copies_updates_formula_cache(tmp_path: Path) -> None:
    from scripts.ops.verify_inventory_snapshot import calculate_workbook_copies

    path = tmp_path / "formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = 2
    sheet["A2"] = 3
    sheet["A3"] = "=SUM(A1:A2)"
    workbook.save(path)
    workbook.close()

    calculate_workbook_copies((path,))

    calculated = load_workbook(path, data_only=True, read_only=True)
    try:
        assert calculated.active["A3"].value == 5
    finally:
        calculated.close()


def test_parse_args_rejects_database_write_flags() -> None:
    from scripts.ops.verify_inventory_snapshot import _parse_args

    with pytest.raises(SystemExit):
        _parse_args(
            [
                "--department-dir",
                "department",
                "--warehouse-workbook",
                "warehouse.xlsx",
                "--apply",
            ]
        )


def _write_inventory_source(
    path: Path,
    *,
    sheet_name: str,
    code: str,
    mes_name: str,
    quantity: int,
    quantity_header: str = "현재고",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet["B2"] = "품목"
    sheet["D2"] = quantity_header
    sheet["F2"] = "MES 품목명"
    sheet["H2"] = "MES 코드"
    sheet["J2"] = "담당자 확인"
    sheet["B3"] = f"기존 {mes_name}"
    sheet["D3"] = quantity
    sheet["F3"] = mes_name
    sheet["H3"] = code
    sheet["J3"] = "O"
    workbook.save(path)
    workbook.close()


def test_run_verification_preserves_sources_and_db_and_writes_reports(
    tmp_path: Path,
) -> None:
    from scripts.ops.verify_inventory_snapshot import run_verification

    department_dir = tmp_path / "departments"
    department_dir.mkdir()
    sources = {
        "2026.09_생산부 자재_고압,진공,튜닝파트.xlsx": (
            "고압",
            "6-HR-0001",
            "고압 품목",
            4,
            "현재고",
        ),
        "2026.09_생산부 자재_조립,출하파트.xlsx": (
            "조립 자재",
            "6-AR-0001",
            "조립 품목",
            5,
            "현재고",
        ),
        "2026.09_생산부 자재_튜브 파트.xlsx": (
            "튜브",
            "6-TR-0001",
            "튜브 품목",
            6,
            "현재고",
        ),
        "2026.09_출하_완제품.xlsx": (
            "완제품",
            "6-PR-0001",
            "출하 품목",
            7,
            "총 합",
        ),
    }
    for name, (sheet, code, mes_name, quantity, quantity_header) in sources.items():
        _write_inventory_source(
            department_dir / name,
            sheet_name=sheet,
            code=code,
            mes_name=mes_name,
            quantity=quantity,
            quantity_header=quantity_header,
        )
    warehouse = tmp_path / "warehouse.xlsx"
    _write_inventory_source(
        warehouse,
        sheet_name="26.09월",
        code="6-AR-0001",
        mes_name="조립 품목",
        quantity=10,
    )
    employee_db = tmp_path / "employee.db"
    with sqlite3.connect(employee_db) as connection:
        connection.executescript(
            """
            CREATE TABLE items (
                item_id TEXT PRIMARY KEY,
                mes_code TEXT NOT NULL,
                item_name TEXT NOT NULL,
                process_type_code TEXT NOT NULL,
                deleted_at TEXT
            );
            CREATE TABLE inventory (item_id TEXT PRIMARY KEY, warehouse_qty INTEGER NOT NULL);
            CREATE TABLE inventory_locations (
                item_id TEXT NOT NULL,
                department TEXT NOT NULL,
                status TEXT NOT NULL,
                quantity INTEGER NOT NULL
            );
            INSERT INTO items VALUES ('h', '6-HR-0001', '고압 품목', 'HR', NULL);
            INSERT INTO items VALUES ('a', '6-AR-0001', '조립 품목', 'AR', NULL);
            INSERT INTO items VALUES ('t', '6-TR-0001', '튜브 품목', 'TR', NULL);
            INSERT INTO items VALUES ('p', '6-PR-0001', '출하 품목', 'PR', NULL);
            INSERT INTO inventory VALUES ('h', 0), ('a', 10), ('t', 0), ('p', 0);
            INSERT INTO inventory_locations VALUES ('h', '고압', 'PRODUCTION', 4);
            INSERT INTO inventory_locations VALUES ('a', '조립', 'PRODUCTION', 5);
            INSERT INTO inventory_locations VALUES ('t', '튜브', 'PRODUCTION', 6);
            INSERT INTO inventory_locations VALUES ('p', '출하', 'PRODUCTION', 7);
            """
        )
    original_paths = [*(department_dir.glob("*.xlsx")), warehouse, employee_db]
    before_hashes = {
        path: hashlib.sha256(path.read_bytes()).hexdigest() for path in original_paths
    }

    run = run_verification(
        department_dir=department_dir,
        warehouse_workbook=warehouse,
        warehouse_sheet=None,
        employee_db=employee_db,
        output_dir=tmp_path / "reports",
        calculator=lambda paths: None,
        timestamp=datetime(2026, 9, 4, 17, 0, 0),
    )

    assert all(row.status == "일치" for row in run.result.comparisons)
    assert run.metadata.warehouse_sheet == "26.09월"
    assert run.report_paths.xlsx_path.is_file()
    assert run.report_paths.json_path.is_file()
    assert {
        path: hashlib.sha256(path.read_bytes()).hexdigest() for path in original_paths
    } == before_hashes


def test_compare_inventory_aggregates_duplicates_and_never_matches_by_name() -> None:
    from scripts.ops.verify_inventory_snapshot import MesItem, SourceRow, compare_inventory

    item = MesItem(
        item_id="a",
        mes_code="6-AR-0001",
        item_name="현재 품명",
        process_type_code="AR",
        deleted_at=None,
        warehouse_quantity=0,
        department_quantities={"조립": 5},
    )
    rows = (
        SourceRow(
            "department",
            "assembly_shipping",
            Path("assembly.xlsx"),
            "조립 자재",
            3,
            "기존 품명 1",
            "현재 품명",
            "6-AR-0001",
            "O",
            2,
        ),
        SourceRow(
            "department",
            "assembly_shipping",
            Path("assembly.xlsx"),
            "조립 자재",
            4,
            "기존 품명 2",
            "현재 품명",
            "6-AR-0001",
            "O",
            3,
        ),
        SourceRow(
            "warehouse",
            "warehouse",
            Path("warehouse.xlsx"),
            "26.08월",
            4,
            "현재 품명",
            "현재 품명",
            "",
            "O",
            9,
        ),
    )

    result = compare_inventory(rows, (item,))

    assembly = next(row for row in result.comparisons if row.bucket == "조립")
    assert assembly.excel_quantity == 5
    assert assembly.status == "일치"
    assert any(issue.issue_type == "MES 코드 없음" for issue in result.issues)
    assert not any(row.bucket == "창고" for row in result.comparisons)


def test_compare_inventory_reports_row_errors_and_continues_other_rows() -> None:
    from scripts.ops.verify_inventory_snapshot import MesItem, SourceRow, compare_inventory

    item = MesItem(
        item_id="t",
        mes_code="6-TR-0001",
        item_name="튜브 품목",
        process_type_code="TR",
        deleted_at=None,
        warehouse_quantity=0,
        department_quantities={"튜브": 4},
    )
    rows = (
        SourceRow(
            "department",
            "tube",
            Path("tube.xlsx"),
            "튜브",
            3,
            "오류 행",
            "튜브 품목",
            "6-TR-0001",
            "O",
            "#VALUE!",
        ),
        SourceRow(
            "department",
            "tube",
            Path("tube.xlsx"),
            "튜브",
            4,
            "정상 행",
            "튜브 품목",
            "6-TR-0001",
            "O",
            4,
        ),
    )

    result = compare_inventory(rows, (item,))

    assert result.comparisons[0].status == "일치"
    assert any(
        issue.issue_type == "원본 수량 오류" and issue.row_number == 3
        for issue in result.issues
    )


def test_compare_inventory_reports_name_file_deleted_and_mes_only_problems() -> None:
    from scripts.ops.verify_inventory_snapshot import MesItem, SourceRow, compare_inventory

    active = MesItem(
        item_id="a",
        mes_code="6-AR-0001",
        item_name="현재 조립 품명",
        process_type_code="AR",
        deleted_at=None,
        warehouse_quantity=0,
        department_quantities={"조립": 3},
    )
    deleted_with_stock = MesItem(
        item_id="d",
        mes_code="6-HR-0002",
        item_name="삭제 고압 품목",
        process_type_code="HR",
        deleted_at="2026-08-01",
        warehouse_quantity=7,
        department_quantities={},
    )
    deleted_zero = MesItem(
        item_id="z",
        mes_code="6-HR-0003",
        item_name="삭제 0 품목",
        process_type_code="HR",
        deleted_at="2026-08-01",
        warehouse_quantity=0,
        department_quantities={},
    )
    row = SourceRow(
        "department",
        "tube",
        Path("tube.xlsx"),
        "튜브",
        3,
        "기존 이름",
        "옛 조립 품명",
        "6-AR-0001",
        "O",
        3,
    )

    result = compare_inventory((row,), (active, deleted_with_stock, deleted_zero))
    issue_types = [issue.issue_type for issue in result.issues]

    assert "품명 불일치" in issue_types
    assert "잘못된 부서 파일" in issue_types
    assert "삭제 품목 재고 있음" in issue_types
    assert not any(issue.mes_code == "6-HR-0003" for issue in result.issues)
    missing = next(row for row in result.comparisons if row.mes_code == "6-HR-0002")
    assert missing.bucket == "창고"
    assert missing.status == "Excel 누락"
