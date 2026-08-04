from __future__ import annotations

from datetime import datetime
import os
from pathlib import Path
import sqlite3
import subprocess
import sys
from typing import Any
import uuid

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import text


def _write_workbook(path: Path, rows: list[tuple[str, Any, str, str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "26.07월"
    sheet["D3"] = "품명"
    sheet["O3"] = "현재고"
    sheet["BF3"] = "MES 품목명"
    sheet["BH3"] = "MES 코드"
    sheet["BI3"] = "담당자 확인"
    sheet["BJ3"] = "확인"
    for row_number, (code, quantity, mes_name, confirmation, state) in enumerate(rows, start=4):
        sheet.cell(row=row_number, column=4, value=f"원본 {mes_name}")
        sheet.cell(row=row_number, column=15, value=quantity)
        sheet.cell(row=row_number, column=58, value=mes_name)
        sheet.cell(row=row_number, column=60, value=code)
        sheet.cell(row=row_number, column=61, value=confirmation)
        sheet.cell(row=row_number, column=62, value=state)
    workbook.save(path)
    workbook.close()


def test_parse_workbook_aggregates_duplicate_mes_codes(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(
        path,
        [
            ("3-TR-0001", 5, "전극", "O", "확인완료"),
            ("3-TR-0001", 7, "전극", "O", "확인완료"),
            ("3-TR-0002", 0, "필라멘트", "O", "확인완료"),
        ],
    )

    snapshot = parse_workbook(path, "26.07월")

    assert snapshot.row_count == 3
    assert snapshot.quantities == {"3-TR-0001": 12, "3-TR-0002": 0}
    assert snapshot.mes_names == {"3-TR-0001": "전극", "3-TR-0002": "필라멘트"}
    assert snapshot.duplicate_groups == 1


def test_parse_workbook_treats_blank_quantity_as_zero(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [("3-TR-0001", " ", "전극", "O", "확인완료")])

    snapshot = parse_workbook(path, "26.07월")

    assert snapshot.quantities == {"3-TR-0001": 0}


def test_parse_workbook_rejects_empty_inventory_sheet(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [])

    with pytest.raises(WarehouseSyncError, match="no inventory rows"):
        parse_workbook(path, "26.07월")


@pytest.mark.parametrize(
    ("confirmation", "state"),
    [("", "확인완료"), ("O", "확인대기")],
)
def test_parse_workbook_rejects_unconfirmed_rows(
    tmp_path: Path,
    confirmation: str,
    state: str,
) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [("3-TR-0001", 5, "전극", confirmation, state)])

    with pytest.raises(WarehouseSyncError, match="row 4.*확인"):
        parse_workbook(path, "26.07월")


@pytest.mark.parametrize("quantity", [-1, 1.5, "not-a-number"])
def test_parse_workbook_rejects_invalid_quantities(tmp_path: Path, quantity: Any) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [("3-TR-0001", quantity, "전극", "O", "확인완료")])

    with pytest.raises(WarehouseSyncError, match="row 4.*quantity"):
        parse_workbook(path, "26.07월")


def test_parse_workbook_rejects_conflicting_names_for_duplicate_code(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(
        path,
        [
            ("3-TR-0001", 5, "전극", "O", "확인완료"),
            ("3-TR-0001", 7, "다른 전극", "O", "확인완료"),
        ],
    )

    with pytest.raises(WarehouseSyncError, match="conflicting MES names"):
        parse_workbook(path, "26.07월")


def test_parse_workbook_rejects_blank_mes_code(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [("", 5, "전극", "O", "확인완료")])

    with pytest.raises(WarehouseSyncError, match="row 4.*MES 코드"):
        parse_workbook(path, "26.07월")


def test_parse_workbook_rejects_shifted_columns(tmp_path: Path) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, parse_workbook

    path = tmp_path / "warehouse.xlsx"
    _write_workbook(path, [("3-TR-0001", 5, "전극", "O", "확인완료")])
    workbook = load_workbook(path)
    sheet = workbook["26.07월"]
    sheet["BH3"] = "잘못된 열"
    workbook.save(path)
    workbook.close()

    with pytest.raises(WarehouseSyncError, match="expected header.*BH3"):
        parse_workbook(path, "26.07월")


def _write_employee_db(path: Path, rows: list[dict[str, Any]]) -> None:
    with sqlite3.connect(path) as connection:
        connection.executescript(
            """
            CREATE TABLE items (
                item_id TEXT PRIMARY KEY,
                mes_code TEXT NOT NULL UNIQUE,
                item_name TEXT NOT NULL,
                sort_order INTEGER,
                unit TEXT NOT NULL,
                legacy_part TEXT,
                legacy_item_type TEXT,
                supplier TEXT,
                min_stock INTEGER,
                model_symbol TEXT NOT NULL,
                process_type_code TEXT NOT NULL,
                serial_no INTEGER NOT NULL,
                bom_completed_at TEXT,
                sales_review_required INTEGER NOT NULL DEFAULT 0,
                deleted_at TEXT
            );
            """
        )
        connection.executemany(
            """
            INSERT INTO items (
                item_id, mes_code, item_name, sort_order, unit, legacy_part,
                legacy_item_type, supplier, min_stock, model_symbol,
                process_type_code, serial_no, bom_completed_at,
                sales_review_required, deleted_at
            ) VALUES (
                :item_id, :mes_code, :item_name, :sort_order, :unit, :legacy_part,
                :legacy_item_type, :supplier, :min_stock, :model_symbol,
                :process_type_code, :serial_no, :bom_completed_at,
                :sales_review_required, :deleted_at
            )
            """,
            rows,
        )


def _employee_item(
    *,
    item_id: str,
    mes_code: str,
    item_name: str,
    supplier: str | None = None,
) -> dict[str, Any]:
    model_symbol, process_type_code, serial_text = mes_code.split("-")
    return {
        "item_id": item_id,
        "mes_code": mes_code,
        "item_name": item_name,
        "sort_order": 1,
        "unit": "EA",
        "legacy_part": "자재창고",
        "legacy_item_type": "원자재",
        "supplier": supplier,
        "min_stock": 10,
        "model_symbol": model_symbol,
        "process_type_code": process_type_code,
        "serial_no": int(serial_text),
        "bom_completed_at": None,
        "sales_review_required": 0,
        "deleted_at": None,
    }


def test_run_sync_updates_only_warehouse_and_related_item_master(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from app.models import (
        BoxSizeEnum,
        Department,
        Employee,
        EmployeeLevelEnum,
        Inventory,
        InventoryLocation,
        LocationStatusEnum,
        StockRequest,
        StockRequestStatusEnum,
        StockRequestTypeEnum,
        TransactionLog,
        TransactionTypeEnum,
        WarehouseAngle,
        WarehouseBox,
        WarehouseBoxItem,
        WarehouseSpecialZone,
        WarehouseSpecialZoneItem,
    )
    from scripts.ops.sync_warehouse_inventory import WorkbookSnapshot, run_sync

    department = Department(name="조립", display_order=1, is_active=True, io_enabled=True)
    employee = Employee(
        employee_code="SYNC001",
        name="동기화 담당자",
        role="ops",
        department="조립",
        level=EmployeeLevelEnum.STAFF,
        warehouse_role="primary",
        department_role="primary",
        display_order=0,
        is_active=True,
    )
    db_session.add_all([department, employee])
    db_session.flush()

    renamed = make_item(
        name="예전 품명",
        process_type_code="TR",
        warehouse_qty=10,
        pending=2,
        model_symbol="3",
        serial_no=1,
    )
    absent = make_item(
        name="엑셀 미포함",
        process_type_code="TR",
        warehouse_qty=7,
        model_symbol="3",
        serial_no=2,
    )
    deleted = make_item(
        name="삭제 품목",
        process_type_code="TR",
        warehouse_qty=4,
        model_symbol="3",
        serial_no=3,
    )
    deleted.deleted_at = datetime(2026, 8, 1)
    db_session.add_all(
        [
            InventoryLocation(
                item_id=renamed.item_id,
                department="조립",
                status=LocationStatusEnum.PRODUCTION,
                quantity=5,
            ),
            InventoryLocation(
                item_id=absent.item_id,
                department="조립",
                status=LocationStatusEnum.DEFECTIVE,
                quantity=1,
            ),
            TransactionLog(
                item_id=renamed.item_id,
                transaction_type=TransactionTypeEnum.RECEIVE,
                quantity_change=10,
                quantity_before=0,
                quantity_after=10,
            ),
            StockRequest(
                requester_employee_id=employee.employee_id,
                requester_name=employee.name,
                requester_department=employee.department,
                request_type=StockRequestTypeEnum.RAW_SHIP,
                status=StockRequestStatusEnum.RESERVED,
            ),
        ]
    )
    db_session.query(Inventory).filter(Inventory.item_id == renamed.item_id).update({"quantity": 15})
    db_session.query(Inventory).filter(Inventory.item_id == absent.item_id).update({"quantity": 8})

    angle = WarehouseAngle(label="A", rows=1, layers=1, jaris_per_cell=1)
    db_session.add(angle)
    db_session.flush()
    box = WarehouseBox(
        angle_id=angle.id,
        row_no=1,
        layer_no=1,
        jari_index=0,
        size=BoxSizeEnum.SMALL,
        stack_order=0,
    )
    zone = WarehouseSpecialZone(
        label="통로",
        zone_type="aisle",
        pos_x=0,
        pos_y=0,
        width=80,
        height=40,
        display_order=0,
        is_active=True,
    )
    db_session.add_all([box, zone])
    db_session.flush()
    db_session.add_all(
        [
            WarehouseBoxItem(box_id=box.box_id, item_id=renamed.item_id, quantity=10),
            WarehouseSpecialZoneItem(zone_id=zone.id, item_id=absent.item_id, quantity=7),
        ]
    )
    db_session.commit()

    new_item_id = uuid.uuid4().hex
    second_new_item_id = uuid.uuid4().hex
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [
            _employee_item(
                item_id=str(renamed.item_id),
                mes_code="34-TR-0001",
                item_name="현재 품명",
                supplier="현재 공급처",
            ),
            _employee_item(
                item_id=new_item_id,
                mes_code="3-TR-0004",
                item_name="신규 품목",
            ),
            _employee_item(
                item_id=second_new_item_id,
                mes_code="3-TR-0005",
                item_name="두 번째 신규 품목",
            ),
        ],
    )
    snapshot = WorkbookSnapshot(
        row_count=3,
        quantities={"34-TR-0001": 20, "3-TR-0004": 3, "3-TR-0005": 2},
        mes_names={
            "34-TR-0001": "현재 품명",
            "3-TR-0004": "신규 품목",
            "3-TR-0005": "두 번째 신규 품목",
        },
        duplicate_groups=0,
    )
    location_before = [
        tuple(row)
        for row in db_session.execute(
            text(
                "SELECT location_id, item_id, department, status, quantity, updated_at, defective_at "
                "FROM inventory_locations ORDER BY location_id"
            )
        )
    ]
    history_before = (
        db_session.query(TransactionLog).count(),
        db_session.query(StockRequest).count(),
    )

    summary = run_sync(db_session, employee_db, snapshot, apply=True)

    assert summary.applied is True
    assert summary.master_items_added == 2
    assert summary.master_items_updated == 1
    assert summary.absent_nonzero_items_zeroed == 2
    assert summary.absent_quantity_zeroed == 11
    assert summary.source_zero_nonzero_items == 0
    assert summary.source_zero_quantity == 0
    assert summary.warehouse_box_items_deleted == 1
    assert summary.special_zone_items_deleted == 1
    assert db_session.query(WarehouseAngle).count() == 1
    assert db_session.query(WarehouseBox).count() == 1
    assert db_session.query(WarehouseSpecialZone).count() == 1
    assert db_session.query(WarehouseBoxItem).count() == 0
    assert db_session.query(WarehouseSpecialZoneItem).count() == 0

    db_session.refresh(renamed)
    assert renamed.mes_code == "34-TR-0001"
    assert renamed.item_name == "현재 품명"
    assert renamed.supplier is None
    renamed_inventory = db_session.query(Inventory).filter_by(item_id=renamed.item_id).one()
    assert int(renamed_inventory.warehouse_qty) == 20
    assert int(renamed_inventory.quantity) == 25
    assert int(renamed_inventory.pending_quantity) == 2

    absent_inventory = db_session.query(Inventory).filter_by(item_id=absent.item_id).one()
    deleted_inventory = db_session.query(Inventory).filter_by(item_id=deleted.item_id).one()
    new_inventory = db_session.query(Inventory).filter_by(item_id=new_item_id).one()
    second_new_inventory = db_session.query(Inventory).filter_by(item_id=second_new_item_id).one()
    assert (int(absent_inventory.warehouse_qty), int(absent_inventory.quantity)) == (0, 1)
    assert (int(deleted_inventory.warehouse_qty), int(deleted_inventory.quantity)) == (0, 0)
    assert (int(new_inventory.warehouse_qty), int(new_inventory.quantity)) == (3, 3)
    assert (int(second_new_inventory.warehouse_qty), int(second_new_inventory.quantity)) == (2, 2)

    location_after = [
        tuple(row)
        for row in db_session.execute(
            text(
                "SELECT location_id, item_id, department, status, quantity, updated_at, defective_at "
                "FROM inventory_locations ORDER BY location_id"
            )
        )
    ]
    assert location_after == location_before
    assert (
        db_session.query(TransactionLog).count(),
        db_session.query(StockRequest).count(),
    ) == history_before


def test_run_sync_dry_run_does_not_mutate_dev_or_employee_db(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from app.models import Inventory
    from scripts.ops.sync_warehouse_inventory import WorkbookSnapshot, run_sync

    item = make_item(
        name="예전 품명",
        process_type_code="TR",
        warehouse_qty=10,
        pending=2,
        model_symbol="3",
        serial_no=1,
    )
    db_session.commit()
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [_employee_item(item_id=str(item.item_id), mes_code=item.mes_code, item_name="현재 품명")],
    )
    employee_before = employee_db.read_bytes()
    snapshot = WorkbookSnapshot(
        row_count=1,
        quantities={item.mes_code: 20},
        mes_names={item.mes_code: "현재 품명"},
        duplicate_groups=0,
    )

    summary = run_sync(db_session, employee_db, snapshot, apply=False)

    db_session.refresh(item)
    inventory = db_session.query(Inventory).filter_by(item_id=item.item_id).one()
    assert summary.applied is False
    assert summary.master_items_updated == 1
    assert summary.inventory_items_changed == 1
    assert item.item_name == "예전 품명"
    assert int(inventory.warehouse_qty) == 10
    assert int(inventory.pending_quantity) == 2
    assert employee_db.read_bytes() == employee_before


def test_run_sync_reports_source_zero_separately_from_absent_items(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from scripts.ops.sync_warehouse_inventory import WorkbookSnapshot, run_sync

    item = make_item(name="실사 0 품목", warehouse_qty=5, model_symbol="3", serial_no=1)
    db_session.commit()
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [_employee_item(item_id=str(item.item_id), mes_code=item.mes_code, item_name=item.item_name)],
    )
    snapshot = WorkbookSnapshot(
        row_count=1,
        quantities={item.mes_code: 0},
        mes_names={item.mes_code: item.item_name},
        duplicate_groups=0,
    )

    summary = run_sync(db_session, employee_db, snapshot, apply=False)

    assert summary.absent_nonzero_items_zeroed == 0
    assert summary.absent_quantity_zeroed == 0
    assert summary.source_zero_nonzero_items == 1
    assert summary.source_zero_quantity == 5


def test_run_sync_rejects_missing_employee_item_without_mutation(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from app.models import Inventory
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, WorkbookSnapshot, run_sync

    item = make_item(name="보존 품목", warehouse_qty=10, model_symbol="3", serial_no=1)
    db_session.commit()
    employee_db = tmp_path / "employee.db"
    _write_employee_db(employee_db, [])
    snapshot = WorkbookSnapshot(
        row_count=1,
        quantities={item.mes_code: 20},
        mes_names={item.mes_code: item.item_name},
        duplicate_groups=0,
    )

    with pytest.raises(WarehouseSyncError, match="no active items"):
        run_sync(db_session, employee_db, snapshot, apply=True)

    inventory = db_session.query(Inventory).filter_by(item_id=item.item_id).one()
    assert int(inventory.warehouse_qty) == 10


def test_run_sync_reports_name_mismatch_and_rejects_pending_conflict(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, WorkbookSnapshot, run_sync

    item = make_item(
        name="개발 품목",
        warehouse_qty=10,
        pending=5,
        model_symbol="3",
        serial_no=1,
    )
    db_session.commit()
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [_employee_item(item_id=str(item.item_id), mes_code=item.mes_code, item_name="직원 품목")],
    )
    mismatched = WorkbookSnapshot(
        row_count=1,
        quantities={item.mes_code: 10},
        mes_names={item.mes_code: "다른 품목"},
        duplicate_groups=0,
    )
    mismatch_summary = run_sync(db_session, employee_db, mismatched, apply=False)
    assert mismatch_summary.employee_name_mismatches == 1

    pending_conflict = WorkbookSnapshot(
        row_count=1,
        quantities={item.mes_code: 4},
        mes_names={item.mes_code: "직원 품목"},
        duplicate_groups=0,
    )
    with pytest.raises(WarehouseSyncError, match="pending quantity exceeds"):
        run_sync(db_session, employee_db, pending_conflict, apply=False)


def test_run_sync_rejects_mes_code_owned_by_another_dev_item(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, WorkbookSnapshot, run_sync

    source_item = make_item(name="원본", model_symbol="3", serial_no=1)
    collision = make_item(name="충돌", model_symbol="34", serial_no=1)
    db_session.commit()
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [
            _employee_item(
                item_id=str(source_item.item_id),
                mes_code=collision.mes_code,
                item_name="현재 품목",
            )
        ],
    )
    snapshot = WorkbookSnapshot(
        row_count=1,
        quantities={collision.mes_code: 1},
        mes_names={collision.mes_code: "현재 품목"},
        duplicate_groups=0,
    )

    with pytest.raises(WarehouseSyncError, match="MES code collision"):
        run_sync(db_session, employee_db, snapshot, apply=True)


def test_execute_sync_requires_confirmation_and_backs_up_before_apply(
    tmp_path: Path,
    db_session,
    make_item,
) -> None:
    from scripts.ops.sync_warehouse_inventory import WarehouseSyncError, execute_sync

    item = make_item(name="현재 품명", warehouse_qty=10, model_symbol="3", serial_no=1)
    db_session.commit()
    dev_db = tmp_path / "dev.db"
    with sqlite3.connect(dev_db) as target:
        db_session.connection().connection.driver_connection.backup(target)
    employee_db = tmp_path / "employee.db"
    _write_employee_db(
        employee_db,
        [_employee_item(item_id=str(item.item_id), mes_code=item.mes_code, item_name=item.item_name)],
    )
    workbook = tmp_path / "warehouse.xlsx"
    _write_workbook(workbook, [(item.mes_code, 20, item.item_name, "O", "확인완료")])
    dev_before = dev_db.read_bytes()
    backup_calls: list[str] = []

    preview = execute_sync(
        workbook_path=workbook,
        sheet_name="26.07월",
        dev_db_path=dev_db,
        employee_db_path=employee_db,
        apply=False,
        confirm=None,
        backup_fn=backup_calls.append,
    )
    assert preview.applied is False
    assert dev_db.read_bytes() == dev_before
    assert backup_calls == []

    with pytest.raises(WarehouseSyncError, match="SYNC-WAREHOUSE"):
        execute_sync(
            workbook_path=workbook,
            sheet_name="26.07월",
            dev_db_path=dev_db,
            employee_db_path=employee_db,
            apply=True,
            confirm="WRONG",
            backup_fn=backup_calls.append,
        )
    assert backup_calls == []

    applied = execute_sync(
        workbook_path=workbook,
        sheet_name="26.07월",
        dev_db_path=dev_db,
        employee_db_path=employee_db,
        apply=True,
        confirm="SYNC-WAREHOUSE",
        backup_fn=backup_calls.append,
    )

    assert applied.applied is True
    assert backup_calls == [str(dev_db.resolve())]
    with sqlite3.connect(dev_db) as connection:
        assert connection.execute(
            "SELECT warehouse_qty FROM inventory WHERE item_id = ?", (item.item_id.hex,)
        ).fetchone() == (20,)

    _write_workbook(workbook, [(item.mes_code, 21, item.item_name, "O", "확인완료")])
    environment = os.environ.copy()
    environment["MES_RUNTIME_ROOT"] = str(tmp_path / "runtime")
    result = subprocess.run(
        [
            sys.executable,
            str(Path(__file__).resolve().parents[3] / "scripts" / "ops" / "sync_warehouse_inventory.py"),
            str(workbook),
            "--dev-db",
            str(dev_db),
            "--employee-db",
            str(employee_db),
            "--apply",
            "--confirm",
            "SYNC-WAREHOUSE",
        ],
        cwd=tmp_path,
        env=environment,
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    with sqlite3.connect(dev_db) as connection:
        assert connection.execute(
            "SELECT warehouse_qty FROM inventory WHERE item_id = ?", (item.item_id.hex,)
        ).fetchone() == (21,)
