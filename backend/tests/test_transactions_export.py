"""거래 이력 export(csv/xlsx) 의 요청자/승인자명 동기화 (D1, history-overhaul-fixup).

목록 조회와 동일하게 export 도 IoBatch outerjoin → requester_name + approver_name 컬럼을
채우고, search 가 requester_name 까지 닿는지 검증한다.
"""

from __future__ import annotations

import csv
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.models import (
    DepartmentEnum,
    Employee,
    EmployeeLevelEnum,
    IoBatch,
    Item,
    StockRequest,
    StockRequestStatusEnum,
    StockRequestTypeEnum,
    TransactionLog,
    TransactionTypeEnum,
)


def _emp(db, code: str, name: str) -> Employee:
    e = Employee(
        employee_code=code, name=name, role="조립/staff",
        department=DepartmentEnum.ASSEMBLY, level=EmployeeLevelEnum.STAFF,
        display_order=0,
    )
    db.add(e)
    db.flush()
    return e


def _seed_batch_transaction(db) -> Item:
    """요청자≠승인자인 결재 배치 + 그 결과 TransactionLog 1건."""
    requester = _emp(db, "EXP_REQ", "요청자A")
    approver = _emp(db, "EXP_APP", "승인자B")
    item = Item(
        item_name="ExportItem", process_type_code="TR", model_symbol="9", serial_no=1
    )
    db.add(item)
    db.flush()

    sr = StockRequest(
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department="조립",
        request_type=StockRequestTypeEnum.WAREHOUSE_TO_DEPT,
        status=StockRequestStatusEnum.COMPLETED,
        approved_by_name=approver.name,
        approved_by_employee_id=approver.employee_id,
    )
    db.add(sr)
    db.flush()

    batch = IoBatch(
        work_type="warehouse_io",
        sub_type="warehouse_to_dept",
        status="completed",
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department="조립",
        stock_request_id=sr.request_id,
    )
    db.add(batch)
    db.flush()

    db.add(
        TransactionLog(
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_TO_PROD,
            quantity_change=Decimal("5"),
            operation_batch_id=batch.batch_id,
            created_at=datetime.utcnow(),
        )
    )
    db.commit()
    return item


def _seed_multi_request_batch_transactions(db) -> dict[str, str]:
    requester = _emp(db, "MULTI_REQ", "복수요청자")
    warehouse_approver = _emp(db, "MULTI_WH_APP", "창고승인자")
    department_approver = _emp(db, "MULTI_DEPT_APP", "부서승인자")
    warehouse_item = Item(
        item_name="복수창고로그", process_type_code="TR", model_symbol="9", serial_no=11
    )
    department_item = Item(
        item_name="복수부서로그", process_type_code="HF", model_symbol="9", serial_no=12
    )
    db.add_all([warehouse_item, department_item])
    db.flush()
    batch = IoBatch(
        work_type="internal_use",
        sub_type="internal_use_out",
        status="completed",
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department=DepartmentEnum.AS.value,
        stock_request_id=None,
        submitted_at=datetime.utcnow(),
    )
    db.add(batch)
    db.flush()
    warehouse_request = StockRequest(
        request_code="SR-MULTI-WH",
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department=DepartmentEnum.AS.value,
        request_type=StockRequestTypeEnum.INTERNAL_USE,
        status=StockRequestStatusEnum.COMPLETED,
        operation_batch_id=batch.batch_id,
        requires_warehouse_approval=True,
        approved_by_name=warehouse_approver.name,
        approved_by_employee_id=warehouse_approver.employee_id,
        approved_at=datetime.utcnow(),
        submitted_at=datetime.utcnow(),
    )
    department_request = StockRequest(
        request_code="SR-MULTI-DEPT",
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department=DepartmentEnum.AS.value,
        request_type=StockRequestTypeEnum.INTERNAL_USE,
        status=StockRequestStatusEnum.COMPLETED,
        operation_batch_id=batch.batch_id,
        requires_warehouse_approval=False,
        requires_department_approval=True,
        department_approved_by_name=department_approver.name,
        department_approved_by_employee_id=department_approver.employee_id,
        department_approved_at=datetime.utcnow(),
        submitted_at=datetime.utcnow(),
    )
    db.add_all([warehouse_request, department_request])
    db.flush()
    db.add_all(
        [
            TransactionLog(
                item_id=warehouse_item.item_id,
                transaction_type=TransactionTypeEnum.INTERNAL_USE,
                quantity_change=Decimal("-1"),
                quantity_before=Decimal("2"),
                quantity_after=Decimal("1"),
                department=DepartmentEnum.AS.value,
                reference_no=warehouse_request.request_code,
                operation_batch_id=batch.batch_id,
                created_at=datetime.utcnow(),
            ),
            TransactionLog(
                item_id=department_item.item_id,
                transaction_type=TransactionTypeEnum.INTERNAL_USE,
                quantity_change=Decimal("-1"),
                quantity_before=Decimal("2"),
                quantity_after=Decimal("1"),
                department=DepartmentEnum.AS.value,
                reference_no=department_request.request_code,
                operation_batch_id=batch.batch_id,
                created_at=datetime.utcnow(),
            ),
        ]
    )
    db.commit()
    return {
        warehouse_item.item_name: warehouse_approver.name,
        department_item.item_name: department_approver.name,
    }


def _range():
    # 시드 거래는 created_at=datetime.utcnow() (UTC). 필터 범위도 UTC 날짜로 맞춰야
    # KST 자정~오전9시 구간에서 로컬 날짜(today)와 어긋나 빈 결과가 나오지 않는다.
    today = datetime.utcnow().date().isoformat()
    return f"start_date={today}&end_date={today}"


def test_export_csv_includes_requester_and_approver(client, db_session):
    _seed_batch_transaction(db_session)
    resp = client.get(f"/api/inventory/transactions/export.csv?{_range()}")
    assert resp.status_code == 200, resp.text
    body = resp.text
    header = body.splitlines()[0]
    assert "requester_name" in header and "approver_name" in header
    # 요청자명 + 승인자명(요청자≠승인자)이 행에 채워진다.
    assert "요청자A" in body
    assert "승인자B" in body


def test_multi_request_history_and_exports_use_each_log_request_approver(
    client, db_session
):
    expected = _seed_multi_request_batch_transactions(db_session)

    response = client.get("/api/inventory/transactions", params={"limit": 2000})
    assert response.status_code == 200, response.text
    listed = {
        row["item_name"]: row["approver_name"]
        for row in response.json()
        if row["item_name"] in expected
    }
    assert listed == expected

    csv_response = client.get(f"/api/inventory/transactions/export.csv?{_range()}")
    assert csv_response.status_code == 200, csv_response.text
    csv_rows = csv.DictReader(StringIO(csv_response.text))
    exported_csv = {
        row["item_name"]: row["approver_name"]
        for row in csv_rows
        if row["item_name"] in expected
    }
    assert exported_csv == expected

    xlsx_response = client.get(f"/api/inventory/transactions/export.xlsx?{_range()}")
    assert xlsx_response.status_code == 200, xlsx_response.text
    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    exported_xlsx = {
        row[3]: row[11]
        for row in rows[1:]
        if row[3] in expected
    }
    assert exported_xlsx == expected


def test_export_csv_search_matches_requester_name(client, db_session):
    _seed_batch_transaction(db_session)
    resp = client.get(f"/api/inventory/transactions/export.csv?search=요청자A&{_range()}")
    assert resp.status_code == 200, resp.text
    assert "ExportItem" in resp.text  # requester_name 검색으로 매칭됨


def test_export_csv_search_ignores_spaces_and_separators(client, db_session):
    item = _seed_batch_transaction(db_session)
    db_session.query(TransactionLog).filter(TransactionLog.item_id == item.item_id).one().reference_no = "TX- 12/3"
    db_session.commit()

    response = client.get(f"/api/inventory/transactions/export.csv?search=tx123&{_range()}")

    assert response.status_code == 200, response.text
    assert "ExportItem" in response.text


def test_export_csv_includes_requester_and_approver_from_stock_request_reference(client, db_session):
    requester = _emp(db_session, "EXP_DREQ", "DirectRequester")
    approver = _emp(db_session, "EXP_DAPP", "DirectApprover")
    item = Item(
        item_name="DirectExportItem", process_type_code="TR", model_symbol="9", serial_no=1
    )
    db_session.add(item)
    db_session.flush()

    sr = StockRequest(
        request_code="SR-DIRECT-EXPORT",
        requester_employee_id=requester.employee_id,
        requester_name=requester.name,
        requester_department=DepartmentEnum.ASSEMBLY.value,
        request_type=StockRequestTypeEnum.WAREHOUSE_TO_DEPT,
        status=StockRequestStatusEnum.COMPLETED,
        approved_by_name=approver.name,
        approved_by_employee_id=approver.employee_id,
        approved_at=datetime.utcnow(),
        submitted_at=datetime.utcnow(),
    )
    db_session.add(sr)
    db_session.flush()

    db_session.add(
        TransactionLog(
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_TO_PROD,
            quantity_change=Decimal("0"),
            reference_no=sr.request_code,
            created_at=datetime.utcnow(),
        )
    )
    db_session.commit()

    resp = client.get(f"/api/inventory/transactions/export.csv?{_range()}")
    assert resp.status_code == 200, resp.text
    assert "DirectRequester" in resp.text
    assert "DirectApprover" in resp.text


def test_export_xlsx_ok(client, db_session):
    _seed_batch_transaction(db_session)
    resp = client.get(f"/api/inventory/transactions/export.xlsx?{_range()}")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats"
    )


def test_export_xlsx_search_ignores_spaces_and_separators(client, db_session):
    item = _seed_batch_transaction(db_session)
    db_session.query(TransactionLog).filter(TransactionLog.item_id == item.item_id).one().reference_no = "XLSX- 45/6"
    other = Item(
        item_name="OtherExportItem",
        process_type_code="TR",
        model_symbol="9",
        serial_no=2,
    )
    db_session.add(other)
    db_session.flush()
    db_session.add(
        TransactionLog(
            item_id=other.item_id,
            transaction_type=TransactionTypeEnum.RECEIVE,
            quantity_change=Decimal("1"),
            reference_no="OTHER-999",
            created_at=datetime.utcnow(),
        )
    )
    db_session.commit()

    response = client.get(f"/api/inventory/transactions/export.xlsx?search=xlsx456&{_range()}")

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert [(row[3], row[8]) for row in rows[1:]] == [("ExportItem", "XLSX- 45/6")]


def test_export_xlsx_uses_dynamic_internal_use_label(client, db_session):
    item = Item(
        item_name="연구 반출품", process_type_code="TR", model_symbol="9", serial_no=1
    )
    db_session.add(item)
    db_session.flush()
    db_session.add(
        TransactionLog(
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.INTERNAL_USE,
            quantity_change=Decimal("-1"),
            quantity_before=Decimal("2"),
            quantity_after=Decimal("1"),
            department=DepartmentEnum.RESEARCH.value,
            created_at=datetime.utcnow(),
        )
    )
    db_session.commit()

    resp = client.get(f"/api/inventory/transactions/export.xlsx?{_range()}")
    assert resp.status_code == 200, resp.text
    workbook = load_workbook(BytesIO(resp.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any(row[1] == "연구소 반출" for row in rows[1:])
