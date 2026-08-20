from __future__ import annotations

import csv
import uuid
from contextlib import contextmanager
from datetime import datetime
from io import BytesIO, StringIO
from types import SimpleNamespace
from urllib.parse import quote

from sqlalchemy import event
from openpyxl import load_workbook
from fastapi import FastAPI
from fastapi.responses import Response
from starlette.requests import Request


ADMIN_HEADERS = {"X-Admin-Pin": "0000"}
EXPORT_HEADERS = [
    "일시(KST)",
    "직원명",
    "사번",
    "단말명",
    "접속유형",
    "화면",
    "작업",
    "결과",
    "대상/변경 요약",
    "세션 ID",
    "요청 ID",
    "관련 ID",
]


def _add_employee(db_session, *, code: str = "E22", name: str = "홍길동") -> None:
    from app.models import Employee
    from app.services.pin_auth import hash_pin

    db_session.add(
        Employee(
            employee_code=code,
            name=name,
            role="조립",
            department="조립",
            is_active=True,
            pin_hash=hash_pin("2468"),
            pin_requires_change=False,
        )
    )
    db_session.commit()


def _observe_realtime_commits(db_session):
    from app.models import DataRevision
    from app.services import realtime

    db_session.add(DataRevision(id=1, revision=0))
    db_session.commit()
    event.listen(db_session, "before_commit", realtime._advance_revision_before_commit)


def _stop_observing_realtime_commits(db_session):
    from app.services import realtime

    event.remove(db_session, "before_commit", realtime._advance_revision_before_commit)


def test_client_event_persists_actor_and_terminal_snapshots(client, db_session):
    from app.models import ActivityAuditLog, AuditTerminal, Employee

    _add_employee(db_session)
    employee = db_session.query(Employee).filter(Employee.employee_code == "E22").one()
    terminal_id = str(uuid.uuid4())
    db_session.add(AuditTerminal(terminal_id=terminal_id, name="조립 PC"))
    db_session.commit()

    verified = client.post(
        f"/api/employees/{employee.employee_id}/verify-pin",
        headers={"X-MES-Audit-Session": "session-1"},
        json={"pin": "2468"},
    )
    assert verified.status_code == 200

    response = client.post(
        "/api/client-events",
        headers={"X-Request-Id": "req-client-1"},
        json={
            "event": "ui_nav",
            "source": "desktop",
            "session_id": "session-1",
            "terminal_id": terminal_id,
            "screen_key": "inventory",
            "screen_label": "재고 현황",
            "action_key": "open_history",
            "action_label": "이력 열기",
            "target_summary": "품목 DX-1",
            "related_id": "item-1",
        },
    )

    assert response.status_code == 204
    rows = (
        db_session.query(ActivityAuditLog)
        .filter(ActivityAuditLog.action_key == "open_history")
        .all()
    )
    assert len(rows) == 1
    row = rows[0]
    assert row.actor_employee_name == "홍길동"
    assert row.actor_employee_code == "E22"
    assert row.terminal_id == terminal_id
    assert row.terminal_name == "조립 PC"
    assert row.source == "desktop"
    assert row.session_id == "session-1"
    assert row.screen_key == "inventory"
    assert row.action_label == "이력 열기"
    assert row.outcome == "success"
    assert row.request_id == "req-client-1"
    assert row.related_id == "item-1"


def test_client_event_persists_without_advancing_realtime_revision(client, db_session):
    from app.models import ActivityAuditLog, DataRevision

    _observe_realtime_commits(db_session)
    try:
        response = client.post(
            "/api/client-events",
            json={"event": "ui_nav", "source": "desktop", "to": "desktop.inventory"},
        )
    finally:
        _stop_observing_realtime_commits(db_session)

    assert response.status_code == 204
    assert db_session.query(ActivityAuditLog).count() == 1
    assert db_session.get(DataRevision, 1).revision == 0


def test_client_events_require_session_and_reject_forged_employee_headers(
    auth_client,
    db_session,
):
    from app.models import ActivityAuditLog, Employee

    _add_employee(db_session, code="E22", name="탭 A 작업자")
    _add_employee(db_session, code="E23", name="탭 B 작업자")
    employee = db_session.query(Employee).filter(Employee.employee_code == "E23").one()
    session_id = str(uuid.uuid4())

    anonymous = auth_client.post(
        "/api/client-events",
        headers={"X-MES-Employee-Code": "E23", "X-MES-Audit-Session": session_id},
        json={"event": "ui_nav", "source": "desktop", "session_id": session_id},
    )
    assert anonymous.status_code == 401
    assert db_session.query(ActivityAuditLog).count() == 0

    verified = auth_client.post(
        "/api/operator-session",
        json={"employee_id": str(employee.employee_id), "pin": "2468"},
    )
    assert verified.status_code == 200
    before_forged = db_session.query(ActivityAuditLog).count()

    forged = auth_client.post(
        "/api/client-events",
        headers={
            "X-MES-Employee-Code": "E22",
            "X-MES-Audit-Session": session_id,
        },
        json={"event": "ui_nav", "source": "desktop", "session_id": session_id},
    )
    assert forged.status_code == 403
    assert db_session.query(ActivityAuditLog).count() == before_forged

    recorded = auth_client.post(
        "/api/client-events",
        headers={
            "X-MES-Employee-Code": "E23",
            "X-MES-Audit-Session": session_id,
        },
        json={"event": "ui_nav", "source": "desktop", "session_id": session_id},
    )
    assert recorded.status_code == 204
    row = (
        db_session.query(ActivityAuditLog)
        .filter(ActivityAuditLog.action_key == "ui_nav")
        .order_by(ActivityAuditLog.occurred_at.desc())
        .first()
    )
    assert row.actor_employee_code == "E23"
    assert row.actor_employee_name == "탭 B 작업자"

    logged_out = auth_client.post(
        "/api/client-events",
        headers={
            "X-MES-Employee-Code": "E23",
            "X-MES-Audit-Session": session_id,
        },
        json={"event": "ui_logout", "source": "desktop", "session_id": session_id},
    )
    assert logged_out.status_code == 204
    logout_row = (
        db_session.query(ActivityAuditLog)
        .filter(ActivityAuditLog.action_key == "ui_logout")
        .one()
    )
    assert logout_row.actor_employee_code == "E23"
    assert logout_row.actor_employee_name == "탭 B 작업자"
    deleted = auth_client.delete("/api/operator-session")
    assert deleted.status_code == 204
    after_logout = auth_client.post(
        "/api/client-events",
        headers={"X-MES-Employee-Code": "E23", "X-MES-Audit-Session": session_id},
        json={"event": "ui_nav", "source": "desktop", "session_id": session_id},
    )
    assert after_logout.status_code == 401
    assert (
        db_session.query(ActivityAuditLog)
        .filter(ActivityAuditLog.action_key == "ui_nav")
        .count()
        == 1
    )


def test_client_event_rejects_secret_without_db_persistence(client, db_session):
    from app.models import ActivityAuditLog

    response = client.post(
        "/api/client-events",
        json={"event": "ui_action_cancel", "source": "mobile", "access_token": "never-store-me"},
    )

    assert response.status_code == 422
    assert db_session.query(ActivityAuditLog).count() == 0


def test_client_event_without_source_keeps_desktop_compatibility(client, db_session):
    from app.models import ActivityAuditLog

    response = client.post("/api/client-events", json={"event": "ui_logout"})

    assert response.status_code == 204
    assert db_session.query(ActivityAuditLog).one().source == "desktop"


def test_navigation_event_persists_the_destination_screen(client, db_session):
    from app.models import ActivityAuditLog

    response = client.post(
        "/api/client-events",
        json={
            "event": "ui_nav",
            "source": "desktop",
            "from": "desktop.history",
            "to": "desktop.weekly",
            "screen_key": "desktop.weekly",
            "screen_label": "weekly",
        },
    )

    assert response.status_code == 204
    row = db_session.query(ActivityAuditLog).one()
    assert row.screen_key == "desktop.weekly"
    assert row.screen_label == "weekly"


def test_write_audit_skips_unverified_and_unregistered_requests(
    client, db_session
):
    from app.models import ActivityAuditLog

    common_headers = {
        "X-MES-Audit-Session": "write-session",
        "X-MES-Audit-Screen": "inventory",
        "X-MES-Audit-Screen-Label": quote("재고 현황"),
    }
    ok = client.post("/api/health/write-check", headers=common_headers)
    missing = client.post("/api/not-a-real-write", headers=common_headers)
    invalid = client.post("/api/io/submit", headers=common_headers, json={})
    long_missing = client.post(f"/api/{'x' * 300}", headers=common_headers)
    event = client.post(
        "/api/client-events",
        json={"event": "ui_logout", "source": "desktop", "session_id": "write-session"},
    )

    assert ok.status_code == 200
    assert missing.status_code == 404
    assert invalid.status_code == 422
    assert long_missing.status_code == 404
    assert event.status_code == 204
    rows = db_session.query(ActivityAuditLog).order_by(ActivityAuditLog.occurred_at).all()
    assert len(rows) == 2
    generic = [row for row in rows if row.action_key != "ui_logout"]
    assert [row.outcome for row in generic] == ["failed"]
    assert all(row.session_id == "write-session" for row in generic)
    assert all(row.screen_label == "재고 현황" for row in generic)
    assert all(len(row.action_key) <= 160 for row in generic)
    assert sum(row.action_key == "ui_logout" for row in rows) == 1


def test_anonymous_write_failures_do_not_mutate_activity_audit(
    auth_client,
    db_session,
) -> None:
    from app.models import ActivityAuditLog

    before = db_session.query(ActivityAuditLog).count()

    missing = auth_client.post("/api/not-a-real-write")
    invalid = auth_client.post("/api/io/submit", json={})

    assert missing.status_code == 404
    assert invalid.status_code == 401
    assert db_session.query(ActivityAuditLog).count() == before


def test_unhandled_write_error_is_persisted_as_failed_audit(
    client, db_session, monkeypatch
):
    from fastapi.testclient import TestClient
    from app.main import app
    from app.models import ActivityAuditLog, Employee
    from app.routers import io as io_router

    _add_employee(db_session, code="AUDIT-ERR-01", name="감사 오류 작업자")
    actor = (
        db_session.query(Employee)
        .filter(Employee.employee_code == "AUDIT-ERR-01")
        .one()
    )

    def raise_unhandled(_db, _payload, *, requester):
        assert requester.employee_id == actor.employee_id
        raise RuntimeError("unexpected failure")

    monkeypatch.setattr(io_router.io_actions_svc, "submit", raise_unhandled)
    with TestClient(app, raise_server_exceptions=False) as nonraising_client:
        response = nonraising_client.post(
            "/api/io/submit",
            headers={"X-MES-Audit-Session": "unhandled-session"},
            json={
                "requester_employee_id": str(actor.employee_id),
                "work_type": "process",
                "sub_type": "produce",
                "bundles": [],
            },
        )

    assert response.status_code == 500
    row = db_session.query(ActivityAuditLog).one()
    assert row.outcome == "failed"
    assert row.session_id == "unhandled-session"


def test_terminal_upsert_requires_pin_and_records_current_name(client, db_session):
    from app.models import ActivityAuditLog, AuditTerminal

    terminal_id = str(uuid.uuid4())
    payload = {"terminal_id": terminal_id, "name": "출하 PC"}

    denied = client.put("/api/admin/activity-audit/terminals/current", json=payload)
    created = client.put(
        "/api/admin/activity-audit/terminals/current",
        headers=ADMIN_HEADERS,
        json=payload,
    )
    updated = client.put(
        "/api/admin/activity-audit/terminals/current",
        headers=ADMIN_HEADERS,
        json={"terminal_id": terminal_id, "name": "출하 PC 2"},
    )

    assert denied.status_code == 400
    assert created.status_code == 200
    assert updated.status_code == 200
    assert updated.json() == {"terminal_id": terminal_id, "name": "출하 PC 2"}
    terminal = db_session.get(AuditTerminal, terminal_id)
    assert terminal is not None
    assert terminal.name == "출하 PC 2"
    successful = (
        db_session.query(ActivityAuditLog)
        .filter(ActivityAuditLog.action_key == "http.put.admin.activity-audit.terminals.current")
        .filter(ActivityAuditLog.outcome == "success")
        .all()
    )
    assert len(successful) == 2
    assert successful[-1].terminal_name == "출하 PC 2"


def test_terminal_upsert_does_not_advance_realtime_revision(client, db_session):
    from app.models import AuditTerminal, DataRevision, SystemSetting
    from app.services.pin_auth import hash_pin

    db_session.add(SystemSetting(setting_key="admin_pin", setting_value=hash_pin("0000")))
    db_session.commit()
    _observe_realtime_commits(db_session)
    terminal_id = str(uuid.uuid4())
    try:
        response = client.put(
            "/api/admin/activity-audit/terminals/current",
            headers=ADMIN_HEADERS,
            json={"terminal_id": terminal_id, "name": "감사 전용 PC"},
        )
    finally:
        _stop_observing_realtime_commits(db_session)

    assert response.status_code == 200, response.text
    assert db_session.get(AuditTerminal, terminal_id) is not None
    assert db_session.get(DataRevision, 1).revision == 0


def test_background_write_audit_does_not_advance_realtime_revision(
    db_session,
    monkeypatch,
):
    from app import _access_log
    from app.models import ActivityAuditLog, DataRevision

    _observe_realtime_commits(db_session)

    @contextmanager
    def _fixed_audit_session(_app):
        yield db_session

    monkeypatch.setattr(_access_log, "_audit_session", _fixed_audit_session)
    entry = _access_log._WriteAuditEntry(
        method="POST",
        path="/api/items",
        status=201,
        source="desktop",
        terminal_id=None,
        session_id="audit-session",
        screen_key="items",
        screen_label="품목",
        actor_employee_code=None,
        request_id="audit-request",
        action_key="http.post.items",
        action_label="POST /api/items",
        target_summary="/api/items",
        related_id=None,
    )
    try:
        _access_log._record_write_audit(object(), entry)
    finally:
        _stop_observing_realtime_commits(db_session)

    assert db_session.query(ActivityAuditLog).count() == 1
    assert db_session.get(DataRevision, 1).revision == 0


def test_write_action_metadata_normalizes_path_ids():
    from app._access_log import _write_action_metadata

    action_key, action_label, related_id, target_summary = _write_action_metadata(
        "DELETE",
        "/api/notifications/36f7d993-b2d9-4f5d-a0d4-6cc8b4e9a685",
    )

    assert action_key == "http.delete.notifications.id"
    assert action_label == "DELETE /api/notifications/:id"
    assert related_id == "36f7d993-b2d9-4f5d-a0d4-6cc8b4e9a685"
    assert target_summary == "/api/notifications/:id"


def test_write_audit_is_attached_to_the_response_background(client):
    from app._access_log import _schedule_write_audit

    app = FastAPI()
    request = Request(
        {
            "type": "http",
            "app": app,
            "method": "POST",
            "path": "/api/health/write-check",
            "headers": [],
        }
    )
    response = Response()
    request.state.actor_emp = "AUDIT-BACKGROUND-ACTOR"

    _schedule_write_audit(response, request, method="POST", path="/api/health/write-check", status=200)

    assert response.background is not None


def test_io_submit_exposes_the_related_batch_for_request_audit(monkeypatch):
    from app.routers import io as io_router

    request = SimpleNamespace(state=SimpleNamespace())
    actor = SimpleNamespace(
        employee_id=uuid.uuid4(),
        employee_code="AUDIT-DIRECT-01",
        name="직접 호출 작업자",
    )
    payload = SimpleNamespace(requester_employee_id=actor.employee_id)
    batch_id = str(uuid.uuid4())
    monkeypatch.setattr(
        io_router.io_actions_svc,
        "submit",
        lambda _db, _payload, *, requester: {
            "batch": {
                "batch_id": batch_id,
                "work_type": "process",
                "sub_type": "produce",
                "bundles": [],
            },
            "requires_approval": False,
        },
    )
    monkeypatch.setattr(io_router, "_evt_emit", lambda *_args, **_kwargs: None)

    io_router.submit_io(payload, request, actor, object())

    assert request.state.activity_audit_related_id == batch_id
    assert request.state.activity_audit_target_summary == "process · produce"


def test_monthly_exports_are_kst_sorted_and_have_exact_columns(client, db_session):
    from app.models import ActivityAuditLog

    db_session.add_all(
        [
            ActivityAuditLog(
                occurred_at=datetime(2026, 4, 30, 15, 1),
                actor_employee_name="두번째",
                actor_employee_code="E02",
                terminal_name="미등록 단말",
                source="mobile",
                screen_label="출하",
                action_key="later",
                action_label="완료",
                outcome="success",
                session_id="s2",
                request_id="r2",
            ),
            ActivityAuditLog(
                occurred_at=datetime(2026, 4, 30, 15, 0),
                actor_employee_name="첫번째",
                actor_employee_code="E01",
                terminal_name="미등록 단말",
                source="desktop",
                screen_label="재고",
                action_key="first",
                action_label="조회",
                outcome="failed",
                target_summary="품목 A",
                session_id="s1",
                request_id="r1",
                related_id="item-a",
            ),
            ActivityAuditLog(
                occurred_at=datetime(2026, 5, 31, 15, 0),
                terminal_name="미등록 단말",
                source="desktop",
                action_key="next-month",
                action_label="다음 달",
                outcome="success",
            ),
        ]
    )
    db_session.commit()

    listing = client.get("/api/admin/activity-audit/files", headers=ADMIN_HEADERS)
    csv_response = client.get(
        "/api/admin/activity-audit/2026-05.csv", headers=ADMIN_HEADERS
    )
    xlsx_response = client.get(
        "/api/admin/activity-audit/2026-05.xlsx", headers=ADMIN_HEADERS
    )

    assert listing.status_code == 200
    may = next(row for row in listing.json() if row["month"] == "2026-05")
    assert may["row_count"] == 2
    rows = list(csv.reader(StringIO(csv_response.content.decode("utf-8-sig"))))
    assert rows[0] == EXPORT_HEADERS
    assert [row[1] for row in rows[1:]] == ["첫번째", "두번째"]
    assert rows[1][0] == "2026-05-01 00:00:00"
    assert rows[1][4] == "데스크톱"
    assert rows[1][7] == "실패"

    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
    xlsx_rows = list(workbook.active.iter_rows(values_only=True))
    assert list(xlsx_rows[0]) == EXPORT_HEADERS
    assert [row[1] for row in xlsx_rows[1:]] == ["첫번째", "두번째"]


def test_activity_audit_exports_escape_spreadsheet_formulas(client, db_session):
    from app.models import ActivityAuditLog

    db_session.add(
        ActivityAuditLog(
            occurred_at=datetime(2026, 4, 30, 15, 0),
            actor_employee_name="=HYPERLINK(\"https://unsafe.example\")",
            terminal_name="미등록 단말",
            source="desktop",
            action_key="ui_nav",
            action_label="=2+2",
            outcome="success",
            target_summary="@unsafe",
        )
    )
    db_session.commit()

    csv_response = client.get(
        "/api/admin/activity-audit/2026-05.csv", headers=ADMIN_HEADERS
    )
    xlsx_response = client.get(
        "/api/admin/activity-audit/2026-05.xlsx", headers=ADMIN_HEADERS
    )

    csv_row = list(csv.reader(StringIO(csv_response.content.decode("utf-8-sig"))))[1]
    assert csv_row[1] == "'=HYPERLINK(\"https://unsafe.example\")"
    assert csv_row[6] == "'=2+2"
    assert csv_row[8] == "'@unsafe"

    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=False)
    assert workbook.active.cell(2, 2).data_type == "s"
    assert workbook.active.cell(2, 2).value == "'=HYPERLINK(\"https://unsafe.example\")"


def test_activity_audit_exports_require_pin_and_have_no_backfill(client):
    assert client.get("/api/admin/activity-audit/files").status_code == 400
    assert client.get("/api/admin/activity-audit/2026-05.csv").status_code == 400
    assert client.get("/api/admin/activity-audit/2026-05.xlsx").status_code == 400
    assert (
        client.post("/api/admin/activity-audit/backfill", headers=ADMIN_HEADERS).status_code
        == 404
    )
