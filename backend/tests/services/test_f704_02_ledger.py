"""F704-02 XML 패키지 렌더링 단위 테스트."""

from __future__ import annotations

import os
import re
import subprocess
import zipfile
from datetime import date, datetime
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from app.services.f704_02_ledger import F704LedgerEntry, TEMPLATE_PATH, _remark, _requester, render_workbook


_IGNORABLE_ATTRIBUTE_RE = re.compile(rb"(?:[A-Za-z_][\w.-]*:)?Ignorable=[\"']([^\"']+)[\"']")
_NAMESPACE_DECLARATION_RE = re.compile(rb"xmlns:([A-Za-z_][\w.-]*)=[\"']")


def _missing_ignorable_namespace_prefixes(worksheet_xml: bytes) -> set[str]:
    worksheet_start = worksheet_xml.index(b"<worksheet")
    root_tag = worksheet_xml[worksheet_start : worksheet_xml.index(b">", worksheet_start) + 1]
    ignored = _IGNORABLE_ATTRIBUTE_RE.search(root_tag)
    assert ignored is not None
    declared = {prefix.decode("ascii") for prefix in _NAMESPACE_DECLARATION_RE.findall(root_tag)}
    return set(ignored.group(1).decode("ascii").split()) - declared


def _entry(sequence: int) -> F704LedgerEntry:
    return F704LedgerEntry(
        occurred_on=date(2026, 1, 1),
        created_at=datetime(2026, 1, 1, 9, 0),
        log_id=f"log-{sequence}",
        item_code=f"9-TR-{sequence:04d}",
        item_name=f"확장 품목 {sequence}",
        quantity=1,
        direction="입고",
        counterpart="외부입고",
        requester="요청자",
        remark="",
    )


def test_render_workbook_extends_template_rows_with_same_data_style():
    """원본 양식의 사전 서식 행을 넘으면 마지막 행 서식으로 연장한다."""
    rendered = render_workbook(_entry(index) for index in range(1, 1935))

    worksheet = load_workbook(BytesIO(rendered), data_only=False)["양식"]
    assert worksheet["A1937"].value == 1934
    assert worksheet["E1937"].value == "확장 품목 1934"
    assert worksheet["F1937"].value == 1
    assert worksheet.auto_filter.ref == "A3:K1937"
    assert worksheet["F1937"].style_id == worksheet["F1936"].style_id


def test_render_workbook_opens_ledger_at_first_data_row():
    rendered = render_workbook([_entry(1)])

    worksheet = load_workbook(BytesIO(rendered), data_only=False)["양식"]
    assert worksheet.sheet_view.pane is not None
    assert worksheet.sheet_view.pane.ySplit == 3
    assert worksheet.sheet_view.pane.topLeftCell == "A4"
    bottom_selection = next(
        selection for selection in worksheet.sheet_view.selection if selection.pane == "bottomLeft"
    )
    assert bottom_selection.activeCell == "A4"
    assert bottom_selection.sqref == "A4"


def test_template_and_rendered_worksheets_declare_ignored_namespaces():
    rendered = render_workbook([_entry(1)])

    for package in (TEMPLATE_PATH, BytesIO(rendered)):
        with zipfile.ZipFile(package) as archive:
            for worksheet in ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"):
                assert _missing_ignorable_namespace_prefixes(archive.read(worksheet)) == set()


def test_requester_uses_io_batch_then_stock_request_then_shipping_request():
    batch = SimpleNamespace(requester_name="IO requester")
    stock_request = SimpleNamespace(requester_name="stock requester")
    shipping_request = SimpleNamespace(requested_by_name="shipping requester")

    assert _requester(batch, stock_request, shipping_request) == "IO requester"
    assert _requester(SimpleNamespace(requester_name=" "), stock_request, shipping_request) == "stock requester"
    assert _requester(None, None, shipping_request) == "shipping requester"


def test_remark_uses_original_work_memo_fields_and_ignores_system_note():
    assert _remark(SimpleNamespace(notes=" 작업 메모\n"), None, None) == "작업 메모"
    assert _remark(None, SimpleNamespace(notes=" 요청 메모 "), None) == "요청 메모"
    assert _remark(None, None, SimpleNamespace(notes=" 출하 메모 ")) == "출하 메모"
    assert _remark(None, None, None) == ""


def test_remark_hides_daily_development_notes():
    hidden_notes = (
        "DEV-DAILY-20260803: assembly direct warehouse issue",
        "DEV-DAILY-20260803: assembly BOM component issue",
        "DEV-DAILY-20260803: high voltage direct warehouse issue",
        "DEV-DAILY-20260803: high voltage direct warehouse return",
    )

    for note in hidden_notes:
        assert _remark(SimpleNamespace(notes=note), None, None) == ""


def test_template_keeps_both_forms_but_has_no_ledger_values():
    workbook = load_workbook(TEMPLATE_PATH, data_only=False)

    assert workbook.sheetnames == ["양식", "양식_출력용"]
    for worksheet in workbook.worksheets:
        assert [cell.value for cell in worksheet[3][:11]] == [
            "No",
            "날 짜",
            "발주번호",
            "품번",
            "품  명",
            "수량",
            "입고\n출고",
            "입/출고처",
            "담당자",
            "검수",
            "비고",
        ]
        assert all(cell.value is None for cell in worksheet[3][11:])
        assert worksheet["F3"].alignment.horizontal == "center"
        assert worksheet["A61"].fill.fill_type is None
        assert worksheet["L61"].fill.fill_type is None
        assert not any(
            worksheet[f"E{row}"].alignment.shrinkToFit
            for row in range(4, worksheet.max_row + 1)
        )
        assert all(
            worksheet[f"F{row}"].number_format == "#,##0"
            for row in range(4, worksheet.max_row + 1)
        )
        for row in worksheet.iter_rows(min_row=4, max_col=11):
            assert all(cell.value is None for cell in row[1:])


@pytest.mark.skipif(os.name != "nt", reason="Microsoft Excel COM 검증은 Windows에서만 실행한다.")
def test_excel_opens_template_and_rendered_workbook(tmp_path):
    """Excel이 원본 템플릿과 생성 결과를 복구 없이 직접 열 수 있어야 한다."""
    rendered_path = tmp_path / "f704-rendered.xlsx"
    rendered_path.write_bytes(render_workbook([_entry(1)]))
    script = """
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
try {
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open($env:F704_XLSX_PATH)
    $book.Close($false)
} finally {
    $excel.Quit()
    [Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) | Out-Null
}
"""
    for path in (TEMPLATE_PATH, rendered_path):
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", script],
            check=False,
            capture_output=True,
            env={**os.environ, "F704_XLSX_PATH": str(path)},
        )
        assert result.returncode == 0, (result.stderr or b"").decode("utf-8", "replace")


@pytest.mark.skipif(os.name != "nt", reason="Microsoft Excel COM 검증은 Windows에서만 실행한다.")
def test_excel_opens_later_template_rows_when_quantity_is_written(tmp_path):
    """빈 수량 셀이 있는 후반 행에도 Excel 순서대로 셀을 써야 한다."""
    rendered_path = tmp_path / "f704-later-row.xlsx"
    rendered_path.write_bytes(render_workbook(_entry(index) for index in range(1, 242)))
    script = """
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
try {
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open($env:F704_XLSX_PATH)
    $book.Close($false)
} finally {
    $excel.Quit()
    [Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) | Out-Null
}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        check=False,
        capture_output=True,
        env={**os.environ, "F704_XLSX_PATH": str(rendered_path)},
    )
    assert result.returncode == 0, (result.stderr or b"").decode("utf-8", "replace")


@pytest.mark.skipif(os.name != "nt", reason="Microsoft Excel COM 검증은 Windows에서만 실행한다.")
def test_excel_template_has_no_fill_past_last_visible_column():
    """61행의 행 서식은 O열 이후 빈 영역에도 채우기를 남기면 안 된다."""
    script = """
$ErrorActionPreference = 'Stop'
$excel = New-Object -ComObject Excel.Application
try {
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $book = $excel.Workbooks.Open($env:F704_XLSX_PATH)
    foreach ($sheet in @($book.Worksheets)) {
        foreach ($cell in @('O61', 'XFD61')) {
            if ($sheet.Range($cell).Interior.Pattern -ne -4142) {
                throw "Unexpected fill at $($sheet.Name)!$cell"
            }
        }
    }
    $book.Close($false)
} finally {
    $excel.Quit()
    [Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel) | Out-Null
}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        check=False,
        capture_output=True,
        env={**os.environ, "F704_XLSX_PATH": str(TEMPLATE_PATH)},
    )
    assert result.returncode == 0, (result.stderr or b"").decode("utf-8", "replace")
