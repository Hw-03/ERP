"""F705-02 daily production aggregation and template rendering tests."""

from __future__ import annotations

import uuid
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.models import ProductSymbol, TransactionLog, TransactionTypeEnum
from app.services.f705_02_production_log import collect_daily_quantities, render_workbook


def _add_log(
    db_session,
    item,
    *,
    quantity: int,
    occurred_at: datetime,
    transaction_type: TransactionTypeEnum = TransactionTypeEnum.PRODUCE,
) -> None:
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=transaction_type,
            quantity_change=Decimal(str(quantity)),
            quantity_before=Decimal("0"),
            quantity_after=Decimal(str(quantity)),
            created_at=occurred_at,
        )
    )


def test_collect_daily_quantities_uses_produce_only_and_abs_item_day_sum(db_session, make_item):
    db_session.add_all(
        [
            ProductSymbol(slot=3, symbol="3", model_name="DX3000"),
            ProductSymbol(slot=8, symbol="8", model_name="SOLO"),
        ]
    )
    hf_dx = make_item(name="HF DX", process_type_code="HF", model_symbol="3")
    af_solo = make_item(name="AF SOLO", process_type_code="AF", model_symbol="8")
    tf_dx = make_item(name="TF DX", process_type_code="TF", model_symbol="3")
    unknown = make_item(name="unknown", process_type_code="HF", model_symbol="X")
    _add_log(db_session, hf_dx, quantity=10, occurred_at=datetime(2026, 1, 3, 9, 0))
    _add_log(db_session, hf_dx, quantity=-3, occurred_at=datetime(2026, 1, 3, 14, 0))
    _add_log(db_session, hf_dx, quantity=99, occurred_at=datetime(2026, 1, 3, 15, 0), transaction_type=TransactionTypeEnum.RECEIVE)
    _add_log(db_session, af_solo, quantity=4, occurred_at=datetime(2026, 1, 3, 10, 0))
    _add_log(db_session, tf_dx, quantity=88, occurred_at=datetime(2026, 1, 3, 10, 0))
    _add_log(db_session, unknown, quantity=77, occurred_at=datetime(2026, 1, 3, 10, 0))
    db_session.commit()

    quantities = collect_daily_quantities(db_session, 2026)

    assert quantities == {
        date(2026, 1, 3): {
            ("HF", "DX3000"): 7,
            ("AF", "SOLO"): 4,
        }
    }


def test_render_workbook_keeps_compact_template_formulas_and_blank_zero_inputs():
    rendered = render_workbook(
        2026,
        {
            date(2026, 3, 2): {
                ("HF", "DX3000"): 8,
                ("AF", "SOLO"): 4,
                ("PF", "COCOON"): 3,
            }
        },
    )

    workbook = load_workbook(BytesIO(rendered), data_only=False)
    worksheet = workbook["26.03"]

    assert workbook.sheetnames == [f"26.{month:02d}" for month in range(1, 13)]
    assert worksheet["D1"].value.date() == date(2026, 3, 1)
    assert worksheet["E3"].value == 8
    assert worksheet["E25"].value == 4
    assert worksheet["E30"].value == 3
    assert worksheet["D3"].value is None
    assert worksheet["AI3"].value == "=SUM(D3:AH3)"
    assert worksheet["D8"].value == "=SUM(D3:D7)"
    assert worksheet["AI32"].value == "=SUM(D32:AH32)"
    assert worksheet["AJ21"].value == 250
    assert "A21:A26" in {str(cell_range) for cell_range in worksheet.merged_cells.ranges}
    assert "A27:A32" in {str(cell_range) for cell_range in worksheet.merged_cells.ranges}

    cached_workbook = load_workbook(BytesIO(rendered), data_only=True)
    cached_worksheet = cached_workbook["26.03"]
    assert cached_worksheet["AI3"].value == 8
    assert cached_worksheet["D8"].value == 0
    assert cached_worksheet["E8"].value == 8
    assert cached_worksheet["AI8"].value == 8
    assert cached_workbook["26.02"]["AF3"].value is None
    assert cached_workbook["26.02"]["AG3"].value is None
    assert cached_workbook["26.02"]["AH3"].value is None

    holiday_formulas = [
        formula
        for conditional in worksheet.conditional_formatting
        for rule in conditional.rules
        for formula in rule.formula
        if "DATE(2026,3,2)" in formula
    ]
    assert holiday_formulas


def test_render_workbook_maps_all_processes_and_models_to_the_contract_rows():
    models = ("DX3000", "ADX4000W", "ADX6000FB", "COCOON", "SOLO")
    expected_rows = {
        "HF": (3, 4, 5, 6, 7, 8),
        "VF": (9, 10, 11, 12, 13, 14),
        "NF": (15, 16, 17, 18, 19, 20),
        "AF": (21, 22, 23, 24, 25, 26),
        "PF": (27, 28, 29, 30, 31, 32),
    }
    quantities = {
        (process, model): process_index * 10 + model_index
        for process_index, process in enumerate(expected_rows, start=1)
        for model_index, model in enumerate(models, start=1)
    }

    rendered = render_workbook(2026, {date(2026, 1, 2): quantities})
    workbook = load_workbook(BytesIO(rendered), data_only=False)
    cached_workbook = load_workbook(BytesIO(rendered), data_only=True)
    worksheet = workbook["26.01"]
    cached_worksheet = cached_workbook["26.01"]

    for process_index, (process, rows) in enumerate(expected_rows.items(), start=1):
        model_rows, total_row = rows[:5], rows[5]
        expected_values = [process_index * 10 + model_index for model_index in range(1, 6)]
        assert [worksheet.cell(row, 5).value for row in model_rows] == expected_values
        assert worksheet.cell(total_row, 5).value == f"=SUM(E{model_rows[0]}:E{model_rows[-1]})"
        assert cached_worksheet.cell(total_row, 5).value == sum(expected_values)

    assert worksheet["B5"].value == "ADX6000 시리즈"
    assert worksheet["B23"].value == worksheet["B29"].value == "ADX6000FB"
