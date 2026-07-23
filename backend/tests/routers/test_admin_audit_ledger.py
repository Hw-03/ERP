"""F704-02 연간 자재 입출고관리대장 다운로드 API 테스트."""

from __future__ import annotations

import hashlib
import uuid
import zipfile
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from urllib.parse import quote

from openpyxl import load_workbook

from app.models import (
    Employee,
    IoBatch,
    IoBundle,
    IoLine,
    ProductSymbol,
    TransactionLog,
    TransactionTypeEnum,
)
from app.services import f704_02_ledger


ADMIN_HEADERS = {"X-Admin-Pin": "0000"}


def _row_layout(worksheet):
    return {
        row: (dimension.height, dimension.hidden, dimension.outlineLevel, dimension.collapsed, dimension.style)
        for row, dimension in worksheet.row_dimensions.items()
    }


def _header_footer_layout(worksheet):
    return tuple(
        tuple((part.text, part.font, part.size, part.color) for part in (header.left, header.center, header.right))
        for header in (
            worksheet.oddHeader,
            worksheet.oddFooter,
            worksheet.evenHeader,
            worksheet.evenFooter,
            worksheet.firstHeader,
            worksheet.firstFooter,
        )
    )


def test_f704_download_requires_admin_pin(client):
    response = client.get("/api/admin/audit-ledger/f704-02.xlsx?year=2026")

    assert response.status_code == 400


def _add_requester_batch_and_line(db_session, item, *, requester: str = "요청자 김") -> IoBatch:
    employee = Employee(
        employee_id=uuid.uuid4(),
        employee_code="REQ-001",
        name=requester,
        role="생산/사원",
        department="생산",
    )
    db_session.add(employee)
    db_session.flush()
    batch = IoBatch(
        batch_id=uuid.uuid4(),
        work_type="move",
        sub_type="transfer",
        status="completed",
        requester_employee_id=employee.employee_id,
        requester_name=requester,
        requester_department="생산",
        from_department="창고",
        to_department="조립",
    )
    db_session.add(batch)
    db_session.flush()
    bundle = IoBundle(
        bundle_id=uuid.uuid4(),
        batch_id=batch.batch_id,
        source_kind="item",
        source_item_id=item.item_id,
        title_snapshot=item.item_name,
        quantity=7,
    )
    db_session.add(bundle)
    db_session.flush()
    db_session.add(
        IoLine(
            line_id=uuid.uuid4(),
            bundle_id=bundle.bundle_id,
            item_id=item.item_id,
            item_name_snapshot=item.item_name,
            mes_code_snapshot=item.mes_code,
            unit="EA",
            direction="move",
            from_bucket="warehouse",
            to_bucket="production",
            to_department="조립",
            quantity=7,
            origin="manual",
        )
    )
    return batch


def test_f704_download_includes_kst_year_warehouse_receive(client, db_session, make_item):
    """UTC 전년도 시각이라도 KST 새해 창고 입고는 해당 연도 대장에 담긴다."""
    item = make_item(name="KST 입고 품목")
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.RECEIVE,
            quantity_change=Decimal("5"),
            quantity_before=Decimal("0"),
            quantity_after=Decimal("5"),
            reference_no="RCV-2026-001",
            produced_by="처리자여야-아님",
            inventory_effect=[{"scope": "warehouse", "delta": 5}],
            created_at=datetime(2025, 12, 31, 15, 30),
        )
    )
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.RECEIVE,
            quantity_change=Decimal("4"),
            quantity_before=Decimal("5"),
            quantity_after=Decimal("9"),
            inventory_effect=[{"scope": "warehouse", "delta": 4}],
            created_at=datetime(2026, 12, 31, 15, 0),
        )
    )
    db_session.commit()

    response = client.get(
        "/api/admin/audit-ledger/f704-02.xlsx?year=2026",
        headers=ADMIN_HEADERS,
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        "attachment; filename=\"F704-02-R01-2026.xlsx\"; "
        f"filename*=UTF-8''{quote('F704-02 (R01) 2026년 자재 입출고관리대장.xlsx')}"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["양식", "양식_출력용"]
    worksheet = workbook["양식"]
    assert worksheet["A4"].value == 1
    assert worksheet["B4"].value.date().isoformat() == "2026-01-01"
    assert worksheet["F4"].value == "KST 입고 품목"
    assert worksheet["H4"].value == 5
    assert worksheet["I4"].value == "입고"
    assert worksheet["F5"].value is None


def test_f704_download_uses_requester_for_warehouse_to_production(client, db_session, make_item):
    """수량변화가 0인 창고→생산 이동도 실제 라인과 요청자를 따라 한 행으로 남긴다."""
    item = make_item(name="생산 이동 품목")
    db_session.add(ProductSymbol(slot=9, symbol="9", model_name="COCOON"))
    batch = _add_requester_batch_and_line(db_session, item)
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_TO_PROD,
            quantity_change=Decimal("0"),
            quantity_before=Decimal("10"),
            quantity_after=Decimal("10"),
            transfer_qty=Decimal("7"),
            operation_batch_id=batch.batch_id,
            produced_by="처리자 홍",
            notes="생산 투입",
            created_at=datetime(2026, 4, 1, 1, 0),
        )
    )
    db_session.commit()

    response = client.get("/api/admin/audit-ledger/f704-02.xlsx?year=2026", headers=ADMIN_HEADERS)

    assert response.status_code == 200, response.text
    worksheet = load_workbook(BytesIO(response.content), data_only=False)["양식"]
    assert worksheet["E4"].value == "COCOON"
    assert worksheet["H4"].value == 7
    assert worksheet["I4"].value == "출고"
    assert worksheet["J4"].value == "조립"
    assert worksheet["K4"].value == "요청자 김"
    assert "처리자 홍" not in (worksheet["M4"].value or "")


def test_f704_download_excludes_recorded_nonwarehouse_effect(client, db_session, make_item):
    """inventory_effect가 창고 변동 없음을 명시하면 거래 유형만으로 되살리지 않는다."""
    item = make_item(name="부서 내부 품목")
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.TRANSFER_DEPT,
            quantity_change=Decimal("0"),
            quantity_before=Decimal("3"),
            quantity_after=Decimal("3"),
            transfer_qty=Decimal("3"),
            inventory_effect=[
                {"scope": "location", "department": "조립", "status": "PRODUCTION", "delta": -3},
                {"scope": "location", "department": "튜브", "status": "PRODUCTION", "delta": 3},
            ],
            created_at=datetime(2026, 4, 2, 1, 0),
        )
    )
    db_session.commit()

    response = client.get("/api/admin/audit-ledger/f704-02.xlsx?year=2026", headers=ADMIN_HEADERS)

    assert response.status_code == 200, response.text
    worksheet = load_workbook(BytesIO(response.content), data_only=False)["양식"]
    assert worksheet["B4"].value is None
    assert worksheet["F4"].value is None


def test_f704_download_excludes_cancelled_warehouse_transaction(client, db_session, make_item):
    item = make_item(name="cancelled warehouse transaction")
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item.item_id,
            transaction_type=TransactionTypeEnum.SHIP,
            quantity_change=Decimal("-4"),
            quantity_before=Decimal("4"),
            quantity_after=Decimal("0"),
            cancelled=True,
            inventory_effect=[{"scope": "warehouse", "delta": -4}],
            created_at=datetime(2026, 4, 3, 1, 0),
        )
    )
    db_session.commit()

    response = client.get("/api/admin/audit-ledger/f704-02.xlsx?year=2026", headers=ADMIN_HEADERS)

    assert response.status_code == 200, response.text
    worksheet = load_workbook(BytesIO(response.content), data_only=False)["양식"]
    assert worksheet["B4"].value is None
    assert worksheet["F4"].value is None


def test_f704_download_preserves_template_supporting_parts(client, db_session):
    """동적 데이터와 K열 표시 외의 원본 XLSX 구성요소는 그대로 유지한다."""
    response = client.get("/api/admin/audit-ledger/f704-02.xlsx?year=2026", headers=ADMIN_HEADERS)

    assert response.status_code == 200, response.text
    critical_parts = (
        "xl/printerSettings/printerSettings1.bin",
        "xl/printerSettings/printerSettings2.bin",
        "xl/comments1.xml",
        "xl/comments2.xml",
        "xl/drawings/vmlDrawing1.vml",
        "xl/drawings/vmlDrawing2.vml",
        "xl/externalLinks/externalLink1.xml",
        "xl/externalLinks/externalLink2.xml",
        "xl/externalLinks/_rels/externalLink1.xml.rels",
        "xl/externalLinks/_rels/externalLink2.xml.rels",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/_rels/sheet2.xml.rels",
        "xl/worksheets/sheet2.xml",
    )
    with zipfile.ZipFile(f704_02_ledger.TEMPLATE_PATH) as template, zipfile.ZipFile(BytesIO(response.content)) as rendered:
        for part in critical_parts:
            assert hashlib.sha256(rendered.read(part)).digest() == hashlib.sha256(template.read(part)).digest()

    template_workbook = load_workbook(f704_02_ledger.TEMPLATE_PATH, data_only=False)
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == template_workbook.sheetnames == ["양식", "양식_출력용"]

    template_worksheet = template_workbook["양식"]
    worksheet = workbook["양식"]
    assert worksheet.freeze_panes == "A4"
    assert worksheet.page_setup.orientation == "landscape"
    assert worksheet.page_setup.scale == 61
    assert worksheet["K3"].value == template_worksheet["K3"].value == "담당자"
    assert worksheet.column_dimensions["K"].hidden is False
    assert worksheet.column_dimensions["L"].hidden is True
    assert worksheet.auto_filter.ref == template_worksheet.auto_filter.ref
    assert worksheet.page_setup == template_worksheet.page_setup
    assert worksheet.page_margins == template_worksheet.page_margins
    assert worksheet.print_options == template_worksheet.print_options
    assert worksheet.sheet_view.showGridLines == template_worksheet.sheet_view.showGridLines
    assert worksheet.sheet_view.zoomScale == template_worksheet.sheet_view.zoomScale
    assert worksheet.sheet_view.zoomScaleNormal == template_worksheet.sheet_view.zoomScaleNormal
    assert worksheet.sheet_view.pane is not None
    assert template_worksheet.sheet_view.pane is not None
    assert worksheet.sheet_view.pane.ySplit == template_worksheet.sheet_view.pane.ySplit == 3
    assert worksheet.sheet_view.pane.state == template_worksheet.sheet_view.pane.state == "frozen"
    assert worksheet.sheet_view.pane.topLeftCell == "A4"
    rendered_bottom_selection = next(
        selection for selection in worksheet.sheet_view.selection if selection.pane == "bottomLeft"
    )
    assert rendered_bottom_selection.activeCell == "A4"
    assert rendered_bottom_selection.sqref == "A4"
    assert worksheet.sheet_properties == template_worksheet.sheet_properties
    assert _header_footer_layout(worksheet) == _header_footer_layout(template_worksheet)
    assert worksheet.print_area == template_worksheet.print_area
    assert worksheet.print_title_rows == template_worksheet.print_title_rows
    assert worksheet.print_title_cols == template_worksheet.print_title_cols
    assert _row_layout(worksheet) == _row_layout(template_worksheet)

    for column, template_dimension in template_worksheet.column_dimensions.items():
        rendered_dimension = worksheet.column_dimensions[column]
        assert (
            rendered_dimension.width,
            rendered_dimension.outlineLevel,
            rendered_dimension.collapsed,
            rendered_dimension.style,
            rendered_dimension.bestFit,
            rendered_dimension.min,
            rendered_dimension.max,
        ) == (
            template_dimension.width,
            template_dimension.outlineLevel,
            template_dimension.collapsed,
            template_dimension.style,
            template_dimension.bestFit,
            template_dimension.min,
            template_dimension.max,
        )
        assert rendered_dimension.hidden is (False if column == "K" else template_dimension.hidden)
