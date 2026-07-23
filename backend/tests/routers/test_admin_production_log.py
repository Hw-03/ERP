"""F705-02 annual production-log download API tests."""

import hashlib
import zipfile
from io import BytesIO

from openpyxl import load_workbook

from app.services import f705_02_production_log


ADMIN_HEADERS = {"X-Admin-Pin": "0000"}


def _row_layout(worksheet):
    return {
        row: (dimension.height, dimension.hidden, dimension.outlineLevel, dimension.collapsed, dimension.style)
        for row, dimension in worksheet.row_dimensions.items()
    }


def test_f705_download_requires_admin_pin(client):
    response = client.get("/api/admin/production-log/f705-02.xlsx?year=2026")

    assert response.status_code == 400


def test_f705_download_returns_selected_year_template(client):
    response = client.get(
        "/api/admin/production-log/f705-02.xlsx?year=2025",
        headers=ADMIN_HEADERS,
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "F705-02-R01-2025.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == [f"25.{month:02d}" for month in range(1, 13)]
    assert workbook["25.02"]["D1"].value.year == 2025


def test_f705_download_rejects_year_outside_template_range(client):
    response = client.get(
        "/api/admin/production-log/f705-02.xlsx?year=2100",
        headers=ADMIN_HEADERS,
    )

    assert response.status_code == 422


def test_f705_download_preserves_template_layout_and_supporting_package_parts(client):
    response = client.get("/api/admin/production-log/f705-02.xlsx?year=2026", headers=ADMIN_HEADERS)

    assert response.status_code == 200, response.text
    changed_parts = {f705_02_production_log.WORKBOOK_XML, *f705_02_production_log.WORKSHEET_XMLS}
    with zipfile.ZipFile(f705_02_production_log.TEMPLATE_PATH) as template, zipfile.ZipFile(
        BytesIO(response.content)
    ) as rendered:
        assert set(rendered.namelist()) == set(template.namelist())
        for part in set(template.namelist()) - changed_parts:
            assert hashlib.sha256(rendered.read(part)).digest() == hashlib.sha256(template.read(part)).digest()

    template_workbook = load_workbook(f705_02_production_log.TEMPLATE_PATH, data_only=False)
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    template_worksheet = template_workbook["26.03"]
    worksheet = workbook["26.03"]

    assert worksheet.merged_cells == template_worksheet.merged_cells
    assert worksheet.page_setup == template_worksheet.page_setup
    assert worksheet.page_margins == template_worksheet.page_margins
    assert worksheet.print_options == template_worksheet.print_options
    assert worksheet.sheet_properties == template_worksheet.sheet_properties
    assert worksheet.print_area == template_worksheet.print_area
    assert worksheet.print_title_rows == template_worksheet.print_title_rows
    assert worksheet.print_title_cols == template_worksheet.print_title_cols
    assert _row_layout(worksheet) == _row_layout(template_worksheet)

    for column, template_dimension in template_worksheet.column_dimensions.items():
        dimension = worksheet.column_dimensions[column]
        assert (
            dimension.width,
            dimension.hidden,
            dimension.outlineLevel,
            dimension.collapsed,
            dimension.style,
            dimension.bestFit,
            dimension.min,
            dimension.max,
        ) == (
            template_dimension.width,
            template_dimension.hidden,
            template_dimension.outlineLevel,
            template_dimension.collapsed,
            template_dimension.style,
            template_dimension.bestFit,
            template_dimension.min,
            template_dimension.max,
        )

    for row in range(1, 33):
        for column in range(1, 37):
            assert worksheet.cell(row, column).style_id == template_worksheet.cell(row, column).style_id
