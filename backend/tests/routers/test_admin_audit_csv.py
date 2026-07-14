"""admin_audit_csv 라우터 smoke 테스트."""

from __future__ import annotations

import csv as _csv
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models import DepartmentEnum, TransactionLog, TransactionTypeEnum

ADMIN_HEADERS = {"X-Admin-Pin": "0000"}


@pytest.fixture()
def csv_env(tmp_path, monkeypatch):
    monkeypatch.setenv("AUDIT_CSV_DIR", str(tmp_path))
    return tmp_path


def _add_log(db_session, item_id, *, tx, qty, when):
    db_session.add(
        TransactionLog(
            log_id=uuid.uuid4(),
            item_id=item_id,
            transaction_type=tx,
            quantity_change=qty,
            quantity_before=Decimal("0"),
            quantity_after=qty,
            reference_no="R-1",
            produced_by="tester",
            notes="api smoke",
            created_at=when,
        )
    )


def test_files_endpoint_empty(client, csv_env):
    res = client.get("/api/admin/audit-csv/files", headers=ADMIN_HEADERS)
    assert res.status_code == 200
    assert res.json() == []


def test_backfill_then_list_and_download(client, db_session, make_item, csv_env):
    item = make_item(name="API 품목")
    _add_log(db_session, item.item_id, tx=TransactionTypeEnum.RECEIVE, qty=Decimal("4"),
             when=datetime(2026, 5, 10, 9, 0))
    _add_log(db_session, item.item_id, tx=TransactionTypeEnum.SHIP, qty=Decimal("-2"),
             when=datetime(2026, 5, 11, 10, 0))
    db_session.commit()

    backfill = client.post("/api/admin/audit-csv/backfill", headers=ADMIN_HEADERS)
    assert backfill.status_code == 200, backfill.text
    body = backfill.json()
    assert body["total_rows"] == 2
    assert body["months"] == ["2026-05"]

    listing = client.get("/api/admin/audit-csv/files", headers=ADMIN_HEADERS).json()
    assert len(listing) == 1
    assert listing[0]["month"] == "2026-05"
    assert listing[0]["row_count"] == 2

    csv_resp = client.get("/api/admin/audit-csv/2026-05.csv", headers=ADMIN_HEADERS)
    assert csv_resp.status_code == 200
    text = csv_resp.content.decode("utf-8-sig")
    rows = list(_csv.reader(text.splitlines()))
    assert rows[0][0] == "일시"
    assert len(rows) == 3  # header + 2

    xlsx_resp = client.get("/api/admin/audit-csv/2026-05.xlsx", headers=ADMIN_HEADERS)
    assert xlsx_resp.status_code == 200
    assert xlsx_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert len(xlsx_resp.content) > 0


def test_internal_use_label_is_reused_by_audit_xlsx(
    client, db_session, make_item, csv_env
):
    item = make_item(name="연구 감사 반출품")
    _add_log(
        db_session,
        item.item_id,
        tx=TransactionTypeEnum.INTERNAL_USE,
        qty=Decimal("-2"),
        when=datetime(2026, 7, 14, 9, 0),
    )
    db_session.flush()
    log = db_session.query(TransactionLog).one()
    log.department = DepartmentEnum.RESEARCH.value
    db_session.commit()

    backfill = client.post("/api/admin/audit-csv/backfill", headers=ADMIN_HEADERS)
    assert backfill.status_code == 200, backfill.text
    csv_resp = client.get("/api/admin/audit-csv/2026-07.csv", headers=ADMIN_HEADERS)
    assert "연구소 반출" in csv_resp.content.decode("utf-8-sig")

    xlsx_resp = client.get("/api/admin/audit-csv/2026-07.xlsx", headers=ADMIN_HEADERS)
    assert xlsx_resp.status_code == 200, xlsx_resp.text
    workbook = load_workbook(BytesIO(xlsx_resp.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any(row[1] == "연구소 반출" for row in rows[1:])


def test_invalid_month_format_rejected(client, csv_env):
    res = client.get("/api/admin/audit-csv/2026-5.csv", headers=ADMIN_HEADERS)
    assert res.status_code == 400


def test_missing_month_returns_404(client, csv_env):
    res = client.get("/api/admin/audit-csv/2099-12.csv", headers=ADMIN_HEADERS)
    assert res.status_code == 404
