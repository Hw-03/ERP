"""서버 사용자 텍스트 검색의 정규화 회귀 테스트."""

from io import BytesIO

from openpyxl import load_workbook


def test_normalize_search_text_removes_whitespace_and_separators():
    from app.utils.search import normalize_search_text

    assert normalize_search_text(" Search\t\f- Item.01/\v\r\n") == "searchitem01"


def test_build_normalized_search_filter_matches_null_safe_sql_columns(db_session, make_item):
    from app.models import Inventory, Item
    from app.utils.search import build_normalized_search_filter

    matching = make_item(name="Null-safe", serial_no=1)
    missing_code = make_item(name="Other", serial_no=2)
    db_session.query(Inventory).filter(Inventory.item_id == matching.item_id).one().location = "AB- 12.3/4"
    db_session.query(Inventory).filter(Inventory.item_id == missing_code.item_id).one().location = None
    db_session.commit()

    search_filter = build_normalized_search_filter("ab1234", Inventory.location)

    assert search_filter is not None
    assert [row.item_id for row in db_session.query(Item).join(Inventory).filter(search_filter).all()] == [matching.item_id]
    assert build_normalized_search_filter(" \t\f-./\v\r\n", Inventory.location) is None


def test_items_list_search_ignores_spaces_and_separators(client, db_session, make_item):
    item = make_item(name="Search- Item/01")
    db_session.commit()

    response = client.get("/api/items", params={"search": "searchitem01"})

    assert response.status_code == 200, response.text
    assert [row["item_id"] for row in response.json()] == [str(item.item_id)]


def test_items_list_keeps_mes_code_and_location_search_targets(client, db_session, make_item):
    from app.models import Inventory

    item = make_item(name="Target item")
    db_session.query(Inventory).filter(Inventory.item_id == item.item_id).one().location = "Rack- 7/1"
    db_session.commit()

    code_response = client.get("/api/items", params={"search": item.mes_code.lower()})
    location_response = client.get("/api/items", params={"search": "rack71"})

    assert code_response.status_code == 200, code_response.text
    assert [row["item_id"] for row in code_response.json()] == [str(item.item_id)]
    assert location_response.status_code == 200, location_response.text
    assert [row["item_id"] for row in location_response.json()] == [str(item.item_id)]


def test_items_xlsx_export_search_uses_normalized_location(client, db_session, make_item):
    from app.models import Inventory

    item = make_item(name="Export location item")
    db_session.query(Inventory).filter(Inventory.item_id == item.item_id).one().location = "Export- Rack/10"
    db_session.commit()

    response = client.get("/api/items/export.xlsx", params={"search": "exportrack10"})

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    assert "Export location item" in {
        value
        for row in workbook.active.iter_rows(values_only=True)
        for value in row
    }
