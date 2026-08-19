#!/usr/bin/env python3
"""Prepare a clean inventory baseline for DEXCOWIN MES operation.

Default mode is dry-run. Applying changes:

    python scripts/ops/inventory_cutover.py real_inventory.csv --apply --confirm START-OVER

Input file format, one row per stock bucket:

    mes_code,bucket,department,quantity,location
    3-TR-0001,warehouse,,100,WH-A
    3-AA-0001,production,Assembly,7,Line-1
    3-AA-0001,defective,Assembly,2,Line-1

The script keeps item/BOM/employee/master data, clears operational history and
warehouse map quantities, then overwrites inventory quantities from the file.
"""

from __future__ import annotations

import argparse
import csv
import enum
import json
import os
import subprocess
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Any

from sqlalchemy import String, cast, create_engine, text
from sqlalchemy.orm import Session, sessionmaker


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = PROJECT_ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from dotenv import load_dotenv  # noqa: E402

from app.models import (  # noqa: E402
    Department,
    Inventory,
    InventoryLocation,
    IoBatch,
    IoBundle,
    IoLine,
    Item,
    LocationStatusEnum,
    ShippingAllocation,
    ShippingRequest,
    StockRequest,
    StockRequestLine,
    TransactionEditLog,
    TransactionLog,
    TransactionTypeEnum,
    WarehouseBoxItem,
)

load_dotenv(BACKEND_DIR / ".env")


VALID_BUCKETS = {"warehouse", "production", "defective"}
BUCKET_ALIASES = {
    "wh": "warehouse",
    "warehouse": "warehouse",
    "prod": "production",
    "production": "production",
    "defect": "defective",
    "defective": "defective",
}


class CutoverInputError(ValueError):
    """Raised when the cutover input is unsafe to apply."""

    def __init__(
        self,
        message: str,
        *,
        shipping_report: tuple[ShippingCutoverReportEntry, ...] = (),
    ) -> None:
        super().__init__(message)
        self.shipping_report = shipping_report


class ShippingCutoverDisposition(str, enum.Enum):
    """Whether a historical shipping request can still change stock."""

    TERMINAL_SAFE = "TERMINAL_SAFE"
    FUTURE_DELTA = "FUTURE_DELTA"
    INCONSISTENT = "INCONSISTENT"


def classify_shipping_cutover_state(
    *,
    status: str,
    allocation_statuses: Iterable[str],
    active_log_phases: Iterable[str],
    active_pickup_log_count: int,
    active_pickup_effect_log_count: int,
    malformed_active_log_count: int = 0,
    malformed_allocation_count: int = 0,
) -> ShippingCutoverDisposition:
    """Classify one shipping request from persisted state only."""
    normalized_status = str(status).upper()
    allocations = frozenset(str(value).upper() for value in allocation_statuses)
    phases = frozenset(str(value).upper() for value in active_log_phases)
    known_allocations = {"RESERVED", "CONSUMED", "RELEASED"}

    if malformed_active_log_count or malformed_allocation_count:
        return ShippingCutoverDisposition.INCONSISTENT
    if not allocations <= known_allocations:
        return ShippingCutoverDisposition.INCONSISTENT
    if active_pickup_log_count != active_pickup_effect_log_count:
        return ShippingCutoverDisposition.INCONSISTENT
    if ("PICKUP" in phases or active_pickup_log_count) and normalized_status != "PICKED_UP":
        return ShippingCutoverDisposition.INCONSISTENT

    if normalized_status == "REQUESTED":
        return (
            ShippingCutoverDisposition.FUTURE_DELTA
            if not allocations
            else ShippingCutoverDisposition.INCONSISTENT
        )
    if normalized_status == "PREPARING":
        return (
            ShippingCutoverDisposition.FUTURE_DELTA
            if allocations in {frozenset(), frozenset({"RELEASED"})}
            else ShippingCutoverDisposition.INCONSISTENT
        )
    if normalized_status == "PREPARED":
        allowed = {
            frozenset(),
            frozenset({"RESERVED"}),
            frozenset({"RESERVED", "RELEASED"}),
        }
        return (
            ShippingCutoverDisposition.FUTURE_DELTA
            if allocations in allowed
            else ShippingCutoverDisposition.INCONSISTENT
        )
    if normalized_status == "PICKED_UP":
        allowed = {
            frozenset(),
            frozenset({"CONSUMED"}),
            frozenset({"CONSUMED", "RELEASED"}),
        }
        if allocations not in allowed or active_pickup_effect_log_count < 1:
            return ShippingCutoverDisposition.INCONSISTENT
        return ShippingCutoverDisposition.FUTURE_DELTA
    if normalized_status == "CANCELLED":
        return (
            ShippingCutoverDisposition.TERMINAL_SAFE
            if allocations in {frozenset(), frozenset({"RELEASED"})}
            else ShippingCutoverDisposition.INCONSISTENT
        )
    return ShippingCutoverDisposition.INCONSISTENT


def _has_effective_inventory_delta(value: Any) -> bool:
    """Return whether an effect can prove a real, reversible inventory delta."""
    if not isinstance(value, list) or not value:
        return False
    for cell in value:
        if not isinstance(cell, dict):
            return False
        scope = cell.get("scope")
        delta = cell.get("delta")
        if type(delta) is not int or delta == 0 or abs(delta) > 2_147_483_647:
            return False
        if scope == "location":
            if set(cell) != {"scope", "department", "status", "delta"}:
                return False
            if not isinstance(cell.get("department"), str) or not cell["department"].strip():
                return False
            try:
                LocationStatusEnum(cell.get("status"))
            except (TypeError, ValueError):
                return False
        elif scope == "warehouse_box":
            if set(cell) != {"scope", "box_id", "delta"}:
                return False
            if not isinstance(cell.get("box_id"), str) or not cell["box_id"].strip():
                return False
        elif scope == "warehouse":
            if set(cell) != {"scope", "delta"}:
                return False
        else:
            return False
    return True


@dataclass(frozen=True)
class CutoverRow:
    mes_code: str
    bucket: str
    quantity: int
    source_row: int
    department: str | None = None
    location: str | None = None


@dataclass(frozen=True)
class CutoverOptions:
    apply: bool = False
    clear_history: bool = True
    clear_warehouse_map: bool = True
    require_all_items: bool = True


@dataclass(frozen=True)
class CutoverSummary:
    applied: bool
    items_updated: int
    inventory_locations_deleted: int
    inventory_locations_inserted: int
    transaction_logs_deleted: int
    transaction_edit_logs_deleted: int
    stock_requests_deleted: int
    stock_request_lines_deleted: int
    io_batches_deleted: int
    io_bundles_deleted: int
    io_lines_deleted: int
    warehouse_box_items_deleted: int
    missing_items: tuple[str, ...]
    shipping_report: tuple[ShippingCutoverReportEntry, ...] = ()


@dataclass(frozen=True)
class ShippingAllocationQuantity:
    status: str
    quantity: int


@dataclass(frozen=True)
class ShippingCutoverReportEntry:
    request_id: str
    status: str
    allocation_quantities: tuple[ShippingAllocationQuantity, ...]
    active_log_phases: tuple[str, ...]
    active_pickup_log_count: int
    active_pickup_effect_log_count: int
    malformed_active_log_count: int
    malformed_allocation_count: int
    disposition: ShippingCutoverDisposition

    @property
    def allocation_statuses(self) -> tuple[str, ...]:
        return tuple(row.status for row in self.allocation_quantities)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _canonical(row: dict[str, Any], *names: str) -> str:
    normalized = {_normalize_header(k): v for k, v in row.items()}
    for name in names:
        key = _normalize_header(name)
        if key in normalized:
            return _clean(normalized[key])
    return ""


def _parse_quantity(raw: Any, source_row: int) -> int:
    text = _clean(raw).replace(",", "")
    if not text:
        return 0
    try:
        value = Decimal(text)
    except InvalidOperation as exc:
        raise CutoverInputError(f"row {source_row}: invalid quantity {raw!r}") from exc
    if value < 0:
        raise CutoverInputError(f"row {source_row}: negative quantity is not allowed")
    if value != value.to_integral_value():
        raise CutoverInputError(f"row {source_row}: quantity must be an integer")
    return int(value)


def _row_from_mapping(row: dict[str, Any], source_row: int) -> CutoverRow | None:
    mes_code = _canonical(row, "mes_code", "code", "item_code").upper()
    bucket_raw = _canonical(row, "bucket", "stock_bucket").lower()
    department = _canonical(row, "department", "dept") or None
    location = _canonical(row, "location") or None
    quantity_raw = _canonical(row, "quantity", "qty")

    if not any([mes_code, bucket_raw, department, location, quantity_raw]):
        return None
    if not mes_code:
        raise CutoverInputError(f"row {source_row}: mes_code is required")

    bucket = BUCKET_ALIASES.get(bucket_raw, bucket_raw)
    if bucket not in VALID_BUCKETS:
        raise CutoverInputError(f"row {source_row}: invalid bucket {bucket_raw!r}")

    if bucket == "warehouse" and department:
        raise CutoverInputError(f"row {source_row}: warehouse bucket must not have department")
    if bucket in {"production", "defective"} and not department:
        raise CutoverInputError(f"row {source_row}: department is required for {bucket}")

    return CutoverRow(
        mes_code=mes_code,
        bucket=bucket,
        department=department,
        quantity=_parse_quantity(quantity_raw, source_row),
        location=location,
        source_row=source_row,
    )


def _read_csv(path: Path) -> list[CutoverRow]:
    rows: list[CutoverRow] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise CutoverInputError("input file has no header row")
        for source_row, raw in enumerate(reader, start=2):
            parsed = _row_from_mapping(raw, source_row)
            if parsed is not None:
                rows.append(parsed)
    return rows


def _read_xlsx(path: Path) -> list[CutoverRow]:
    try:
        import openpyxl
    except ImportError as exc:
        raise CutoverInputError("openpyxl is required to read .xlsx files") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [_clean(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if not any(headers):
        raise CutoverInputError("input file has no header row")

    rows: list[CutoverRow] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        parsed = _row_from_mapping(raw, source_row)
        if parsed is not None:
            rows.append(parsed)
    workbook.close()
    return rows


def parse_cutover_file(path: Path) -> list[CutoverRow]:
    source = path.resolve()
    if not source.exists():
        raise CutoverInputError(f"input file not found: {source}")
    suffix = source.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv(source)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx(source)
    else:
        raise CutoverInputError("input file must be .csv, .xlsx, or .xlsm")
    if not rows:
        raise CutoverInputError("input file has no stock rows")
    return rows


def _active_items_by_code(db: Session) -> dict[str, Item]:
    items = db.query(Item).filter(Item.deleted_at.is_(None)).all()
    return {str(item.mes_code).upper(): item for item in items if item.mes_code}


def _validate_rows(db: Session, rows: Iterable[CutoverRow], options: CutoverOptions) -> tuple[dict[str, dict[str, Any]], tuple[str, ...]]:
    items_by_code = _active_items_by_code(db)
    active_codes = set(items_by_code)
    targets: dict[str, dict[str, Any]] = {}
    seen_keys: set[tuple[str, str, str | None]] = set()

    active_department_names = {
        str(row.name)
        for row in db.query(Department).filter(Department.is_active.is_(True)).all()
    }

    for row in rows:
        code = row.mes_code.upper()
        if code not in items_by_code:
            raise CutoverInputError(f"row {row.source_row}: unknown mes_code {row.mes_code}")
        key = (code, row.bucket, row.department if row.bucket != "warehouse" else None)
        if key in seen_keys:
            raise CutoverInputError(f"row {row.source_row}: duplicate stock bucket for {row.mes_code}")
        seen_keys.add(key)

        if row.bucket in {"production", "defective"} and active_department_names and row.department not in active_department_names:
            raise CutoverInputError(f"row {row.source_row}: unknown department {row.department!r}")

        target = targets.setdefault(
            code,
            {"item": items_by_code[code], "warehouse": 0, "locations": {}, "location": None},
        )
        if row.location and target["location"] is None:
            target["location"] = row.location
        if row.bucket == "warehouse":
            target["warehouse"] = row.quantity
        else:
            status = LocationStatusEnum.PRODUCTION if row.bucket == "production" else LocationStatusEnum.DEFECTIVE
            target["locations"][(row.department, status)] = row.quantity

    missing_codes = tuple(sorted(active_codes - set(targets)))
    if missing_codes and options.require_all_items:
        preview = ", ".join(missing_codes[:10])
        suffix = "..." if len(missing_codes) > 10 else ""
        raise CutoverInputError(f"missing mes_code rows for {len(missing_codes)} active item(s): {preview}{suffix}")

    if not options.require_all_items:
        for code in missing_codes:
            targets[code] = {"item": items_by_code[code], "warehouse": 0, "locations": {}, "location": None}

    return targets, missing_codes


def _count_rows_deleted(db: Session, model: type) -> int:
    return int(db.query(model).count())


def _delete_rows(db: Session, model: type) -> int:
    count = _count_rows_deleted(db, model)
    db.query(model).delete(synchronize_session=False)
    return count


def _preview_counts(db: Session, clear_history: bool, clear_warehouse_map: bool) -> dict[str, int]:
    return {
        "transaction_edit_logs_deleted": _count_rows_deleted(db, TransactionEditLog) if clear_history else 0,
        "transaction_logs_deleted": _count_rows_deleted(db, TransactionLog) if clear_history else 0,
        "stock_request_lines_deleted": _count_rows_deleted(db, StockRequestLine) if clear_history else 0,
        "stock_requests_deleted": _count_rows_deleted(db, StockRequest) if clear_history else 0,
        "io_lines_deleted": _count_rows_deleted(db, IoLine) if clear_history else 0,
        "io_bundles_deleted": _count_rows_deleted(db, IoBundle) if clear_history else 0,
        "io_batches_deleted": _count_rows_deleted(db, IoBatch) if clear_history else 0,
        "warehouse_box_items_deleted": _count_rows_deleted(db, WarehouseBoxItem) if clear_warehouse_map else 0,
        "inventory_locations_deleted": _count_rows_deleted(db, InventoryLocation),
    }


def _clear_operational_state(db: Session, options: CutoverOptions) -> dict[str, int]:
    counts = {
        "transaction_edit_logs_deleted": 0,
        "transaction_logs_deleted": 0,
        "stock_request_lines_deleted": 0,
        "stock_requests_deleted": 0,
        "io_lines_deleted": 0,
        "io_bundles_deleted": 0,
        "io_batches_deleted": 0,
        "warehouse_box_items_deleted": 0,
        "inventory_locations_deleted": 0,
    }
    if options.clear_history:
        counts["transaction_edit_logs_deleted"] = _delete_rows(db, TransactionEditLog)
        counts["transaction_logs_deleted"] = _delete_rows(db, TransactionLog)
        counts["stock_request_lines_deleted"] = _delete_rows(db, StockRequestLine)
        counts["stock_requests_deleted"] = _delete_rows(db, StockRequest)
        counts["io_lines_deleted"] = _delete_rows(db, IoLine)
        counts["io_bundles_deleted"] = _delete_rows(db, IoBundle)
        counts["io_batches_deleted"] = _delete_rows(db, IoBatch)
    if options.clear_warehouse_map:
        counts["warehouse_box_items_deleted"] = _delete_rows(db, WarehouseBoxItem)
    counts["inventory_locations_deleted"] = _delete_rows(db, InventoryLocation)
    return counts


def _apply_targets(db: Session, targets: dict[str, dict[str, Any]]) -> int:
    inserted_locations = 0
    for target in targets.values():
        item = target["item"]
        inv = db.query(Inventory).filter(Inventory.item_id == item.item_id).first()
        if inv is None:
            inv = Inventory(item_id=item.item_id, quantity=0, warehouse_qty=0, pending_quantity=0)
            db.add(inv)
            db.flush()

        warehouse_qty = int(target["warehouse"])
        loc_total = 0
        for (department, status), quantity in target["locations"].items():
            qty = int(quantity)
            if qty <= 0:
                continue
            db.add(
                InventoryLocation(
                    item_id=item.item_id,
                    department=department,
                    status=status,
                    quantity=qty,
                )
            )
            loc_total += qty
            inserted_locations += 1

        inv.warehouse_qty = warehouse_qty
        inv.pending_quantity = 0
        inv.last_reserver_employee_id = None
        inv.last_reserver_name = None
        inv.quantity = warehouse_qty + loc_total
        if target["location"] is not None:
            inv.location = target["location"]
    return inserted_locations


def _acquire_cutover_write_lock(db: Session) -> None:
    """Start the apply transaction while excluding shipping writers."""
    bind = db.get_bind()
    bind_in_transaction = getattr(bind, "in_transaction", None)
    if db.in_transaction() or (callable(bind_in_transaction) and bind_in_transaction()):
        raise CutoverInputError("cutover apply requires a fresh Session without an open transaction")
    dialect = bind.dialect.name
    try:
        if dialect == "sqlite":
            db.execute(text("BEGIN IMMEDIATE"))
            return
        if dialect == "postgresql":
            db.execute(
                text(
                    "LOCK TABLE shipping_requests, shipping_allocations, transaction_logs "
                    "IN ACCESS EXCLUSIVE MODE"
                )
            )
            return
    except Exception:
        db.rollback()
        raise
    raise CutoverInputError(f"cutover apply does not support database dialect {dialect!r}")


def _decode_inventory_effect(raw_value: str | None) -> tuple[Any, bool]:
    if raw_value is None:
        return None, False
    try:
        return json.loads(raw_value), False
    except (TypeError, json.JSONDecodeError):
        return None, True


def _active_log_flag(raw_value: str | None) -> tuple[bool, bool]:
    normalized = str(raw_value).strip().lower()
    if normalized in {"0", "false"}:
        return True, False
    if normalized in {"1", "true"}:
        return False, False
    return True, True


def inspect_shipping_cutover(db: Session) -> tuple[ShippingCutoverReportEntry, ...]:
    """Return request, allocation, and raw log evidence without enum ORM decoding."""
    request_statuses = {
        str(request_id): str(raw_status).upper()
        for request_id, raw_status in (
            db.query(
                ShippingRequest.request_id,
                cast(ShippingRequest.status, String).label("status"),
            )
            .all()
        )
    }
    allocation_totals: dict[str, dict[str, int]] = {}
    malformed_allocations: dict[str, int] = {}
    for request_id, raw_status, raw_quantity in db.query(
        ShippingAllocation.request_id,
        cast(ShippingAllocation.status, String).label("status"),
        ShippingAllocation.quantity,
    ).all():
        key = str(request_id)
        status = str(raw_status).upper()
        totals = allocation_totals.setdefault(key, {})
        if type(raw_quantity) is not int or raw_quantity <= 0 or raw_quantity > 2_147_483_647:
            malformed_allocations[key] = malformed_allocations.get(key, 0) + 1
            continue
        totals[status] = totals.get(status, 0) + raw_quantity

    log_evidence: dict[str, dict[str, Any]] = {}
    known_phases = {"PICKUP", "PREPARE", "COMPONENT_CHANGE"}
    known_transaction_types = {member.value for member in TransactionTypeEnum}
    log_rows = db.query(
        TransactionLog.shipping_request_id,
        cast(TransactionLog.shipping_phase, String).label("shipping_phase"),
        cast(TransactionLog.transaction_type, String).label("transaction_type"),
        cast(TransactionLog.inventory_effect, String).label("inventory_effect"),
        cast(TransactionLog.cancelled, String).label("cancelled"),
    ).filter(TransactionLog.shipping_request_id.is_not(None)).all()
    for request_id, raw_phase, raw_transaction_type, raw_effect, raw_cancelled in log_rows:
        active, malformed_cancelled = _active_log_flag(raw_cancelled)
        if not active:
            continue
        key = str(request_id)
        evidence = log_evidence.setdefault(
            key,
            {"phases": set(), "pickup_logs": 0, "pickup_effects": 0, "malformed_logs": 0},
        )
        phase = str(raw_phase).upper()
        transaction_type = str(raw_transaction_type).upper()
        evidence["phases"].add(phase)
        effect, malformed_json = _decode_inventory_effect(raw_effect)
        malformed_log = (
            malformed_cancelled
            or phase not in known_phases
            or transaction_type not in known_transaction_types
            or malformed_json
        )
        if phase == "PICKUP":
            evidence["pickup_logs"] += 1
            if _has_effective_inventory_delta(effect):
                evidence["pickup_effects"] += 1
        if malformed_log:
            evidence["malformed_logs"] += 1

    report: list[ShippingCutoverReportEntry] = []
    request_ids = set(request_statuses) | set(allocation_totals) | set(log_evidence)
    for request_id in sorted(request_ids):
        totals = allocation_totals.get(request_id, {})
        log = log_evidence.get(
            request_id,
            {"phases": set(), "pickup_logs": 0, "pickup_effects": 0, "malformed_logs": 0},
        )
        status = request_statuses.get(request_id, "MISSING_REQUEST")
        malformed_allocation_count = malformed_allocations.get(request_id, 0)
        disposition = classify_shipping_cutover_state(
            status=status,
            allocation_statuses=totals,
            active_log_phases=log["phases"],
            active_pickup_log_count=log["pickup_logs"],
            active_pickup_effect_log_count=log["pickup_effects"],
            malformed_active_log_count=log["malformed_logs"],
            malformed_allocation_count=malformed_allocation_count,
        )
        report.append(
            ShippingCutoverReportEntry(
                request_id=request_id,
                status=status,
                allocation_quantities=tuple(
                    ShippingAllocationQuantity(status=allocation_status, quantity=quantity)
                    for allocation_status, quantity in sorted(totals.items())
                ),
                active_log_phases=tuple(sorted(log["phases"])),
                active_pickup_log_count=log["pickup_logs"],
                active_pickup_effect_log_count=log["pickup_effects"],
                malformed_active_log_count=log["malformed_logs"],
                malformed_allocation_count=malformed_allocation_count,
                disposition=disposition,
            )
        )
    return tuple(report)


def _ensure_shipping_cutover_safe(db: Session) -> tuple[ShippingCutoverReportEntry, ...]:
    report = inspect_shipping_cutover(db)
    unsafe = [entry for entry in report if entry.disposition is not ShippingCutoverDisposition.TERMINAL_SAFE]
    if unsafe:
        details = "; ".join(
            f"{entry.request_id}={entry.disposition.value}"
            f"(status={entry.status}, "
            f"allocations={[(row.status, row.quantity) for row in entry.allocation_quantities]}, "
            f"active_phases={list(entry.active_log_phases)}, pickup_logs={entry.active_pickup_log_count}, "
            f"effective_pickup_effects={entry.active_pickup_effect_log_count}, "
            f"malformed_logs={entry.malformed_active_log_count}, "
            f"malformed_allocations={entry.malformed_allocation_count})"
            for entry in unsafe
        )
        raise CutoverInputError(
            f"shipping cutover preflight rejected: {details}",
            shipping_report=report,
        )
    return report


def run_cutover(db: Session, rows: list[CutoverRow], options: CutoverOptions) -> CutoverSummary:
    if not options.clear_history:
        raise CutoverInputError("clear_history=False is unsafe; cutover must clear operational history")
    if not options.apply:
        targets, missing = _validate_rows(db, rows, options)
        shipping_report = _ensure_shipping_cutover_safe(db)
        counts = _preview_counts(db, options.clear_history, options.clear_warehouse_map)
        return CutoverSummary(
            applied=False,
            items_updated=len(targets),
            inventory_locations_inserted=sum(
                1
                for target in targets.values()
                for quantity in target["locations"].values()
                if int(quantity) > 0
            ),
            missing_items=missing,
            shipping_report=shipping_report,
            **counts,
        )

    _acquire_cutover_write_lock(db)
    try:
        targets, missing = _validate_rows(db, rows, options)
        shipping_report = _ensure_shipping_cutover_safe(db)
        counts = _clear_operational_state(db, options)
        inserted_locations = _apply_targets(db, targets)
        db.commit()
        return CutoverSummary(
            applied=True,
            items_updated=len(targets),
            inventory_locations_inserted=inserted_locations,
            missing_items=missing,
            shipping_report=shipping_report,
            **counts,
        )
    except Exception:
        db.rollback()
        raise


def _database_url(args: argparse.Namespace) -> str:
    return args.db_url or os.getenv("DATABASE_URL") or f"sqlite:///{(BACKEND_DIR / 'mes.db').as_posix()}"


def _sqlite_path_from_url(url: str) -> Path | None:
    if not url.startswith("sqlite:///"):
        return None
    return Path(url.removeprefix("sqlite:///")).resolve()


def _backup_sqlite(db_path: Path) -> None:
    result = subprocess.run(
        [sys.executable, str(PROJECT_ROOT / "scripts" / "ops" / "backup_db.py"), "--sqlite", str(db_path)],
        cwd=PROJECT_ROOT,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise SystemExit(result.returncode)


def _make_session(url: str):
    engine = create_engine(
        url,
        connect_args={"check_same_thread": False} if url.startswith("sqlite") else {},
    )
    return sessionmaker(bind=engine)


def _print_shipping_report(
    report: tuple[ShippingCutoverReportEntry, ...],
    *,
    stream: Any = None,
) -> None:
    target = stream or sys.stdout
    print("  shipping preflight:", file=target)
    if not report:
        print("    no shipping request/allocation/log evidence", file=target)
        return
    for entry in report:
        allocation_text = ",".join(
            f"{row.status}={row.quantity}"
            for row in entry.allocation_quantities
        ) or "none"
        phases = ",".join(entry.active_log_phases) or "none"
        print(
            f"    request={entry.request_id} status={entry.status} "
            f"disposition={entry.disposition.value} allocations={allocation_text} "
            f"active_phases={phases} pickup_logs={entry.active_pickup_log_count} "
            f"effective_pickup_effects={entry.active_pickup_effect_log_count} "
            f"malformed_logs={entry.malformed_active_log_count} "
            f"malformed_allocations={entry.malformed_allocation_count}",
            file=target,
        )


def _print_summary(summary: CutoverSummary) -> None:
    mode = "APPLY" if summary.applied else "DRY-RUN"
    print(f"[{mode}] inventory cutover summary")
    print(f"  items updated              : {summary.items_updated}")
    print(f"  inventory locations deleted: {summary.inventory_locations_deleted}")
    print(f"  inventory locations inserted: {summary.inventory_locations_inserted}")
    print(f"  transaction logs deleted   : {summary.transaction_logs_deleted}")
    print(f"  stock requests deleted     : {summary.stock_requests_deleted}")
    print(f"  io batches deleted         : {summary.io_batches_deleted}")
    print(f"  warehouse box items deleted: {summary.warehouse_box_items_deleted}")
    if summary.missing_items:
        print(f"  missing items zeroed       : {len(summary.missing_items)}")
    _print_shipping_report(summary.shipping_report)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Clean DEXCOWIN MES operational stock state and load a new baseline")
    parser.add_argument("source", type=Path, help="CSV/XLSX cutover file")
    parser.add_argument("--db-url", default=None, help="SQLAlchemy database URL")
    parser.add_argument("--apply", action="store_true", help="Apply changes. Default is dry-run.")
    parser.add_argument("--confirm", default="", help="Required value for --apply: START-OVER")
    parser.add_argument("--no-backup", action="store_true", help="Skip SQLite backup before --apply")
    parser.add_argument(
        "--missing-items-zero",
        action="store_true",
        help="Set active items missing from the input file to zero instead of failing.",
    )
    parser.add_argument(
        "--keep-history",
        action="store_true",
        help="Legacy unsafe option; always rejected because cutover must clear operational history.",
    )
    parser.add_argument("--keep-warehouse-map", action="store_true", help="Do not clear warehouse map item quantities")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.apply and args.confirm != "START-OVER":
        print("[CUTOVER] --apply requires --confirm START-OVER", file=sys.stderr)
        return 2
    if args.keep_history:
        print("[CUTOVER] input rejected: --keep-history is unsafe and no longer supported", file=sys.stderr)
        return 1

    try:
        rows = parse_cutover_file(args.source)
        url = _database_url(args)
        if args.apply and not args.no_backup:
            sqlite_path = _sqlite_path_from_url(url)
            if sqlite_path is not None:
                _backup_sqlite(sqlite_path)
            else:
                print("[CUTOVER] non-SQLite DB detected; run an external DB backup before applying", file=sys.stderr)
                return 2

        SessionLocal = _make_session(url)
        with SessionLocal() as db:
            summary = run_cutover(
                db,
                rows,
                CutoverOptions(
                    apply=args.apply,
                    clear_history=not args.keep_history,
                    clear_warehouse_map=not args.keep_warehouse_map,
                    require_all_items=not args.missing_items_zero,
                ),
            )
        _print_summary(summary)
        return 0
    except CutoverInputError as exc:
        print(f"[CUTOVER] input rejected: {exc}", file=sys.stderr)
        _print_shipping_report(exc.shipping_report, stream=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
