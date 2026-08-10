"""Read-only Excel-to-MES inventory verification.

The verifier intentionally has no database write path.  It treats the current
MES code as the identity of an item and compares warehouse and production
department buckets independently.
"""

from __future__ import annotations

import argparse
import base64
from collections import defaultdict
from contextlib import closing
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
import json
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import warnings
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BUCKET_ORDER = ("창고", "튜브", "고압", "진공", "튜닝", "조립", "출하")
PROCESS_DEPARTMENTS = {
    "T": "튜브",
    "H": "고압",
    "V": "진공",
    "N": "튜닝",
    "A": "조립",
    "P": "출하",
}
SOURCE_DEPARTMENTS = {
    "high_vacuum_tuning": frozenset({"고압", "진공", "튜닝"}),
    "assembly_shipping": frozenset({"조립", "출하"}),
    "tube": frozenset({"튜브"}),
    "finished": frozenset({"출하"}),
}
DEPARTMENT_FILE_MARKERS = {
    "high_vacuum_tuning": "고압,진공,튜닝파트",
    "assembly_shipping": "조립,출하파트",
    "tube": "튜브파트",
    "finished": "출하완제품",
}
DEPARTMENT_SHEETS = {
    "high_vacuum_tuning": "고압",
    "assembly_shipping": "조립 자재",
    "tube": "튜브",
    "finished": "완제품",
}
HEADER_ALIASES = {
    "original_name": frozenset({"품목", "품명"}),
    "quantity": frozenset({"현재고", "총합"}),
    "mes_name": frozenset({"MES품목명"}),
    "mes_code": frozenset({"MES코드"}),
    "confirmation": frozenset({"담당자확인"}),
}
MONTH_SHEET_PATTERN = re.compile(r"^(\d{2})\.(\d{2})월$")
PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EMPLOYEE_DB = Path(r"C:\ERP-dev\backend\mes.db")
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "_attic" / "runtime" / "inventory_verification"

_EXCEL_CALCULATOR = r"""
$ErrorActionPreference = 'Stop'
$pathText = [Text.Encoding]::UTF8.GetString([Convert]::FromBase64String($env:MES_INVENTORY_WORKBOOKS))
$paths = ConvertFrom-Json $pathText
$excel = $null
$workbook = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.AskToUpdateLinks = $false
    foreach ($path in $paths) {
        $workbook = $excel.Workbooks.Open([string]$path, 0, $false)
        if ($workbook.ReadOnly) {
            throw "Temporary workbook opened read-only: $path"
        }
        $excel.CalculateFullRebuild()
        $deadline = [DateTime]::UtcNow.AddSeconds(120)
        while ($excel.CalculationState -ne 0) {
            if ([DateTime]::UtcNow -ge $deadline) {
                throw "Excel calculation timed out: $path"
            }
            Start-Sleep -Milliseconds 200
        }
        $workbook.Save()
        $workbook.Close($false)
        [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook)
        $workbook = $null
    }
}
finally {
    if ($workbook) { $workbook.Close($false) }
    if ($excel) { $excel.Quit() }
    if ($workbook) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) }
    if ($excel) { [void][Runtime.InteropServices.Marshal]::ReleaseComObject($excel) }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""


class InventoryVerificationError(RuntimeError):
    """Raised when a trustworthy comparison cannot be produced."""


@dataclass(frozen=True)
class SourceRow:
    """One inventory row read from a calculated temporary workbook copy."""

    source_kind: str
    source_group: str
    source_path: Path
    sheet_name: str
    row_number: int
    original_name: str
    mes_name: str
    mes_code: str
    confirmation: str
    quantity: object


@dataclass(frozen=True)
class MesItem:
    """Read-only item master and stock snapshot required for comparison."""

    item_id: str
    mes_code: str
    item_name: str
    process_type_code: str
    deleted_at: str | None
    warehouse_quantity: int
    department_quantities: dict[str, int]


@dataclass(frozen=True)
class VerificationIssue:
    """A mapping, source-data, or confirmation problem requiring attention."""

    severity: str
    issue_type: str
    message: str
    source_path: str = ""
    sheet_name: str = ""
    row_number: int | None = None
    mes_code: str = ""
    excel_name: str = ""
    mes_name: str = ""
    quantity: int | None = None


@dataclass(frozen=True)
class ComparisonRow:
    """Aggregated Excel and MES quantities for one bucket and MES code."""

    bucket: str
    mes_code: str
    mes_name: str
    excel_quantity: int
    mes_quantity: int
    difference: int
    status: str
    confirmed: bool
    source_rows: tuple[str, ...]


@dataclass(frozen=True)
class VerificationResult:
    """All comparisons and issues produced from one consistent snapshot."""

    comparisons: tuple[ComparisonRow, ...]
    issues: tuple[VerificationIssue, ...]


@dataclass(frozen=True)
class RunMetadata:
    """Source identity and point-in-time information embedded in both reports."""

    verified_at: str
    source_hashes: dict[str, str]
    employee_db_path: str
    employee_db_hash: str
    warehouse_sheet: str


@dataclass(frozen=True)
class ReportPaths:
    """Final Excel and JSON report paths sharing one timestamped stem."""

    xlsx_path: Path
    json_path: Path


@dataclass(frozen=True)
class SourceSnapshot:
    """Rows and identities extracted from stable source workbook copies."""

    rows: tuple[SourceRow, ...]
    source_hashes: dict[str, str]
    warehouse_sheet: str


@dataclass(frozen=True)
class VerificationRun:
    """Return value for programmatic and CLI verification callers."""

    result: VerificationResult
    metadata: RunMetadata
    report_paths: ReportPaths


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _compact(value: object) -> str:
    return re.sub(r"[\s_]+", "", _text(value)).upper()


def file_sha256(path: Path) -> str:
    """Return the source identity used to prove read-only execution."""

    import hashlib

    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def discover_department_workbooks(source_dir: Path) -> dict[str, Path]:
    """Find exactly one workbook for each required production source group."""

    source_dir = Path(source_dir)
    if not source_dir.is_dir():
        raise InventoryVerificationError(f"부서 폴더가 없습니다: {source_dir}")
    candidates = [
        path
        for path in source_dir.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    discovered: dict[str, Path] = {}
    for group, marker in DEPARTMENT_FILE_MARKERS.items():
        matches = [path for path in candidates if _compact(marker) in _compact(path.stem)]
        if len(matches) != 1:
            names = ", ".join(sorted(path.name for path in matches)) or "없음"
            raise InventoryVerificationError(
                f"{group} 파일은 정확히 1개여야 합니다: {names}"
            )
        discovered[group] = matches[0]
    return discovered


def assert_sources_unlocked(paths: Iterable[Path]) -> None:
    """Reject Excel owner files so a weekly snapshot has an unambiguous cutoff."""

    locked = []
    for raw_path in paths:
        path = Path(raw_path)
        owner_file = path.with_name(f"~${path.name}")
        if owner_file.exists():
            locked.append(path.name)
    if locked:
        raise InventoryVerificationError(
            f"Excel 파일이 열려 있습니다: {', '.join(sorted(locked))}"
        )


def calculate_workbook_copies(paths: Iterable[Path]) -> None:
    """Calculate and save only caller-provided temporary workbook copies."""

    workbook_paths = [str(Path(path).resolve()) for path in paths]
    if not workbook_paths:
        raise InventoryVerificationError("계산할 Excel 사본이 없습니다")
    if os.name != "nt":
        raise InventoryVerificationError("Microsoft Excel 계산은 Windows에서만 지원합니다")
    encoded_command = base64.b64encode(
        _EXCEL_CALCULATOR.encode("utf-16le")
    ).decode("ascii")
    environment = os.environ.copy()
    environment["MES_INVENTORY_WORKBOOKS"] = base64.b64encode(
        json.dumps(workbook_paths, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")
    try:
        completed = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-EncodedCommand",
                encoded_command,
            ],
            env=environment,
            capture_output=True,
            text=True,
            timeout=180,
            check=False,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise InventoryVerificationError(f"Excel 계산 실행 실패: {exc}") from exc
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "unknown error").strip()
        raise InventoryVerificationError(f"Excel 계산 실패: {detail}")


def collect_source_rows(
    *,
    department_dir: Path,
    warehouse_workbook: Path,
    warehouse_sheet: str | None,
    calculator: Callable[[Iterable[Path]], None] = calculate_workbook_copies,
) -> SourceSnapshot:
    """Calculate local copies, extract current stock, and prove sources unchanged."""

    department_paths = discover_department_workbooks(department_dir)
    warehouse_workbook = Path(warehouse_workbook).resolve()
    if not warehouse_workbook.is_file():
        raise InventoryVerificationError(f"창고 파일이 없습니다: {warehouse_workbook}")
    source_paths = (*department_paths.values(), warehouse_workbook)
    assert_sources_unlocked(source_paths)
    hashes_before = {
        str(Path(path).resolve()): file_sha256(Path(path)) for path in source_paths
    }

    rows: list[SourceRow] = []
    selected_warehouse_sheet = ""
    with tempfile.TemporaryDirectory(prefix="mes-inventory-verification-") as temp_name:
        temp_dir = Path(temp_name)
        copied_departments: dict[str, Path] = {}
        for index, (group, source_path) in enumerate(
            department_paths.items(), start=1
        ):
            copied = temp_dir / f"department_{index:02d}.xlsx"
            shutil.copy2(source_path, copied)
            copied_departments[group] = copied
        copied_warehouse = temp_dir / "warehouse.xlsx"
        shutil.copy2(warehouse_workbook, copied_warehouse)
        calculator((*copied_departments.values(), copied_warehouse))

        for group, copied in copied_departments.items():
            rows.extend(
                extract_rows_from_workbook(
                    copied,
                    source_kind="department",
                    source_group=group,
                    sheet_name=DEPARTMENT_SHEETS[group],
                    source_display_path=department_paths[group].resolve(),
                )
            )
        warehouse_rows = extract_rows_from_workbook(
            copied_warehouse,
            source_kind="warehouse",
            source_group="warehouse",
            sheet_name=warehouse_sheet,
            source_display_path=warehouse_workbook,
        )
        selected_warehouse_sheet = warehouse_rows[0].sheet_name
        rows.extend(warehouse_rows)

    hashes_after = {
        str(Path(path).resolve()): file_sha256(Path(path)) for path in source_paths
    }
    if hashes_after != hashes_before:
        raise InventoryVerificationError("검증 중 원본 Excel 파일이 변경되었습니다")
    return SourceSnapshot(
        rows=tuple(rows),
        source_hashes=hashes_before,
        warehouse_sheet=selected_warehouse_sheet,
    )


def _latest_month_sheet(sheet_names: Iterable[str]) -> str:
    months: list[tuple[int, int, str]] = []
    for name in sheet_names:
        match = MONTH_SHEET_PATTERN.fullmatch(name)
        if match:
            months.append((int(match.group(1)), int(match.group(2)), name))
    if not months:
        raise InventoryVerificationError("YY.MM월 형식의 창고 시트를 찾지 못했습니다")
    return max(months)[2]


def _find_header_layout(
    sheet: Any, *, preferred_quantity_header: str
) -> tuple[int, dict[str, int]]:
    """Locate all required semantic columns on one of the first ten rows."""

    max_column = max(1, sheet.max_column or 1)
    for row_number in range(1, min(10, sheet.max_row or 0) + 1):
        found: dict[str, int] = {}
        for column in range(1, max_column + 1):
            value = _compact(sheet.cell(row=row_number, column=column).value)
            if not value:
                continue
            for role, aliases in HEADER_ALIASES.items():
                if value in aliases:
                    if role in found:
                        if role == "original_name":
                            found[role] = max(found[role], column)
                            continue
                        if role == "quantity":
                            previous_value = _compact(
                                sheet.cell(row=row_number, column=found[role]).value
                            )
                            if value == preferred_quantity_header:
                                if previous_value == preferred_quantity_header:
                                    raise InventoryVerificationError(
                                        f"{sheet.title} {row_number}행에 {role} 헤더가 중복되었습니다"
                                    )
                                found[role] = column
                                continue
                            if previous_value == preferred_quantity_header:
                                continue
                        raise InventoryVerificationError(
                            f"{sheet.title} {row_number}행에 {role} 헤더가 중복되었습니다"
                        )
                    found[role] = column
        if set(found) == set(HEADER_ALIASES):
            return row_number, found
    raise InventoryVerificationError(
        f"{sheet.title} 첫 10행에서 필수 재고/MES 헤더를 찾지 못했습니다"
    )


def extract_rows_from_workbook(
    path: Path,
    *,
    source_kind: str,
    source_group: str,
    sheet_name: str | None = None,
    source_display_path: Path | None = None,
) -> tuple[SourceRow, ...]:
    """Read calculated values by discovered headers without saving the workbook."""

    path = Path(path)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="wmf image format is not supported.*",
            category=UserWarning,
        )
        workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        selected_sheet = sheet_name
        if selected_sheet is None:
            if source_kind != "warehouse":
                raise InventoryVerificationError("부서 파일은 시트명을 지정해야 합니다")
            selected_sheet = _latest_month_sheet(workbook.sheetnames)
        if selected_sheet not in workbook.sheetnames:
            raise InventoryVerificationError(
                f"시트를 찾지 못했습니다: {path.name}!{selected_sheet}"
            )
        sheet = workbook[selected_sheet]
        preferred_quantity_header = (
            "총합" if source_group == "finished" else "현재고"
        )
        header_row, columns = _find_header_layout(
            sheet, preferred_quantity_header=preferred_quantity_header
        )
        rows: list[SourceRow] = []
        for row_number in range(header_row + 1, (sheet.max_row or header_row) + 1):
            original_name = _text(
                sheet.cell(row=row_number, column=columns["original_name"]).value
            )
            mes_name = _text(sheet.cell(row=row_number, column=columns["mes_name"]).value)
            mes_code = _text(sheet.cell(row=row_number, column=columns["mes_code"]).value)
            if not (original_name or mes_name or mes_code):
                continue
            rows.append(
                SourceRow(
                    source_kind=source_kind,
                    source_group=source_group,
                    source_path=source_display_path or path,
                    sheet_name=selected_sheet,
                    row_number=row_number,
                    original_name=original_name,
                    mes_name=mes_name,
                    mes_code=mes_code.upper(),
                    confirmation=_text(
                        sheet.cell(row=row_number, column=columns["confirmation"]).value
                    ),
                    quantity=sheet.cell(
                        row=row_number, column=columns["quantity"]
                    ).value,
                )
            )
        if not rows:
            raise InventoryVerificationError(f"재고 행이 없습니다: {path.name}!{selected_sheet}")
        return tuple(rows)
    finally:
        workbook.close()


def copy_sqlite_snapshot(source: Path, target: Path) -> None:
    """Copy a consistent SQLite snapshot from a strictly read-only source."""

    source = Path(source).resolve()
    target = Path(target).resolve()
    if not source.is_file():
        raise InventoryVerificationError(f"직원 DB가 없습니다: {source}")
    if target.exists():
        raise InventoryVerificationError(f"DB 스냅샷 대상이 이미 존재합니다: {target}")
    target.parent.mkdir(parents=True, exist_ok=True)
    uri = f"file:{source.as_posix()}?mode=ro"
    try:
        with closing(sqlite3.connect(uri, uri=True)) as source_connection:
            source_connection.execute("PRAGMA query_only=ON")
            with closing(sqlite3.connect(target)) as target_connection:
                source_connection.backup(target_connection)
    except sqlite3.Error as exc:
        raise InventoryVerificationError(f"직원 DB 스냅샷 생성 실패: {exc}") from exc


def load_mes_items(snapshot_path: Path) -> tuple[MesItem, ...]:
    """Load item master and bucket quantities from a local read-only snapshot."""

    snapshot_path = Path(snapshot_path).resolve()
    uri = f"file:{snapshot_path.as_posix()}?mode=ro"
    try:
        with closing(sqlite3.connect(uri, uri=True)) as connection:
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA query_only=ON")
            item_rows = connection.execute(
                """
                SELECT i.item_id, i.mes_code, i.item_name, i.process_type_code,
                       i.deleted_at, COALESCE(inv.warehouse_qty, 0) AS warehouse_qty
                FROM items AS i
                LEFT JOIN inventory AS inv ON inv.item_id = i.item_id
                ORDER BY i.mes_code
                """
            ).fetchall()
            location_rows = connection.execute(
                """
                SELECT item_id, department, SUM(quantity) AS quantity
                FROM inventory_locations
                WHERE status = 'PRODUCTION'
                GROUP BY item_id, department
                """
            ).fetchall()
    except sqlite3.Error as exc:
        raise InventoryVerificationError(f"직원 DB 조회 실패: {exc}") from exc

    quantities: defaultdict[str, dict[str, int]] = defaultdict(dict)
    for row in location_rows:
        quantities[str(row["item_id"])][str(row["department"])] = int(
            row["quantity"] or 0
        )
    return tuple(
        MesItem(
            item_id=str(row["item_id"]),
            mes_code=_text(row["mes_code"]).upper(),
            item_name=_text(row["item_name"]),
            process_type_code=_text(row["process_type_code"]).upper(),
            deleted_at=_text(row["deleted_at"]) or None,
            warehouse_quantity=int(row["warehouse_qty"] or 0),
            department_quantities=quantities.get(str(row["item_id"]), {}),
        )
        for row in item_rows
    )


def _quantity(value: object) -> int:
    """Return a non-negative integer or raise for a row-level source error."""

    if isinstance(value, bool) or value is None or _text(value) == "":
        raise ValueError("현재고가 비어 있습니다")
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("현재고가 숫자가 아닙니다") from exc
    if not number.is_finite() or number < 0 or number != number.to_integral_value():
        raise ValueError("현재고는 0 이상의 정수여야 합니다")
    return int(number)


def _department_for_item(item: MesItem) -> str | None:
    code = _text(item.process_type_code).upper()
    return PROCESS_DEPARTMENTS.get(code[:1]) if code else None


def _mes_quantity(item: MesItem, bucket: str) -> int:
    if bucket == "창고":
        return int(item.warehouse_quantity or 0)
    return int(item.department_quantities.get(bucket, 0) or 0)


def _issue_for_row(
    row: SourceRow,
    *,
    severity: str,
    issue_type: str,
    message: str,
    quantity: int | None = None,
    mes_name: str = "",
) -> VerificationIssue:
    return VerificationIssue(
        severity=severity,
        issue_type=issue_type,
        message=message,
        source_path=str(row.source_path),
        sheet_name=row.sheet_name,
        row_number=row.row_number,
        mes_code=_text(row.mes_code).upper(),
        excel_name=_text(row.mes_name or row.original_name),
        mes_name=mes_name,
        quantity=quantity,
    )


def compare_inventory(
    source_rows: Iterable[SourceRow],
    mes_items: Iterable[MesItem],
) -> VerificationResult:
    """Compare valid codes only; never infer identity from an item name."""

    items_by_code = {_text(item.mes_code).upper(): item for item in mes_items}
    grouped: defaultdict[
        tuple[str, str], list[tuple[SourceRow, int, MesItem]]
    ] = defaultdict(list)
    issues: list[VerificationIssue] = []

    for row in source_rows:
        code = _text(row.mes_code).upper()
        try:
            quantity = _quantity(row.quantity)
        except ValueError as exc:
            issues.append(
                _issue_for_row(
                    row,
                    severity="오류",
                    issue_type="원본 수량 오류",
                    message=str(exc),
                )
            )
            continue
        if not code:
            issues.append(
                _issue_for_row(
                    row,
                    severity="오류",
                    issue_type="MES 코드 없음",
                    message="MES 코드가 없어 수량을 비교할 수 없습니다",
                    quantity=quantity,
                )
            )
            continue
        item = items_by_code.get(code)
        if item is None:
            issues.append(
                _issue_for_row(
                    row,
                    severity="오류",
                    issue_type="MES 코드 미등록",
                    message="직원 MES에 등록된 코드가 아닙니다",
                    quantity=quantity,
                )
            )
            continue
        bucket = "창고" if row.source_kind == "warehouse" else _department_for_item(item)
        if bucket is None:
            issues.append(
                _issue_for_row(
                    row,
                    severity="오류",
                    issue_type="지원하지 않는 공정코드",
                    message=f"공정코드 {item.process_type_code!r}의 부서를 결정할 수 없습니다",
                    quantity=quantity,
                    mes_name=item.item_name,
                )
            )
            continue
        if _text(row.confirmation).upper() != "O":
            issues.append(
                _issue_for_row(
                    row,
                    severity="경고",
                    issue_type="담당자 미확인",
                    message="담당자 확인란에 O가 없지만 코드 기준으로 비교했습니다",
                    quantity=quantity,
                    mes_name=item.item_name,
                )
            )
        if _text(row.mes_name) and _text(row.mes_name) != _text(item.item_name):
            issues.append(
                _issue_for_row(
                    row,
                    severity="경고",
                    issue_type="품명 불일치",
                    message="Excel MES 품목명이 직원 MES의 현재 품명과 다릅니다",
                    quantity=quantity,
                    mes_name=item.item_name,
                )
            )
        allowed = SOURCE_DEPARTMENTS.get(row.source_group)
        if row.source_kind == "department" and allowed is not None and bucket not in allowed:
            issues.append(
                _issue_for_row(
                    row,
                    severity="경고",
                    issue_type="잘못된 부서 파일",
                    message=f"품목코드 기준 부서는 {bucket}입니다",
                    quantity=quantity,
                    mes_name=item.item_name,
                )
            )
        grouped[(bucket, code)].append((row, quantity, item))

    comparisons: list[ComparisonRow] = []
    seen_keys = set(grouped)
    for (bucket, code), group in grouped.items():
        item = group[0][2]
        excel_quantity = sum(quantity for _, quantity, _ in group)
        mes_quantity = _mes_quantity(item, bucket)
        difference = mes_quantity - excel_quantity
        confirmed = all(_text(row.confirmation).upper() == "O" for row, _, _ in group)
        comparisons.append(
            ComparisonRow(
                bucket=bucket,
                mes_code=code,
                mes_name=item.item_name,
                excel_quantity=excel_quantity,
                mes_quantity=mes_quantity,
                difference=difference,
                status="일치" if difference == 0 else "수량 불일치",
                confirmed=confirmed,
                source_rows=tuple(
                    f"{row.source_path.name}!{row.sheet_name}!{row.row_number}"
                    for row, _, _ in group
                ),
            )
        )
        if item.deleted_at and (excel_quantity > 0 or mes_quantity > 0):
            issues.append(
                VerificationIssue(
                    severity="오류",
                    issue_type="삭제 품목 재고 있음",
                    message="MES에서 삭제된 품목에 Excel 또는 MES 재고가 있습니다",
                    mes_code=code,
                    excel_name=_text(group[0][0].mes_name or group[0][0].original_name),
                    mes_name=item.item_name,
                    quantity=max(excel_quantity, mes_quantity),
                )
            )

    for item in items_by_code.values():
        for bucket in BUCKET_ORDER:
            mes_quantity = _mes_quantity(item, bucket)
            key = (bucket, item.mes_code.upper())
            if mes_quantity <= 0 or key in seen_keys:
                continue
            comparisons.append(
                ComparisonRow(
                    bucket=bucket,
                    mes_code=item.mes_code.upper(),
                    mes_name=item.item_name,
                    excel_quantity=0,
                    mes_quantity=mes_quantity,
                    difference=mes_quantity,
                    status="Excel 누락",
                    confirmed=False,
                    source_rows=(),
                )
            )
            if item.deleted_at:
                issues.append(
                    VerificationIssue(
                        severity="오류",
                        issue_type="삭제 품목 재고 있음",
                        message="MES에서 삭제된 품목에 MES 재고가 있습니다",
                        mes_code=item.mes_code.upper(),
                        mes_name=item.item_name,
                        quantity=mes_quantity,
                    )
                )

    order = {bucket: index for index, bucket in enumerate(BUCKET_ORDER)}
    comparisons.sort(key=lambda row: (order.get(row.bucket, 99), row.mes_code))
    issues.sort(
        key=lambda issue: (
            0 if issue.severity == "오류" else 1,
            issue.source_path,
            issue.sheet_name,
            issue.row_number or 0,
            issue.mes_code,
            issue.issue_type,
        )
    )
    return VerificationResult(comparisons=tuple(comparisons), issues=tuple(issues))


TITLE_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="FFD9EAF7")
MATCH_FILL = PatternFill("solid", fgColor="FFC6EFCE")
WARNING_FILL = PatternFill("solid", fgColor="FFFFF2CC")
ERROR_FILL = PatternFill("solid", fgColor="FFF4CCCC")


def _summary_payload(result: VerificationResult) -> dict[str, Any]:
    bucket_totals = []
    for bucket in BUCKET_ORDER:
        rows = [row for row in result.comparisons if row.bucket == bucket]
        bucket_totals.append(
            {
                "bucket": bucket,
                "excel_quantity": sum(row.excel_quantity for row in rows),
                "mes_quantity": sum(row.mes_quantity for row in rows),
                "difference": sum(row.difference for row in rows),
                "match_count": sum(row.status == "일치" for row in rows),
                "difference_count": sum(row.status != "일치" for row in rows),
            }
        )
    return {
        "comparison_count": len(result.comparisons),
        "difference_count": sum(
            row.status != "일치" for row in result.comparisons
        ),
        "issue_count": len(result.issues),
        "error_count": sum(issue.severity == "오류" for issue in result.issues),
        "warning_count": sum(issue.severity == "경고" for issue in result.issues),
        "bucket_totals": bucket_totals,
    }


def _style_table(sheet: Any, widths: tuple[int, ...]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="1F1F1F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_reports(
    result: VerificationResult,
    metadata: RunMetadata,
    output_dir: Path,
    *,
    timestamp: datetime | None = None,
) -> ReportPaths:
    """Write human-readable Excel and machine-readable JSON from one result."""

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = timestamp or datetime.now()
    stem = f"inventory_verification_{timestamp:%Y%m%d_%H%M%S}"
    xlsx_path = output_dir / f"{stem}.xlsx"
    json_path = output_dir / f"{stem}.json"
    summary = _summary_payload(result)
    payload = {
        "metadata": asdict(metadata),
        "summary": summary,
        "comparisons": [asdict(row) for row in result.comparisons],
        "issues": [asdict(issue) for issue in result.issues],
    }
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "요약"
    summary_sheet.merge_cells("A1:F1")
    summary_sheet["A1"] = "DEXCOWIN MES 주간 재고 정합성 검증"
    summary_sheet["A1"].fill = TITLE_FILL
    summary_sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    metadata_rows = (
        ("검증 시각", metadata.verified_at),
        ("직원 DB", metadata.employee_db_path),
        ("직원 DB SHA-256", metadata.employee_db_hash),
        ("창고 시트", metadata.warehouse_sheet),
        ("전체 비교", summary["comparison_count"]),
        ("수량 차이", summary["difference_count"]),
        ("품목 오류·경고", summary["issue_count"]),
    )
    for row_number, (label, value) in enumerate(metadata_rows, start=3):
        summary_sheet.cell(row_number, 1, label).font = Font(bold=True)
        summary_sheet.cell(row_number, 2, value)
    summary_start = 12
    summary_headers = ("구분", "Excel 합계", "MES 합계", "차이", "일치", "불일치")
    summary_sheet.append([])
    for column, value in enumerate(summary_headers, start=1):
        summary_sheet.cell(summary_start, column, value)
    for offset, bucket in enumerate(summary["bucket_totals"], start=1):
        values = (
            bucket["bucket"],
            bucket["excel_quantity"],
            bucket["mes_quantity"],
            bucket["difference"],
            bucket["match_count"],
            bucket["difference_count"],
        )
        for column, value in enumerate(values, start=1):
            summary_sheet.cell(summary_start + offset, column, value)
    for cell in summary_sheet[summary_start]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 28
    for column in "CDEF":
        summary_sheet.column_dimensions[column].width = 16
    source_start = summary_start + len(BUCKET_ORDER) + 3
    summary_sheet.merge_cells(
        start_row=source_start, start_column=1, end_row=source_start, end_column=2
    )
    summary_sheet.merge_cells(
        start_row=source_start, start_column=3, end_row=source_start, end_column=6
    )
    summary_sheet.cell(source_start, 1, "원본 파일")
    summary_sheet.cell(source_start, 3, "SHA-256")
    for cell in summary_sheet[source_start]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for offset, (path, digest) in enumerate(
        sorted(metadata.source_hashes.items()), start=1
    ):
        row_number = source_start + offset
        summary_sheet.merge_cells(
            start_row=row_number, start_column=1, end_row=row_number, end_column=2
        )
        summary_sheet.merge_cells(
            start_row=row_number, start_column=3, end_row=row_number, end_column=6
        )
        summary_sheet.cell(row_number, 1, path).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        summary_sheet.cell(row_number, 3, digest).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        summary_sheet.row_dimensions[row_number].height = 45

    difference_sheet = workbook.create_sheet("수량차이")
    difference_sheet.append(
        ["구분", "MES 코드", "MES 품명", "Excel 수량", "MES 수량", "차이", "원본 행"]
    )
    for row in result.comparisons:
        if row.status == "일치":
            continue
        difference_sheet.append(
            [
                row.bucket,
                row.mes_code,
                row.mes_name,
                row.excel_quantity,
                row.mes_quantity,
                row.difference,
                "\n".join(row.source_rows),
            ]
        )
        for cell in difference_sheet[difference_sheet.max_row]:
            cell.fill = ERROR_FILL
    _style_table(difference_sheet, (12, 18, 42, 14, 14, 12, 50))

    issue_sheet = workbook.create_sheet("품목오류")
    issue_sheet.append(
        [
            "심각도",
            "유형",
            "내용",
            "원본 파일",
            "시트",
            "행",
            "MES 코드",
            "Excel MES 품명",
            "현재 MES 품명",
            "수량",
        ]
    )
    for issue in result.issues:
        issue_sheet.append(
            [
                issue.severity,
                issue.issue_type,
                issue.message,
                issue.source_path,
                issue.sheet_name,
                issue.row_number,
                issue.mes_code,
                issue.excel_name,
                issue.mes_name,
                issue.quantity,
            ]
        )
        fill = ERROR_FILL if issue.severity == "오류" else WARNING_FILL
        for cell in issue_sheet[issue_sheet.max_row]:
            cell.fill = fill
    _style_table(issue_sheet, (10, 22, 42, 46, 18, 8, 18, 38, 38, 12))

    all_sheet = workbook.create_sheet("전체대조")
    all_sheet.append(
        [
            "구분",
            "MES 코드",
            "MES 품명",
            "Excel 수량",
            "MES 수량",
            "차이",
            "판정",
            "담당자 확인",
            "원본 행",
        ]
    )
    for row in result.comparisons:
        all_sheet.append(
            [
                row.bucket,
                row.mes_code,
                row.mes_name,
                row.excel_quantity,
                row.mes_quantity,
                row.difference,
                row.status,
                "O" if row.confirmed else "",
                "\n".join(row.source_rows),
            ]
        )
        status_cell = all_sheet.cell(all_sheet.max_row, 7)
        status_cell.fill = MATCH_FILL if row.status == "일치" else ERROR_FILL
    _style_table(all_sheet, (12, 18, 42, 14, 14, 12, 18, 14, 50))

    workbook.save(xlsx_path)
    workbook.close()
    return ReportPaths(xlsx_path=xlsx_path, json_path=json_path)


def run_verification(
    *,
    department_dir: Path,
    warehouse_workbook: Path,
    warehouse_sheet: str | None,
    employee_db: Path,
    output_dir: Path,
    calculator: Callable[[Iterable[Path]], None] = calculate_workbook_copies,
    timestamp: datetime | None = None,
) -> VerificationRun:
    """Run one complete comparison without exposing any source write operation."""

    timestamp = timestamp or datetime.now().astimezone()
    source_snapshot = collect_source_rows(
        department_dir=department_dir,
        warehouse_workbook=warehouse_workbook,
        warehouse_sheet=warehouse_sheet,
        calculator=calculator,
    )
    employee_db = Path(employee_db).resolve()
    employee_hash_before = file_sha256(employee_db)
    with tempfile.TemporaryDirectory(prefix="mes-db-verification-") as temp_name:
        snapshot_path = Path(temp_name) / "employee_snapshot.db"
        copy_sqlite_snapshot(employee_db, snapshot_path)
        mes_items = load_mes_items(snapshot_path)
    employee_hash_after = file_sha256(employee_db)
    if employee_hash_after != employee_hash_before:
        raise InventoryVerificationError("검증 중 직원 DB가 변경되었습니다")

    result = compare_inventory(source_snapshot.rows, mes_items)
    metadata = RunMetadata(
        verified_at=timestamp.astimezone().isoformat(timespec="seconds"),
        source_hashes=source_snapshot.source_hashes,
        employee_db_path=str(employee_db),
        employee_db_hash=employee_hash_before,
        warehouse_sheet=source_snapshot.warehouse_sheet,
    )
    report_paths = write_reports(
        result,
        metadata,
        output_dir,
        timestamp=timestamp,
    )
    return VerificationRun(
        result=result,
        metadata=metadata,
        report_paths=report_paths,
    )


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="부서·창고 Excel 현재고와 직원 DEXCOWIN MES 재고를 읽기 전용으로 비교합니다."
    )
    parser.add_argument("--department-dir", type=Path, required=True)
    parser.add_argument("--warehouse-workbook", type=Path, required=True)
    parser.add_argument("--warehouse-sheet")
    parser.add_argument("--employee-db", type=Path, default=DEFAULT_EMPLOYEE_DB)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI boundary: comparison differences are reports, not process failures."""

    args = _parse_args(argv)
    try:
        run = run_verification(
            department_dir=args.department_dir,
            warehouse_workbook=args.warehouse_workbook,
            warehouse_sheet=args.warehouse_sheet,
            employee_db=args.employee_db,
            output_dir=args.output_dir,
        )
    except InventoryVerificationError as exc:
        print(f"[INVENTORY VERIFICATION] ERROR: {exc}", file=sys.stderr)
        return 2
    summary = _summary_payload(run.result)
    print("[INVENTORY VERIFICATION] COMPLETED")
    print(json.dumps(summary, ensure_ascii=True, indent=2))
    print(f"EXCEL_REPORT={run.report_paths.xlsx_path.resolve()}")
    print(f"JSON_REPORT={run.report_paths.json_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
