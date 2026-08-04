#!/usr/bin/env python3
"""Synchronize the warehouse inventory baseline without changing department stock."""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path
import sqlite3
import sys
from typing import Any, Callable
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import create_engine, func
from sqlalchemy.orm import Session, sessionmaker


PROJECT_ROOT = Path(__file__).resolve().parents[2]
BACKEND_DIR = PROJECT_ROOT / "backend"

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.models import (  # noqa: E402
    Inventory,
    InventoryLocation,
    Item,
    ProcessType,
    WarehouseBoxItem,
    WarehouseSpecialZoneItem,
)


class WarehouseSyncError(ValueError):
    """Raised when the warehouse baseline cannot be applied safely."""


@dataclass(frozen=True)
class WorkbookSnapshot:
    """Validated warehouse quantities aggregated by current MES code."""

    row_count: int
    quantities: dict[str, int]
    mes_names: dict[str, str]
    duplicate_groups: int


@dataclass(frozen=True)
class EmployeeItem:
    """Current employee-server item fields required by the warehouse sync."""

    item_id: str
    mes_code: str
    item_name: str
    sort_order: int | None
    unit: str
    legacy_part: str | None
    legacy_item_type: str | None
    supplier: str | None
    min_stock: int | None
    model_symbol: str
    process_type_code: str
    serial_no: int
    bom_completed_at: datetime | None
    sales_review_required: bool


@dataclass(frozen=True)
class SyncSummary:
    """Dry-run or applied warehouse synchronization totals."""

    applied: bool
    source_rows: int
    source_codes: int
    source_quantity: int
    duplicate_groups: int
    employee_name_mismatches: int
    master_items_added: int
    master_items_updated: int
    inventory_items_changed: int
    absent_nonzero_items_zeroed: int
    absent_quantity_zeroed: int
    source_zero_nonzero_items: int
    source_zero_quantity: int
    warehouse_box_items_deleted: int
    special_zone_items_deleted: int


def _parse_quantity(value: object, row_number: int) -> int:
    """Convert one cached Excel quantity while rejecting unsafe values."""
    text = str(value if value is not None else 0).strip().replace(",", "")
    if not text:
        return 0
    try:
        quantity = Decimal(text)
    except InvalidOperation as exc:
        raise WarehouseSyncError(f"row {row_number}: invalid quantity {value!r}") from exc
    if quantity < 0 or quantity != quantity.to_integral_value():
        raise WarehouseSyncError(f"row {row_number}: quantity must be a non-negative integer")
    return int(quantity)


def parse_workbook(path: Path, sheet_name: str) -> WorkbookSnapshot:
    """Read the confirmed warehouse rows and aggregate duplicate MES codes."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise WarehouseSyncError(f"worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(min_row=3, max_col=62, values_only=True)
        header_row = next(rows)
        for column_index, cell, expected in _EXPECTED_HEADERS:
            actual = str(header_row[column_index] or "").strip()
            if actual != expected:
                raise WarehouseSyncError(
                    f"expected header {expected!r} at {cell}, found {actual!r}"
                )
        quantities: defaultdict[str, int] = defaultdict(int)
        mes_names: dict[str, str] = {}
        row_count = 0
        row_counts: defaultdict[str, int] = defaultdict(int)
        for row_number, values in enumerate(rows, start=4):
            excel_name = str(values[3] or "").strip()
            mes_name = str(values[57] or "").strip()
            mes_code = str(values[59] or "").strip().upper()
            if not any((excel_name, mes_name, mes_code)):
                continue
            if not mes_code:
                raise WarehouseSyncError(f"row {row_number}: MES 코드 is required")
            if not mes_name:
                raise WarehouseSyncError(f"row {row_number}: MES 품목명 is required")
            confirmation = str(values[60] or "").strip().upper()
            state = str(values[61] or "").strip()
            if confirmation != "O" or state != "확인완료":
                raise WarehouseSyncError(f"row {row_number}: 확인 상태 is not complete")
            quantity = _parse_quantity(values[14], row_number)
            previous_name = mes_names.get(mes_code)
            if previous_name is not None and previous_name != mes_name:
                raise WarehouseSyncError(
                    f"conflicting MES names for {mes_code}: {previous_name!r} vs {mes_name!r}"
                )
            quantities[mes_code] += quantity
            mes_names[mes_code] = mes_name
            row_counts[mes_code] += 1
            row_count += 1
        if row_count == 0:
            raise WarehouseSyncError("worksheet has no inventory rows")
        return WorkbookSnapshot(
            row_count=row_count,
            quantities=dict(quantities),
            mes_names=mes_names,
            duplicate_groups=sum(1 for count in row_counts.values() if count > 1),
        )
    finally:
        workbook.close()


def _optional_datetime(value: object) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    return datetime.fromisoformat(str(value))


def _normalize_item_id(value: object) -> str:
    """Return the canonical 32-character UUID used by SQLite UUIDString columns."""
    try:
        return UUID(str(value)).hex
    except (TypeError, ValueError, AttributeError) as exc:
        raise WarehouseSyncError(f"invalid item_id in employee DB: {value!r}") from exc


def _load_employee_items(path: Path, codes: set[str]) -> dict[str, EmployeeItem]:
    """Load active employee-server items through a read-only SQLite connection."""
    if not path.is_file():
        raise WarehouseSyncError(f"employee DB not found: {path}")
    placeholders = ",".join("?" for _ in codes)
    uri = f"file:{path.resolve().as_posix()}?mode=ro"
    with sqlite3.connect(uri, uri=True) as connection:
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA query_only=ON")
        rows = connection.execute(
            f"""
            SELECT item_id, mes_code, item_name, sort_order, unit, legacy_part,
                   legacy_item_type, supplier, min_stock, model_symbol,
                   process_type_code, serial_no, bom_completed_at,
                   sales_review_required
            FROM items
            WHERE deleted_at IS NULL AND mes_code IN ({placeholders})
            """,
            sorted(codes),
        ).fetchall()
    return {
        str(row["mes_code"]): EmployeeItem(
            item_id=_normalize_item_id(row["item_id"]),
            mes_code=str(row["mes_code"]),
            item_name=str(row["item_name"]),
            sort_order=row["sort_order"],
            unit=str(row["unit"]),
            legacy_part=row["legacy_part"],
            legacy_item_type=row["legacy_item_type"],
            supplier=row["supplier"],
            min_stock=row["min_stock"],
            model_symbol=str(row["model_symbol"]),
            process_type_code=str(row["process_type_code"]),
            serial_no=int(row["serial_no"]),
            bom_completed_at=_optional_datetime(row["bom_completed_at"]),
            sales_review_required=bool(row["sales_review_required"]),
        )
        for row in rows
    }


_ITEM_FIELDS = (
    "item_name",
    "sort_order",
    "unit",
    "legacy_part",
    "legacy_item_type",
    "supplier",
    "min_stock",
    "model_symbol",
    "process_type_code",
    "serial_no",
    "bom_completed_at",
    "sales_review_required",
)

_EXISTING_ITEM_FIELDS = (
    "item_name",
    "model_symbol",
    "process_type_code",
    "serial_no",
)


_EXPECTED_HEADERS = (
    (3, "D3", "품명"),
    (14, "O3", "현재고"),
    (57, "BF3", "MES 품목명"),
    (59, "BH3", "MES 코드"),
    (60, "BI3", "담당자 확인"),
    (61, "BJ3", "확인"),
)


def _item_values(item: EmployeeItem) -> dict[str, Any]:
    return {field: getattr(item, field) for field in _ITEM_FIELDS}


def _item_differs(item: Item, employee_item: EmployeeItem) -> bool:
    return any(
        getattr(item, field) != getattr(employee_item, field)
        for field in _EXISTING_ITEM_FIELDS
    )


def _location_totals(db: Session) -> dict[str, int]:
    rows = db.query(
        InventoryLocation.item_id,
        func.coalesce(func.sum(InventoryLocation.quantity), 0),
    ).group_by(InventoryLocation.item_id)
    return {_normalize_item_id(item_id): int(quantity) for item_id, quantity in rows}


def _validate_sync(
    db: Session,
    employee_items: dict[str, EmployeeItem],
    snapshot: WorkbookSnapshot,
) -> tuple[list[EmployeeItem], list[tuple[Item, EmployeeItem]], dict[str, str], int]:
    missing = sorted(set(snapshot.quantities) - set(employee_items))
    if missing:
        raise WarehouseSyncError(f"employee DB has no active items for: {', '.join(missing)}")
    mismatched_names = sorted(
        code
        for code, item in employee_items.items()
        if snapshot.mes_names[code] != item.item_name
    )

    dev_items = db.query(Item).all()
    by_id = {_normalize_item_id(item.item_id): item for item in dev_items}
    by_code = {item.mes_code: item for item in dev_items}
    additions: list[EmployeeItem] = []
    updates: list[tuple[Item, EmployeeItem]] = []
    projected_codes = {_normalize_item_id(item.item_id): item.mes_code for item in dev_items}
    for employee_item in employee_items.values():
        existing_by_id = by_id.get(employee_item.item_id)
        existing_by_code = by_code.get(employee_item.mes_code)
        if existing_by_code is not None and _normalize_item_id(existing_by_code.item_id) != employee_item.item_id:
            raise WarehouseSyncError(
                f"MES code collision: {employee_item.mes_code} belongs to another dev item"
            )
        if existing_by_id is None:
            additions.append(employee_item)
        else:
            projected_codes[employee_item.item_id] = employee_item.mes_code
            if _item_differs(existing_by_id, employee_item):
                updates.append((existing_by_id, employee_item))

    known_process_types = {
        code for (code,) in db.query(ProcessType.code).filter(
            ProcessType.code.in_({item.process_type_code for item in employee_items.values()})
        )
    }
    missing_process_types = sorted(
        {item.process_type_code for item in employee_items.values()} - known_process_types
    )
    if missing_process_types:
        raise WarehouseSyncError(f"dev DB is missing process types: {', '.join(missing_process_types)}")

    projected_codes.update({item.item_id: item.mes_code for item in additions})
    code_owners: dict[str, str] = {}
    for item_id, code in projected_codes.items():
        owner = code_owners.get(code)
        if owner is not None and owner != item_id:
            raise WarehouseSyncError(f"projected MES code collision: {code}")
        code_owners[code] = item_id

    inventory_by_id = {_normalize_item_id(row.item_id): row for row in db.query(Inventory).all()}
    for item_id, code in projected_codes.items():
        inventory = inventory_by_id.get(item_id)
        pending = int(inventory.pending_quantity or 0) if inventory is not None else 0
        target = snapshot.quantities.get(code, 0)
        if target < pending:
            raise WarehouseSyncError(
                f"pending quantity exceeds warehouse target: {code} target={target} pending={pending}"
            )
    return additions, updates, projected_codes, len(mismatched_names)


def run_sync(
    db: Session,
    employee_db_path: Path,
    snapshot: WorkbookSnapshot,
    *,
    apply: bool = False,
) -> SyncSummary:
    """Validate and optionally apply one atomic warehouse-only baseline sync."""
    employee_items = _load_employee_items(employee_db_path, set(snapshot.quantities))
    additions, updates, projected_codes, name_mismatch_count = _validate_sync(
        db, employee_items, snapshot
    )
    inventory_by_id = {_normalize_item_id(row.item_id): row for row in db.query(Inventory).all()}
    changed = 0
    absent_zeroed = 0
    absent_zeroed_quantity = 0
    source_zeroed = 0
    source_zeroed_quantity = 0
    for item_id, code in projected_codes.items():
        current = int(inventory_by_id[item_id].warehouse_qty or 0) if item_id in inventory_by_id else 0
        target = snapshot.quantities.get(code, 0)
        if current != target:
            changed += 1
            if target == 0 and current > 0:
                if code in snapshot.quantities:
                    source_zeroed += 1
                    source_zeroed_quantity += current
                else:
                    absent_zeroed += 1
                    absent_zeroed_quantity += current

    box_count = db.query(WarehouseBoxItem).count()
    zone_count = db.query(WarehouseSpecialZoneItem).count()
    summary = SyncSummary(
        applied=apply,
        source_rows=snapshot.row_count,
        source_codes=len(snapshot.quantities),
        source_quantity=sum(snapshot.quantities.values()),
        duplicate_groups=snapshot.duplicate_groups,
        employee_name_mismatches=name_mismatch_count,
        master_items_added=len(additions),
        master_items_updated=len(updates),
        inventory_items_changed=changed,
        absent_nonzero_items_zeroed=absent_zeroed,
        absent_quantity_zeroed=absent_zeroed_quantity,
        source_zero_nonzero_items=source_zeroed,
        source_zero_quantity=source_zeroed_quantity,
        warehouse_box_items_deleted=box_count,
        special_zone_items_deleted=zone_count,
    )
    if not apply:
        return summary

    try:
        for item, employee_item in updates:
            for field in _EXISTING_ITEM_FIELDS:
                setattr(item, field, getattr(employee_item, field))
        for employee_item in additions:
            db.add(Item(item_id=UUID(employee_item.item_id), **_item_values(employee_item)))
        db.flush()

        location_totals = _location_totals(db)
        inventories = {_normalize_item_id(row.item_id): row for row in db.query(Inventory).all()}
        for item in db.query(Item).all():
            item_id = _normalize_item_id(item.item_id)
            inventory = inventories.get(item_id)
            if inventory is None:
                inventory = Inventory(
                    item_id=item.item_id,
                    warehouse_qty=0,
                    quantity=0,
                    pending_quantity=0,
                )
                db.add(inventory)
            target = snapshot.quantities.get(item.mes_code, 0)
            inventory.warehouse_qty = target
            inventory.quantity = target + location_totals.get(item_id, 0)

        db.query(WarehouseBoxItem).delete(synchronize_session=False)
        db.query(WarehouseSpecialZoneItem).delete(synchronize_session=False)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return summary


def execute_sync(
    *,
    workbook_path: Path,
    sheet_name: str,
    dev_db_path: Path,
    employee_db_path: Path,
    apply: bool,
    confirm: str | None,
    backup_fn: Callable[[str], object] | None = None,
) -> SyncSummary:
    """Run a dry-run or backed-up apply against one explicit development DB."""
    if apply and confirm != "SYNC-WAREHOUSE":
        raise WarehouseSyncError("apply requires --confirm SYNC-WAREHOUSE")
    if not dev_db_path.is_file():
        raise WarehouseSyncError(f"development DB not found: {dev_db_path}")

    snapshot = parse_workbook(workbook_path, sheet_name)
    engine = create_engine(
        f"sqlite:///{dev_db_path.resolve().as_posix()}",
        connect_args={"check_same_thread": False},
    )
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    try:
        with SessionLocal() as db:
            preview = run_sync(db, employee_db_path, snapshot, apply=False)
        if not apply:
            return preview

        if backup_fn is None:
            from scripts.ops.backup_db import backup_sqlite

            backup_fn = backup_sqlite
        backup_fn(str(dev_db_path.resolve()))
        with SessionLocal() as db:
            return run_sync(db, employee_db_path, snapshot, apply=True)
    finally:
        engine.dispose()


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Synchronize warehouse inventory while preserving department stock and history."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet", default="26.07월")
    parser.add_argument("--dev-db", type=Path, default=BACKEND_DIR / "mes.db")
    parser.add_argument("--employee-db", type=Path, default=Path(r"C:\ERP-dev\backend\mes.db"))
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    try:
        summary = execute_sync(
            workbook_path=args.workbook,
            sheet_name=args.sheet,
            dev_db_path=args.dev_db,
            employee_db_path=args.employee_db,
            apply=args.apply,
            confirm=args.confirm,
        )
    except WarehouseSyncError as exc:
        print(f"[WAREHOUSE SYNC] ERROR: {exc}", file=sys.stderr)
        return 2
    mode = "APPLIED" if summary.applied else "DRY-RUN"
    print(f"[WAREHOUSE SYNC] {mode}")
    print(json.dumps(asdict(summary), ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
