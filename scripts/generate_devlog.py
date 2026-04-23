"""
개발 현황 엑셀 생성 스크립트
실행: python scripts/generate_devlog.py
출력: data/개발현황.xlsx
"""
import subprocess
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("data/개발현황.xlsx")

# ── 색상 상수 ──────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
DONE_BG     = "E2EFDA"
TODO_BG     = "F2F2F2"
STRIPE_BG   = "EBF3FB"
WHITE_BG    = "FFFFFF"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text):
    cell.value = text
    cell.fill = make_fill(HEADER_BG)
    cell.font = Font(bold=True, color=HEADER_FG, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def apply_data(cell, text, bg=WHITE_BG, bold=False, align="left"):
    cell.value = text
    cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()

def auto_col_width(ws, extra=2):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

# ── Sheet 1: 개발 현황 요약 ───────────────────────────────
def build_summary(ws):
    ws.title = "개발 현황 요약"
    ws.row_dimensions[1].height = 22

    headers = ["기간", "분류", "주요 작업", "커밋 수", "상태"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    rows = [
        ("2026-04-10 ~ 04-11", "기반 구축",
         "DB 스키마 설계, FastAPI 백엔드 초기 구축\n자재 통합 스크립트, SQLite 연동", 12, "완료"),
        ("2026-04-12 ~ 04-14", "UI 개발",
         "모바일·데스크톱 UI 전면 개발\n직원 관리, 바코드/QR 스캔 기반, 다크모드\n레거시 데스크톱 UI 기본값 설정", 47, "완료"),
        ("2026-04-17 ~ 04-18", "기능 고도화",
         "4-파트 ERP 코드 체계 (M1~M7)\nQueue(생산/분해/반품), Pending/Available 분리\n안전재고 알림, 실사 라우터", 8, "완료"),
        ("2026-04-21",          "UI 정제",
         "레거시 데스크톱 레이아웃 Figma 맞춤 정제", 2, "완료"),
        ("2026-04-22 ~ 04-23", "이번 주 정비",
         "ERP 코드 전사 기준 통일 (item_code → erp_code)\n재고 현황 막대 시각화, 부서 배지 표시\n모바일 UX 전면 재설계 (위저드 흐름)\n재고 계산 일원화·N+1 제거\nObsidian 인수인계 문서 정리", 31, "완료"),
    ]

    for r, (period, cat, work, cnt, status) in enumerate(rows, 2):
        ws.row_dimensions[r].height = 60
        apply_data(ws.cell(r, 1), period, align="center")
        apply_data(ws.cell(r, 2), cat,    align="center")
        apply_data(ws.cell(r, 3), work)
        apply_data(ws.cell(r, 4), cnt,    align="center")
        apply_data(ws.cell(r, 5), status, align="center", bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

# ── Sheet 2: 기능 전체 목록 ──────────────────────────────
def build_features(ws):
    ws.title = "기능 전체 목록"
    ws.row_dimensions[1].height = 22

    headers = ["No", "기능명", "분류", "상태", "비고"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    features = [
        # (기능명, 분류, 완료여부, 비고)
        ("품목 마스터 관리",           "재고",   True,  ""),
        ("ERP 자재 코드 체계 통일",     "재고",   True,  "담당자 협의 필요"),
        ("재고 입고 / 출고 / 조정",     "재고",   True,  ""),
        ("재고 이력 관리",             "재고",   True,  ""),
        ("창고 / 생산 / 불량 버킷 분리", "재고",   True,  ""),
        ("안전재고 기준 이하 알림",      "재고",   True,  ""),
        ("재고 현황 막대 시각화",       "재고",   True,  ""),
        ("BOM 등록 / 조회",            "BOM",   True,  ""),
        ("BOM 웹 수정 기능",           "BOM",   True,  ""),
        ("직원 명단 관리",             "직원",   True,  ""),
        ("부서별 담당 색상 배지 표시",   "직원",   True,  ""),
        ("관리자 화면",                "관리",   True,  ""),
        ("데이터 CSV 내보내기",         "관리",   True,  ""),
        ("PC 화면 (데스크톱)",          "UI",    True,  ""),
        ("모바일 화면",                "UI",    True,  ""),
        ("다크 / 라이트 모드",          "UI",    True,  ""),
        ("원클릭 실행 (start.bat)",     "인프라", True,  ""),
        # ── 예정 ─────────────────────────────────────────
        ("총재고 / 가용재고 / 예약재고 분리", "재고",  False, ""),
        ("생산 / 분해 / 반품 처리 Queue",   "생산",  False, ""),
        ("QR · 바코드 스캔 완성",           "입출고", False, "모바일 카메라 연동"),
        ("실제 운영 데이터 입력",            "데이터", False, "권동환 사원 협의 후"),
        ("로그인 및 권한 관리",              "보안",  False, ""),
        ("외부 접근 환경 구성",              "인프라", False, "PC 또는 NAS 서버"),
        ("발주 관리",                       "구매",  False, ""),
        ("생산 실적 / 원가 관리",            "생산",  False, ""),
        ("부서별 진행현황 화면",             "관리",  False, "사장님 요청 — 검토 중"),
    ]

    for r, (name, cat, done, note) in enumerate(features, 2):
        no = r - 1
        status = "✅ 완료" if done else "🔲 예정"
        bg = DONE_BG if done else TODO_BG
        apply_data(ws.cell(r, 1), no,     bg, align="center")
        apply_data(ws.cell(r, 2), name,   bg)
        apply_data(ws.cell(r, 3), cat,    bg, align="center")
        apply_data(ws.cell(r, 4), status, bg, bold=True, align="center")
        apply_data(ws.cell(r, 5), note,   bg)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 28

# ── Sheet 3: 커밋 이력 ───────────────────────────────────
def fetch_commits():
    result = subprocess.run(
        ["git", "log", "--format=%ad|%h|%s", "--date=short"],
        capture_output=True, text=True, encoding="utf-8"
    )
    commits = []
    for line in result.stdout.strip().splitlines():
        parts = line.split("|", 2)
        if len(parts) != 3:
            continue
        date, hash7, msg = parts
        m = re.match(r"^(feat|fix|refactor|docs|chore|design|style|test|ci|build|perf)(\(.+?\))?!?:", msg)
        ctype = m.group(1) if m else "기타"
        commits.append((date, hash7, msg, ctype))
    return commits

def build_commits(ws, commits):
    ws.title = "커밋 이력"
    ws.row_dimensions[1].height = 22

    headers = ["날짜", "해시", "커밋 메시지", "타입"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    for r, (date, hash7, msg, ctype) in enumerate(commits, 2):
        bg = STRIPE_BG if r % 2 == 0 else WHITE_BG
        apply_data(ws.cell(r, 1), date,  bg, align="center")
        apply_data(ws.cell(r, 2), hash7, bg, align="center")
        apply_data(ws.cell(r, 3), msg,   bg)
        apply_data(ws.cell(r, 4), ctype, bg, align="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12

# ── 메인 ─────────────────────────────────────────────────
def main():
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    build_summary(wb.create_sheet())
    build_features(wb.create_sheet())

    commits = fetch_commits()
    build_commits(wb.create_sheet(), commits)

    wb.save(OUTPUT_PATH)
    print(f"저장 완료: {OUTPUT_PATH}  (커밋 {len(commits)}개)")

if __name__ == "__main__":
    main()
