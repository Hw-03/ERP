"""A 파일(_매칭품명추가.xlsx)에 MES 코드 칸(F열) 끼워넣고 DB와 정합성 검증.

- A의 E열 '매칭 확정 품명' → 마스터 K(확정 품명) → 마스터 P(MES 코드) → A 새 F열
- 메타 표기 행 (예: '(매칭작업에서 삭제됨)') 은 F열도 비움
- 작업 후 DB(items 테이블)와 대조: MES 코드 존재 여부 + 품명/규격 일치 여부 리포트
"""
import re
import shutil
import sqlite3
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
A_SRC = ROOT / '_attic' / 'data' / '0520 권동환 사원님 재고' / 'F704-03__R00__자재_재고_현황_매칭품명추가.xlsx'
A_BAK = A_SRC.with_name(A_SRC.stem + '_원본백업_20260520.xlsx')
MATCH = ROOT / '_attic' / 'data' / '생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx'
DB = ROOT / 'backend' / 'mes.db'

FILL_UNMATCH = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_DB_MISS = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 연빨강 = DB와 불일치


def main():
    # 1) 백업
    if not A_BAK.exists():
        shutil.copy(A_SRC, A_BAK)
        print(f'백업 생성: {A_BAK.name}')
    else:
        print(f'백업 이미 존재: {A_BAK.name}')

    # 2) 마스터 K → MES 코드 매핑
    print('마스터 K(확정품명) → MES 코드 인덱스 ...')
    wb_m = openpyxl.load_workbook(MATCH, data_only=True)
    ws_master = wb_m['마스터_품목']
    k_to_mes = {}
    k_to_master_row = {}
    for r in range(2, ws_master.max_row + 1):
        k = ws_master.cell(row=r, column=11).value
        mes = ws_master.cell(row=r, column=16).value
        if k and mes:
            k_to_mes[str(k).strip()] = str(mes).strip()
            k_to_master_row[str(k).strip()] = r
    print(f'  인덱스: {len(k_to_mes)}개')

    # 3) DB items 인덱스
    print('DB items 인덱스 ...')
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('SELECT item_code, item_name, spec FROM items')
    db_by_code = {}
    for code, name, spec in cur.fetchall():
        if code:
            db_by_code[code.strip()] = (name or '', spec or '')
    conn.close()
    print(f'  DB items: {len(db_by_code)}개')

    # 4) A 열기, 두 시트 모두 작업
    wb = openpyxl.load_workbook(A_SRC)

    # 통계
    report = defaultdict(lambda: {
        'total': 0, 'meta': 0, 'matched': 0, 'mes_assigned': 0,
        'db_hit_exact': 0, 'db_hit_name_diff': 0, 'db_miss': 0,
        'master_miss': 0,
    })
    db_issues_by_sheet = defaultdict(list)

    for sheet_name in ['전체', '26.03월']:
        ws = wb[sheet_name]
        # E열(5번) 오른쪽에 한 칸 삽입 → 기존 F열 이후가 G열로 밀림
        ws.insert_cols(6, amount=1)
        ws.cell(row=3, column=6, value='MES 코드')

        for r in range(4, ws.max_row + 1):
            prod = ws.cell(row=r, column=4).value  # 품명
            e_val = ws.cell(row=r, column=5).value  # 매칭 확정 품명
            spec = ws.cell(row=r, column=7).value  # 규격 (E열 삽입 후 G열)
            if not prod:
                continue
            report[sheet_name]['total'] += 1
            ev = str(e_val).strip() if e_val else ''
            if not ev:
                continue
            # 메타 표기?
            if ev.startswith('(') and ev.endswith(')'):
                report[sheet_name]['meta'] += 1
                cell_mes = ws.cell(row=r, column=6)
                cell_mes.fill = FILL_UNMATCH
                continue
            report[sheet_name]['matched'] += 1

            mes_code = k_to_mes.get(ev)
            cell_mes = ws.cell(row=r, column=6)
            if not mes_code:
                report[sheet_name]['master_miss'] += 1
                cell_mes.fill = FILL_UNMATCH
                continue
            cell_mes.value = mes_code
            report[sheet_name]['mes_assigned'] += 1

            # DB 대조
            db_entry = db_by_code.get(mes_code)
            if not db_entry:
                report[sheet_name]['db_miss'] += 1
                cell_mes.fill = FILL_DB_MISS
                db_issues_by_sheet[sheet_name].append((r, prod, ev, mes_code, 'DB 없음', '', ''))
                continue
            db_name, db_spec = db_entry
            if db_name.strip() == ev:
                report[sheet_name]['db_hit_exact'] += 1
            else:
                report[sheet_name]['db_hit_name_diff'] += 1
                db_issues_by_sheet[sheet_name].append(
                    (r, prod, ev, mes_code, '품명 다름', db_name, db_spec))
                cell_mes.fill = FILL_DB_MISS

    wb.save(A_SRC)
    print(f'저장 완료: {A_SRC.name}')

    # 5) 리포트
    print()
    print('=== 결과 리포트 ===')
    for sn, st in report.items():
        print(f'[{sn}]')
        print(f'  데이터 행: {st["total"]}')
        print(f'  메타 표기 (회색): {st["meta"]}')
        print(f'  실제 매칭값: {st["matched"]}')
        print(f'    └ 마스터 K 누락: {st["master_miss"]}')
        print(f'    └ MES 코드 부여: {st["mes_assigned"]}')
        print(f'  DB 대조 (MES 코드 기준):')
        print(f'    └ 일치 (품명까지 동일): {st["db_hit_exact"]}')
        print(f'    └ 코드는 있으나 품명 다름 (빨강): {st["db_hit_name_diff"]}')
        print(f'    └ DB에 없음 (빨강): {st["db_miss"]}')
        print()

    print('=== DB 불일치 상세 (sheet=26.03월 기준 첫 20건) ===')
    for r, prod, ev, code, status, db_name, db_spec in db_issues_by_sheet['26.03월'][:20]:
        print(f'  r{r}: 권동환={prod!r} | A품명={ev!r} | 코드={code}')
        print(f'        [{status}] DB품명={db_name!r} DB규격={db_spec!r}')


if __name__ == '__main__':
    main()
