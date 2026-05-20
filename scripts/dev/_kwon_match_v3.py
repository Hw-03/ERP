"""권동환 파일 매칭 v3 — 마스터 F+G + 토큰 자카드 점수"""
import openpyxl, re
from collections import defaultdict

KWON = r'_attic/data/0520 권동환 사원님 재고/F704-03 (R00) 자재 재고 현황_통합.xlsx'
MATCH = r'_attic/data/생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx'

wb_k = openpyxl.load_workbook(KWON, data_only=True)
wb_m = openpyxl.load_workbook(MATCH, data_only=True)
ws_k = wb_k['26.05월_수정본']
ws_master = wb_m['마스터_품목']


def norm(s):
    if s is None: return ''
    return re.sub(r'\s+', ' ', str(s).strip()).lower()


def strip_brackets(s):
    return re.sub(r'\[[^\]]*\]', '', str(s)).strip()


def norm_loose(s):
    if s is None: return ''
    s = re.sub(r'[_]+', ' ', str(s))
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()


def get_tokens(s):
    if s is None: return set()
    parts = re.findall(r'[A-Za-z0-9가-힣]{2,}', str(s))
    return set(p.lower() for p in parts)


# 마스터 행
master_rows = []
for r in range(2, ws_master.max_row + 1):
    f = ws_master.cell(row=r, column=6).value
    g = ws_master.cell(row=r, column=7).value
    k = ws_master.cell(row=r, column=11).value
    mes = ws_master.cell(row=r, column=16).value
    model = ws_master.cell(row=r, column=3).value
    g_clean = strip_brackets(g) if g else ''
    mt = ''
    if g:
        m = re.search(r'\[([^\]]+)\]', str(g))
        if m: mt = m.group(1)
    if not mt and model: mt = str(model)
    master_rows.append((r, f, g, g_clean, mt, k, mes))

# 부서별 시트
extras = []
for sn in ['튜브파트_재고현황', '고압진공_재고현황', '조립자재_재고현황', '조립완제품_재고현황', '출하_재고현황']:
    ws = wb_m[sn]
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        mr = ws.cell(row=r, column=18).value
        model = ws.cell(row=r, column=3).value
        mt = ''
        if e:
            m = re.search(r'\[([^\]]+)\]', str(e))
            if m: mt = m.group(1)
        if not mt and model: mt = str(model)
        if d and mr:
            extras.append((str(d), mr, mt, sn))

exact_idx = defaultdict(list)
loose_idx = defaultdict(list)
token_db = []

for r, f, g, gc, mt, k, mes in master_rows:
    for src_name, label in [(f, 'F'), (g, 'G'), (gc, 'G_clean')]:
        if src_name:
            exact_idx[norm(src_name)].append((r, mt, label, src_name))
            loose_idx[norm_loose(src_name)].append((r, mt, label, src_name))
            token_db.append((get_tokens(src_name), r, mt, label, src_name))
    combined = f' {f or ""} {gc or ""} '
    token_db.append((get_tokens(combined), r, mt, 'F+G', combined.strip()))

for nm, mr, mt, sh in extras:
    exact_idx[norm(nm)].append((mr, mt, sh, nm))
    loose_idx[norm_loose(nm)].append((mr, mt, sh, nm))
    token_db.append((get_tokens(nm), mr, mt, sh, nm))


def gen_kwon_variants(name, spec):
    variants = []
    if not name: return variants
    variants.append(name)
    if spec:
        variants.append(f'{name} ({spec})')
        variants.append(f'{name} {spec}')
    if '_' in name:
        parts = [p.strip() for p in name.split('_') if p.strip()]
        variants.extend(parts)
        if len(parts) >= 2:
            variants.append(parts[-1])
            variants.append(parts[0])
    inside = re.findall(r'\(([^)]+)\)', name)
    outside = re.sub(r'\([^)]*\)', '', name).strip()
    if outside and outside != name: variants.append(outside)
    for ins in inside:
        variants.append(ins.strip())
    n2 = re.sub(r'^(\d+kv|F\d+\w*|adx\w+|dx\w+|dxdr-\w+|cocoon|solo|SOLO|COCOON)\s+', '', name, flags=re.I).strip()
    if n2 and n2 != name: variants.append(n2)
    seen, out = set(), []
    for v in variants:
        v = v.strip()
        if v and v.lower() not in seen:
            out.append(v)
            seen.add(v.lower())
    return out


def models_match(km_list, master_tag):
    if not km_list or not master_tag: return False
    mt = master_tag.lower()
    for km in km_list:
        km = km.strip().lower()
        if not km: continue
        if km in mt or mt in km: return True
    return False


def pick_best(candidates, kwon_models):
    if not candidates: return None
    if len(candidates) == 1: return candidates[0]
    for c in candidates:
        if models_match(kwon_models, c[1]):
            return c
    return candidates[0]


hits = {'L1': 0, 'L2': 0, 'L3': 0, 'L4': 0, 'L5': 0, 'L6': 0, 'fail': 0}
fail_list = []
matched = []

for r in range(4, ws_k.max_row + 1):
    code = ws_k.cell(row=r, column=3).value
    prod = ws_k.cell(row=r, column=4).value
    spec = ws_k.cell(row=r, column=5).value
    if not prod: continue
    pname = str(prod).strip()
    pspec = str(spec).strip() if spec else ''
    pcode = str(code).strip() if code else ''
    pmodels = [m.strip() for m in pcode.split(',')] if pcode else []

    variants = gen_kwon_variants(pname, pspec)

    found = None
    for v in variants:
        c = exact_idx.get(norm(v))
        if c:
            found = pick_best(c, pmodels)
            break
    if found:
        matched.append((r, 'L1', found))
        hits['L1'] += 1
        continue

    for v in variants:
        c = loose_idx.get(norm_loose(v))
        if c:
            found = pick_best(c, pmodels)
            break
    if found:
        matched.append((r, 'L2', found))
        hits['L2'] += 1
        continue

    kwon_tokens = get_tokens(f'{pname} {pspec}')
    if len(kwon_tokens) >= 1:
        scored = []
        for mtokens, mr, mt, src, raw in token_db:
            if not mtokens: continue
            inter = kwon_tokens & mtokens
            if not inter: continue
            union = kwon_tokens | mtokens
            jacc = len(inter) / len(union)
            cov = len(inter) / len(kwon_tokens) if kwon_tokens else 0
            scored.append((jacc, cov, mr, mt, src, raw))
        if scored:
            scored.sort(reverse=True)
            best = scored[0]
            jacc, cov, mr, mt, src, raw = best
            if jacc >= 0.6 and models_match(pmodels, mt) and cov >= 0.8:
                matched.append((r, 'L3', (mr, mt, src, raw)))
                hits['L3'] += 1
                continue
            if jacc >= 0.5 and cov >= 0.7:
                matched.append((r, 'L4', (mr, mt, src, raw)))
                hits['L4'] += 1
                continue
            if jacc >= 0.35 and models_match(pmodels, mt):
                matched.append((r, 'L5', (mr, mt, src, raw)))
                hits['L5'] += 1
                continue
            if jacc >= 0.4:
                matched.append((r, 'L6', (mr, mt, src, raw)))
                hits['L6'] += 1
                continue

    hits['fail'] += 1
    fail_list.append((r, pname, pspec, pcode))

print('=== 매칭 결과 (v3) ===')
total = sum(hits.values())
total_hit = total - hits['fail']
for k, v in hits.items():
    print(f'  {k}: {v}')
print(f'  → 매칭 성공: {total_hit}/{total} ({total_hit * 100 / total:.1f}%)')

print()
print(f'=== 미매칭 {len(fail_list)}건 전체 ===')
for r, n, s, c in fail_list:
    print(f'  r{r}: 품명={n!r} 규격={s!r} 모델={c!r}')

print()
print('=== L3 표본(매우 강한 자동 매칭) ===')
n3 = 0
for r, level, info in matched:
    if level == 'L3' and n3 < 8:
        mr, mt, src, raw = info
        prow = next((x for x in master_rows if x[0] == mr), None)
        Kval = prow[5] if prow else '?'
        kp = ws_k.cell(row=r, column=4).value
        ks = ws_k.cell(row=r, column=5).value
        kc = ws_k.cell(row=r, column=3).value
        print(f'  r{r}: 권동환 {kp!r}+{ks!r}+{kc!r}')
        print(f'        → {src}: {raw!r} [모델={mt!r}] → 마스터R{mr} 확정={Kval!r}')
        n3 += 1

print()
print('=== L4 표본(강함, 모델무관) ===')
n4 = 0
for r, level, info in matched:
    if level == 'L4' and n4 < 8:
        mr, mt, src, raw = info
        prow = next((x for x in master_rows if x[0] == mr), None)
        Kval = prow[5] if prow else '?'
        kp = ws_k.cell(row=r, column=4).value
        ks = ws_k.cell(row=r, column=5).value
        kc = ws_k.cell(row=r, column=3).value
        print(f'  r{r}: 권동환 {kp!r}+{ks!r}+{kc!r}')
        print(f'        → {src}: {raw!r} [모델={mt!r}] → 마스터R{mr} 확정={Kval!r}')
        n4 += 1

print()
print('=== L5 표본(모델일치+자카드중간) ===')
n5 = 0
for r, level, info in matched:
    if level == 'L5' and n5 < 8:
        mr, mt, src, raw = info
        prow = next((x for x in master_rows if x[0] == mr), None)
        Kval = prow[5] if prow else '?'
        kp = ws_k.cell(row=r, column=4).value
        ks = ws_k.cell(row=r, column=5).value
        kc = ws_k.cell(row=r, column=3).value
        print(f'  r{r}: 권동환 {kp!r}+{ks!r}+{kc!r}')
        print(f'        → {src}: {raw!r} [모델={mt!r}] → 마스터R{mr} 확정={Kval!r}')
        n5 += 1

print()
print('=== L6 표본(약함, 모델무관) ===')
n6 = 0
for r, level, info in matched:
    if level == 'L6' and n6 < 12:
        mr, mt, src, raw = info
        prow = next((x for x in master_rows if x[0] == mr), None)
        Kval = prow[5] if prow else '?'
        kp = ws_k.cell(row=r, column=4).value
        ks = ws_k.cell(row=r, column=5).value
        kc = ws_k.cell(row=r, column=3).value
        print(f'  r{r}: 권동환 {kp!r}+{ks!r}+{kc!r}')
        print(f'        → {src}: {raw!r} [모델={mt!r}] → 마스터R{mr} 확정={Kval!r}')
        n6 += 1
