"""PA(부모) × AF(자식) Jaccard 토큰 매칭 후보 리포트.

backend/mes.db 에서 PA(process_type_code='PA') 30개와 AF(process_type_code='AF') 44개를
읽어 토큰 Jaccard 유사도 상위 N개 AF 후보를 xlsx 로 출력한다.

목적: 사용자가 워크벤치에서 PA-AF BOM 부모-자식 연결을 결정할 때 후보를 좁혀준다.
자동 INSERT 는 하지 않는다.

실행:
    python scripts/dev/pa_af_match_candidates.py
산출:
    pa_af_candidates_20260522.xlsx (repo root)
"""
from __future__ import annotations

import re
import sqlite3
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception as exc:
    print(f"openpyxl 가 필요합니다: {exc}", file=sys.stderr)
    sys.exit(1)

DB_PATH = Path(__file__).resolve().parents[2] / "backend" / "mes.db"
OUTPUT_PATH = Path(__file__).resolve().parents[2] / "pa_af_candidates_20260522.xlsx"
TOP_N = 5

# 매칭 시 의미 없는 단어 — 빠지면 점수가 깎이지 않도록 사전 제거.
STOPWORDS = {
    "가방", "포장", "완료",
    "ass", "ass'y", "for",
    "the", "and", "or",
    "ea", "kit", "set",
}

# 액세서리·거치대 PA: AF(본체) 자식 불가. AR/AA 부품이 자식이 된다.
# 사용자 도메인 룰 (2026-05-22 확정).
ACCESSORY_PA_NAMES = {
    "ADX6000 스탠드 브라켓 조인트 ASS'Y",
    "AllMyT For ADX6000 ASS'Y",
    "AllMyT For DX3000 ASS'Y",
    "Chest & Bed For ADX6000 ASS'Y",
    "Podiatry For ADX6000 ASS'Y",
}


def tokenize(name: str) -> set[str]:
    """item_name 을 정규화된 토큰 집합으로."""
    if not name:
        return set()
    s = name.lower()
    # 숫자 다음에 영문이 붙은 경우 분리 (60kv -> 60 kv, 1.7ma -> 1.7 ma)
    s = re.sub(r"(\d+(?:\.\d+)?)([a-z가-힣]+)", r"\1 \2", s)
    # 영문 다음에 숫자 분리
    s = re.sub(r"([a-z가-힣]+)(\d+)", r"\1 \2", s)
    # 구분자 → 공백
    s = re.sub(r"[_/,\[\]\(\)'~\-\.\+]", " ", s)
    tokens = {t for t in s.split() if t}
    return tokens - STOPWORDS


def jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 0.0
    inter = a & b
    union = a | b
    return len(inter) / len(union)


def main() -> int:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    pa_rows = con.execute(
        "SELECT item_id, item_name, model_symbol, unit FROM items "
        "WHERE process_type_code='PA' ORDER BY item_name"
    ).fetchall()
    af_rows = con.execute(
        "SELECT item_id, item_name, model_symbol, unit FROM items "
        "WHERE process_type_code='AF' ORDER BY item_name"
    ).fetchall()
    # 이미 등록된 BOM (PA->AF) 쌍 표시용
    existing_pairs = {
        (r[0], r[1])
        for r in con.execute(
            "SELECT parent_item_id, child_item_id FROM bom"
        ).fetchall()
    }
    con.close()

    print(f"PA {len(pa_rows)}개  AF {len(af_rows)}개  → 후보 상위 {TOP_N}개씩 매칭")

    # 토큰 사전 계산
    pa_tokens = [(r, tokenize(r["item_name"])) for r in pa_rows]
    af_tokens = [(r, tokenize(r["item_name"])) for r in af_rows]

    # PA 별 점수 매트릭스
    rows_out: list[dict] = []
    for pa, pa_tok in pa_tokens:
        if pa["item_name"] in ACCESSORY_PA_NAMES:
            rows_out.append({
                "pa_name": pa["item_name"],
                "pa_id": pa["item_id"],
                "pa_model": pa["model_symbol"] or "",
                "rank": "",
                "score": "",
                "af_name": "(AF 자식 불가 - 액세서리. AR/AA 부품을 자식으로)",
                "af_id": "",
                "af_model": "",
                "common_tokens": "",
                "already_linked": "",
                "accessory": True,
            })
            continue
        scored = []
        for af, af_tok in af_tokens:
            score = jaccard(pa_tok, af_tok)
            if score <= 0:
                continue
            scored.append((score, af, af_tok))
        scored.sort(key=lambda t: -t[0])
        top = scored[:TOP_N]
        if not top:
            rows_out.append({
                "pa_name": pa["item_name"],
                "pa_id": pa["item_id"],
                "pa_model": pa["model_symbol"] or "",
                "rank": "",
                "score": "",
                "af_name": "(후보 없음)",
                "af_id": "",
                "af_model": "",
                "common_tokens": "",
                "already_linked": "",
                "accessory": False,
            })
            continue
        for rank, (score, af, af_tok) in enumerate(top, 1):
            common = pa_tok & af_tok
            already = "Y" if (pa["item_id"], af["item_id"]) in existing_pairs else ""
            rows_out.append({
                "pa_name": pa["item_name"],
                "pa_id": pa["item_id"],
                "pa_model": pa["model_symbol"] or "",
                "rank": rank,
                "score": round(score, 3),
                "af_name": af["item_name"],
                "af_id": af["item_id"],
                "af_model": af["model_symbol"] or "",
                "common_tokens": " ".join(sorted(common)),
                "already_linked": already,
                "accessory": False,
            })

    # xlsx 출력
    wb = Workbook()
    ws = wb.active
    ws.title = "PA-AF 후보"

    headers = [
        ("PA(부모)", 50),
        ("PA model", 8),
        ("순위", 5),
        ("Jaccard", 8),
        ("AF(자식 후보)", 55),
        ("AF model", 8),
        ("공통 토큰", 40),
        ("이미 등록", 8),
        ("결정 (Y/N)", 10),  # 사용자가 직접 표시할 수 있도록 빈 열
        ("PA item_id", 36),
        ("AF item_id", 36),
    ]
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)
    for col_i, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.freeze_panes = "A2"

    pa_zebra = PatternFill("solid", fgColor="F2F2F2")
    accessory_fill = PatternFill("solid", fgColor="FFE4E1")  # 연한 살구 - 액세서리 PA
    current_pa = None
    zebra = False
    for ridx, row in enumerate(rows_out, 2):
        if row["pa_name"] != current_pa:
            current_pa = row["pa_name"]
            zebra = not zebra
        if row.get("accessory"):
            fill = accessory_fill
        else:
            fill = pa_zebra if zebra else None
        values = [
            row["pa_name"], row["pa_model"], row["rank"], row["score"],
            row["af_name"], row["af_model"], row["common_tokens"],
            row["already_linked"], "", row["pa_id"], row["af_id"],
        ]
        for col_i, v in enumerate(values, 1):
            cell = ws.cell(row=ridx, column=col_i, value=v)
            if fill:
                cell.fill = fill

    # 가이드 시트
    ws2 = wb.create_sheet("사용법")
    guide = [
        "PA(부모) × AF(자식) 매칭 후보 리포트",
        "",
        "각 PA(조립완료 품목) 마다 Jaccard 토큰 유사도가 높은 AF(부품) 상위 5개씩 표시.",
        "",
        "활용 흐름:",
        "1. '결정 (Y/N)' 열에 Y / N 표시 (옵션, 시각화용)",
        "2. 워크벤치에서 PA 선택 → '자식 추가'로 AF 입력",
        "3. xlsx 의 'AF item_id' 열을 그대로 복사해서 사용 가능",
        "",
        "Jaccard 점수 해석:",
        "  ≥ 0.6  높은 일치  (모델·사양·옵션 다수 공유)",
        "  0.3~0.6 중간     (모델·사양 공유, 옵션 다름)",
        "  < 0.3  낮음     (참고만)",
        "",
        "주의:",
        "  - Jaccard 만으로는 옵션(색상, 콘 길이 등) 일치 보장 안 됨.",
        "  - 사람이 최종 결정. 자동 INSERT 는 일어나지 않음.",
        "  - '이미 등록' 열에 Y 표시된 행은 이미 BOM 에 존재.",
        "",
        "액세서리 PA (살구색 행, 5개):",
        "  - ADX6000 스탠드 브라켓 조인트 ASS'Y",
        "  - AllMyT For ADX6000 ASS'Y / AllMyT For DX3000 ASS'Y",
        "  - Chest & Bed For ADX6000 ASS'Y / Podiatry For ADX6000 ASS'Y",
        "  - AF 본체를 자식으로 받지 않음. AR/AA 부품을 자식으로 워크벤치에서 직접 입력.",
    ]
    for i, line in enumerate(guide, 1):
        c = ws2.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 80

    wb.save(OUTPUT_PATH)
    print(f"\n[저장] {OUTPUT_PATH}")
    print(f"  PA 행수: {len(rows_out)}  (PA {len(pa_rows)} × 후보 ~{TOP_N})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
