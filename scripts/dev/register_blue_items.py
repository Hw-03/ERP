"""진한 파랑(FF00B0F0) 21건을 DB에 신규 등록하고 권동환 엑셀에 반영.

흐름:
1. 진한파랑_등록대상_20260522.xlsx 에서 [MES_등록명, model, process_type, min_stock] 읽기
2. 권동환 엑셀의 진한 파랑 부모 행과 매핑 (행번호 기준)
3. 각 자재에 item_code 자동 부여 (max(serial_no)+1)
4. DB 트랜잭션: items + inventory + inventory_locations 등록
5. 권동환 엑셀: 부모 G열에 멀티코드, 부모 바로 아래 자식 N행 추가 (진한 초록 패턴)
"""
import shutil
import sqlite3
import sys
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.ops.maintenance_backup import create_sqlite_snapshot  # noqa: E402

DB = Path('backend/mes.db')

PLAN = Path('_attic/data/0520 권동환 사원님 재고/진한파랑_등록대상_20260522.xlsx')
EXCEL = Path('_attic/data/0520 권동환 사원님 재고/확정품명, 코드 추가.xlsx')
EXCEL_BAK = Path('_attic/data/0520 권동환 사원님 재고/백업/확정품명, 코드 추가_파랑등록전_20260522.xlsx')

BLUE = 'FF00B0F0'


def create_db_backup(source_path: Path = DB) -> Path:
    """Create the pre-registration DB snapshot in MES_RUNTIME_ROOT."""
    return create_sqlite_snapshot(source_path, "blue-register")


def fc(cell):
    f = cell.fill
    if not f or f.patternType is None:
        return None
    try:
        rgb = f.fgColor.rgb
        return str(rgb) if rgb else None
    except Exception:
        return None


def read_plan():
    wb = openpyxl.load_workbook(PLAN, data_only=True)
    ws = wb.active
    plan = []
    for r in range(2, ws.max_row + 1):
        row_label = ws.cell(row=r, column=1).value
        if not row_label:
            continue
        plan.append({
            'parent_row_str': str(row_label),  # 'r635' 등
            'item_name': str(ws.cell(row=r, column=2).value).strip(),
            'model': str(ws.cell(row=r, column=3).value).strip(),
            'process_type': str(ws.cell(row=r, column=4).value).strip(),
            'min_stock': ws.cell(row=r, column=5).value,
        })
    return plan


def get_parent_rows(ws):
    rows = []
    for r in range(4, ws.max_row + 1):
        if fc(ws.cell(row=r, column=6)) == BLUE:
            rows.append(r)
    return rows


def register_in_db(plan_with_codes):
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    now = datetime.now().isoformat(sep=' ', timespec='microseconds')
    depts = ['조립', '고압', '진공', '튜닝', '튜브']

    for entry in plan_with_codes:
        item_id = uuid.uuid4().hex
        entry['item_id'] = item_id
        cur.execute('''
            INSERT INTO items (
                item_id, item_name, sort_order, unit,
                created_at, updated_at,
                min_stock, model_symbol, process_type_code, serial_no, item_code
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            item_id, entry['item_name'], entry['sort_order'], 'EA',
            now, now,
            entry['min_stock'], entry['model'], entry['process_type'],
            entry['serial_no'], entry['item_code'],
        ))
        cur.execute('''
            INSERT INTO inventory (
                inventory_id, item_id, quantity, warehouse_qty, pending_quantity, updated_at
            ) VALUES (?, ?, 0, 0, 0, ?)
        ''', (uuid.uuid4().hex, item_id, now))
        for d in depts:
            cur.execute('''
                INSERT INTO inventory_locations (
                    location_id, item_id, department, status, quantity, updated_at
                ) VALUES (?, ?, ?, 'PRODUCTION', 0, ?)
            ''', (uuid.uuid4().hex, item_id, d, now))

    conn.commit()
    conn.close()


def main():
    print('=== 1. DB 백업 ===')
    db_backup = create_db_backup()
    print(f'  {db_backup}')

    print()
    print('=== 2. 등록 계획 읽기 ===')
    plan = read_plan()
    print(f'  {len(plan)}건')

    # 행 매핑: parent_row_str 'r635' → 등록할 자재들 리스트
    by_parent = {}
    for e in plan:
        by_parent.setdefault(e['parent_row_str'], []).append(e)
    print(f'  부모 행 수: {len(by_parent)}')

    print()
    print('=== 3. 권동환 엑셀에서 진한 파랑 부모 행 확인 ===')
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['26.05월_수정본']
    parent_rows = get_parent_rows(ws)
    print(f'  진한 파랑 행: {len(parent_rows)}')
    parent_labels = [f'r{r}' for r in parent_rows]
    missing_in_excel = set(by_parent.keys()) - set(parent_labels)
    missing_in_plan = set(parent_labels) - set(by_parent.keys())
    if missing_in_excel:
        print(f'  ⚠ 계획엔 있는데 엑셀에 없음: {missing_in_excel}')
    if missing_in_plan:
        print(f'  ⚠ 엑셀엔 있는데 계획에 없음: {missing_in_plan}')

    print()
    print('=== 4. item_code 시리얼 부여 ===')
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    # model+process 조합별 max(serial_no) 캐시
    max_serial = {}
    for e in plan:
        key = (e['model'], e['process_type'])
        if key not in max_serial:
            cur.execute('SELECT MAX(serial_no) FROM items WHERE model_symbol=? AND process_type_code=?', key)
            row = cur.fetchone()
            max_serial[key] = row[0] or 0
        max_serial[key] += 1
        e['serial_no'] = max_serial[key]
        e['item_code'] = f"{e['model']}-{e['process_type']}-{e['serial_no']:04d}"
        # sort_order: 같은 코드 패밀리의 마지막 sort_order + 1
        cur.execute('''
            SELECT MAX(sort_order) FROM items
            WHERE model_symbol=? AND process_type_code=?
        ''', key)
        ms = cur.fetchone()[0] or 0
        e['sort_order'] = ms + 1
    conn.close()
    for e in plan:
        print(f"  {e['parent_row_str']} → {e['item_code']} | {e['item_name']!r}")

    print()
    print('=== 5. DB 등록 ===')
    register_in_db(plan)
    print(f'  items {len(plan)}건 + inventory {len(plan)}건 + inventory_locations {len(plan)*5}건')

    print()
    print('=== 6. 권동환 엑셀 백업 ===')
    if not EXCEL_BAK.exists():
        shutil.copy(EXCEL, EXCEL_BAK)
        print(f'  {EXCEL_BAK.name}')

    print()
    print('=== 7. 권동환 엑셀 업데이트 ===')
    # 부모 행 역순으로 처리 (자식 삽입이 위 행 인덱스 안 흔들도록)
    parents_sorted_desc = sorted(by_parent.keys(),
                                  key=lambda s: -int(s.lstrip('r')))
    for label in parents_sorted_desc:
        pr = int(label.lstrip('r'))
        entries = by_parent[label]
        codes = [e['item_code'] for e in entries]
        # 부모 G열에 멀티코드 ('&'로 연결)
        ws.cell(row=pr, column=7, value=' & '.join(codes))

        # 자식 행 N개 삽입
        n = len(entries)
        ws.insert_rows(pr + 1, amount=n)
        parent_height = ws.row_dimensions[pr].height
        parent_prod = ws.cell(row=pr, column=4).value
        parent_spec = ws.cell(row=pr, column=5).value

        for i, e in enumerate(entries, start=1):
            cr = pr + i
            ws.cell(row=cr, column=4, value=parent_prod)
            ws.cell(row=cr, column=5, value=parent_spec)
            ws.cell(row=cr, column=6, value=e['item_name'])
            ws.cell(row=cr, column=7, value=e['item_code'])
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                pcell = ws.cell(row=pr, column=col)
                ccell = ws.cell(row=cr, column=col)
                if pcell.has_style:
                    ccell.font = copy(pcell.font)
                    ccell.border = copy(pcell.border)
                    ccell.alignment = copy(pcell.alignment)
                    ccell.number_format = pcell.number_format
                ccell.fill = PatternFill(fill_type=None)
            if parent_height is not None:
                ws.row_dimensions[cr].height = parent_height

        print(f'  {label} → 자식 {n}행 추가 + G열 멀티코드')

    wb.save(EXCEL)
    print(f'저장: {EXCEL.name}')


if __name__ == '__main__':
    main()
