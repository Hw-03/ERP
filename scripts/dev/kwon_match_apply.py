"""권동환 사원님 재고 파일에 MES 품명·MES 코드 칸 끼워넣기.

입력
  - _attic/data/0520 권동환 사원님 재고/F704-03 (R00) 자재 재고 현황_통합.xlsx
  - _attic/data/생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx

출력
  - _attic/data/0520 권동환 사원님 재고/F704-03 (R00) 자재 재고 현황_통합_매칭반영_20260520.xlsx

대상 시트: 26.05월_수정본 (다른 시트는 손대지 않음)

매칭: L1~L3 자동 채택, L4~L6 추정(노란색 음영), 미매칭(회색 음영).
"""
import re
import shutil
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
KWON_SRC = ROOT / '_attic' / 'data' / '0520 권동환 사원님 재고' / 'F704-03 (R00) 자재 재고 현황_통합.xlsx'
MATCH_SRC = ROOT / '_attic' / 'data' / '생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx'
OUT_PATH = ROOT / '_attic' / 'data' / '0520 권동환 사원님 재고' / 'F704-03 (R00) 자재 재고 현황_통합_매칭반영_20260520.xlsx'

TARGET_SHEET = '26.05월_수정본'

FILL_ESTIMATE = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 노란색
FILL_UNMATCHED = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # 회색


def norm(s):
    if s is None:
        return ''
    return re.sub(r'\s+', ' ', str(s).strip()).lower()


def strip_brackets(s):
    return re.sub(r'\[[^\]]*\]', '', str(s)).strip()


def norm_loose(s):
    if s is None:
        return ''
    s = re.sub(r'[_]+', ' ', str(s))
    s = re.sub(r'\s+', ' ', s)
    return s.strip().lower()


def get_tokens(s):
    if s is None:
        return set()
    parts = re.findall(r'[A-Za-z0-9가-힣]{2,}', str(s))
    return set(p.lower() for p in parts)


def build_match_index(wb_m):
    """마스터 파일에서 매칭 인덱스 구축. 반환: (master_rows_by_id, exact_idx, loose_idx, token_db)."""
    ws_master = wb_m['마스터_품목']

    master_rows = {}
    for r in range(2, ws_master.max_row + 1):
        f = ws_master.cell(row=r, column=6).value
        g = ws_master.cell(row=r, column=7).value
        k = ws_master.cell(row=r, column=11).value
        mes = ws_master.cell(row=r, column=16).value
        model = ws_master.cell(row=r, column=3).value
        master_rows[r] = {
            'F': f, 'G': g, 'K': k, 'MES': mes, 'model': model,
            'G_clean': strip_brackets(g) if g else '',
        }

    exact_idx = defaultdict(list)
    loose_idx = defaultdict(list)
    token_db = []

    for r, info in master_rows.items():
        f, g, gc = info['F'], info['G'], info['G_clean']
        mt = ''
        if g:
            m = re.search(r'\[([^\]]+)\]', str(g))
            if m:
                mt = m.group(1)
        if not mt and info['model']:
            mt = str(info['model'])
        for src_name, label in [(f, 'F'), (g, 'G'), (gc, 'G_clean')]:
            if src_name:
                exact_idx[norm(src_name)].append((r, mt, label))
                loose_idx[norm_loose(src_name)].append((r, mt, label))
                token_db.append((get_tokens(src_name), r, mt, label))
        combined = f' {f or ""} {gc or ""} '
        token_db.append((get_tokens(combined), r, mt, 'F+G'))

    # 부서별 시트 (보조 인덱스)
    for sn in ['튜브파트_재고현황', '고압진공_재고현황', '조립자재_재고현황',
               '조립완제품_재고현황', '출하_재고현황']:
        if sn not in wb_m.sheetnames:
            continue
        ws = wb_m[sn]
        for r in range(2, ws.max_row + 1):
            d = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            mr = ws.cell(row=r, column=18).value
            model = ws.cell(row=r, column=3).value
            mt = ''
            if e:
                m = re.search(r'\[([^\]]+)\]', str(e))
                if m:
                    mt = m.group(1)
            if not mt and model:
                mt = str(model)
            if d and mr:
                exact_idx[norm(d)].append((mr, mt, sn))
                loose_idx[norm_loose(d)].append((mr, mt, sn))
                token_db.append((get_tokens(d), mr, mt, sn))

    # 창고_재고현황
    ws_wh = wb_m['창고_재고현황']
    for r in range(2, ws_wh.max_row + 1):
        d = ws_wh.cell(row=r, column=4).value
        mr = ws_wh.cell(row=r, column=19).value
        if not d or not mr:
            continue
        parts = str(d).split('\n')
        raw = parts[0].strip()
        mt = ''
        for p in parts[1:]:
            p = p.strip()
            if p.startswith('[') and p.endswith(']'):
                mt = p[1:-1]
                break
        exact_idx[norm(raw)].append((mr, mt, '창고_재고현황'))
        loose_idx[norm_loose(raw)].append((mr, mt, '창고_재고현황'))
        token_db.append((get_tokens(raw), mr, mt, '창고_재고현황'))

    return master_rows, exact_idx, loose_idx, token_db


def gen_variants(name, spec):
    """권동환 표현의 변형 후보 생성."""
    variants = []
    if not name:
        return variants
    variants.append(name)
    if spec:
        variants.append(f'{name} ({spec})')
        variants.append(f'{name} {spec}')
    if '_' in name:
        parts = [p.strip() for p in name.split('_') if p.strip()]
        variants.extend(parts)
        if len(parts) >= 2:
            variants.append(parts[-1])
            variants.append(parts[0])
    inside = re.findall(r'\(([^)]+)\)', name)
    outside = re.sub(r'\([^)]*\)', '', name).strip()
    if outside and outside != name:
        variants.append(outside)
    for ins in inside:
        variants.append(ins.strip())
    n2 = re.sub(r'^(\d+kv|F\d+\w*|adx\w+|dx\w+|dxdr-\w+|cocoon|solo)\s+', '', name, flags=re.I).strip()
    if n2 and n2 != name:
        variants.append(n2)
    seen, out = set(), []
    for v in variants:
        v = v.strip()
        if v and v.lower() not in seen:
            out.append(v)
            seen.add(v.lower())
    return out


def models_match(km_list, master_tag):
    if not km_list or not master_tag:
        return False
    mt = master_tag.lower()
    for km in km_list:
        km = km.strip().lower()
        if km and (km in mt or mt in km):
            return True
    return False


def pick_best(candidates, kwon_models):
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    for c in candidates:
        if models_match(kwon_models, c[1]):
            return c
    return candidates[0]


def match_row(pname, pspec, pmodels, exact_idx, loose_idx, token_db):
    """한 권동환 행에 대해 매칭 단계 결정. 반환: (level, master_row) or (None, None)."""
    variants = gen_variants(pname, pspec)

    # L1: 정확
    for v in variants:
        c = exact_idx.get(norm(v))
        if c:
            best = pick_best(c, pmodels)
            return 'L1', best[0]

    # L2: loose
    for v in variants:
        c = loose_idx.get(norm_loose(v))
        if c:
            best = pick_best(c, pmodels)
            return 'L2', best[0]

    # L3~L6: 토큰 자카드
    kwon_tokens = get_tokens(f'{pname} {pspec}')
    if not kwon_tokens:
        return None, None

    scored = []
    for mtokens, mr, mt, src in token_db:
        if not mtokens:
            continue
        inter = kwon_tokens & mtokens
        if not inter:
            continue
        union = kwon_tokens | mtokens
        jacc = len(inter) / len(union)
        cov = len(inter) / len(kwon_tokens)
        scored.append((jacc, cov, mr, mt, src))
    if not scored:
        return None, None
    scored.sort(reverse=True)
    jacc, cov, mr, mt, src = scored[0]
    if jacc >= 0.6 and models_match(pmodels, mt) and cov >= 0.8:
        return 'L3', mr
    if jacc >= 0.5 and cov >= 0.7:
        return 'L4', mr
    if jacc >= 0.35 and models_match(pmodels, mt):
        return 'L5', mr
    if jacc >= 0.4:
        return 'L6', mr
    return None, None


def main():
    print(f'원본 복사: {KWON_SRC.name} → {OUT_PATH.name}')
    shutil.copy(KWON_SRC, OUT_PATH)

    print(f'마스터 인덱스 구축 중 ...')
    wb_m = openpyxl.load_workbook(MATCH_SRC, data_only=True)
    master_rows, exact_idx, loose_idx, token_db = build_match_index(wb_m)
    print(f'  마스터 행: {len(master_rows)}, 정확 인덱스: {len(exact_idx)}, 토큰 후보: {len(token_db)}')

    wb_out = openpyxl.load_workbook(OUT_PATH)
    if TARGET_SHEET not in wb_out.sheetnames:
        raise SystemExit(f'대상 시트 없음: {TARGET_SHEET}')
    ws = wb_out[TARGET_SHEET]

    # E열 앞에 두 개의 빈 열 삽입 → 기존 E열 이후 모든 칸이 G열로 밀림
    print('E열 앞에 빈 열 2개 삽입 (MES 품명·MES 코드)')
    ws.insert_cols(5, amount=2)

    # 헤더 (행 3) 작성
    ws.cell(row=3, column=5, value='MES 품명')
    ws.cell(row=3, column=6, value='MES 코드')

    # 매칭 + 채우기
    counts = {'L1': 0, 'L2': 0, 'L3': 0, 'L4': 0, 'L5': 0, 'L6': 0, 'fail': 0}
    auto_levels = {'L1', 'L2', 'L3'}
    estimate_levels = {'L4', 'L5', 'L6'}

    for r in range(4, ws.max_row + 1):
        # 새 칸 삽입 후 권동환 원본의 열 매핑:
        # A(1)=No, B(2)=부품종류, C(3)=MAKER품번(모델), D(4)=품명, E(5)=MES품명[새], F(6)=MES코드[새], G(7)=규격(원래 E)
        prod = ws.cell(row=r, column=4).value
        spec = ws.cell(row=r, column=7).value  # 규격은 이제 G열
        code = ws.cell(row=r, column=3).value
        if not prod:
            continue
        pname = str(prod).strip()
        pspec = str(spec).strip() if spec else ''
        pcode = str(code).strip() if code else ''
        pmodels = [m.strip() for m in pcode.split(',')] if pcode else []

        level, master_r = match_row(pname, pspec, pmodels, exact_idx, loose_idx, token_db)

        cell_mes_name = ws.cell(row=r, column=5)
        cell_mes_code = ws.cell(row=r, column=6)

        if level is None:
            counts['fail'] += 1
            cell_mes_name.fill = FILL_UNMATCHED
            cell_mes_code.fill = FILL_UNMATCHED
            continue

        counts[level] += 1
        info = master_rows.get(master_r, {})
        cell_mes_name.value = info.get('K') or ''
        cell_mes_code.value = info.get('MES') or ''

        if level in estimate_levels:
            cell_mes_name.fill = FILL_ESTIMATE
            cell_mes_code.fill = FILL_ESTIMATE
        # auto_levels는 음영 없음

    # 저장
    wb_out.save(OUT_PATH)

    total = sum(counts.values())
    hits = total - counts['fail']
    print()
    print(f'=== 결과 ===')
    for k in ['L1', 'L2', 'L3', 'L4', 'L5', 'L6', 'fail']:
        print(f'  {k}: {counts[k]}')
    print(f'  총 행: {total}, 매칭: {hits} ({hits * 100 / total:.1f}%)')
    print(f'  자동 채택 (L1~L3): {counts["L1"] + counts["L2"] + counts["L3"]}')
    print(f'  추정 노란색 (L4~L6): {counts["L4"] + counts["L5"] + counts["L6"]}')
    print(f'  미매칭 회색: {counts["fail"]}')
    print()
    print(f'산출물: {OUT_PATH}')


if __name__ == '__main__':
    main()
