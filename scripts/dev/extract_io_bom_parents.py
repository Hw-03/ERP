"""
입출고 관리대장 3개년(2024/2025/2026) 분석 - PA/PF BOM 부모품 후보 추출

실행: python scripts/dev/extract_io_bom_parents.py
입력: data/입출고 관리대장/F704-04 (R00) {YEAR}년 제품 입출고 관리대장.xlsx
출력: data/입출고_BOM부모후보.xlsx (6시트)

산출물 시트:
  1. raw_all          : 3파일×6시트 모든 행 (원본 보존)
  2. raw_normalized   : raw_all + 정규화 4컬럼
  3. unique_5keys     : 고유 (모델·스펙·국가·거래처·Packing) 조합 (PA 후보)
  4. pf_parents       : 고유 (모델·스펙·국가·거래처) 4-key (PF 부모 후보)
  5. pa_variants      : 각 PF 부모별 Packing 변형 목록
  6. discarded        : 모델명/거래처 누락 등 제외된 행 (누락 검증용)
"""

from __future__ import annotations

import re
import sys
import io
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).resolve().parents[2]
INPUT_DIR = ROOT / "data" / "입출고 관리대장"
OUTPUT_PATH = ROOT / "data" / "입출고_BOM부모후보.xlsx"

YEARS = [2024, 2025, 2026]


# 메인 제품 시트 (시트명 고정). 컬럼 인덱스는 헤더(R3)에서 자동 탐지.
PRODUCT_SHEETS = ["DX3000", "DX3000M", "ADX4000W", "ADX6000시리즈", "COCOON", "solo"]

# 헤더 텍스트 → 의미 컬럼 매핑 규칙 (소문자 비교, 첫 매칭 채택)
HEADER_RULES = {
    "model":    [lambda s: s == "모델명"],
    "country":  [lambda s: s == "출고국가"],
    "customer": [lambda s: s == "출고처명"],
    "kv":       [lambda s: s == "kv, ma"],
    "packing":  [lambda s: s == "packing list"],
    "serial":   [lambda s: s == "제품시리얼"],
    "shipdate": [lambda s: s == "출고일"],
}

def detect_columns(header_row) -> dict:
    """헤더 행에서 의미 컬럼 인덱스를 자동 탐지."""
    idx = {}
    for col_i, val in enumerate(header_row):
        if val is None:
            continue
        s = str(val).strip().lower()
        for key, rules in HEADER_RULES.items():
            if key in idx:
                continue
            for rule in rules:
                if rule(s):
                    idx[key] = col_i
                    break
    return idx


# ── 정규화 매핑 ─────────────────────────────────────────────────────────────
COUNTRY_MAP = {
    "USA": "USA",
    "미국": "USA",
}

def norm_model(s: str) -> str:
    return (s or "").strip()

_kv_re = re.compile(r"\s+")
def norm_spec(s: str) -> str:
    if not s:
        return ""
    t = s.replace("\n", "").replace("\r", "")
    t = _kv_re.sub("", t)            # 모든 공백 제거
    return t.lower()

def norm_country(s: str) -> tuple[str, str]:
    """returns (normalized, flag) - flag 'odd' if 줄바꿈 등 이상 표기."""
    if not s:
        return "", ""
    raw = s.strip()
    flag = ""
    if "\n" in raw or "(" in raw:
        flag = "odd"
    clean = raw.replace("\n", " ").strip()
    mapped = COUNTRY_MAP.get(clean, clean)
    return mapped, flag

_packing_ws = re.compile(r"[\s ]+")
def norm_packing(s: str) -> str:
    if not s:
        return ""
    return _packing_ws.sub(" ", s).strip()


# ── 스타일 (generate_devlog.py 패턴 차용) ────────────────────────────────────
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
STRIPE_BG = "EBF3FB"
WHITE_BG = "FFFFFF"
GOLD_BG = "FFF2CC"

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row: int, headers: list[str]):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(HEADER_BG)
        c.font = Font(bold=True, color=HEADER_FG, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate

def write_row(ws, row: int, values: list, stripe: bool = False, highlight: bool = False):
    bg = GOLD_BG if highlight else (STRIPE_BG if stripe else WHITE_BG)
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = fill(bg)
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = thin_border()

def auto_col_width(ws, max_width: int = 60):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for piece in str(cell.value).split("\n"):
                # 한글 폭 가중치
                w = sum(2 if ord(ch) > 127 else 1 for ch in piece)
                if w > longest:
                    longest = w
        ws.column_dimensions[letter].width = min(max_width, max(8, longest + 2))


# ── 데이터 추출 ─────────────────────────────────────────────────────────────
def fmt_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return ""
    return str(v).strip()

def extract_rows():
    rows = []        # 정상 행
    discarded = []   # 제외된 행

    for year in YEARS:
        path = INPUT_DIR / f"F704-04 (R00) {year}년 제품 입출고 관리대장.xlsx"
        if not path.exists():
            print(f"[WARN] not found: {path}")
            continue

        wb = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in PRODUCT_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            # 헤더(R3)에서 자동 인식
            hdr_iter = ws.iter_rows(values_only=True, min_row=3, max_row=3)
            header_row = next(hdr_iter, None)
            if header_row is None:
                print(f"[WARN] {year}/{sheet_name} 헤더 없음, skip")
                continue
            idx = detect_columns(header_row)
            missing = [k for k in ("model", "country", "customer", "kv", "packing") if k not in idx]
            if missing:
                print(f"[WARN] {year}/{sheet_name} 컬럼 누락 {missing}, skip")
                continue

            extracted = 0
            for excel_row_idx, row in enumerate(ws.iter_rows(values_only=True, min_row=4), start=4):
                def cell(i):
                    return row[i] if (i is not None and i < len(row)) else None

                model = cell(idx.get("model"))
                country = cell(idx.get("country"))
                customer = cell(idx.get("customer"))
                kv = cell(idx.get("kv"))
                packing = cell(idx.get("packing"))
                serial = cell(idx.get("serial"))
                ship_date = cell(idx.get("shipdate"))

                model_s = norm_model(str(model)) if model else ""
                customer_s = (str(customer).strip() if customer else "")
                country_s = (str(country).strip() if country else "")
                kv_s = (str(kv).strip() if kv else "")
                pack_s = (str(packing).strip() if packing else "")
                serial_s = (str(serial).strip() if serial else "")
                ship_s = fmt_date(ship_date)

                # 누락 검증 - 모델/거래처 없으면 discarded 행으로 분리
                if not model_s:
                    if any([country_s, customer_s, kv_s, pack_s, serial_s, ship_s]):
                        discarded.append({
                            "year": year, "sheet": sheet_name, "excel_row": excel_row_idx,
                            "reason": "no_model", "model": "", "country": country_s,
                            "customer": customer_s, "spec": kv_s, "packing": pack_s,
                            "serial": serial_s, "ship_date": ship_s,
                        })
                    continue

                if not customer_s and not country_s:
                    discarded.append({
                        "year": year, "sheet": sheet_name, "excel_row": excel_row_idx,
                        "reason": "no_country_and_customer", "model": model_s,
                        "country": "", "customer": "", "spec": kv_s, "packing": pack_s,
                        "serial": serial_s, "ship_date": ship_s,
                    })
                    continue

                country_norm, country_flag = norm_country(country_s)
                rec = {
                    "year": year, "sheet": sheet_name, "excel_row": excel_row_idx,
                    "model": model_s,
                    "spec": kv_s,
                    "country": country_s,
                    "customer": customer_s,
                    "packing": pack_s,
                    "serial": serial_s,
                    "ship_date": ship_s,
                    "model_norm": model_s,
                    "spec_norm": norm_spec(kv_s),
                    "country_norm": country_norm,
                    "country_flag": country_flag,
                    "packing_clean": norm_packing(pack_s),
                }
                rows.append(rec)
                extracted += 1

            print(f"[{year}/{sheet_name}] extracted={extracted} (max_row={ws.max_row})")
        wb.close()

    return rows, discarded


# ── 집계 ─────────────────────────────────────────────────────────────────────
def aggregate(rows: list[dict]):
    # 5-keys: (model_norm, spec_norm, country_norm, customer, packing_clean)
    five = defaultdict(lambda: {"count": 0, "first": "", "last": "",
                                "years": set(), "sample_spec": "",
                                "sample_country": ""})
    for r in rows:
        key = (r["model_norm"], r["spec_norm"], r["country_norm"], r["customer"], r["packing_clean"])
        d = five[key]
        d["count"] += 1
        sd = r["ship_date"]
        if sd:
            if not d["first"] or sd < d["first"]:
                d["first"] = sd
            if not d["last"] or sd > d["last"]:
                d["last"] = sd
        d["years"].add(r["year"])
        if not d["sample_spec"]:
            d["sample_spec"] = r["spec"]
        if not d["sample_country"]:
            d["sample_country"] = r["country"]

    # 4-keys: PF 부모
    four = defaultdict(lambda: {"count": 0, "packings": set(),
                                "first": "", "last": "", "years": set(),
                                "sample_spec": "", "sample_country": ""})
    for r in rows:
        k = (r["model_norm"], r["spec_norm"], r["country_norm"], r["customer"])
        d = four[k]
        d["count"] += 1
        d["packings"].add(r["packing_clean"])
        sd = r["ship_date"]
        if sd:
            if not d["first"] or sd < d["first"]:
                d["first"] = sd
            if not d["last"] or sd > d["last"]:
                d["last"] = sd
        d["years"].add(r["year"])
        if not d["sample_spec"]:
            d["sample_spec"] = r["spec"]
        if not d["sample_country"]:
            d["sample_country"] = r["country"]

    # pa_variants: PF별 packing 변형 (빈도 포함)
    pa_var = defaultdict(Counter)
    for r in rows:
        k = (r["model_norm"], r["spec_norm"], r["country_norm"], r["customer"])
        pa_var[k][r["packing_clean"]] += 1

    return five, four, pa_var


# ── 엑셀 출력 ────────────────────────────────────────────────────────────────
def write_xlsx(rows, discarded, five, four, pa_var):
    wb = Workbook()
    wb.remove(wb.active)

    # 1) raw_all - 사용자 요청 6컬럼 + 중복 제거
    ws1 = wb.create_sheet("raw_all")
    headers1 = ["sheet", "모델명", "kV/mA(원본)", "출고국가(원본)",
                "출고처명", "Packing List(원본)"]
    write_header(ws1, 1, headers1)
    seen = set()
    write_i = 2
    for r in rows:
        key = (r["sheet"], r["model"], r["spec"], r["country"],
               r["customer"], r["packing"])
        if key in seen:
            continue
        seen.add(key)
        write_row(ws1, write_i, list(key), stripe=(write_i % 2 == 0))
        write_i += 1
    auto_col_width(ws1)

    # 2) raw_normalized
    ws2 = wb.create_sheet("raw_normalized")
    headers2 = headers1 + ["model_norm", "spec_norm", "country_norm", "country_flag", "packing_clean"]
    write_header(ws2, 1, headers2)
    for i, r in enumerate(rows, start=2):
        write_row(ws2, i, [r["year"], r["sheet"], r["excel_row"], r["model"],
                           r["spec"], r["country"], r["customer"], r["packing"],
                           r["serial"], r["ship_date"],
                           r["model_norm"], r["spec_norm"], r["country_norm"],
                           r["country_flag"], r["packing_clean"]],
                  stripe=(i % 2 == 0),
                  highlight=bool(r["country_flag"]))
    auto_col_width(ws2)

    # 3) unique_5keys (PA 후보) - count desc
    ws3 = wb.create_sheet("unique_5keys")
    headers3 = ["model", "spec_norm", "country_norm", "customer", "packing_clean",
                "등장횟수", "첫출고", "마지막출고", "발생연도", "샘플 spec(원본)",
                "샘플 country(원본)"]
    write_header(ws3, 1, headers3)
    items5 = sorted(five.items(), key=lambda kv: (-kv[1]["count"], kv[0]))
    for i, (k, d) in enumerate(items5, start=2):
        m, s, co, cu, pk = k
        write_row(ws3, i, [m, s, co, cu, pk, d["count"], d["first"], d["last"],
                           ",".join(str(y) for y in sorted(d["years"])),
                           d["sample_spec"], d["sample_country"]],
                  stripe=(i % 2 == 0))
    auto_col_width(ws3)

    # 4) pf_parents (모델·스펙·국가·거래처) - PF 부모 후보
    ws4 = wb.create_sheet("pf_parents")
    headers4 = ["model", "spec_norm", "country_norm", "customer",
                "등장횟수", "Packing 변형 수", "첫출고", "마지막출고", "발생연도",
                "샘플 spec(원본)", "샘플 country(원본)",
                "PF 부모 ID(제안)"]
    write_header(ws4, 1, headers4)
    items4 = sorted(four.items(), key=lambda kv: (-kv[1]["count"], kv[0]))
    for i, (k, d) in enumerate(items4, start=2):
        m, s, co, cu = k
        pf_id = f"{m}_{s}_{co}_{cu}"
        write_row(ws4, i, [m, s, co, cu, d["count"], len(d["packings"]),
                           d["first"], d["last"],
                           ",".join(str(y) for y in sorted(d["years"])),
                           d["sample_spec"], d["sample_country"], pf_id],
                  stripe=(i % 2 == 0))
    auto_col_width(ws4)

    # 5) pa_variants: 각 PF 부모별 packing 변형
    ws5 = wb.create_sheet("pa_variants")
    headers5 = ["model", "spec_norm", "country_norm", "customer",
                "packing_clean", "변형#", "변형내빈도",
                "PA 후보 ID(제안)"]
    write_header(ws5, 1, headers5)
    row_i = 2
    keys_sorted = sorted(pa_var.keys(),
                         key=lambda k: (-sum(pa_var[k].values()), k))
    for k in keys_sorted:
        m, s, co, cu = k
        variants = pa_var[k].most_common()
        for idx, (pk, cnt) in enumerate(variants, start=1):
            pf_id = f"{m}_{s}_{co}_{cu}"
            pa_id = f"{pf_id}_가방포장완료" if idx == 1 else f"{pf_id}_가방포장완료_v{idx}"
            write_row(ws5, row_i, [m, s, co, cu, pk, idx, cnt, pa_id],
                      stripe=(row_i % 2 == 0))
            row_i += 1
    auto_col_width(ws5)

    # 6) discarded
    ws6 = wb.create_sheet("discarded")
    headers6 = ["year", "sheet", "excel_row", "사유", "모델명",
                "출고국가", "출고처명", "kV/mA", "Packing", "제품시리얼", "출고일"]
    write_header(ws6, 1, headers6)
    for i, r in enumerate(discarded, start=2):
        write_row(ws6, i, [r["year"], r["sheet"], r["excel_row"], r["reason"],
                           r["model"], r["country"], r["customer"],
                           r["spec"], r["packing"], r["serial"], r["ship_date"]],
                  stripe=(i % 2 == 0))
    auto_col_width(ws6)

    wb.save(OUTPUT_PATH)


def main():
    print(f"입력 폴더: {INPUT_DIR}")
    print(f"출력 파일: {OUTPUT_PATH}")
    print("=" * 70)

    rows, discarded = extract_rows()

    print("=" * 70)
    print(f"총 추출(유효): {len(rows)}행 / 제외: {len(discarded)}행")

    five, four, pa_var = aggregate(rows)
    print(f"고유 5키 조합 (PA 후보):  {len(five)}")
    print(f"고유 4키 조합 (PF 부모): {len(four)}")
    print(f"PF 부모 중 Packing 변형 ≥2 :"
          f" {sum(1 for k, v in four.items() if len(v['packings']) >= 2)}")

    write_xlsx(rows, discarded, five, four, pa_var)
    print("=" * 70)
    print(f"저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
