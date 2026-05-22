"""mes.db 전체를 엑셀 한 파일에 시트별로 덤프"""
import sqlite3
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = Path('backend/mes.db')
OUT = Path('_attic/data/db_dump_20260522.xlsx')

HDR_FILL = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
HDR_FONT = Font(color='FFFFFF', bold=True)
ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=False)


def to_cell(v):
    if v is None:
        return ''
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, bytes):
        return v.hex()
    return v


def main():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
    tables = [r[0] for r in cur.fetchall()]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet('_index')
    summary.append(['테이블', '행수', '컬럼수', '비고'])
    for c in range(1, 5):
        cell = summary.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT

    for t in tables:
        sheet_name = t[:31]
        ws = wb.create_sheet(sheet_name)

        cur.execute(f'PRAGMA table_info("{t}")')
        cols_info = cur.fetchall()
        col_names = [c[1] for c in cols_info]

        for c, name in enumerate(col_names, start=1):
            cell = ws.cell(row=1, column=c, value=name)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = ALIGN

        cur.execute(f'SELECT * FROM "{t}"')
        rows = cur.fetchall()

        for r_i, row in enumerate(rows, start=2):
            for c_i, v in enumerate(row, start=1):
                ws.cell(row=r_i, column=c_i, value=to_cell(v))

        ws.freeze_panes = 'A2'
        for c_i, name in enumerate(col_names, start=1):
            w = max(len(name) + 2, 12)
            w = min(w, 40)
            ws.column_dimensions[get_column_letter(c_i)].width = w

        note = ''
        if len(rows) == 0:
            note = '비어있음'
        summary.append([t, len(rows), len(col_names), note])
        print(f'  {t}: {len(rows)}행, {len(col_names)}컬럼')

    summary.freeze_panes = 'A2'
    summary.column_dimensions['A'].width = 30
    summary.column_dimensions['B'].width = 10
    summary.column_dimensions['C'].width = 10
    summary.column_dimensions['D'].width = 20

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print()
    print(f'저장: {OUT}')
    conn.close()


if __name__ == '__main__':
    main()
