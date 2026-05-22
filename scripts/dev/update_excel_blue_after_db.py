"""DB에 이미 등록된 진한 파랑 21건을 권동환 엑셀에 반영.

행번호가 아닌 F열 메모(따옴표 안 품명) 기준으로 매칭. DB는 건드리지 않음.
"""
import re
import shutil
import sqlite3
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

DB = Path('backend/mes.db')
EXCEL = Path('_attic/data/0520 권동환 사원님 재고/확정품명, 코드 추가.xlsx')
EXCEL_BAK = Path(f'_attic/data/0520 권동환 사원님 재고/백업/확정품명, 코드 추가_파랑반영전_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
BLUE = 'FF00B0F0'


def fc(cell):
    f = cell.fill
    if not f or f.patternType is None:
        return None
    try:
        rgb = f.fgColor.rgb
        return str(rgb) if rgb else None
    except Exception:
        return None


def load_db_index():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('SELECT item_code, item_name FROM items WHERE item_code IS NOT NULL')
    by_name = {}
    for code, name in cur.fetchall():
        if name:
            by_name[name.strip()] = code.strip()
    conn.close()
    return by_name


def parse_memo_items(f_val):
    """F열 메모에서 따옴표 안 텍스트 추출, ','로 split하여 자재 이름 리스트 반환."""
    m = re.search(r"'([^']+)'", str(f_val or ''))
    if not m:
        return []
    raw = m.group(1)
    return [n.strip() for n in raw.split(',') if n.strip()]


def main():
    by_name = load_db_index()
    print(f'DB item_code 인덱스: {len(by_name)}')

    shutil.copy(EXCEL, EXCEL_BAK)
    print(f'백업: {EXCEL_BAK.name}')

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['26.05월_수정본']

    # 진한 파랑 부모 행 수집 (현재 위치)
    parents = []
    for r in range(4, ws.max_row + 1):
        if fc(ws.cell(row=r, column=6)) == BLUE:
            f_val = ws.cell(row=r, column=6).value
            names = parse_memo_items(f_val)
            codes = []
            missing = []
            for nm in names:
                code = by_name.get(nm)
                if code:
                    codes.append((nm, code))
                else:
                    missing.append(nm)
            parents.append({
                'r': r, 'names': names, 'codes': codes, 'missing': missing,
                'prod': ws.cell(row=r, column=4).value,
                'spec': ws.cell(row=r, column=5).value,
            })

    print(f'진한 파랑 부모 행: {len(parents)}')
    for p in parents:
        marker = '⚠' if p['missing'] else '✓'
        print(f"  {marker} r{p['r']}: {p['prod']!r}  ({len(p['codes'])}/{len(p['names'])} 매칭)")
        if p['missing']:
            for m in p['missing']:
                print(f'      미매칭: {m!r}')

    if any(p['missing'] for p in parents):
        print('\n⚠ 미매칭 자재 있음. 중단.')
        return

    # 부모 역순 처리
    for parent in sorted(parents, key=lambda p: -p['r']):
        pr = parent['r']
        items = parent['codes']
        if not items:
            continue
        # 부모 G열에 멀티코드
        ws.cell(row=pr, column=7, value=' & '.join(c for _, c in items))

        n = len(items)
        ws.insert_rows(pr + 1, amount=n)
        parent_height = ws.row_dimensions[pr].height

        for i, (item_name, item_code) in enumerate(items, start=1):
            cr = pr + i
            ws.cell(row=cr, column=4, value=parent['prod'])
            ws.cell(row=cr, column=5, value=parent['spec'])
            ws.cell(row=cr, column=6, value=item_name)
            ws.cell(row=cr, column=7, value=item_code)
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                pcell = ws.cell(row=pr, column=col)
                ccell = ws.cell(row=cr, column=col)
                if pcell.has_style:
                    ccell.font = copy(pcell.font)
                    ccell.border = copy(pcell.border)
                    ccell.alignment = copy(pcell.alignment)
                    ccell.number_format = pcell.number_format
                ccell.fill = PatternFill(fill_type=None)
            if parent_height is not None:
                ws.row_dimensions[cr].height = parent_height

        print(f'  r{pr} → 자식 {n}행 (G멀티코드)')

    wb.save(EXCEL)
    print(f'저장: {EXCEL.name}')


if __name__ == '__main__':
    main()
