"""db_dump 파일의 시트명을 한국어로 변경 (DB는 건드리지 않음)"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

OUT = Path('_attic/data/db_dump_20260522.xlsx')

RENAME = {
    'admin_audit_logs': '관리자_감사로그',
    'bom': 'BOM_트리',
    'departments': '부서',
    'employee_assigned_models': '사원_담당모델',
    'employees': '사원',
    'inventory': '재고_요약',
    'inventory_locations': '재고_위치별',
    'io_batches': '입출고_배치',
    'io_bundles': '입출고_묶음',
    'io_lines': '입출고_라인',
    'item_models': '품목_모델슬롯',
    'items': '품목',
    'option_codes': '옵션코드',
    'process_flow_rules': '공정흐름_규칙',
    'process_types': '공정유형',
    'product_symbols': '제품기호_슬롯',
    'ship_package_items': '출하패키지_품목(빈)',
    'ship_packages': '출하패키지(빈)',
    'stock_request_lines': '자재요청_라인',
    'stock_requests': '자재요청',
    'system_settings': '시스템설정',
    'transaction_edit_logs': '거래수정_로그',
    'transaction_logs': '거래로그',
    'variance_logs': '재고차이_로그(빈)',
}

HDR_FILL = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
HDR_FONT = Font(color='FFFFFF', bold=True)


def main():
    wb = openpyxl.load_workbook(OUT)

    for old, new in RENAME.items():
        if old in wb.sheetnames:
            ws = wb[old]
            ws.title = new
            print(f'  {old} → {new}')

    if '_index' in wb.sheetnames:
        idx = wb['_index']
        idx.title = '_목차'
        idx.delete_rows(1, idx.max_row)
        idx.append(['한국어 시트명', '원본 테이블명', '행수', '컬럼수', '비고'])
        for c in range(1, 6):
            cell = idx.cell(row=1, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        for old, new in RENAME.items():
            if new in wb.sheetnames:
                ws = wb[new]
                rows = ws.max_row - 1
                cols = ws.max_column
                note = '비어있음' if rows == 0 else ''
                idx.append([new, old, rows, cols, note])
        idx.column_dimensions['A'].width = 24
        idx.column_dimensions['B'].width = 28
        idx.column_dimensions['C'].width = 8
        idx.column_dimensions['D'].width = 8
        idx.column_dimensions['E'].width = 16
        idx.freeze_panes = 'A2'
        wb.move_sheet('_목차', offset=-len(wb.sheetnames))

    wb.save(OUT)
    print(f'저장: {OUT}')


if __name__ == '__main__':
    main()
