"""(process_type, serial) 시리얼 충돌 5건 수정 + 권동환 엑셀 G열 갱신."""
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.ops.maintenance_backup import create_sqlite_snapshot  # noqa: E402

DB = Path('backend/mes.db')

EXCEL = Path('_attic/data/0520 권동환 사원님 재고/확정품명, 코드 추가.xlsx')
EXCEL_BAK = Path(f'_attic/data/0520 권동환 사원님 재고/백업/확정품명, 코드 추가_시리얼수정전_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

# 재부여 대상
FIXES = [
    ('46-AA-0001', '46-AA-0080', 80),
    ('46-AA-0002', '46-AA-0081', 81),
    ('348-AR-0001', '348-AR-0722', 722),
    ('348-AR-0002', '348-AR-0723', 723),
    ('34678-NR-0001', '34678-NR-0011', 11),
]


def create_db_backup(source_path: Path = DB) -> Path:
    """Create the pre-conflict-fix DB snapshot in MES_RUNTIME_ROOT."""
    return create_sqlite_snapshot(source_path, "fix-conflicts")


def fix_db():
    print('=== DB 시리얼 수정 ===')
    db_backup = create_db_backup()
    print(f'백업: {db_backup}')
    c = sqlite3.connect(DB)
    cur = c.cursor()
    cur.execute('BEGIN')
    for old, new, serial in FIXES:
        cur.execute('SELECT item_name FROM items WHERE item_code=?', (old,))
        row = cur.fetchone()
        if not row:
            print(f'  ⚠ {old} 없음 — skip')
            continue
        cur.execute('UPDATE items SET item_code=?, serial_no=? WHERE item_code=?',
                    (new, serial, old))
        print(f'  {old} → {new}  | {row[0]!r}')
    c.commit()
    c.close()


def fix_excel():
    print()
    print('=== 권동환 엑셀 G열 갱신 ===')
    shutil.copy(EXCEL, EXCEL_BAK)
    print(f'백업: {EXCEL_BAK.name}')
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb['26.05월_수정본']

    code_map = {old: new for old, new, _ in FIXES}
    updated = 0
    for r in range(4, ws.max_row + 1):
        g_cell = ws.cell(row=r, column=7)
        v = g_cell.value
        if not v: continue
        s = str(v).strip()
        if s in code_map:
            new = code_map[s]
            g_cell.value = new
            print(f'  r{r}: {s} → {new}')
            updated += 1
        elif '&' in s or ',' in s:
            # 멀티코드도 갱신
            parts = [p.strip() for p in s.replace('&', ',').split(',') if p.strip()]
            new_parts = [code_map.get(p, p) for p in parts]
            if new_parts != parts:
                # 원래 구분자 보존 (& vs ,)
                if '&' in s:
                    new_v = ' & '.join(new_parts)
                else:
                    new_v = ', '.join(new_parts)
                g_cell.value = new_v
                print(f'  r{r}(multi): {s} → {new_v}')
                updated += 1
    wb.save(EXCEL)
    print(f'  엑셀 갱신: {updated}건')


if __name__ == '__main__':
    fix_db()
    fix_excel()

    # 검증
    print()
    print('=== 검증 ===')
    c = sqlite3.connect(DB)
    cur = c.cursor()
    for _, new, _ in FIXES:
        cur.execute('SELECT item_code, item_name FROM items WHERE item_code=?', (new,))
        print(f'  {new}: {cur.fetchone()}')
    # 충돌 재확인
    cur.execute("""
        SELECT process_type_code, serial_no, COUNT(*) FROM items
        WHERE item_code IS NOT NULL
        GROUP BY process_type_code, serial_no
        HAVING COUNT(*) > 1
    """)
    conflicts = cur.fetchall()
    print(f'  남은 시리얼 충돌: {len(conflicts)}건')
    for pt, sn, n in conflicts:
        print(f'    {pt} serial={sn}: {n}건')
    c.close()
