"""연녹/없음/연파랑 41건 후보 보조표 생성"""
import openpyxl
import re
import sqlite3
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

ROOT = Path('_attic/data/0520 권동환 사원님 재고')
CUR = ROOT / '확정품명, 코드 추가.xlsx'
MASTER = Path('_attic/data/생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx')
A_FILE = ROOT / '원본' / 'F704-03__R00__자재_재고_현황_매칭품명추가.xlsx'
DB = Path('backend/erp.db')
OUT = ROOT / '판단필요_후보표_20260521.xlsx'


def fc(cell):
    f = cell.fill
    if not f or f.patternType is None:
        return None
    try:
        return f.fgColor.rgb
    except Exception:
        return None


def tokens(s):
    if not s:
        return set()
    parts = re.findall(r'[A-Za-z0-9가-힣]{2,}', str(s))
    return set(p.lower() for p in parts)


def score(qt, text):
    tt = tokens(text)
    if not tt:
        return 0.0
    inter = qt & tt
    if not inter:
        return 0.0
    return len(inter) / len(qt | tt)


def main():
    wb_k = openpyxl.load_workbook(CUR)
    ws_k = wb_k['26.05월_수정본']

    target_rows = {'GREEN': [], 'NONE': [], 'BLUE': []}
    for r in range(4, ws_k.max_row + 1):
        prod = ws_k.cell(row=r, column=4).value
        if not prod:
            continue
        color = fc(ws_k.cell(row=r, column=6))
        spec = ws_k.cell(row=r, column=5).value
        cur_match = ws_k.cell(row=r, column=6).value
        code_col = ws_k.cell(row=r, column=8).value
        rec = (
            r,
            str(prod).strip(),
            str(spec).strip() if spec else '',
            str(cur_match).strip() if cur_match else '',
            str(code_col).strip() if code_col else '',
        )
        if color == 'FFE2EFDA':
            target_rows['GREEN'].append(rec)
        elif color is None:
            target_rows['NONE'].append(rec)
        elif color == 'FFCCE5FF':
            target_rows['BLUE'].append(rec)

    print(f'연녹 {len(target_rows["GREEN"])} / 없음 {len(target_rows["NONE"])} / 연파랑 {len(target_rows["BLUE"])}')

    wb_m = openpyxl.load_workbook(MASTER, data_only=True)
    ws_master = wb_m['마스터_품목']
    master_rows = []
    for r in range(2, ws_master.max_row + 1):
        k = ws_master.cell(row=r, column=11).value
        f = ws_master.cell(row=r, column=6).value
        g = ws_master.cell(row=r, column=7).value
        p = ws_master.cell(row=r, column=16).value
        cmodel = ws_master.cell(row=r, column=3).value
        m_spec = ws_master.cell(row=r, column=13).value
        if k:
            master_rows.append({
                'k': str(k).strip(),
                'f': str(f).strip() if f else '',
                'g': str(g).strip() if g else '',
                'mes': str(p).strip() if p else '',
                'model': str(cmodel).strip() if cmodel else '',
                'spec': str(m_spec).strip() if m_spec else '',
            })

    wb_a = openpyxl.load_workbook(A_FILE, data_only=True)
    a_rows = []
    for sn in ['전체', '26.03월']:
        ws_a = wb_a[sn]
        for r in range(4, ws_a.max_row + 1):
            prod = ws_a.cell(row=r, column=4).value
            ev = ws_a.cell(row=r, column=5).value
            spec = ws_a.cell(row=r, column=7).value
            if prod:
                a_rows.append({
                    'sheet': sn,
                    'r': r,
                    'prod': str(prod).strip(),
                    'spec': str(spec).strip() if spec else '',
                    'ev': str(ev).strip() if ev else '',
                })

    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('SELECT erp_code, item_name, spec, model_symbol FROM items')
    db_rows = []
    for c, n, s, m in cur.fetchall():
        if c:
            db_rows.append({'code': c, 'name': n or '', 'spec': s or '', 'model': m or ''})
    conn.close()

    def find_candidates(prod, spec, code, topn=3):
        q = tokens(f'{prod} {spec} {code}')
        if not q:
            return []
        cands = []
        for mr in master_rows:
            text = f'{mr["k"]} {mr["f"]} {mr["g"]} {mr["spec"]} {mr["model"]}'
            s = score(q, text)
            if s > 0:
                cands.append(('M', s, mr['k'], mr['mes'], mr['model'], mr['spec']))
        for dr in db_rows:
            text = f'{dr["name"]} {dr["spec"]} {dr["model"]}'
            s = score(q, text)
            if s > 0:
                cands.append(('D', s, dr['name'], dr['code'], dr['model'], dr['spec']))
        for ar in a_rows:
            text = f'{ar["prod"]} {ar["spec"]} {ar["ev"]}'
            s = score(q, text)
            if s > 0:
                cands.append(('A', s, ar['ev'] or ar['prod'], '', '', ar['spec']))
        cands.sort(key=lambda x: -x[1])
        seen, dedup = set(), []
        for c in cands:
            key = (c[0], c[2])
            if key in seen:
                continue
            seen.add(key)
            dedup.append(c)
            if len(dedup) >= topn:
                break
        return dedup

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    FILL_HDR = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    FONT_HDR = Font(color='FFFFFF', bold=True)
    ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

    GROUP_FILLS = {
        'GREEN': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'NONE': PatternFill(fill_type=None),
        'BLUE': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
    }

    GROUPS = [
        ('GREEN', '연녹_매칭헷갈림'),
        ('NONE', '없음_잘모르겠는것'),
        ('BLUE', '연파랑_5월신규의심'),
    ]
    HEADERS = [
        '행', '권동환 품명', '권동환 규격', '권동환 모델/코드', '현재 입력된 확정품명',
        '후보1 출처/점수', '후보1 품명', '후보1 MES코드', '후보1 모델', '후보1 규격',
        '후보2 출처/점수', '후보2 품명', '후보2 MES코드',
        '후보3 출처/점수', '후보3 품명', '후보3 MES코드',
    ]

    for group, title in GROUPS:
        ws = wb_out.create_sheet(title)
        for c, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = FILL_HDR
            cell.font = FONT_HDR
            cell.alignment = ALIGN

        for row_i, (r, prod, spec, cur_match, code) in enumerate(target_rows[group], start=2):
            ws.cell(row=row_i, column=1, value=r)
            ws.cell(row=row_i, column=2, value=prod)
            ws.cell(row=row_i, column=3, value=spec)
            ws.cell(row=row_i, column=4, value=code)
            ws.cell(row=row_i, column=5, value=cur_match)
            for col in range(1, 6):
                ws.cell(row=row_i, column=col).fill = GROUP_FILLS[group]
                ws.cell(row=row_i, column=col).alignment = ALIGN

            cands = find_candidates(prod, spec, code, topn=3)
            cols_map = {
                0: (6, 7, 8, 9, 10),
                1: (11, 12, 13, None, None),
                2: (14, 15, 16, None, None),
            }
            for i, c in enumerate(cands):
                src, sc, name, mes, mdl, sp = c
                cm = cols_map[i]
                ws.cell(row=row_i, column=cm[0], value=f'{src} / {sc:.2f}')
                ws.cell(row=row_i, column=cm[1], value=name)
                ws.cell(row=row_i, column=cm[2], value=mes)
                if cm[3]:
                    ws.cell(row=row_i, column=cm[3], value=mdl)
                    ws.cell(row=row_i, column=cm[4], value=sp)

        widths = [6, 28, 22, 16, 28, 12, 30, 14, 14, 18, 12, 30, 14, 12, 30, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'B2'

    wb_out.save(OUT)
    print(f'저장: {OUT}')


if __name__ == '__main__':
    main()
