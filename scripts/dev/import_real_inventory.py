"""
data/재고_입력_양식.xlsx 를 읽어서 items / inventory 테이블에 반영.

기본 동작:
    py scripts/dev/import_real_inventory.py
        → dry-run. 몇 건 추가/업데이트/스킵될지 리포트만 출력.

실제 반영:
    py scripts/dev/import_real_inventory.py --apply
        → 기존 품목은 품번(item_code) 기준 upsert, 없으면 새 품목 등록.

기존 971개 테스트 데이터 삭제 후 완전 교체:
    py scripts/dev/import_real_inventory.py --apply --wipe-existing
        (주의: 되돌릴 수 없음. 사용자 확인 필수)
"""

from __future__ import annotations

import os
import sys
from argparse import ArgumentParser
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_XLSX = ROOT / "data" / "재고_입력_양식.xlsx"
BACKEND_DIR = ROOT / "backend"
SQLITE_PATH = BACKEND_DIR / "erp.db"

sys.path.insert(0, str(BACKEND_DIR))
os.environ.setdefault("DATABASE_URL", f"sqlite:///{SQLITE_PATH.as_posix()}")

from app.database import Base, SessionLocal, engine  # noqa: E402
from app.models import Inventory, Item  # noqa: E402


# 양식과 1:1 매칭되는 헤더 (순서 무관, 헤더명으로 찾음)
REQUIRED_HEADERS = ["품목명", "공정코드", "현재수량"]
OPTIONAL_HEADERS = ["규격", "단위", "부서", "모델", "품번", "자재분류", "공급사", "안전재고", "바코드"]
ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS

DATA_START_ROW = 6  # 1=헤더, 2~4=예시, 5=힌트, 6~=실입력

VALID_PROCESS_TYPE_CODES = {
    "TR", "HR", "VR", "NR", "AR", "PR",
    "TA", "HA", "VA", "NA", "AA", "PA",
    "TF", "HF", "VF", "NF", "AF", "PF",
}

PROCESS_TYPE_TO_FILE_TYPE: dict[str, str] = {
    "TR": "원자재", "HR": "원자재", "VR": "원자재",
    "NR": "원자재", "AR": "원자재", "PR": "원자재",
    "TA": "조립자재", "TF": "조립자재",
    "HA": "발생부자재", "HF": "발생부자재",
    "VA": "발생부자재", "VF": "발생부자재",
    "NA": "조립자재", "NF": "조립자재",
    "AA": "조립자재", "AF": "조립자재",
    "PA": "조립자재", "PF": "완제품",
}
PROCESS_TYPE_TO_PART: dict[str, str] = {
    "TR": "자재창고", "HR": "자재창고", "VR": "자재창고",
    "NR": "자재창고", "AR": "자재창고", "PR": "자재창고",
    "TA": "튜브파트", "TF": "튜브파트",
    "HA": "고압파트", "HF": "고압파트",
    "VA": "진공파트", "VF": "진공파트",
    "NA": "튜닝파트", "NF": "튜닝파트",
    "AA": "조립출하", "AF": "조립출하",
    "PA": "출하", "PF": "출하",
}


def parse_decimal(value) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in {"", "None", "N/A", "-"}:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except InvalidOperation:
        return None


def read_rows(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    """양식 xlsx 를 읽어서 dict 리스트와 에러 메시지 리스트 반환."""
    wb = load_workbook(xlsx_path, data_only=True)
    if "재고입력" not in wb.sheetnames:
        raise RuntimeError(f"'재고입력' 시트가 없습니다: {xlsx_path}")
    ws = wb["재고입력"]

    header_row = [ws.cell(row=1, column=c).value for c in range(1, 30)]
    header_to_col: dict[str, int] = {}
    for idx, h in enumerate(header_row, start=1):
        if h and str(h).strip() in ALL_HEADERS:
            header_to_col[str(h).strip()] = idx

    missing = [h for h in REQUIRED_HEADERS if h not in header_to_col]
    if missing:
        raise RuntimeError(f"필수 헤더 누락: {missing}")

    rows: list[dict] = []
    errors: list[str] = []

    for excel_row in range(DATA_START_ROW, ws.max_row + 1):
        raw = {h: ws.cell(row=excel_row, column=col).value for h, col in header_to_col.items()}
        # 완전히 빈 행은 스킵 (에러 아님)
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw.values()):
            continue

        item_name = (str(raw.get("품목명") or "")).strip()
        if not item_name:
            errors.append(f"행 {excel_row}: 품목명 비어 있음 → 스킵")
            continue

        category = (str(raw.get("공정코드") or "")).strip().upper()
        if category not in VALID_PROCESS_TYPE_CODES:
            errors.append(f"행 {excel_row}: 공정코드 '{category}' 유효하지 않음 → 스킵 "
                          f"(가능값: {sorted(VALID_PROCESS_TYPE_CODES)})")
            continue

        qty = parse_decimal(raw.get("현재수량"))
        if qty is None:
            qty = Decimal("0")

        rows.append({
            "excel_row": excel_row,
            "품목명": item_name,
            "공정코드": category,
            "현재수량": qty,
            "규격": (str(raw.get("규격") or "")).strip() or None,
            "단위": (str(raw.get("단위") or "")).strip() or "EA",
            "부서": (str(raw.get("부서") or "")).strip() or None,
            "모델": (str(raw.get("모델") or "")).strip() or "공용",
            "품번": (str(raw.get("품번") or "")).strip() or None,
            "자재분류": (str(raw.get("자재분류") or "")).strip() or None,
            "공급사": (str(raw.get("공급사") or "")).strip() or None,
            "안전재고": parse_decimal(raw.get("안전재고")),
            "바코드": (str(raw.get("바코드") or "")).strip() or None,
        })

    return rows, errors


def assign_item_codes(db, rows: list[dict]) -> None:
    """품번 비어있는 행에 공정코드별 자동 품번 부여 (TR-00001 형식). in-place."""
    existing = {}
    for item in db.query(Item).all():
        code = item.item_code or ""
        if "-" in code:
            prefix, _, tail = code.partition("-")
            try:
                existing[prefix] = max(existing.get(prefix, 0), int(tail))
            except ValueError:
                pass

    next_serial: dict[str, int] = {pt: existing.get(pt, 0) for pt in VALID_PROCESS_TYPE_CODES}

    for row in rows:
        if row["품번"]:
            continue
        pt = row["공정코드"]
        next_serial[pt] = next_serial.get(pt, 0) + 1
        row["품번"] = f"{pt}-{next_serial[pt]:05d}"


def apply_rows(db, rows: list[dict], wipe: bool) -> dict:
    now = datetime.utcnow()

    if wipe:
        deleted_inv = db.query(Inventory).delete()
        deleted_item = db.query(Item).delete()
        db.commit()
        print(f"  [WIPE] 기존 Inventory {deleted_inv}건, Item {deleted_item}건 삭제")

    stats = Counter()

    for row in rows:
        existing = db.query(Item).filter(Item.item_code == row["품번"]).first()
        pt = row["공정코드"]

        if existing:
            existing.item_name = row["품목명"]
            existing.spec = row["규격"]
            existing.process_type_code = pt
            existing.unit = row["단위"]
            existing.barcode = row["바코드"] or row["품번"]
            existing.legacy_file_type = PROCESS_TYPE_TO_FILE_TYPE.get(pt, "미분류")
            existing.legacy_part = PROCESS_TYPE_TO_PART.get(pt, "자재창고")
            existing.legacy_item_type = row["자재분류"]
            existing.legacy_model = row["모델"]
            existing.supplier = row["공급사"]
            existing.min_stock = row["안전재고"]
            existing.updated_at = now

            inv = db.query(Inventory).filter(Inventory.item_id == existing.item_id).first()
            if inv:
                inv.quantity = row["현재수량"]
                inv.warehouse_qty = row["현재수량"]
                inv.location = row["부서"]
                inv.updated_at = now
            else:
                db.add(Inventory(
                    item_id=existing.item_id,
                    quantity=row["현재수량"],
                    warehouse_qty=row["현재수량"],
                    location=row["부서"],
                    updated_at=now,
                ))
            stats["updated"] += 1
        else:
            item = Item(
                item_code=row["품번"],
                item_name=row["품목명"],
                spec=row["규격"],
                process_type_code=pt,
                unit=row["단위"],
                barcode=row["바코드"] or row["품번"],
                legacy_file_type=PROCESS_TYPE_TO_FILE_TYPE.get(pt, "미분류"),
                legacy_part=PROCESS_TYPE_TO_PART.get(pt, "자재창고"),
                legacy_item_type=row["자재분류"],
                legacy_model=row["모델"],
                supplier=row["공급사"],
                min_stock=row["안전재고"],
                created_at=now,
                updated_at=now,
            )
            db.add(item)
            db.flush()
            db.add(Inventory(
                item_id=item.item_id,
                quantity=row["현재수량"],
                warehouse_qty=row["현재수량"],
                location=row["부서"],
                updated_at=now,
            ))
            stats["inserted"] += 1

    db.commit()
    return stats


def main() -> int:
    parser = ArgumentParser(description="실재고 엑셀 양식을 DB에 반영")
    parser.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX),
                        help=f"입력 xlsx 경로 (기본: {DEFAULT_XLSX})")
    parser.add_argument("--apply", action="store_true",
                        help="실제로 DB 반영 (지정 없으면 dry-run)")
    parser.add_argument("--wipe-existing", action="store_true",
                        help="기존 Item/Inventory 전체 삭제 후 반영 (주의)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).resolve()
    if not xlsx_path.exists():
        print(f"[ERROR] 파일 없음: {xlsx_path}")
        return 1

    print(f"[READ] {xlsx_path}")
    rows, errors = read_rows(xlsx_path)

    print(f"[PARSE] 유효 {len(rows)}건, 스킵 {len(errors)}건")
    if errors:
        print("  스킵 사유 (최대 10건):")
        for e in errors[:10]:
            print(f"    - {e}")

    if not rows:
        print("유효한 행이 없습니다. 작업 종료.")
        return 0

    # 공정코드 분포
    cat_dist = Counter(r["공정코드"] for r in rows)
    print(f"[PROCESS_TYPE] {dict(cat_dist)}")

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        # 품번 자동 부여 (dry-run/apply 공통)
        before_codes = sum(1 for r in rows if r["품번"])
        assign_item_codes(db, rows)
        auto_assigned = len(rows) - before_codes
        if auto_assigned:
            print(f"[AUTO-CODE] 품번 자동 부여: {auto_assigned}건")

        # 업데이트/신규 예측
        would_update = 0
        would_insert = 0
        for row in rows:
            if db.query(Item).filter(Item.item_code == row["품번"]).first():
                would_update += 1
            else:
                would_insert += 1

        if not args.apply:
            print()
            print("=" * 50)
            print("  DRY-RUN (실제 DB 변경 없음)")
            print("=" * 50)
            print(f"  신규 추가 예정  : {would_insert}건")
            print(f"  기존 업데이트    : {would_update}건")
            if args.wipe_existing:
                existing_item = db.query(Item).count()
                print(f"  [WIPE] 삭제 예정: 기존 {existing_item}건 전부")
            print()
            print("  실제 반영: --apply 추가")
            print("  기존 971개 전체 교체: --apply --wipe-existing")
            return 0

        # --- apply 분기 ---
        if args.wipe_existing:
            print()
            print("!! 주의: 기존 품목/재고를 전부 삭제한 뒤 새로 반영합니다.")
            print("   되돌리려면 `py backend/seed.py` 로 테스트 데이터 복구.")
            print()

        stats = apply_rows(db, rows, wipe=args.wipe_existing)
        print()
        print("=" * 50)
        print("  APPLY 완료")
        print("=" * 50)
        print(f"  신규 추가 : {stats.get('inserted', 0)}건")
        print(f"  업데이트  : {stats.get('updated', 0)}건")
        print(f"  총 DB Item: {db.query(Item).count()}건")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
