"""담당자 엑셀(생산부_재고_매칭작업_최종.xlsx) ↔ DB 13항목 재동기화.

DB 정본 → Excel 단방향. 물리 삭제 + 시트간 참조(N↔R·Q·수식) 전체 재계산.

핵심: openpyxl delete_rows 는 수식을 재번역하지 않으므로, 삭제 후
모든 수식을 '현재 행번호 템플릿'으로 재생성해 결정적으로 복구한다.

비멱등: 신규 코드가 이미 마스터 P 에 있으면 중단(백업 복원 후 재실행).
"""
import datetime
import re
import shutil
import sqlite3
import sys
from copy import copy

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule  # noqa: F401  (호환용 import)

sys.stdout.reconfigure(encoding="utf-8")

XLSX = r"data/생산부_재고_매칭작업_최종.xlsx"
DB = r"backend/erp.db"
MASTER = "마스터_품목"
ORIGIN_SHEETS = [
    "튜브파트_재고현황",
    "고압진공_재고현황",
    "조립자재_재고현황",
    "조립완제품_재고현황",
    "창고_재고현황",
    "출하_재고현황",
]

RENAMES = {
    "8-VA-0011": "발생부 고압B/D+튜브 최종 작업完 [DXDR-070]",
    "6-AA-0046": "ADX6000FB BODY RIGHT ASS'Y",
    "46-AR-0100": "ADX4000W, ADX6000 16핀 FFC Cable (사파리 공용)",
    "6-AR-0113": "ADX6000 BOTTOM BLOCK",
}
DELETE_CODES = ["6-AR-0355", "6-AR-0185"]
NEW_CODES = [
    "6-AR-0356", "7-AR-0357", "7-AR-0358", "7-AR-0359",
    "4-AR-0360", "4-AR-0361", "4-AR-0362", "4-AR-0363",
    "4-AA-0077", "6-AA-0078", "4-AF-0043", "8-AF-0044",
    "6-PR-0182", "6-PR-0183", "6-PR-0184",
    "3-PA-0027", "6-PA-0028", "6-PA-0029", "6-PA-0030",
]
MODEL_NAME = {"3": "DX3000", "4": "ADX4000W", "6": "ADX6000", "7": "COCOON", "8": "SOLO"}
DIGIT_COL = {"3": 10, "4": 11, "6": 12, "7": 13, "8": 14, "?": 15}  # 원본 J~O
SHEET_BY_PT0 = {"A": "조립", "P": "출하"}  # 부서(A열)
MARK_COL = {"R": 6, "A": 7, "F": 8}  # 원본 F/G/H


def norm(v):
    return re.sub(r"\s+", "", str(v)).upper() if v not in (None, "") else ""


def core(sheet_name):
    return sheet_name.replace("_재고현황", "")


def link_cols(ws):
    """원본시트 헤더에서 (바로가기, 마스터행, 마스터시트) 컬럼 인덱스 탐지.

    창고_재고현황만 '공급처'가 P 뒤에 삽입돼 1칸 밀림(R/S/T) — 헤더로 자동 판별.
    """
    cl = cmr = cms = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "")
        if "바로가기" in h:
            cl = c
        elif h.strip() == "마스터행":
            cmr = c
        elif "마스터시트" in h:
            cms = c
    if not (cl and cmr and cms):
        raise SystemExit(f"[{ws.title}] 링크 컬럼 탐지 실패 cl={cl} cmr={cmr} cms={cms}")
    return cl, cmr, cms


def clone_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


# ---- 수식 템플릿 (기검증, sync_excel_from_db.py 와 동일) ----
def m_d(r):
    return ('=IF(INDIRECT("\'"&$O{r}&"\'!F"&$N{r})="O","원자재",'
            'IF(INDIRECT("\'"&$O{r}&"\'!G"&$N{r})="O","반제품",'
            'IF(INDIRECT("\'"&$O{r}&"\'!H"&$N{r})="O","완제품","")))').format(r=r)


def m_e(r):
    return ('=IFERROR(IF(INDIRECT("\'"&$O{r}&"\'!I"&$N{r})=0,"",'
            'INDIRECT("\'"&$O{r}&"\'!I"&$N{r})),"")').format(r=r)


def m_h(r):
    return ('=IF(AND($F{r}="",$G{r}=""),"",IF($F{r}="","창고전용",'
            'IF($G{r}<>"","매칭완료",IF($D{r}="반제품","반제품(창고없음)",'
            'IF($D{r}="완제품","완제품(창고없음)",'
            'IF($D{r}="원자재","⚠ 원자재_부서보관","미분류"))))))').format(r=r)


def m_i(r):
    return ('=IFERROR(IF(INDIRECT("\'"&$O{r}&"\'!P"&$N{r})=0,"",'
            'INDIRECT("\'"&$O{r}&"\'!P"&$N{r})),"")').format(r=r)


def m_l(r, sheet):
    return ('=HYPERLINK("#\'"&$O{r}&"\'!A"&$N{r},"→ {c} R"&$N{r})'
            ).format(r=r, c=core(sheet))


def o_i(r):
    return ('=IFERROR(IF(OR($F{r}="O",$G{r}="O",$H{r}="O"),'
            'CHOOSE(MATCH($A{r},{{"튜브","고압","진공","튜닝","조립","출하"}},0),'
            '"T","H","V","N","A","P")'
            '&IF($F{r}="O","R",IF($G{r}="O","A",IF($H{r}="O","F",""))),""),"")'
            ).format(r=r)


def o_p(r):
    return ('=IF($I{r}="","",IF($J{r}="O","3","")&IF($K{r}="O","4","")'
            '&IF($L{r}="O","6","")&IF($M{r}="O","7","")&IF($N{r}="O","8","")'
            '&IF($O{r}="O","?","")'
            '&IF(OR($J{r}="O",$K{r}="O",$L{r}="O",$M{r}="O",$N{r}="O",$O{r}="O"),"-","")'
            "&$I{r})").format(r=r)


def o_q(mrow):
    return '=HYPERLINK("#\'마스터_품목\'!A{m}","→ 마스터 R{m}")'.format(m=mrow)


def main():
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = XLSX.replace(".xlsx", f".bak_memo13_{ts}.xlsx")
    shutil.copy(XLSX, bak)
    print(f"backup -> {bak}")

    wb = openpyxl.load_workbook(XLSX, data_only=False)
    m = wb[MASTER]

    # 멱등 가드
    existing_p = {norm(m.cell(r, 16).value) for r in range(2, m.max_row + 1)}
    for c in NEW_CODES:
        if norm(c) in existing_p:
            raise SystemExit(f"신규 코드 {c} 가 이미 마스터에 존재 — 백업 복원 후 재실행")

    # 삭제 대상 위치 동적 탐색
    del_master = []   # 원본 마스터 행번호
    del_origin = []   # (sheet, origin_row)
    for code in DELETE_CODES:
        mr = next(r for r in range(2, m.max_row + 1)
                  if norm(m.cell(r, 16).value) == norm(code))
        del_master.append(mr)
        del_origin.append((m.cell(mr, 15).value, int(m.cell(mr, 14).value)))
    print(f"삭제: master rows {del_master} / origin {del_origin}")

    # 1) 물리 삭제 (마스터: 큰 행 먼저)
    for mr in sorted(del_master, reverse=True):
        m.delete_rows(mr, 1)
    for sh, orow in del_origin:
        wb[sh].delete_rows(orow, 1)

    # 2) 마스터 N 보정 (원본 삭제로 인한 시프트)
    for sh, orow in del_origin:
        for r in range(2, m.max_row + 1):
            if m.cell(r, 15).value == sh:
                nv = m.cell(r, 14).value
                if isinstance(nv, int) and nv > orow:
                    m.cell(r, 14, nv - 1)

    # 3) 원본 R 보정 (마스터 삭제로 인한 시프트) — 전 원본시트 (컬럼 헤더 자동탐지)
    dm = sorted(del_master)
    for sh in ORIGIN_SHEETS:
        ws = wb[sh]
        _, cmr, _ = link_cols(ws)
        for r in range(2, ws.max_row + 1):
            rv = ws.cell(r, cmr).value  # 마스터행
            if isinstance(rv, int):
                ws.cell(r, cmr, rv - sum(1 for d in dm if d < rv))

    # 4) 수식 전 재생성
    #  4a) 마스터 D/E/H/I/L
    for r in range(2, m.max_row + 1):
        o = m.cell(r, 15).value
        if o in (None, "") or m.cell(r, 14).value in (None, ""):
            continue
        m.cell(r, 4, m_d(r))
        m.cell(r, 5, m_e(r))
        m.cell(r, 8, m_h(r))
        m.cell(r, 9, m_i(r))
        m.cell(r, 12, m_l(r, str(o)))
    #  4b) 삭제발생 2시트 I/P (행 시프트됨)
    for sh in ("창고_재고현황", "조립자재_재고현황"):
        ws = wb[sh]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value not in (None, ""):  # 부서 있는 데이터행
                ws.cell(r, 9, o_i(r))
                ws.cell(r, 16, o_p(r))
    #  4c) 전 원본시트 바로가기 HYPERLINK (보정된 마스터행 기준, 컬럼 자동탐지)
    for sh in ORIGIN_SHEETS:
        ws = wb[sh]
        cl, cmr, _ = link_cols(ws)
        for r in range(2, ws.max_row + 1):
            rv = ws.cell(r, cmr).value
            if isinstance(rv, int):
                ws.cell(r, cl, o_q(rv))

    # 5) A 이름수정 (마스터 K + 원본 D)
    for code, new_name in RENAMES.items():
        mr = next(r for r in range(2, m.max_row + 1)
                  if norm(m.cell(r, 16).value) == norm(code))
        m.cell(mr, 11, new_name)  # K 확정품명
        osheet = m.cell(mr, 15).value
        orow = int(m.cell(mr, 14).value)
        wb[osheet].cell(orow, 4, new_name)  # 원본 D 품목
        print(f"[A] {code} 이름수정")

    # 6) 신규 19 append
    con = sqlite3.connect(DB)
    rows = con.execute(
        "SELECT erp_code,item_name,model_symbol,process_type_code FROM items "
        "WHERE erp_code IN (%s)" % ",".join("?" * len(NEW_CODES)),
        NEW_CODES,
    ).fetchall()
    con.close()
    by_code = {e: (n, ms, pt) for e, n, ms, pt in rows}
    missing = [c for c in NEW_CODES if c not in by_code]
    if missing:
        raise SystemExit(f"DB에 없는 신규코드: {missing}")

    def target_sheet(pt):
        return "조립완제품_재고현황" if pt == "AF" else "조립자재_재고현황"

    src_master_tpl = 2  # 마스터 서식 템플릿행
    for code in NEW_CODES:
        name, msym, pt = by_code[code]
        osheet_name = target_sheet(pt)
        ows = wb[osheet_name]
        orow = ows.max_row + 1
        mrow = m.max_row + 1

        # 원본행
        tpl_o = 2
        for c in range(1, 20):
            clone_style(ows.cell(tpl_o, c), ows.cell(orow, c))
        ows.row_dimensions[orow].height = ows.row_dimensions[tpl_o].height
        ows.cell(orow, 1, SHEET_BY_PT0[pt[0]])          # A 부서
        ows.cell(orow, 2, None)                          # B 분류(담당자 후속)
        ows.cell(orow, 3, MODEL_NAME.get(msym, msym))    # C 모델
        ows.cell(orow, 4, name)                          # D 품목
        ows.cell(orow, MARK_COL[pt[1]], "O")             # F/G/H 마크
        for d in str(msym):
            if d in DIGIT_COL:
                ows.cell(orow, DIGIT_COL[d], "O")        # J~O 모델마크
        ows.cell(orow, 9, o_i(orow))
        ows.cell(orow, 16, o_p(orow))
        ocl, ocmr, ocms = link_cols(ows)  # 대상시트 링크 컬럼 (조립자재/조립완제품=19열)
        ows.cell(orow, ocl, o_q(mrow))
        ows.cell(orow, ocmr, mrow)
        ows.cell(orow, ocms, MASTER)

        # 마스터행
        for c in range(1, 17):
            clone_style(m.cell(src_master_tpl, c), m.cell(mrow, c))
        m.row_dimensions[mrow].height = m.row_dimensions[src_master_tpl].height
        m.cell(mrow, 1, SHEET_BY_PT0[pt[0]])             # A 그룹(부서)
        m.cell(mrow, 2, None)
        m.cell(mrow, 3, MODEL_NAME.get(msym, msym))
        m.cell(mrow, 4, m_d(mrow))
        m.cell(mrow, 5, m_e(mrow))
        m.cell(mrow, 6, name)
        m.cell(mrow, 7, name)
        m.cell(mrow, 8, m_h(mrow))
        m.cell(mrow, 9, m_i(mrow))
        m.cell(mrow, 10, name)
        m.cell(mrow, 11, name)
        m.cell(mrow, 12, m_l(mrow, osheet_name))
        m.cell(mrow, 14, orow)
        m.cell(mrow, 15, osheet_name)
        m.cell(mrow, 16, code)

    # 7) autofilter 확장
    for sh in ("조립자재_재고현황", "조립완제품_재고현황"):
        ws = wb[sh]
        ws.auto_filter.ref = f"A1:S{ws.max_row}"
    m.auto_filter.ref = f"A1:O{m.max_row}"

    wb.save(XLSX)
    print(f"저장 완료: 마스터 {m.max_row}행, 신규 {len(NEW_CODES)}건 append")


if __name__ == "__main__":
    main()
