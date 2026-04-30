"""
import_inventory_cleanup.py — 생산부_재고_매칭작업_정리본.xlsx 722건 DB 적재.

Usage:
    cd backend
    python ../scripts/dev/import_inventory_cleanup.py [--dry-run]

적재 대상: outputs/inventory_cleanup/생산부_재고_매칭작업_정리본.xlsx
필수 헤더: ERP 코드, 품명, 분류, 현재고

ERP 코드 형식: {model_symbol}-{process_type_code}-{serial_no:04d}[-{option_code}]

재고 위치: 현재고는 모두 해당 부서(생산) 위치. warehouse_qty = 0.
안전재고: 전 품목 min_stock = 200.
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parent.parent.parent / "backend"
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

from app.database import SessionLocal
from app.models import Inventory, InventoryLocation, Item, LocationStatusEnum, ProcessType

EXCEL_PATH = REPO_ROOT / "outputs" / "inventory_cleanup" / "생산부_재고_매칭작업_정리본.xlsx"

VALID_PROCESS_TYPE_CODES = {
    "TR", "TA", "TF",
    "HR", "HA", "HF",
    "VR", "VA", "VF",
    "NR", "NA", "NF",
    "AR", "AA", "AF",
    "PR", "PA", "PF",
}

# process_type_code 첫 글자 → 부서명 (DepartmentEnum 값 기준)
DEPT_MAP: dict[str, str] = {
    "T": "튜브",
    "H": "고압",
    "V": "진공",
    "N": "튜닝",
    "A": "조립",
    "P": "출하",
}

EXPECTED_ROWS = 722
EXPECTED_TOTAL_QTY = Decimal("108924")
DEFAULT_MIN_STOCK = Decimal("200")


def parse_erp_code(raw: str) -> tuple[str, str, int, str | None]:
    """ERP 코드를 (model_symbol, process_type_code, serial_no, option_code)로 파싱."""
    parts = str(raw).strip().split("-")
    if len(parts) < 3:
        raise ValueError(f"ERP 코드 형식 오류: {raw!r}")
    model_symbol = parts[0]
    process_type_code = parts[1]
    try:
        serial_no = int(parts[2])
    except ValueError:
        raise ValueError(f"시리얼 번호 파싱 오류: {raw!r}")
    option_code = parts[3] if len(parts) >= 4 else None
    return model_symbol, process_type_code, serial_no, option_code


def load_excel() -> list[dict]:
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    col_erp = next((i for i, h in enumerate(headers) if "ERP" in h.upper() or "코드" in h), None)
    col_name = next((i for i, h in enumerate(headers) if "품명" in h or "name" in h.lower()), None)
    col_type = next((i for i, h in enumerate(headers) if "분류" in h), None)
    col_qty = next((i for i, h in enumerate(headers) if "현재고" in h or "수량" in h), None)

    if any(v is None for v in [col_erp, col_name, col_qty]):
        sys.exit(f"필수 헤더 없음. 감지된 헤더: {headers}")

    rows = []
    for r in range(2, ws.max_row + 1):
        erp_code = ws.cell(r, col_erp + 1).value
        item_name = ws.cell(r, col_name + 1).value
        legacy_item_type = ws.cell(r, col_type + 1).value if col_type is not None else None
        qty_raw = ws.cell(r, col_qty + 1).value

        if not erp_code or not item_name:
            continue

        rows.append({
            "erp_code": str(erp_code).strip(),
            "item_name": str(item_name).strip(),
            "legacy_item_type": str(legacy_item_type).strip() if legacy_item_type else None,
            "quantity": Decimal(str(qty_raw or 0)),
        })
    return rows


def run(dry_run: bool = False) -> None:
    rows = load_excel()
    print(f"엑셀 로드: {len(rows)}행")

    if len(rows) != EXPECTED_ROWS:
        print(f"[경고] 예상 행수 {EXPECTED_ROWS}와 다름: {len(rows)}")

    total_qty = sum(r["quantity"] for r in rows)
    print(f"재고 합계: {total_qty}")
    if total_qty != EXPECTED_TOTAL_QTY:
        print(f"[경고] 예상 합계 {EXPECTED_TOTAL_QTY}와 다름: {total_qty}")

    db = SessionLocal()
    try:
        valid_codes = {pt.code for pt in db.query(ProcessType).all()}
        if not valid_codes:
            sys.exit("[오류] process_types 테이블이 비어있음. bootstrap_db.py --seed 먼저 실행하세요.")

        # 파싱 결과를 먼저 수집 (item_id가 필요한 location은 flush 후 생성)
        parsed: list[dict] = []
        erp_codes_seen: set[str] = set()

        for i, row in enumerate(rows):
            erp = row["erp_code"]
            if erp in erp_codes_seen:
                sys.exit(f"[오류] ERP 코드 중복: {erp}")
            erp_codes_seen.add(erp)

            try:
                model_symbol, pt_code, serial_no, option_code = parse_erp_code(erp)
            except ValueError as e:
                sys.exit(f"[오류] {e}")

            if pt_code not in valid_codes:
                sys.exit(f"[오류] 유효하지 않은 process_type_code: {pt_code!r} (ERP={erp})")

            dept = DEPT_MAP.get(pt_code[0])
            if dept is None:
                sys.exit(f"[오류] 부서 매핑 실패: process_type_code={pt_code!r} (ERP={erp})")

            parsed.append({
                "erp": erp,
                "item_name": row["item_name"],
                "legacy_item_type": row["legacy_item_type"],
                "model_symbol": model_symbol,
                "pt_code": pt_code,
                "serial_no": serial_no,
                "option_code": option_code,
                "dept": dept,
                "quantity": row["quantity"],
                "sort_order": i + 1,
            })

        print(f"\n적재 예정: items={len(parsed)}, inventory={len(parsed)}, locations={sum(1 for p in parsed if p['quantity'] > 0)}")

        if dry_run:
            print("[dry-run] DB 변경 없이 종료.")
            return

        # 1단계: Item 일괄 삽입 후 flush → item_id 확보
        items_to_add = [
            Item(
                item_code=p["erp"],
                erp_code=p["erp"],
                barcode=p["erp"],
                item_name=p["item_name"],
                unit="EA",
                model_symbol=p["model_symbol"],
                process_type_code=p["pt_code"],
                serial_no=p["serial_no"],
                option_code=p["option_code"],
                legacy_item_type=p["legacy_item_type"],
                sort_order=p["sort_order"],
                min_stock=DEFAULT_MIN_STOCK,
            )
            for p in parsed
        ]
        db.add_all(items_to_add)
        db.flush()  # item_id 생성

        # 2단계: Inventory + InventoryLocation (item_id 직접 참조)
        inventories_to_add = []
        locations_to_add = []
        for item_obj, p in zip(items_to_add, parsed):
            qty = p["quantity"]
            inventories_to_add.append(Inventory(
                item_id=item_obj.item_id,
                quantity=qty,
                warehouse_qty=Decimal("0"),  # 현재고는 창고가 아닌 해당 부서에 있음
                pending_quantity=Decimal("0"),
            ))
            if qty > 0:
                locations_to_add.append(InventoryLocation(
                    item_id=item_obj.item_id,
                    department=p["dept"],
                    status=LocationStatusEnum.PRODUCTION,
                    quantity=qty,
                ))

        db.add_all(inventories_to_add)
        db.add_all(locations_to_add)
        db.commit()
        print("커밋 완료.")

        # ── 검증 ──
        item_count = db.query(Item).count()
        inv_count = db.query(Inventory).count()
        loc_count = db.query(InventoryLocation).count()
        from sqlalchemy import func as sqlfunc
        from app.models import Inventory as Inv
        qty_sum = db.query(sqlfunc.sum(Inv.quantity)).scalar() or Decimal("0")
        wh_sum = db.query(sqlfunc.sum(Inv.warehouse_qty)).scalar() or Decimal("0")

        print(f"\n[검증]")
        print(f"  items 수: {item_count} (예상 {EXPECTED_ROWS}): {'OK' if item_count == EXPECTED_ROWS else 'FAIL'}")
        print(f"  inventory 수: {inv_count} (예상 {EXPECTED_ROWS}): {'OK' if inv_count == EXPECTED_ROWS else 'FAIL'}")
        print(f"  locations 수: {loc_count}")
        print(f"  재고 합계(quantity): {qty_sum} (예상 {EXPECTED_TOTAL_QTY}): {'OK' if qty_sum == EXPECTED_TOTAL_QTY else 'FAIL'}")
        print(f"  warehouse_qty 합계: {wh_sum} (예상 0): {'OK' if wh_sum == 0 else 'FAIL'}")

        from sqlalchemy import text
        fk_check = db.execute(text("PRAGMA foreign_key_check")).fetchall()
        integrity = db.execute(text("PRAGMA integrity_check")).fetchone()
        print(f"  FK check: {'OK (0행)' if not fk_check else f'FAIL ({len(fk_check)}행)'}")
        print(f"  integrity_check: {integrity[0]}")

        pt_codes_in_items = {r[0] for r in db.execute(text("SELECT DISTINCT process_type_code FROM items WHERE process_type_code IS NOT NULL")).fetchall()}
        invalid_pt = pt_codes_in_items - VALID_PROCESS_TYPE_CODES
        print(f"  process_type_code 범위: {'OK' if not invalid_pt else f'FAIL {invalid_pt}'}")

        all_ok = (
            item_count == EXPECTED_ROWS
            and inv_count == EXPECTED_ROWS
            and qty_sum == EXPECTED_TOTAL_QTY
            and wh_sum == 0
            and not fk_check
            and integrity[0] == "ok"
            and not invalid_pt
        )
        print(f"\n{'[완료] 모든 검증 통과.' if all_ok else '[경고] 일부 검증 실패.'}")

    finally:
        db.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="DB 변경 없이 파싱/검증만 실행")
    args = parser.parse_args()
    run(dry_run=args.dry_run)
