"""
개발 현황 엑셀 생성 스크립트
실행: python scripts/dev/generate_devlog.py
출력: data/개발현황.xlsx
"""
import subprocess
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("data/개발현황.xlsx")

# ── 색상 상수 ──────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
KPI_LABEL_BG = "2E75B6"
KPI_VALUE_BG = "DEEAF1"
DONE_BG     = "E2EFDA"
TODO_BG     = "F2F2F2"
STRIPE_BG   = "EBF3FB"
WHITE_BG    = "FFFFFF"
GOLD_BG     = "FFF2CC"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text):
    cell.value = text
    cell.fill = make_fill(HEADER_BG)
    cell.font = Font(bold=True, color=HEADER_FG, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def apply_data(cell, text, bg=WHITE_BG, bold=False, align="left"):
    cell.value = text
    cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()

def auto_col_width(ws, extra=2):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

# ── Sheet 1: 대시보드 ──────────────────────────────────────
def build_dashboard(ws, commits):
    ws.title = "대시보드"

    # 제목
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "📊 DEXCOWIN MES 개발 현황"
    title.fill = make_fill(HEADER_BG)
    title.font = Font(bold=True, color=HEADER_FG, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.border = make_border()
    ws.row_dimensions[1].height = 28

    # KPI 영역 (행 3~6)
    ws.row_dimensions[3].height = 22
    kpis = [
        ("개발 기간", "2026-04-10 ~ 04-30 (21일)"),
        ("총 커밋 수", f"{len(commits)}건"),
        ("기능 완료", "23 / 33개 (70%)"),
        ("개발 영역", "Backend · Frontend · Mobile · Admin · Docs"),
    ]

    for r, (label, value) in enumerate(kpis, 3):
        # 레이블
        c1 = ws.cell(r, 1)
        c1.value = label
        c1.fill = make_fill(KPI_LABEL_BG)
        c1.font = Font(bold=True, color=HEADER_FG, size=11)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border = make_border()

        # 값
        ws.merge_cells(f"B{r}:E{r}")
        c2 = ws.cell(r, 2)
        c2.value = value
        c2.fill = make_fill(KPI_VALUE_BG)
        c2.font = Font(bold=True, size=11)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = make_border()
        ws.row_dimensions[r].height = 22

    # 날짜별 커밋 막대 (행 8~)
    ws.row_dimensions[8].height = 20
    ws.merge_cells("A8:E8")
    chart_title = ws["A8"]
    chart_title.value = "📈 날짜별 커밋 수"
    chart_title.fill = make_fill(HEADER_BG)
    chart_title.font = Font(bold=True, color=HEADER_FG, size=11)
    chart_title.alignment = Alignment(horizontal="center", vertical="center")
    chart_title.border = make_border()

    by_date = group_commits_by_date(commits)
    dates_sorted = sorted(by_date.keys())

    r = 9
    for date in dates_sorted:
        count = len(by_date[date])
        bar = "■" * (count // 2 if count >= 2 else 1)

        ws.merge_cells(f"A{r}:E{r}")
        cell = ws.cell(r, 1)
        cell.value = f"{date}  {bar}  ({count})"
        cell.fill = make_fill(GOLD_BG)
        cell.font = Font(size=10, color="000000")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = make_border()
        ws.row_dimensions[r].height = 18
        r += 1

    ws.column_dimensions["A"].width = 50

# ── Sheet 2: 개발 현황 요약 ───────────────────────────────
def build_summary(ws, commits):
    ws.title = "개발 현황 요약"
    ws.row_dimensions[1].height = 22

    headers = ["기간", "분류", "주요 작업", "커밋 수", "상태"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    # 자동 계산된 커밋 수
    c1 = count_commits_in_range(commits, "2026-04-10", "2026-04-11")
    c2 = count_commits_in_range(commits, "2026-04-12", "2026-04-14")
    c3 = count_commits_in_range(commits, "2026-04-17", "2026-04-19")
    c4 = count_commits_in_range(commits, "2026-04-21", "2026-04-21")
    c5 = count_commits_in_range(commits, "2026-04-22", "2026-04-24")
    c6 = count_commits_in_range(commits, "2026-04-25", "2026-04-27")
    c7 = count_commits_in_range(commits, "2026-04-28", "2026-04-30")

    rows = [
        ("2026-04-10 ~ 04-11", "기반 구축",
         "DB 스키마·FastAPI 백엔드·자재 통합 스크립트·SQLite", c1, "완료"),
        ("2026-04-12 ~ 04-14", "UI 개발",
         "모바일·데스크톱 UI·BOM·직원·QR스캔·다크모드\n레거시 UI 패리티", c2, "완료"),
        ("2026-04-17 ~ 04-19", "기능 고도화",
         "M1~M7(4-파트 코드)·Queue·Pending·안전재고\n실사 라우터", c3, "완료"),
        ("2026-04-21",          "UI 정제",
         "레거시 레이아웃 Figma 맞춤 정제", c4, "완료"),
        ("2026-04-22 ~ 04-24", "실 운영 준비",
         "erp_code 통일·재고 시각화·모바일UX\n품목매칭 도구·ESLint·계산 일원화(N+1 제거)", c5, "완료"),
        ("2026-04-25 ~ 04-27", "코드 품질 정비",
         "입출고 단계형 Wizard UX·BOM Where-Used 조회\n코드 품질 대정비(Phase 5)·CI 도입\n로그인 화면 구현·카테고리 코드 정비", c6, "완료"),
        ("2026-04-28 ~ 04-30", "현장 운영 투입",
         "창고 승인 기반 입출고 흐름·직원별 장바구니\n작업자 PIN 로그인·감사 이력\n시스템명 MES 통일·품목 분류 개편(진행 중)", c7, "완료"),
    ]

    for r, (period, cat, work, cnt, status) in enumerate(rows, 2):
        ws.row_dimensions[r].height = 60
        apply_data(ws.cell(r, 1), period, align="center")
        apply_data(ws.cell(r, 2), cat,    align="center")
        apply_data(ws.cell(r, 3), work)
        apply_data(ws.cell(r, 4), cnt,    align="center")
        apply_data(ws.cell(r, 5), status, align="center", bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10

# ── Sheet 2: 기능 전체 목록 ──────────────────────────────
def build_features(ws):
    ws.title = "기능 전체 목록"
    ws.row_dimensions[1].height = 22

    headers = ["No", "기능명", "분류", "상태", "비고"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    features = [
        # (기능명, 분류, 완료여부, 비고)
        ("품목 마스터 관리",           "재고",   True,  ""),
        ("ERP 자재 코드 체계 통일",     "재고",   True,  "담당자 협의 필요"),
        ("재고 입고 / 출고 / 조정",     "재고",   True,  ""),
        ("재고 이력 관리",             "재고",   True,  ""),
        ("창고 / 생산 / 불량 버킷 분리", "재고",   True,  ""),
        ("안전재고 기준 이하 알림",      "재고",   True,  ""),
        ("재고 현황 막대 시각화",       "재고",   True,  ""),
        ("BOM 등록 / 조회",            "BOM",   True,  ""),
        ("BOM 웹 수정 기능",           "BOM",   True,  ""),
        ("BOM Where-Used 조회",       "BOM",   True,  ""),
        ("직원 명단 관리",             "직원",   True,  ""),
        ("부서별 담당 색상 배지 표시",   "직원",   True,  ""),
        ("관리자 화면",                "관리",   True,  ""),
        ("데이터 CSV 내보내기",         "관리",   True,  ""),
        ("PC 화면 (데스크톱)",          "UI",    True,  ""),
        ("모바일 화면",                "UI",    True,  ""),
        ("다크 / 라이트 모드",          "UI",    True,  ""),
        ("입출고 단계형 Wizard UX",     "UI",    True,  ""),
        ("원클릭 실행 (start.bat)",     "인프라", True,  ""),
        ("테스트 자동화 (CI)",          "인프라", True,  ""),
        ("창고 승인 기반 입출고 흐름",   "재고",   True,  "요청→승인→처리 단계 분리"),
        ("직원별 저장형 입출고 장바구니", "재고",   True,  "화면 이탈 후에도 작업 유지"),
        ("작업자 PIN 로그인 + 감사 이력", "보안",  True,  "수정 이력 자동 기록"),
        # ── 예정 ─────────────────────────────────────────
        ("총재고 / 가용재고 / 예약재고 분리", "재고",  False, ""),
        ("생산 / 분해 / 반품 처리 Queue",   "생산",  False, ""),
        ("QR · 바코드 스캔 완성",           "입출고", False, "모바일 카메라 연동"),
        ("실제 운영 데이터 입력",            "데이터", False, "권동환 사원 협의 후"),
        ("로그인 및 권한 관리",              "보안",  False, "화면 구현 완료 — 인증 연동 예정"),
        ("외부 접근 환경 구성",              "인프라", False, "PC 또는 NAS 서버"),
        ("발주 관리",                       "구매",  False, ""),
        ("생산 실적 / 원가 관리",            "생산",  False, ""),
        ("부서별 진행현황 화면",             "관리",  False, "사장님 요청 — 검토 중"),
        ("품목 분류 개편 (process_type_code)", "재고", False, "진행 중"),
    ]

    for r, (name, cat, done, note) in enumerate(features, 2):
        no = r - 1
        status = "✅ 완료" if done else "🔲 예정"
        bg = DONE_BG if done else TODO_BG
        apply_data(ws.cell(r, 1), no,     bg, align="center")
        apply_data(ws.cell(r, 2), name,   bg)
        apply_data(ws.cell(r, 3), cat,    bg, align="center")
        apply_data(ws.cell(r, 4), status, bg, bold=True, align="center")
        apply_data(ws.cell(r, 5), note,   bg)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 28

# ── Sheet 3: 커밋 이력 ───────────────────────────────────
def fetch_commits():
    result = subprocess.run(
        ["git", "log", "--format=%ad|%h|%s", "--date=short"],
        capture_output=True, text=True, encoding="utf-8"
    )
    commits = []
    for line in result.stdout.strip().splitlines():
        parts = line.split("|", 2)
        if len(parts) != 3:
            continue
        date, hash7, msg = parts
        m = re.match(r"^(feat|fix|refactor|docs|chore|design|style|test|ci|build|perf)(\(.+?\))?!?:", msg)
        ctype = m.group(1) if m else "기타"
        commits.append((date, hash7, msg, ctype))
    return list(reversed(commits))  # 오래된 순으로

def group_commits_by_date(commits):
    grouped = defaultdict(list)
    for date, hash7, msg, ctype in commits:
        grouped[date].append((hash7, msg, ctype))
    return grouped

def count_commits_in_range(commits, start_date, end_date):
    count = 0
    for date, _, _, _ in commits:
        if start_date <= date <= end_date:
            count += 1
    return count

def build_daily(ws, commits):
    ws.title = "일일 개발 현황"
    ws.row_dimensions[1].height = 22

    headers = ["날짜", "커밋 수", "주요 작업", "영역"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    by_date = group_commits_by_date(commits)
    dates_sorted = sorted(by_date.keys())

    r = 2
    for date in dates_sorted:
        msgs = by_date[date]
        count = len(msgs)
        work_summary = "\n".join([msg[:60] for hash7, msg, ctype in msgs[:5]])
        area = "Backend" if any("feat" in msg.lower() for _, msg, _ in msgs) else "Frontend"

        ws.row_dimensions[r].height = 50
        bg = STRIPE_BG if r % 2 == 0 else WHITE_BG
        apply_data(ws.cell(r, 1), date,          bg, align="center")
        apply_data(ws.cell(r, 2), count,         bg, align="center")
        apply_data(ws.cell(r, 3), work_summary,  bg)
        apply_data(ws.cell(r, 4), area,          bg, align="center")
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14

def build_commits(ws, commits):
    ws.title = "커밋 이력"
    ws.row_dimensions[1].height = 22

    headers = ["날짜", "해시", "커밋 메시지", "타입"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    for r, (date, hash7, msg, ctype) in enumerate(commits, 2):
        bg = STRIPE_BG if r % 2 == 0 else WHITE_BG
        apply_data(ws.cell(r, 1), date,  bg, align="center")
        apply_data(ws.cell(r, 2), hash7, bg, align="center")
        apply_data(ws.cell(r, 3), msg,   bg)
        apply_data(ws.cell(r, 4), ctype, bg, align="center")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 12

# ── 메인 ─────────────────────────────────────────────────
def main():
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    commits = fetch_commits()

    build_dashboard(wb.create_sheet(), commits)
    build_summary(wb.create_sheet(), commits)
    build_features(wb.create_sheet())
    build_daily(wb.create_sheet(), commits)
    build_commits(wb.create_sheet(), commits)

    wb.save(OUTPUT_PATH)
    print(f"저장 완료: {OUTPUT_PATH}  (커밋 {len(commits)}개)")

if __name__ == "__main__":
    main()
