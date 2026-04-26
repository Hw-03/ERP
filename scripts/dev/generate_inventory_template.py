"""
실제 재고 입력용 엑셀 양식 (data/재고_입력_양식.xlsx) 생성기.

사용법:
    py scripts/dev/generate_inventory_template.py

산출물:
    data/재고_입력_양식.xlsx
      Sheet1 "재고입력"  : 12컬럼 입력표 + 드롭다운 + 예시 3행
      Sheet2 "작성가이드" : 카테고리/모델 설명, 작성 규칙
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_PATH = ROOT / "data" / "재고_입력_양식.xlsx"

# 데이터 행을 얼마나 미리 만들어 둘지 (드롭다운 + 테두리 적용 범위)
PREFILL_ROWS = 500

# 컬럼 정의: (헤더, 필수등급, DB필드, 폭, 예시1, 예시2, 예시3)
#   필수등급: "R"=필수(빨강), "W"=권장(노랑), "O"=선택(회색)
COLUMNS: list[tuple[str, str, str, int, str, str, str]] = [
    ("품목명",   "R", "Item.item_name",         26, "텅스텐 필라멘트", "세라믹 애자 70kV",  "DX3000 완제품"),
    ("카테고리", "R", "Item.category",          12, "RM",             "HA",                "FG"),
    ("현재수량", "R", "Inventory.quantity",     11, 120,              15,                   3),
    ("규격",     "W", "Item.spec",              22, "Ø0.3 × L50",      "70kV 절연",         "DX3000-STD"),
    ("단위",     "W", "Item.unit",              8,  "EA",             "EA",                "SET"),
    ("부서",     "W", "Inventory.location",     10, "튜브",           "고압",              "출하"),
    ("모델",     "W", "Item.legacy_model",      12, "공용",           "DX3000",            "DX3000"),
    ("품번",     "O", "Item.item_code",         14, "",               "",                  ""),
    ("자재분류", "O", "Item.legacy_item_type",  14, "필라멘트",       "애자",              "완제품"),
    ("공급사",   "O", "Item.supplier",          18, "삼성특수금속",   "한성세라믹",        "자체생산"),
    ("안전재고", "O", "Item.min_stock",         10, 20,               5,                   1),
    ("바코드",   "O", "Item.barcode",           14, "",               "",                  ""),
]

# 드롭다운 값 목록 (모델 DB와 일치해야 함: models.py CategoryEnum / DepartmentEnum)
DROPDOWNS: dict[str, list[str]] = {
    "카테고리": ["RM", "TA", "HA", "VA", "BA", "FG", "UK"],  # TF/HF/VF/BF 는 내부용이라 제외
    "단위":     ["EA", "SET", "kg", "g", "m", "mm", "L", "box"],
    "부서":     ["조립", "고압", "진공", "튜닝", "튜브", "AS", "연구", "영업", "출하", "기타"],
    "모델":     ["공용", "DX3000", "ADX4000W", "ADX6000", "COCOON", "SOLO"],
}

# 색상
FILL_REQUIRED    = PatternFill("solid", fgColor="F8B4B4")  # 빨강톤 (필수)
FILL_RECOMMENDED = PatternFill("solid", fgColor="FFF3B0")  # 노랑톤 (권장)
FILL_OPTIONAL    = PatternFill("solid", fgColor="E0E0E0")  # 회색톤 (선택)
FILL_EXAMPLE     = PatternFill("solid", fgColor="E3F2FD")  # 연파랑 (예시행)
FILL_GUIDE_H1    = PatternFill("solid", fgColor="2F5496")  # 가이드 큰제목 배경
FILL_GUIDE_H2    = PatternFill("solid", fgColor="DEEBF7")  # 가이드 소제목 배경

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

FILL_BY_REQ = {"R": FILL_REQUIRED, "W": FILL_RECOMMENDED, "O": FILL_OPTIONAL}


def build_input_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "재고입력"

    # 1행: 헤더
    header_font = Font(bold=True, size=11, color="202020")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, req, _db, width, *_examples) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_align
        cell.fill = FILL_BY_REQ[req]
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 2~4행: 예시 데이터
    example_font = Font(italic=True, color="5A5A5A")
    example_align = Alignment(horizontal="left", vertical="center")
    for example_idx in range(3):
        row_num = 2 + example_idx
        for col_idx, (_h, _r, _d, _w, *examples) in enumerate(COLUMNS, start=1):
            value = examples[example_idx]
            cell = ws.cell(row=row_num, column=col_idx, value=value if value != "" else None)
            cell.font = example_font
            cell.alignment = example_align
            cell.fill = FILL_EXAMPLE
            cell.border = BORDER_THIN

    # 5행: 예시 끝 구분 안내 (병합 후 멘트)
    guide_row = 5
    ws.merge_cells(start_row=guide_row, start_column=1, end_row=guide_row, end_column=len(COLUMNS))
    hint = ws.cell(row=guide_row, column=1,
                   value="↑ 위 3행은 예시입니다. 지우고 실제 데이터를 작성하세요. "
                         "필수(빨강) 3칸만 채우면 등록 가능합니다.")
    hint.font = Font(italic=True, bold=True, color="C62828")
    hint.alignment = Alignment(horizontal="center", vertical="center")
    hint.fill = PatternFill("solid", fgColor="FFF8E1")

    # 6행부터: 빈 데이터행 (드롭다운 + 테두리만 적용)
    data_start = 6
    data_end = data_start + PREFILL_ROWS - 1
    data_align = Alignment(horizontal="left", vertical="center")
    for row_num in range(data_start, data_end + 1):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = data_align
            cell.border = BORDER_THIN

    # 틀 고정 (헤더 + 예시)
    ws.freeze_panes = "A6"

    # 행 높이
    ws.row_dimensions[1].height = 28
    for r in range(2, 5):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[5].height = 26

    # 데이터 검증 (드롭다운)
    header_to_col = {c[0]: idx + 1 for idx, c in enumerate(COLUMNS)}

    for header, options in DROPDOWNS.items():
        col_idx = header_to_col[header]
        col_letter = get_column_letter(col_idx)
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(
            type="list", formula1=formula, allow_blank=True,
            showErrorMessage=True,
            errorTitle="허용되지 않는 값",
            error=f"{header} 값은 드롭다운 목록에서만 선택 가능합니다.\n가능값: {', '.join(options)}",
            showInputMessage=True,
            promptTitle=header,
            prompt=f"드롭다운에서 선택: {', '.join(options)}",
        )
        dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
        ws.add_data_validation(dv)

    # 숫자 검증 (현재수량, 안전재고) — 0 이상 숫자만
    for header in ("현재수량", "안전재고"):
        col_idx = header_to_col[header]
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="decimal", operator="greaterThanOrEqual", formula1=0,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="숫자 입력 필요",
            error=f"{header}은(는) 0 이상의 숫자여야 합니다.",
        )
        dv.add(f"{col_letter}{data_start}:{col_letter}{data_end}")
        ws.add_data_validation(dv)


def build_guide_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("작성가이드")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60

    def h1(row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = FILL_GUIDE_H1
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 26

    def h2(row: int, text: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=12, color="1F3864")
        cell.fill = FILL_GUIDE_H2
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def kv(row: int, key: str, value: str) -> None:
        a = ws.cell(row=row, column=1, value=key)
        a.font = Font(bold=True)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b = ws.cell(row=row, column=2, value=value)
        b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = max(18, 16 * (value.count("\n") + 1))

    r = 1
    h1(r, "재고 입력 양식 작성 가이드");                                                 r += 2

    h2(r, "1. 채우는 순서");                                                              r += 1
    kv(r, "①", "1행 예시 3건 지우기 (6행부터 바로 써도 됨)");                            r += 1
    kv(r, "②", "빨강 칸 3개(품목명/카테고리/현재수량)만 채워도 등록 가능");              r += 1
    kv(r, "③", "노랑 칸(규격/단위/부서/모델)은 알면 채우면 좋음");                        r += 1
    kv(r, "④", "회색 칸은 비워도 됨 (품번/바코드 비우면 자동 부여)");                    r += 2

    h2(r, "2. 카테고리 (드롭다운)");                                                     r += 1
    kv(r, "RM", "원자재 - 전극, 필라멘트, 게터 등 가공 전 단계 부품");                   r += 1
    kv(r, "TA", "튜브 어셈블리 - 튜브 조립 단계 반제품");                                r += 1
    kv(r, "HA", "고압 어셈블리 - 고압부 조립 단계 반제품");                              r += 1
    kv(r, "VA", "진공 어셈블리 - 진공부 조립 단계 반제품");                              r += 1
    kv(r, "BA", "조립 완성 - 고압+진공+튜닝 최종 조립 단계");                            r += 1
    kv(r, "FG", "완제품 - 출하 준비 끝난 최종 제품");                                    r += 1
    kv(r, "UK", "미분류 - 분류 애매하면 UK 선택, 나중에 관리자가 재분류");               r += 2

    h2(r, "3. 모델 (드롭다운)");                                                         r += 1
    kv(r, "공용",      "특정 모델 제품에 묶이지 않는 공용 부품 (디폴트)");               r += 1
    kv(r, "DX3000",    "DX3000 시리즈 전용");                                             r += 1
    kv(r, "ADX4000W",  "ADX4000W 전용");                                                  r += 1
    kv(r, "ADX6000",   "ADX6000 전용");                                                   r += 1
    kv(r, "COCOON",    "COCOON 시리즈 전용");                                             r += 1
    kv(r, "SOLO",      "SOLO 시리즈 전용");                                               r += 2

    h2(r, "4. 부서 (드롭다운)");                                                         r += 1
    kv(r, "조립",   "조립 파트");                                                        r += 1
    kv(r, "고압",   "고압 파트");                                                        r += 1
    kv(r, "진공",   "진공 파트");                                                        r += 1
    kv(r, "튜닝",   "튜닝 파트");                                                        r += 1
    kv(r, "튜브",   "튜브 파트");                                                        r += 1
    kv(r, "AS",     "AS 파트");                                                          r += 1
    kv(r, "출하",   "출하/검사 파트");                                                   r += 1
    kv(r, "기타",   "부서 애매하면 '기타' 선택");                                        r += 2

    h2(r, "5. 수량 관련");                                                               r += 1
    kv(r, "현재수량", "지금 창고에 있는 실물 수량. 0도 괜찮음. 비우면 0 처리.");        r += 1
    kv(r, "안전재고", "이 수량 이하로 떨어지면 경고가 뜰 기준선. 몰라도 됨.");          r += 2

    h2(r, "6. 다 쓴 후");                                                                r += 1
    kv(r, "저장", "엑셀 저장 (파일명 그대로 두거나 날짜 붙여 저장)");                   r += 1
    kv(r, "전달", 'Claude에게 "이 양식으로 재고 교체해줘" 라고 말하기');                r += 1
    kv(r, "검증", "Claude가 먼저 --dry-run으로 몇 건이 추가/변경될지 보여줌");          r += 1
    kv(r, "확정", '"진짜 이걸로 교체" 확인하면 기존 테스트 데이터 삭제 후 반영');       r += 2

    h2(r, "7. 색깔 뜻");                                                                 r += 1
    c = ws.cell(row=r, column=1, value="필수")
    c.fill = FILL_REQUIRED; c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value="반드시 채워야 하는 칸 (비면 그 행은 등록 안 됨)")
    r += 1
    c = ws.cell(row=r, column=1, value="권장")
    c.fill = FILL_RECOMMENDED; c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value="알면 채우는 게 좋지만 비워도 등록은 됨")
    r += 1
    c = ws.cell(row=r, column=1, value="선택")
    c.fill = FILL_OPTIONAL; c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=2, value="있으면 쓰고 없으면 비우기 (품번/바코드 비우면 자동 부여)")


def main() -> None:
    wb = Workbook()
    build_input_sheet(wb)
    build_guide_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] 양식 생성 완료: {OUTPUT_PATH}")
    print(f"     - 입력 시트 : '재고입력' ({len(COLUMNS)}컬럼, 예시 3행 포함, {PREFILL_ROWS}행 프리필)")
    print(f"     - 가이드 시트: '작성가이드'")


if __name__ == "__main__":
    main()
