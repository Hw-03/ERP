"""산출물 26.05월_수정본 시트의 E(MES 품명)/F(MES 코드)를 A 정본 기준으로 갈아끼움.

A의 매칭 확정 품명을 정본으로 보고, 5월 행에 (품명, 규격) 키로 매핑해서 입력.
- A에 매칭값이 있고 정상 품명(메타 표기 아님)이면 → E·F에 입력 (음영 없음)
- A에 메타 표기 (`(매칭작업에서 삭제됨)` 등) 인 경우 → 회색 음영, E·F 비움
- A에 키 자체가 없는 경우 (5월 신규 자재) → 회색 음영, E·F 비움
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
A_FILE = ROOT / '_attic' / 'data' / '0520 권동환 사원님 재고' / 'F704-03__R00__자재_재고_현황_매칭품명추가.xlsx'
OUT = ROOT / '_attic' / 'data' / '0520 권동환 사원님 재고' / 'F704-03 (R00) 자재 재고 현황_통합_매칭반영_20260520.xlsx'
OUT_BAK = OUT.with_name(OUT.stem + '_자동매칭버전백업_20260520.xlsx')
MATCH = ROOT / '_attic' / 'data' / '생산부_재고_매칭작업_최종.bak_memo13_20260518_101641.xlsx'

FILL_UNMATCH = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_NONE = PatternFill(fill_type=None)


def is_meta(v):
    s = str(v or '').strip()
    return s.startswith('(') and s.endswith(')')


def main():
    # 1) 산출물 백업 (자동매칭 버전 보관)
    if not OUT_BAK.exists():
        shutil.copy(OUT, OUT_BAK)
        print(f'산출물 백업 생성: {OUT_BAK.name}')

    # 2) A → (품명, 규격) → 매칭 확정 품명
    wa = openpyxl.load_workbook(A_FILE, data_only=True)
    a_map = {}  # (품명, 규격) → 매칭 확정 품명 (메타 포함)
    for sheet in ['전체', '26.03월']:
        ws = wa[sheet]
        for r in range(4, ws.max_row + 1):
            prod = ws.cell(row=r, column=4).value
            e = ws.cell(row=r, column=5).value
            spec = ws.cell(row=r, column=7).value  # F열은 MES 코드라 규격은 G(7)
            if not prod:
                continue
            key = (str(prod).strip(), str(spec).strip() if spec else '')
            if key not in a_map and e is not None:
                a_map[key] = str(e).strip()
    print(f'A 매핑: {len(a_map)} 키')

    # 3) 마스터 K → MES 코드
    wm = openpyxl.load_workbook(MATCH, data_only=True)
    ws_m = wm['마스터_품목']
    k_to_mes = {}
    for r in range(2, ws_m.max_row + 1):
        k = ws_m.cell(row=r, column=11).value
        mes = ws_m.cell(row=r, column=16).value
        if k and mes:
            k_to_mes[str(k).strip()] = str(mes).strip()
    print(f'마스터 K→MES: {len(k_to_mes)} 매핑')

    # 4) 산출물 열기
    wb = openpyxl.load_workbook(OUT)
    ws = wb['26.05월_수정본']
    # 산출물 칸: A=No, B=부품종류, C=MAKER품번, D=품명, E=MES 품명, F=MES 코드, G=규격, H~...

    stats = {'a_matched': 0, 'a_meta_meta': 0, 'a_miss_5월신규': 0, 'no_prod': 0}
    samples_5월신규 = []

    for r in range(4, ws.max_row + 1):
        prod = ws.cell(row=r, column=4).value
        spec = ws.cell(row=r, column=7).value
        if not prod:
            stats['no_prod'] += 1
            continue
        key = (str(prod).strip(), str(spec).strip() if spec else '')
        cell_e = ws.cell(row=r, column=5)
        cell_f = ws.cell(row=r, column=6)
        # 먼저 둘 다 초기화
        cell_e.value = None
        cell_f.value = None
        cell_e.fill = FILL_NONE
        cell_f.fill = FILL_NONE

        a_val = a_map.get(key)
        if a_val is None:
            # 5월 신규 → 회색
            stats['a_miss_5월신규'] += 1
            cell_e.fill = FILL_UNMATCH
            cell_f.fill = FILL_UNMATCH
            if len(samples_5월신규) < 10:
                samples_5월신규.append((r, prod, spec))
            continue
        if is_meta(a_val):
            # 메타 표기 → 회색
            stats['a_meta_meta'] += 1
            cell_e.fill = FILL_UNMATCH
            cell_f.fill = FILL_UNMATCH
            continue
        # 정상 매칭값 → E·F 채움
        stats['a_matched'] += 1
        cell_e.value = a_val
        mes = k_to_mes.get(a_val)
        cell_f.value = mes or ''

    wb.save(OUT)

    total = sum(stats.values())
    print()
    print('=== 결과 ===')
    print(f'  총 행: {total}')
    print(f'  A에서 정본 매칭값 입력: {stats["a_matched"]} ({stats["a_matched"]*100/total:.1f}%)')
    print(f'  A의 메타 표기(미매칭): {stats["a_meta_meta"]} (회색)')
    print(f'  A에 없음 (5월 신규): {stats["a_miss_5월신규"]} (회색)')
    print(f'  품명 없는 행: {stats["no_prod"]}')
    print()
    print('=== 5월 신규(샘플) ===')
    for r, p, s in samples_5월신규:
        print(f'  r{r}: 품명={p!r} 규격={s!r}')


if __name__ == '__main__':
    main()
