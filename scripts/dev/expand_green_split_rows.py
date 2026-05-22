"""진한 초록(FF00B050) 부모 6행 → 분해된 자식 22행을 부모 바로 아래에 삽입.

권동환 엑셀('확정품명, 코드 추가.xlsx')에만 영향. DB는 건드리지 않음.
"""
import re
import sqlite3
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

CUR = Path('_attic/data/0520 권동환 사원님 재고/확정품명, 코드 추가.xlsx')
DB = Path('backend/mes.db')
GREEN = 'FF00B050'


def fc(cell):
    f = cell.fill
    if not f or f.patternType is None:
        return None
    try:
        rgb = f.fgColor.rgb
        return str(rgb) if rgb else None
    except Exception:
        return None


def load_db_names():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute('SELECT item_code, item_name FROM items WHERE item_code IS NOT NULL')
    by_code = {c.strip(): (n or '').strip() for c, n in cur.fetchall()}
    conn.close()
    return by_code


def parse_parent(ws, r):
    prod = ws.cell(row=r, column=4).value
    spec = ws.cell(row=r, column=5).value
    f_val = ws.cell(row=r, column=6).value or ''
    g_val = ws.cell(row=r, column=7).value or ''
    m = re.search(r"'([^']+)'", str(f_val))
    names = []
    if m:
        names = [n.strip() for n in m.group(1).split('&') if n.strip()]
    codes = [c.strip() for c in str(g_val).split('&') if c.strip()]
    return {'r': r, 'prod': prod, 'spec': spec, 'names': names, 'codes': codes}


def main():
    db_by_code = load_db_names()
    wb = openpyxl.load_workbook(CUR)
    ws = wb['26.05월_수정본']

    parents = []
    for r in range(4, ws.max_row + 1):
        if fc(ws.cell(row=r, column=6)) == GREEN:
            parents.append(parse_parent(ws, r))

    print(f'진한 초록 부모 {len(parents)}건')
    total_children = sum(len(p['codes']) for p in parents)
    print(f'삽입할 자식 행: 총 {total_children}개')

    # 역순 처리 (앞 행 삽입이 뒤 행 인덱스 안 흔들도록)
    for parent in sorted(parents, key=lambda p: -p['r']):
        pr = parent['r']
        n = len(parent['codes'])
        ws.insert_rows(pr + 1, amount=n)
        parent_height = ws.row_dimensions[pr].height

        for i, (mname_in_memo, code) in enumerate(zip(parent['names'], parent['codes']), start=1):
            child_r = pr + i
            db_name = db_by_code.get(code, mname_in_memo)

            ws.cell(row=child_r, column=4, value=parent['prod'])
            ws.cell(row=child_r, column=5, value=parent['spec'])
            ws.cell(row=child_r, column=6, value=db_name)
            ws.cell(row=child_r, column=7, value=code)

            # 부모 행 셀 스타일 복사 후 fill만 해제
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                parent_cell = ws.cell(row=pr, column=col)
                child_cell = ws.cell(row=child_r, column=col)
                if parent_cell.has_style:
                    child_cell.font = copy(parent_cell.font)
                    child_cell.border = copy(parent_cell.border)
                    child_cell.alignment = copy(parent_cell.alignment)
                    child_cell.number_format = parent_cell.number_format
                child_cell.fill = PatternFill(fill_type=None)

            if parent_height is not None:
                ws.row_dimensions[child_r].height = parent_height

        print(f'  r{pr} ({parent["prod"]!r}) → 자식 {n}행 r{pr+1}~r{pr+n}')

    wb.save(CUR)
    print(f'저장: {CUR.name}')


if __name__ == '__main__':
    main()
