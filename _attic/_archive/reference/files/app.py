"""
덱스코윈 재고관리 Flask 서버
실행: python app.py
접속: http://[PC_IP]:5000
"""
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from datetime import datetime, date
import threading
import os
import json

app = Flask(__name__)
CORS(app)

# ── 설정: 엑셀 파일 경로 ───────────────────────────────────────────
BASE_PATH = r"C:\Users\HW\Desktop\재고관리"

FILES = {
    "창고": "F704-03 (R00) 자재 재고 현황.xlsx",
    "조립": "2026.03_생산부 자재_조립,출하파트.xlsx",
    "고압": "2026.03_생산부 자재_고압,진공,튜닝파트.xlsx",
    "출하": "2026.03_출하_완제품.xlsx",
    "데모": "2026.03_데모장비및 테스트장비.xlsx",
}

FILE_CONFIG = {
    "창고": {
        "sheet_main": "26.03월",  # 매월 변경 필요
        "header_row": 3,
        "data_start": 4,
        "col_name":   4,   # D열
        "col_stock":  15,  # O열 (수식)
        "col_in":     16,  # P열
        "col_out":    32,  # AF열
        "type":       "direct",
    },
    "조립": {
        "sheet_main": "조립 자재",
        "sheet_in":   "입고내역",
        "sheet_out":  "출고내역",
        "header_row": 2,
        "data_start": 3,
        "col_name":   4,
        "col_stock":  9,
        "date_start": 6,
        "type":       "daily",
    },
    "고압": {
        "sheet_main": "고압",
        "sheet_in":   "입고내역",
        "sheet_out":  "출고내역",
        "header_row": 2,
        "data_start": 3,
        "col_name":   5,
        "col_stock":  10,
        "date_start": 6,
        "type":       "daily",
    },
    "출하": {
        "sheet_main": "완제품",
        "sheet_in":   "입고내역",
        "sheet_out":  "출고내역",
        "header_row": 2,
        "data_start": 3,
        "col_name":   4,
        "col_stock":  9,
        "date_start": 6,
        "type":       "daily",
    },
    "데모": {
        "sheet_main": "조립 자재",
        "sheet_in":   "입고내역",
        "sheet_out":  "출고내역",
        "header_row": 2,
        "data_start": 3,
        "col_name":   4,
        "col_stock":  9,
        "date_start": 6,
        "type":       "daily",
    },
}

PART_TO_FILE = {
    "자재창고":  "창고",
    "조립/출하": "조립",
    "고압파트":  "고압",
    "진공파트":  "고압",
    "튜닝파트":  "고압",
    "출하":      "출하",
    "데모":      "데모",
}

file_locks = {k: threading.Lock() for k in FILES}


# ── 유틸 ──────────────────────────────────────────────────────────

def get_file_path(file_key):
    return os.path.join(BASE_PATH, FILES[file_key])


def find_row_by_name(ws, cfg, name):
    """품명으로 행 번호 찾기 (D열 단독 또는 D+E열 조합)"""
    col  = cfg["col_name"]
    name = name.strip()
    for r in range(cfg["data_start"], ws.max_row + 1):
        v = ws.cell(r, col).value
        if not v:
            continue
        cell_name = str(v).strip()
        if cell_name == name:
            return r
        e_val = ws.cell(r, col + 1).value
        if e_val:
            if f"{cell_name} ({str(e_val).strip()})" == name:
                return r
            if f"{cell_name} {str(e_val).strip()}" == name:
                return r
    return None


def find_date_col(ws, cfg, target_date):
    """날짜로 열 번호 찾기"""
    for ci in range(cfg["date_start"], ws.max_column + 1):
        v = ws.cell(cfg["header_row"], ci).value
        if v is None:
            break
        if isinstance(v, datetime) and v.date() == target_date:
            return ci
        elif isinstance(v, date) and v == target_date:
            return ci
    return None


# ── 라우트 ────────────────────────────────────────────────────────

@app.route("/")
def index():
    html_path = os.path.join(BASE_PATH, "inventory_v2.html")
    if os.path.exists(html_path):
        return send_file(html_path)
    return "inventory_v2.html 파일이 없습니다. 같은 폴더에 넣어주세요.", 404


@app.route("/api/ping")
def ping():
    return jsonify({"status": "ok", "message": "덱스코윈 재고관리 서버 정상 작동 중"})


@app.route("/api/stock")
def get_stock():
    """단일 품목 현재고 조회"""
    name     = request.args.get("name", "").strip()
    file_key = request.args.get("file", "조립")
    if not name or file_key not in FILES:
        return jsonify({"status": "error", "msg": "파라미터 오류"}), 400

    cfg   = FILE_CONFIG[file_key]
    fpath = get_file_path(file_key)
    if not os.path.exists(fpath):
        return jsonify({"status": "error", "msg": f"파일 없음: {fpath}"}), 404

    try:
        wb  = openpyxl.load_workbook(fpath, data_only=True)
        ws  = wb[cfg["sheet_main"]]
        row = find_row_by_name(ws, cfg, name)
        if row is None:
            return jsonify({"status": "error", "msg": f"품목 없음: {name}"}), 404
        stock = ws.cell(row, cfg["col_stock"]).value or 0
        return jsonify({"status": "ok", "name": name, "stock": stock})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500


@app.route("/api/stock/all")
def get_all_stocks():
    """모든 파일에서 현재고 읽어 반환 {품명: 재고}"""
    result = {}
    for file_key, cfg in FILE_CONFIG.items():
        fpath = get_file_path(file_key)
        if not os.path.exists(fpath):
            continue
        try:
            wb        = openpyxl.load_workbook(fpath, data_only=True)
            ws        = wb[cfg["sheet_main"]]
            col_name  = cfg["col_name"]
            col_stock = cfg["col_stock"]
            for r in range(cfg["data_start"], ws.max_row + 1):
                raw_name = ws.cell(r, col_name).value
                if not raw_name:
                    continue
                name = str(raw_name).strip()
                if not name or name == "-":
                    continue
                raw_stock = ws.cell(r, col_stock).value
                if raw_stock is None:
                    continue
                try:
                    result[name] = int(raw_stock)
                except (ValueError, TypeError):
                    pass
                e_val = ws.cell(r, col_name + 1).value
                if e_val:
                    try:
                        result[f"{name} ({str(e_val).strip()})"] = int(raw_stock)
                    except (ValueError, TypeError):
                        pass
        except Exception as e:
            print(f"[WARN] {file_key} 읽기 실패: {e}")
    print(f"[동기화] 엑셀에서 {len(result)}개 품목 재고 읽기 완료")
    return jsonify({"status": "ok", "stocks": result})


@app.route("/api/stock/compare")
def compare_stock():
    """앱 재고와 엑셀 재고 비교"""
    try:
        raw = request.args.get("stocks", "")
        if not raw:
            return jsonify({"status": "ok", "mismatches": []})

        items      = json.loads(raw)
        mismatches = []

        for item in items:
            name      = item.get("name", "").strip()
            app_stock = item.get("stock", 0)
            part      = item.get("part", "")
            file_key  = PART_TO_FILE.get(part)
            if not file_key:
                continue
            cfg   = FILE_CONFIG[file_key]
            fpath = get_file_path(file_key)
            if not os.path.exists(fpath):
                continue
            try:
                wb          = openpyxl.load_workbook(fpath, data_only=True)
                ws          = wb[cfg["sheet_main"]]
                row         = find_row_by_name(ws, cfg, name)
                if row is None:
                    continue
                excel_stock = ws.cell(row, cfg["col_stock"]).value
                if excel_stock is None:
                    continue
                excel_stock = int(excel_stock)
                if excel_stock != int(app_stock):
                    diff = excel_stock - int(app_stock)
                    sign = "+" if diff > 0 else ""
                    print(f"[불일치] {name}: 앱={int(app_stock)} 엑셀={excel_stock} ({sign}{diff})")
                    mismatches.append({"name": name, "app": int(app_stock), "excel": excel_stock, "diff": diff})
            except Exception:
                continue

        return jsonify({"status": "ok", "mismatches": mismatches})
    except Exception as e:
        return jsonify({"status": "error", "msg": str(e)}), 500


@app.route("/api/io", methods=["POST"])
def process_io():
    """
    입출고 처리
    body: { "name": "품명", "qty": 10, "mode": "in"|"out", "file": "창고"|... }
    """
    data = request.json
    if not data:
        return jsonify({"status": "error", "msg": "요청 데이터 없음"}), 400

    name     = data.get("name", "").strip()
    qty      = int(data.get("qty", 0))
    mode     = data.get("mode", "out")
    file_key = data.get("file", "조립")

    if not name or qty <= 0:
        return jsonify({"status": "error", "msg": "품명 또는 수량 오류"}), 400
    if file_key not in FILES:
        return jsonify({"status": "error", "msg": f"파일 키 오류: {file_key}"}), 400

    cfg   = FILE_CONFIG[file_key]
    fpath = get_file_path(file_key)
    if not os.path.exists(fpath):
        return jsonify({"status": "error", "msg": f"파일 없음: {fpath}"}), 404

    with file_locks[file_key]:
        try:
            print(f"[IO] name='{name}' qty={qty} mode={mode} file={file_key}")
            wb = openpyxl.load_workbook(fpath)

            if cfg["type"] == "direct":
                # ── 자재창고: 입/출고 열에 누적 기입 ────────────────
                ws  = wb[cfg["sheet_main"]]
                row = find_row_by_name(ws, cfg, name)
                if row is None:
                    col    = cfg["col_name"]
                    sample = [str(ws.cell(r, col).value or "").strip()
                              for r in range(cfg["data_start"], min(cfg["data_start"] + 10, ws.max_row + 1))]
                    print(f"[DEBUG] sheet={cfg['sheet_main']} 찾는품명='{name}' 샘플={sample[:5]}")
                    return jsonify({"status": "error", "msg": f"품목 없음: {name}"}), 404

                if mode == "in":
                    cur = ws.cell(row, cfg["col_in"]).value or 0
                    ws.cell(row, cfg["col_in"]).value = cur + qty
                else:
                    cur = ws.cell(row, cfg["col_out"]).value or 0
                    ws.cell(row, cfg["col_out"]).value = cur + qty

            else:
                # ── 생산부 파일: 입/출고 시트 날짜 열에 기입 ─────────
                sheet_name = cfg["sheet_in"] if mode == "in" else cfg["sheet_out"]
                ws         = wb[sheet_name]
                today      = datetime.now().date()

                # data_only 워크북으로 날짜 열 탐색
                wb_ro    = openpyxl.load_workbook(fpath, data_only=True)
                ws_ro    = wb_ro[sheet_name]
                date_col = find_date_col(ws_ro, cfg, today)

                if date_col is None:
                    sample = [str(ws.cell(cfg["header_row"], ci).value)
                              for ci in range(cfg["date_start"], min(cfg["date_start"] + 5, ws.max_column + 1))]
                    print(f"[DEBUG] 오늘={today} 엑셀날짜샘플={sample}")
                    return jsonify({
                        "status": "error",
                        "msg":    f"오늘 날짜({today}) 열을 찾을 수 없습니다. 파일명/시트명을 확인하세요."
                    }), 404

                row = find_row_by_name(ws, cfg, name)
                if row is None:
                    col    = cfg["col_name"]
                    sample = [str(ws.cell(r, col).value or "").strip()
                              for r in range(cfg["data_start"], min(cfg["data_start"] + 10, ws.max_row + 1))]
                    print(f"[DEBUG] sheet={sheet_name} 찾는품명='{name}' 샘플={sample[:5]}")
                    return jsonify({"status": "error", "msg": f"품목 없음: {name}"}), 404

                cur = ws.cell(row, date_col).value or 0
                ws.cell(row, date_col).value = cur + qty

            wb.save(fpath)
            return jsonify({"status": "success", "msg": f"{name} {mode} {qty}개 기록 완료"})

        except Exception as e:
            return jsonify({"status": "error", "msg": str(e)}), 500


# ── 진입점 ────────────────────────────────────────────────────────

if __name__ == "__main__":
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
    except Exception:
        ip = "127.0.0.1"

    print("=" * 52)
    print("  덱스코윈 재고관리 서버 시작")
    print(f"  엑셀 파일 경로: {BASE_PATH}")
    print()
    print(f"  >>> 폰에서 접속: http://{ip}:5000")
    print(f"  >>> 연결 테스트: http://{ip}:5000/api/ping")
    print("=" * 52)
    app.run(host="0.0.0.0", port=5000, debug=False)
