"""
DB restore + ERP code update (sort_order 1:1 match with Excel row index)
- Never matches by item_name
- Excel data row N (row 2 = N=1) -> DB sort_order N
"""
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
DB_PATH = BASE / "backend" / "erp.db"
BACKUP_SRC = BASE / "backend" / "erp_backup_20260430_094340.db"
EXCEL_PATH = BASE / "outputs" / "inventory_cleanup" / "생산부_재고_매칭작업_정리본_공정순번.xlsx"

# ── 1. backup current DB ───────────────────────────────────────────────────────
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
pre_restore_backup = BASE / "backend" / f"erp_before_restore_{ts}.db"
shutil.copy2(DB_PATH, pre_restore_backup)
print(f"[1] current DB backed up: {pre_restore_backup.name}")

# ── 2. restore from backup ─────────────────────────────────────────────────────
shutil.copy2(BACKUP_SRC, DB_PATH)
print(f"[2] DB restored: {BACKUP_SRC.name} -> erp.db")

# ── 3. verify restored DB ─────────────────────────────────────────────────────
conn = sqlite3.connect(str(DB_PATH))
conn.row_factory = sqlite3.Row
cur = conn.cursor()

cur.execute("SELECT COUNT(*) FROM items")
total = cur.fetchone()[0]
assert total == 722, f"items count error: {total} (expected 722)"

cur.execute("SELECT sort_order FROM items ORDER BY sort_order")
sort_orders = [r[0] for r in cur.fetchall()]
expected = list(range(1, 723))
missing = sorted(set(expected) - set(sort_orders))
extra = sorted(set(sort_orders) - set(expected))
assert not missing, f"sort_order missing: {missing}"
assert not extra, f"sort_order unexpected: {extra}"
print(f"[3] restore OK: {total} items, sort_order 1~722 complete")

# ── 4. read Excel ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(str(EXCEL_PATH))
sheet_name = "ERP_코드" if "ERP_코드" in wb.sheetnames else wb.sheetnames[0]
ws = wb[sheet_name]
print(f"[4] Excel sheet: '{sheet_name}', total rows: {ws.max_row}")

# data rows: row 2~723 -> sort_order 1~722
excel_rows = {}  # sort_order -> {erp_code, item_name}
for row_idx in range(2, ws.max_row + 1):
    sort_order = row_idx - 1
    erp_code_val = ws.cell(row_idx, 1).value  # column A
    item_name_val = ws.cell(row_idx, 2).value  # column B (for verification only)
    if erp_code_val is None:
        break
    excel_rows[sort_order] = {
        "erp_code": str(erp_code_val).strip(),
        "item_name": str(item_name_val).strip() if item_name_val else "",
    }

assert len(excel_rows) == 722, f"Excel data row count error: {len(excel_rows)} (expected 722)"
print(f"[4] Excel data loaded: {len(excel_rows)} rows")

# ── 5. item_name verification (reference only, not used for matching) ──────────
cur.execute("SELECT sort_order, item_name FROM items ORDER BY sort_order")
db_rows_by_so = {r[0]: r[1] for r in cur.fetchall()}

name_mismatches = []
for so, ex in excel_rows.items():
    db_name = db_rows_by_so.get(so, "")
    if db_name != ex["item_name"]:
        name_mismatches.append((so, db_name, ex["item_name"]))

if name_mismatches:
    print(f"[5] item_name mismatches (reference only, {len(name_mismatches)} items):")
    for so, db_n, ex_n in name_mismatches[:10]:
        print(f"    sort_order={so}: DB='{db_n}' / Excel='{ex_n}'")
    if len(name_mismatches) > 10:
        print(f"    ... and {len(name_mismatches) - 10} more")
else:
    print("[5] item_name: all match (OK)")

# ── 6. ERP code parse function ─────────────────────────────────────────────────
def parse_erp_code(code: str):
    """
    Format: {model_symbol}-{process_type_code}-{NNNN}
    Returns: (model_symbol, process_type_code, serial_no) or None
    """
    parts = code.split("-")
    if len(parts) != 3:
        return None
    model_sym, proc_type, serial_str = parts
    if not model_sym.isdigit():
        return None
    if not proc_type.isalpha():
        return None
    if not serial_str.isdigit() or len(serial_str) != 4:
        return None
    return model_sym, proc_type.upper(), int(serial_str)

# ── 7. update in transaction (2-phase TEMP approach) ──────────────────────────
print("[7] update started (2-phase TEMP)...")

try:
    cur.execute("BEGIN")

    # phase 1: set all to unique TEMP values to avoid UNIQUE conflicts
    for so in range(1, 723):
        cur.execute(
            "UPDATE items SET erp_code = ? WHERE sort_order = ?",
            (f"__TEMP__{so}", so),
        )

    # phase 2: set final erp_code + derived columns
    parse_fail = []
    for so, ex in excel_rows.items():
        erp_code = ex["erp_code"]
        parsed = parse_erp_code(erp_code)
        if parsed:
            model_sym, proc_type, serial_no = parsed
            cur.execute(
                """UPDATE items
                   SET erp_code = ?,
                       model_symbol = ?,
                       process_type_code = ?,
                       serial_no = ?,
                       updated_at = datetime('now')
                   WHERE sort_order = ?""",
                (erp_code, model_sym, proc_type, serial_no, so),
            )
        else:
            parse_fail.append((so, erp_code))
            cur.execute(
                """UPDATE items
                   SET erp_code = ?,
                       updated_at = datetime('now')
                   WHERE sort_order = ?""",
                (erp_code, so),
            )

    # verify no TEMP remains before commit
    cur.execute("SELECT COUNT(*) FROM items WHERE erp_code LIKE '__TEMP__%'")
    temp_remain = cur.fetchone()[0]
    if temp_remain != 0:
        conn.rollback()
        print(f"[ERROR] {temp_remain} TEMP rows remain -- rolled back")
        sys.exit(1)

    conn.commit()
    print(f"[7] committed successfully")
    if parse_fail:
        print(f"    parse-fail codes (erp_code only updated, {len(parse_fail)} items):")
        for so, code in parse_fail[:5]:
            print(f"      sort_order={so}: {code}")

except Exception as e:
    conn.rollback()
    print(f"[ERROR] exception -- rolled back: {e}")
    raise

# ── 8. final verification ──────────────────────────────────────────────────────
cur.execute("SELECT COUNT(*) FROM items")
final_total = cur.fetchone()[0]

cur.execute("SELECT COUNT(*) FROM items WHERE erp_code IS NOT NULL")
not_null = cur.fetchone()[0]

cur.execute("SELECT COUNT(DISTINCT erp_code) FROM items")
unique_count = cur.fetchone()[0]

cur.execute("SELECT COUNT(*) FROM items WHERE erp_code IS NULL")
null_count = cur.fetchone()[0]

cur.execute("SELECT COUNT(*) FROM items WHERE erp_code LIKE '__TEMP__%'")
temp_count = cur.fetchone()[0]

cur.execute("SELECT sort_order, erp_code FROM items ORDER BY sort_order")
db_erp_by_so = {r[0]: r[1] for r in cur.fetchall()}

code_mismatch = []
for so, ex in excel_rows.items():
    if db_erp_by_so.get(so) != ex["erp_code"]:
        code_mismatch.append((so, db_erp_by_so.get(so), ex["erp_code"]))

conn.close()

# ── 9. summary ────────────────────────────────────────────────────────────────
print()
print("=" * 60)
print("RESULT SUMMARY")
print("=" * 60)
print(f"restore source:       {BACKUP_SRC.name}")
print(f"pre-restore backup:   {pre_restore_backup.name}")
print()
print(f"items total:          {final_total}  (expected: 722)")
print(f"erp_code NOT NULL:    {not_null}  (expected: 722)")
print(f"erp_code UNIQUE:      {unique_count}  (expected: 722)")
print(f"erp_code NULL:        {null_count}  (expected: 0)")
print(f"TEMP remaining:       {temp_count}  (expected: 0)")
print()

ok = (final_total == 722 and not_null == 722 and unique_count == 722
      and null_count == 0 and temp_count == 0 and not code_mismatch)
if ok:
    print("ALL CHECKS PASSED")
else:
    print("CHECKS FAILED:")
    if code_mismatch:
        print(f"  erp_code mismatch: {len(code_mismatch)} items")
        for so, db_c, ex_c in code_mismatch[:5]:
            print(f"    sort_order={so}: DB={db_c!r} / Excel={ex_c!r}")

print()
if name_mismatches:
    print(f"item_name mismatch (reference only): {len(name_mismatches)} items")
else:
    print("item_name mismatch: none")
