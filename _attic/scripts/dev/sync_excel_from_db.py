"""담당자 참조 엑셀(생산부_재고_매칭작업_최종.xlsx) ↔ DB 단방향 동기화.

DB가 정본. 본 스크립트는 DB를 변경하지 않는다.

수행:
  1) 백업 (.bak_<ts>.xlsx)
  2) 방사구 수정: 고압진공_재고현황 R112 모델열(3,4,6,7,8)="O" + 마스터 P129="34678-VR-0039"
  3) 신규 시트 '출하_재고현황' 생성 (조립완제품_재고현황 헤더/서식 복제)
  4) DB의 PA/PF 중 '엑셀 마스터에 아직 없는' 50건을 원본시트+마스터에 append (수식·교차링크 보존)
  5) autofilter 확장 + 저장

비멱등: '출하_재고현황' 가 이미 있으면 중단(백업에서 복원 후 재실행).
"""
import datetime
import re
import shutil
import sqlite3
import sys
from copy import copy

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

XLSX = r"data/생산부_재고_매칭작업_최종.xlsx"
DB = r"backend/erp.db"
SRC_SHEET = "조립완제품_재고현황"
NEW_SHEET = "출하_재고현황"
AFTER_SHEET = "창고_재고현황"
MASTER = "마스터_품목"

DIGIT_TO_COL = {"3": 10, "4": 11, "6": 12, "7": 13, "8": 14, "?": 15}  # 원본시트 J~O


def norm(v):
    return re.sub(r"\s+", "", str(v)).upper() if v not in (None, "") else ""


def clone_cell_style(src, dst):
    """src 셀의 서식(글꼴·채우기·테두리·정렬·표시형식)을 dst 셀에 복제."""
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


# 원본시트 I/P 표준 수식 (행 r)
def i_formula(r):
    return (
        '=IFERROR(IF(OR($F{r}="O",$G{r}="O",$H{r}="O"),'
        'CHOOSE(MATCH($A{r},{{"튜브","고압","진공","튜닝","조립","출하"}},0),'
        '"T","H","V","N","A","P")'
        '&IF($F{r}="O","R",IF($G{r}="O","A",IF($H{r}="O","F",""))),""),"")'
    ).format(r=r)


def p_formula(r):
    return (
        '=IF($I{r}="","",IF($J{r}="O","3","")&IF($K{r}="O","4","")'
        '&IF($L{r}="O","6","")&IF($M{r}="O","7","")&IF($N{r}="O","8","")'
        '&IF($O{r}="O","?","")'
        '&IF(OR($J{r}="O",$K{r}="O",$L{r}="O",$M{r}="O",$N{r}="O",$O{r}="O"),"-","")'
        "&$I{r})"
    ).format(r=r)


# 마스터 표준 수식 (마스터행 m)
def m_d(m):
    return (
        '=IF(INDIRECT("\'"&$O{m}&"\'!F"&$N{m})="O","원자재",'
        'IF(INDIRECT("\'"&$O{m}&"\'!G"&$N{m})="O","반제품",'
        'IF(INDIRECT("\'"&$O{m}&"\'!H"&$N{m})="O","완제품","")))'
    ).format(m=m)


def m_e(m):
    return (
        '=IFERROR(IF(INDIRECT("\'"&$O{m}&"\'!I"&$N{m})=0,"",'
        'INDIRECT("\'"&$O{m}&"\'!I"&$N{m})),"")'
    ).format(m=m)


def m_h(m):
    return (
        '=IF(AND($F{m}="",$G{m}=""),"",IF($F{m}="","창고전용",'
        'IF($G{m}<>"","매칭완료",IF($D{m}="반제품","반제품(창고없음)",'
        'IF($D{m}="완제품","완제품(창고없음)",'
        'IF($D{m}="원자재","⚠ 원자재_부서보관","미분류"))))))'
    ).format(m=m)


def m_i(m):
    return (
        '=IFERROR(IF(INDIRECT("\'"&$O{m}&"\'!P"&$N{m})=0,"",'
        'INDIRECT("\'"&$O{m}&"\'!P"&$N{m})),"")'
    ).format(m=m)


def m_l(m):
    return '=HYPERLINK("#\'"&$O{m}&"\'!A"&$N{m},"→ 출하 R"&$N{m})'.format(m=m)


def q_link(mrow):
    return '=HYPERLINK("#\'마스터_품목\'!A{m}","→ 마스터 R{m}")'.format(m=mrow)


def main():
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = XLSX.replace(".xlsx", f".bak_{ts}.xlsx")
    shutil.copy(XLSX, bak)
    print(f"backup -> {bak}")

    wb = openpyxl.load_workbook(XLSX, data_only=False)
    if NEW_SHEET in wb.sheetnames:
        raise SystemExit(
            f"'{NEW_SHEET}' 가 이미 존재합니다. 백업에서 복원 후 재실행하세요."
        )
    master = wb[MASTER]

    # 기존 마스터 P(MES) 집합 + 실제 마지막 데이터행
    existing = set()
    mlast = 1
    for r in range(2, master.max_row + 1):
        p = master.cell(r, 16).value
        if p not in (None, ""):
            existing.add(norm(p))
            mlast = r

    # --- 2) 방사구 수정 ---
    vr_row = next(
        (r for r in range(2, master.max_row + 1)
         if norm(master.cell(r, 16).value) == "VR-0039"),
        None,
    )
    if vr_row is None:
        raise SystemExit("VR-0039 마스터행을 찾지 못함")
    o_sheet = master.cell(vr_row, 15).value
    o_row = int(master.cell(vr_row, 14).value)
    if o_sheet != "고압진공_재고현황" or o_row != 112:
        raise SystemExit(f"방사구 원본 위치 예상과 다름: {o_sheet}!R{o_row}")
    osh = wb[o_sheet]
    for col in (10, 11, 12, 13, 14):  # J,K,L,M,N = 모델 3,4,6,7,8
        osh.cell(o_row, col, "O")
        clone_cell_style(osh.cell(2, col), osh.cell(o_row, col))  # 같은 컬럼 표준 서식
    master.cell(vr_row, 16, "34678-VR-0039")
    print(f"방사구: {o_sheet}!R{o_row} J~N='O', 마스터 P{vr_row}='34678-VR-0039'")

    # --- 3) DB 조회 + '엑셀에 없는' PA/PF 만 ---
    con = sqlite3.connect(DB)
    db_rows = con.execute(
        "SELECT erp_code,item_name,unit,model_symbol,process_type_code "
        "FROM items WHERE erp_code GLOB '[0-9]*-P[AF]-*' ORDER BY erp_code"
    ).fetchall()
    con.close()
    to_add = [x for x in db_rows if norm(x[0]) not in existing]
    print(f"PA/PF 후보 {len(db_rows)}건 중 신규 {len(to_add)}건")
    if len(to_add) != 50:
        raise SystemExit(f"신규 50건 기대, 실제 {len(to_add)}건 — 중단")

    # --- 4) 출하_재고현황 생성 (조립완제품 헤더/서식 복제) ---
    src = wb[SRC_SHEET]
    idx = wb.sheetnames.index(AFTER_SHEET) + 1
    osheet = wb.create_sheet(NEW_SHEET, idx)
    for c in range(1, 20):  # A:S 19열
        sc = src.cell(1, c)
        clone_cell_style(sc, osheet.cell(1, c, sc.value))
    for col, dim in src.column_dimensions.items():
        if dim.width:
            osheet.column_dimensions[col].width = dim.width
    osheet.row_dimensions[1].height = src.row_dimensions[1].height
    osheet.freeze_panes = "A2"

    # 데이터행 서식 템플릿: 원본=조립완제품 R2, 마스터=마스터 R2
    osrc_h = src.row_dimensions[2].height
    msrc_h = master.row_dimensions[2].height

    r = 2
    m = mlast + 1
    for code, name, unit, msym, ptc in to_add:
        model = name.split("_")[0].strip()
        bunryu = "가방 포장 완료" if ptc == "PA" else "박스 포장 완료"
        digit = re.match(r"^(\d+|\?)", code).group(1)

        # 원본행
        osheet.cell(r, 1, "출하")
        osheet.cell(r, 2, bunryu)
        osheet.cell(r, 3, model)
        osheet.cell(r, 4, name)
        osheet.cell(r, 7 if ptc == "PA" else 8, "O")  # G 반제품 / H 완제품
        for d in digit:
            osheet.cell(r, DIGIT_TO_COL[d], "O")
        osheet.cell(r, 9, i_formula(r))
        osheet.cell(r, 16, p_formula(r))
        osheet.cell(r, 17, q_link(m))
        osheet.cell(r, 18, m)
        osheet.cell(r, 19, MASTER)
        for c in range(1, 20):  # 조립완제품 R2 서식 복제
            clone_cell_style(src.cell(2, c), osheet.cell(r, c))
        osheet.row_dimensions[r].height = osrc_h

        # 마스터행
        master.cell(m, 1, "출하")
        master.cell(m, 2, bunryu)
        master.cell(m, 3, model)
        master.cell(m, 4, m_d(m))
        master.cell(m, 5, m_e(m))
        master.cell(m, 6, name)
        master.cell(m, 7, name)
        master.cell(m, 8, m_h(m))
        master.cell(m, 9, m_i(m))
        master.cell(m, 10, name)
        master.cell(m, 11, name)
        master.cell(m, 12, m_l(m))
        master.cell(m, 14, r)
        master.cell(m, 15, NEW_SHEET)
        master.cell(m, 16, code)
        for c in range(1, 17):  # 마스터 R2 서식 복제
            clone_cell_style(master.cell(2, c), master.cell(m, c))
        master.row_dimensions[m].height = msrc_h

        r += 1
        m += 1

    last_o = r - 1
    last_m = m - 1

    # --- 5) autofilter 확장 + 신규시트 CF(국한·코스메틱) ---
    osheet.auto_filter.ref = f"A1:S{last_o}"
    master.auto_filter.ref = f"A1:O{last_m}"

    blue = PatternFill("solid", fgColor="FFDDEBF7")
    amber = PatternFill("solid", fgColor="FFFFF2CC")
    green = PatternFill("solid", fgColor="FFE2EFDA")
    orange = PatternFill("solid", fgColor="FFFCE4D6")
    osheet.conditional_formatting.add(
        f"A2:A{last_o}",
        FormulaRule(formula=['$A2="출하"'], fill=orange),
    )
    osheet.conditional_formatting.add(
        f"I2:I{last_o}",
        FormulaRule(formula=['RIGHT($I2,1)="R"'], fill=blue),
    )
    osheet.conditional_formatting.add(
        f"I2:I{last_o}",
        FormulaRule(formula=['RIGHT($I2,1)="A"'], fill=amber),
    )
    osheet.conditional_formatting.add(
        f"I2:I{last_o}",
        FormulaRule(formula=['RIGHT($I2,1)="F"'], fill=green),
    )

    wb.save(XLSX)
    print(
        f"저장 완료: 출하_재고현황 R2~R{last_o} ({last_o-1}건), "
        f"마스터 R{mlast+1}~R{last_m} append"
    )


if __name__ == "__main__":
    main()
