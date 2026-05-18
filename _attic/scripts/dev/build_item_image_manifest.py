"""추출 이미지 → 마스터_품목(F열) → MES 코드(P열) 매핑 + frontend/public 배치.

흐름:
  1. 마스터_품목 시트 스캔: sanitize(F 생산부 품명) → [(P MES 코드, 원본 F품명, 그룹), ...]
  2. extracted/{시트}/r{NNN}_{품목명}.{ext} 순회 (-N 접미사는 첫 장만 사용)
  3. 매칭된 이미지를 frontend/public/images/items/{MES코드}.{ext} 로 복사 (1:N이면 N번 복사)
  4. manifest.json 생성: {erp_code: filename}
  5. _unmatched.csv: 매칭 실패 이미지 + 부분일치 후보
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import sys
import warnings
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
MASTER_XLSX = ROOT / "data" / "생산부_재고_매칭작업.xlsx"
EXTRACTED_BASE = ROOT / "data" / "이미지 추출을 위한 원본" / "extracted"
PUBLIC_DIR = ROOT / "frontend" / "public" / "images" / "items"
UNMATCHED_CSV = EXTRACTED_BASE / "_unmatched.csv"

# 마스터_품목 컬럼 (1-based)
COL_그룹 = 1
COL_생산부품명 = 6
COL_MES코드 = 16

FORBIDDEN = re.compile(r'[\\/:*?"<>|]')
WHITESPACE = re.compile(r"\s+")
ROW_PREFIX = re.compile(r"^r\d{3}_")
DUP_SUFFIX = re.compile(r"-\d+$")


def sanitize(name: str) -> str:
    name = FORBIDDEN.sub(" ", str(name))
    name = WHITESPACE.sub(" ", name).strip(" .")
    return name


def load_master_index() -> dict[str, list[tuple]]:
    """sanitize(생산부품명) → [(MES코드, 원본품명, 그룹)]"""
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
    ws = wb["마스터_품목"]
    idx: dict[str, list[tuple]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        pn_raw = ws.cell(r, COL_생산부품명).value
        if not pn_raw:
            continue
        pn = str(pn_raw).strip()
        mes = ws.cell(r, COL_MES코드).value
        group = ws.cell(r, COL_그룹).value
        if not mes:
            continue
        idx[sanitize(pn)].append((str(mes).strip(), pn, str(group or "").strip()))
    return idx


def list_representative_images() -> list[dict]:
    """각 (시트, 행)당 첫 장만 대표로. -N 접미사 파일 제외."""
    reps: list[dict] = []
    for sheet_dir in sorted(EXTRACTED_BASE.iterdir()):
        if not sheet_dir.is_dir():
            continue
        for f in sorted(sheet_dir.iterdir()):
            if f.suffix.lower() == ".csv":
                continue
            stem = f.stem
            # -N 접미사가 붙은 파일은 같은 행의 추가 사진 → skip
            if DUP_SUFFIX.search(stem):
                continue
            m = ROW_PREFIX.match(stem)
            if not m:
                continue
            품목 = ROW_PREFIX.sub("", stem)
            reps.append({
                "sheet": sheet_dir.name,
                "row": int(stem[1:4]),
                "품목": 품목,
                "path": f,
                "ext": f.suffix.lower(),
            })
    return reps


def find_candidates(missing_pn: str, master_idx: dict) -> list[str]:
    """미매칭 이미지 품목명에 대해, 마스터 키 중 양방향 부분일치 후보 최대 3개."""
    out = []
    needle = missing_pn.lower()
    for k in master_idx.keys():
        kl = k.lower()
        if needle in kl or kl in needle:
            out.extend(f"{mes}|{pn}" for mes, pn, _ in master_idx[k])
        if len(out) >= 3:
            break
    return out[:3]


def main() -> int:
    if not MASTER_XLSX.exists():
        print(f"마스터 xlsx 없음: {MASTER_XLSX}", file=sys.stderr)
        return 1

    print(f"마스터: {MASTER_XLSX.name}")
    master_idx = load_master_index()
    print(f"  F열(생산부품명) 인덱스: {len(master_idx)}개 unique 키, "
          f"{sum(len(v) for v in master_idx.values())}개 행")

    reps = list_representative_images()
    print(f"\n대표 이미지: {len(reps)}장 (-N 접미사 제외)")

    # 출력 폴더 준비 (기존 매니페스트 파일은 지움, 폴더는 유지)
    PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
    # 이전 실행 산출물 정리
    for old in PUBLIC_DIR.iterdir():
        if old.is_file():
            old.unlink()

    manifest: dict[str, str] = {}
    copy_count = 0
    matched_imgs = 0
    unmatched: list[dict] = []

    for img in reps:
        candidates = master_idx.get(img["품목"])
        if not candidates:
            unmatched.append(img)
            continue
        matched_imgs += 1
        for mes_code, 원본품명, 그룹 in candidates:
            # MES 코드를 파일명으로 안전화 (혹시 모를 / 등)
            safe = sanitize(mes_code).replace(" ", "")
            dst_name = f"{safe}{img['ext']}"
            dst = PUBLIC_DIR / dst_name
            shutil.copy2(img["path"], dst)
            manifest[mes_code] = dst_name
            copy_count += 1

    # manifest 참조 안 되는 orphan 파일 정리 (예: .png가 .jpg로 덮어쓰여진 잔재)
    referenced = set(manifest.values())
    orphans = 0
    for f in PUBLIC_DIR.iterdir():
        if f.is_file() and f.suffix.lower() != ".json" and f.name not in referenced:
            f.unlink()
            orphans += 1

    # manifest.json
    manifest_path = PUBLIC_DIR / "manifest.json"
    with manifest_path.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2, sort_keys=True)

    # _unmatched.csv
    with UNMATCHED_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["시트", "행", "추출_품목명", "추출_파일명", "마스터_부분일치_후보(MES|품명; 최대3개)"])
        for img in unmatched:
            cands = find_candidates(img["품목"], master_idx)
            w.writerow([
                img["sheet"], img["row"], img["품목"], img["path"].name,
                " || ".join(cands) if cands else "",
            ])

    print()
    print(f"매칭 성공 이미지: {matched_imgs} / {len(reps)}")
    print(f"  → 복사된 파일: {copy_count}개 (1:N 매핑 포함)")
    print(f"  → manifest entries: {len(manifest)}개")
    print(f"미매칭 이미지: {len(unmatched)}개  ({UNMATCHED_CSV.name})")
    print(f"orphan 파일 정리: {orphans}개")
    print(f"\n산출:")
    print(f"  {PUBLIC_DIR}\\manifest.json")
    print(f"  {PUBLIC_DIR}\\*.jpg|png ({copy_count}개)")
    print(f"  {UNMATCHED_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
