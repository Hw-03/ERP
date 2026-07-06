"""
개발 현황 엑셀 생성 스크립트
실행: python _attic/scripts/dev/generate_devlog.py
출력: _attic/docs/개발현황.xlsx
"""
import subprocess
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUTPUT_PATH = Path("_attic/docs/개발현황.xlsx")

# ── 색상 상수 ──────────────────────────────────────────────
HEADER_BG   = "1F4E79"
HEADER_FG   = "FFFFFF"
KPI_LABEL_BG = "2E75B6"
KPI_VALUE_BG = "DEEAF1"
DONE_BG     = "E2EFDA"
TODO_BG     = "F2F2F2"
STRIPE_BG   = "EBF3FB"
WHITE_BG    = "FFFFFF"
GOLD_BG     = "FFF2CC"
OFFHOURS_BG = "FCE4D6"   # 커밋 이력 행 — 근무 외 (연주황)
PERSONAL_BG = "EDEDED"   # 커밋 이력 행 — 개인 선행 (연회색)
BAR_WORK    = "000000"   # 막대 ■ — 근무 (검정)
BAR_OFF     = "ED7D31"   # 막대 ■ — 근무 외 (주황)
BAR_PRE     = "9C9C9C"   # 막대 ■ — 개인 선행 (회색)

# ── 근무 구분 판정 ─────────────────────────────────────────
PERSONAL_PRE_END = "2026-04-20"                  # 이 날짜까지는 회사 착수 이전 (개인 선행)
PUBLIC_HOLIDAYS  = {"2026-05-01", "2026-05-05"}  # 노동절·어린이날
FULL_DAY_LEAVE   = {"2026-05-07"}                # 종일 휴가
HALF_DAY_LEAVE   = {"2026-05-19": 12}            # 오후반차 — 해당 시각(시) 이후 근무 외
WORK_START_HOUR  = 8                             # 근무 시작 08:00
WORK_END_HOUR    = 17                            # 근무 종료 17:00

def classify_work(dt):
    """커밋 datetime → '개인 선행' | '근무' | '근무 외'"""
    d = dt.strftime("%Y-%m-%d")
    if d <= PERSONAL_PRE_END:
        return "개인 선행"
    if d in PUBLIC_HOLIDAYS or d in FULL_DAY_LEAVE:
        return "근무 외"
    if dt.weekday() >= 5:                         # 토(5)·일(6)
        return "근무 외"
    if d in HALF_DAY_LEAVE and dt.hour >= HALF_DAY_LEAVE[d]:
        return "근무 외"
    if WORK_START_HOUR <= dt.hour < WORK_END_HOUR:
        return "근무"
    return "근무 외"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, text):
    cell.value = text
    cell.fill = make_fill(HEADER_BG)
    cell.font = Font(bold=True, color=HEADER_FG, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

def apply_data(cell, text, bg=WHITE_BG, bold=False, align="left"):
    cell.value = text
    cell.fill = make_fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = make_border()

def auto_col_width(ws, extra=2):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

# ── Sheet 1: 대시보드 ──────────────────────────────────────
def _bar_rich(date_str, msgs_of_day):
    """날짜별 커밋 막대 — 근무/근무외/개인선행 색 분할 RichText."""
    n = len(msgs_of_day)
    label = f"{date_str}  "
    suffix = f"  ({n})"

    # 개인 선행 날짜: 전부 회색 ■
    if date_str <= PERSONAL_PRE_END:
        bar = "■" * max(1, n // 2)
        return CellRichText(
            TextBlock(InlineFont(color="000000", sz=10), label),
            TextBlock(InlineFont(color=BAR_PRE, sz=10), bar),
            TextBlock(InlineFont(color="000000", sz=10), suffix),
        )

    n_off  = sum(1 for _, _, _, dt in msgs_of_day if classify_work(dt) == "근무 외")
    n_work = n - n_off
    bar_w  = "■" * (n_work // 2)
    bar_o  = "■" * (n_off  // 2)
    # 1·3건 등 홀수 보정: 둘 다 비면 더 큰 쪽에 1개
    if not bar_w and not bar_o:
        if n_off > n_work:
            bar_o = "■"
        else:
            bar_w = "■"
    parts = [TextBlock(InlineFont(color="000000", sz=10), label)]
    if bar_w:
        parts.append(TextBlock(InlineFont(color=BAR_WORK, sz=10), bar_w))
    if bar_o:
        parts.append(TextBlock(InlineFont(color=BAR_OFF,  sz=10), bar_o))
    parts.append(TextBlock(InlineFont(color="000000", sz=10), suffix))
    return CellRichText(*parts)


def build_dashboard(ws, commits, off_count):
    ws.title = "대시보드"

    # 제목
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "📊 DEXCOWIN MES 개발 현황"
    title.fill = make_fill(HEADER_BG)
    title.font = Font(bold=True, color=HEADER_FG, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.border = make_border()
    ws.row_dimensions[1].height = 28

    # KPI 영역 (행 3~)
    ws.row_dimensions[3].height = 22
    company_commits = count_commits_in_range(commits, "2026-04-21", "2026-05-29")
    kpis = [
        ("개발 기간 (회사)", "2026-04-21 ~ 05-29 (회사 개발 착수 이후)"),
        ("개인 선행 작업", "2026-04-10 ~ 04-20 (회사 개발 착수 이전 · 개인 시간)"),
        ("총 커밋 수", f"{len(commits)}건 (회사 기간 {company_commits}건)"),
        ("근무 외 작업", f"{off_count}건 (평일 야간·주말·휴가 시간 커밋)"),
        ("기능 완료", "91 / 100개 (91%)"),
        ("개발 영역", "Backend · Frontend · Mobile · Admin · Docs"),
    ]

    for r, (label, value) in enumerate(kpis, 3):
        # 레이블
        c1 = ws.cell(r, 1)
        c1.value = label
        c1.fill = make_fill(KPI_LABEL_BG)
        c1.font = Font(bold=True, color=HEADER_FG, size=11)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c1.border = make_border()

        # 값
        ws.merge_cells(f"B{r}:E{r}")
        c2 = ws.cell(r, 2)
        c2.value = value
        c2.fill = make_fill(KPI_VALUE_BG)
        c2.font = Font(bold=True, size=11)
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = make_border()
        ws.row_dimensions[r].height = 22

    # KPI(r=3~8)와 충돌 방지 — row 9는 빈 줄, 차트 타이틀 r=10, 막대 r=11~
    # (이전 r=8 차트 타이틀은 KPI 6번째 "개발 영역"(B8:E8 머지)과 겹쳐 Excel 손상 다이얼로그 유발)
    ws.row_dimensions[10].height = 20
    ws.merge_cells("A10:E10")
    chart_title = ws["A10"]
    chart_title.value = "📈 날짜별 커밋 수"
    chart_title.fill = make_fill(HEADER_BG)
    chart_title.font = Font(bold=True, color=HEADER_FG, size=11)
    chart_title.alignment = Alignment(horizontal="center", vertical="center")
    chart_title.border = make_border()

    by_date = group_commits_by_date(commits)
    dates_sorted = sorted(by_date.keys())

    r = 11
    for date in dates_sorted:
        msgs = by_date[date]
        ws.merge_cells(f"A{r}:E{r}")
        cell = ws.cell(r, 1)
        cell.value = _bar_rich(date, msgs)
        cell.fill = make_fill(GOLD_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = make_border()
        ws.row_dimensions[r].height = 18
        r += 1

    ws.column_dimensions["A"].width = 67.25
    ws.column_dimensions["B"].width = 21.375

# ── Sheet 2: 개발 현황 요약 ───────────────────────────────
def build_summary(ws, commits):
    ws.title = "개발 현황 요약"
    ws.row_dimensions[1].height = 22

    headers = ["기간", "분류", "주요 작업", "커밋 수"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    # 자동 계산된 커밋 수
    c1 = count_commits_in_range(commits, "2026-04-10", "2026-04-11")
    c2 = count_commits_in_range(commits, "2026-04-12", "2026-04-14")
    c3 = count_commits_in_range(commits, "2026-04-17", "2026-04-19")
    c4 = count_commits_in_range(commits, "2026-04-21", "2026-04-21")
    c5 = count_commits_in_range(commits, "2026-04-22", "2026-04-24")
    c6 = count_commits_in_range(commits, "2026-04-25", "2026-04-27")
    c7 = count_commits_in_range(commits, "2026-04-28", "2026-04-30")
    c8 = count_commits_in_range(commits, "2026-05-01", "2026-05-08")
    c9 = count_commits_in_range(commits, "2026-05-09", "2026-05-15")
    c10 = count_commits_in_range(commits, "2026-05-16", "2026-05-22")
    c11 = count_commits_in_range(commits, "2026-05-23", "2026-05-29")
    c12 = count_commits_in_range(commits, "2026-05-30", "2026-06-05")
    c13 = count_commits_in_range(commits, "2026-06-06", "2026-06-11")
    c14 = count_commits_in_range(commits, "2026-06-12", "2026-06-19")
    c15 = count_commits_in_range(commits, "2026-06-20", "2026-06-26")
    c16 = count_commits_in_range(commits, "2026-06-27", "2026-07-03")

    rows = [
        ("2026-04-10 ~ 04-11", "기반 구축 (개인 선행)",
         "DB 스키마·FastAPI 백엔드·자재 통합 스크립트·SQLite", c1, "완료"),
        ("2026-04-12 ~ 04-14", "UI 개발 (개인 선행)",
         "모바일·데스크톱 UI·BOM·직원·QR스캔·다크모드\n레거시 UI 패리티", c2, "완료"),
        ("2026-04-17 ~ 04-19", "기능 고도화 (개인 선행)",
         "M1~M7(4-파트 코드)·Queue·Pending·안전재고\n실사 라우터", c3, "완료"),
        ("2026-04-21",          "UI 정제 (회사 개발 착수)",
         "레거시 레이아웃 Figma 맞춤 정제", c4, "완료"),
        ("2026-04-22 ~ 04-24", "실 운영 준비",
         "erp_code 통일·재고 시각화·모바일UX\n품목매칭 도구·ESLint·계산 일원화(N+1 제거)", c5, "완료"),
        ("2026-04-25 ~ 04-27", "코드 품질 정비",
         "입출고 단계형 Wizard UX·BOM Where-Used 조회\n코드 품질 대정비(Phase 5)·CI 도입\n로그인 화면 구현·카테고리 코드 정비", c6, "완료"),
        ("2026-04-28 ~ 04-30", "현장 운영 투입",
         "창고 승인 흐름·장바구니·PIN 로그인\n입출고탭 개편(부서필터·반품·자가승인)\nBOM 플래너·부서 관리 UX 개선\n품목 분류 개편 완료(DB 722건)", c7, "완료"),
        ("2026-05-01 ~ 05-08", "BOM·UI 정비",
         "자재 코드/품목명 1차 정비(848건)·BOM 세팅 도구\nBOM 관리자 워크벤치·자재 폼 강화\n주간 재고 변화 보고·부서간 조정 위저드\n공통 UI 5종·테스트 커버리지 91%·거대 파일 21개 분해", c8, "완료"),
        ("2026-05-09 ~ 05-15", "입출고 v2 + 내역 재설계 + 부서 결재",
         "입출고 v2 4단계 마법사·멱등성·위저드 마무리\n입출고 내역 화면 전면 재설계(scope·서버필터·묶음·흐름)\n부서 결재 + 직원 권한 분리\nBOM 855건 입력 진행(출하 PA/PF 50건)\n재고 이미지 컬럼·데이터 모델 정리(legacy 제거)\n주간 재고 변화 보고 시각 재정리·UX(Enter 키)·CI 안정화", c9, "완료"),
        ("2026-05-16 ~ 05-22", "인프라·모바일·불량 처리",
         "WS1~WS10 배포 안정화·보안 강화\n모바일 UI/UX 전면 개편(디자이너 PASS)\n불량 처리 흐름 재설계 Phase 1 (라인 결재·격리·처리)\nBOM 워크벤치 3-column·완료 워크플로\n입출고 내역 2·3차 main 머지\n권동환 5월 재고 100% 검증·CSV 외부 감시·ItemCode rename", c10, "완료"),
        ("2026-05-23 ~ 05-29", "사내 시연 + 양식 정비 + 운영 정책 맞춤 + 사번 감사",
         "관리자 화면 4종 한꺼번에 보강 (저장 경고·공통 버튼·BOM 전체 필터)\n직원별 입출고 권한 토글 신설 (개인 단위 차단)\n화면 데이터 갱신 방식 정비 (React Query)\n불량 처리·입출고 흐름 정합 + 5/28 자동결재·재작업 후속 보완\n품목 시리얼(공정 단위) + 김건호 과장님 검토 5월 재고 검증\n품목 관리 소프트 삭제 보강 / 생산 가능 화면 모델별 대표 제품 5종\n현장 사용성 피드백 11건 일괄 반영\n생산부 사내 시연 + 6월 이후 전환 계획 공유 (5/27)\n주간보고 매트릭스 거래타입 5종 보강 + 백엔드 동결 (5/29)\n운영 로그 시스템 + 사번 기반 추적 인프라 (5/29, 권동환 사원님 요청 3)\n부품 코드 자동 갱신 + DB 정합성 정리 (5/29, 어긋남 34건 → 0건)\n다중 사용자 피드백 처리 — PIN 4자리·수량 소수점·담당자 드롭다운 (5/29)", c11, "완료"),
        ("2026-05-30 ~ 06-05", "코드 품질 + 아키텍처 정비",
         "수량 정수 전용(IntQuantity) 아키텍처 정비 — 입력 422·출력·내부 일관화\nItemCode → mes_code 전면 리네임 완료\nlegacy → mes 폴더 개명 (프론트엔드 라우트 전면 재정리)\n클린코드 Phase 1 — schemas/ 분리·_tx_filters·README 정비\n디자인 일관성 정비 — Pretendard 실로딩 확인·EmptyState 중복 정리\nuseToggleSet 훅으로 토글 중복 12개 통합, React Query 패널 3종 이관\nAdmin Depts/Employees 하위 분리, 불량 처리 재설계 착수", c12, "완료"),
        ("2026-06-06 ~ 06-11", "모바일 구성 + 불량 처리 재설계",
         "불량 처리 패널 전면 재설계 + KPI 2개로 압축·부서 범위 명시\n격리 목록 기본 스코프 '전체' 전환·배지·체크박스 UX 개선\n창고 지도 열 클릭·슬롯 번호·박스 편집 기능 — 실사용 기반 마련\n직원별 품목 순서 커스터마이징 (드래그 재정렬·초기화)\n모바일 핵심 화면 구성 — 하단 탭 4개 + 더보기 시트\n모바일 불량 처리 카트 방식 전환 — 다품목 일괄 처리 가능\nio_preview·창고지도·불량 카트 단위 테스트 신설\n코드 품질 정비 — verify_local 전체 통과 확인", c13, "완료"),
        ("2026-06-12 ~ 06-19", "입출고 취소 + 모바일 탭바 + 현장 기준 개선",
         "입출고 취소 기능 신설 — 생산·분해 포함 모든 유형 취소, 원본 기록 보존\n생산 가능 수량 표시 AF 단위 기준 개선 중 (출하준비·빠른조립·총생산 구분)\n인수인계서 임시저장 + 수신 권한을 실제 수령 부서로 제한\n주간보고 화면 집계 오류 수정 (입고 건 생산 실적 혼입 제거)\n데스크톱 빠른 작업 탭 튕김·검색 깜빡임 수정, 창고 관리 탭 통합\n모바일 하단 탭바 슬라이딩 pill 디자인 완성\n모바일 화면 전반 정비 진행 중 (더보기 탭 전환·불량 허브·저장 확인)\nSOLO 테스트 지원으로 개발 시간 제한적", c14, "진행 중"),
        ("2026-06-20 ~ 06-26", "전환 안전·피드백 반영",
         "창고 박스 단위 재고 관리 기반 구축\n생산 가능 수량 기준 및 표시 방식 개선\n재고 전환 전 운영 안전장치 강화\n입출고 승인 전 수정·낱개 출고 집계 정확도 개선\n알림 삭제 기능 추가\n모바일 화면 정비 5·6·7·8차 진행\n부적합품 관리 기획과 오류 보완", c15, "진행 중"),
        ("2026-06-27 ~ 07-03", "출하 관리 신설 + 현장 피드백",
         "출하 관리 화면 신설 및 개선 (진행 중)\n대시보드 재고 표시 개편 — 권동환 사원님 피드백\n불량 처리 개선 — 김건호 과장님 피드백\n입출고 내역 화면 개선\n로그인·탭 이동 방식 개선 — 김건호 과장님 피드백\n창고 지도·승인함 사용성 개선 — 권동환 사원님 피드백\n품목 관리·검색 편의 개선 — 권동환 사원님 피드백\n원자재 입출고 사용성 개선 — 권동환 사원님 피드백", c16, "진행 중"),
    ]

    for r, (period, cat, work, cnt, _status) in enumerate(rows, 2):
        # status 필드는 rows 튜플 호환 위해 받지만 컬럼에 적지 않음 (사용자가 '상태' 컬럼 제거)
        # 행 높이는 work 텍스트 줄 수에 비례 — 사용자가 매번 손으로 키우지 않도록 자동화
        line_count = work.count("\n") + 1
        ws.row_dimensions[r].height = max(60, line_count * 16.2)
        apply_data(ws.cell(r, 1), period, align="center")
        apply_data(ws.cell(r, 2), cat,    align="center")
        apply_data(ws.cell(r, 3), work)
        apply_data(ws.cell(r, 4), cnt,    align="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 10

# ── Sheet 2: 기능 전체 목록 ──────────────────────────────
def build_features(ws):
    ws.title = "기능 전체 목록"
    ws.row_dimensions[1].height = 22

    headers = ["No", "기능명", "분류", "상태", "비고"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    features = [
        # (기능명, 분류, 완료여부, 비고)
        # ── 재고 ─────────────────────────────────────────
        ("품목 마스터 관리",                   "재고",   True,  ""),
        ("자재 코드 체계 통일",                 "재고",   True,  "담당자 협의 필요"),
        ("재고 입고 / 출고 / 조정",             "재고",   True,  ""),
        ("재고 이력 관리",                     "재고",   True,  ""),
        ("창고 / 생산 / 불량 버킷 분리",        "재고",   True,  ""),
        ("안전재고 기준 이하 알림",              "재고",   True,  ""),
        ("재고 현황 막대 시각화",               "재고",   True,  ""),
        ("품목 분류 개편 (process_type_code)", "재고",   True,  "DB 722건 재구성"),
        ("재고 실사 (Physical Count)",         "재고",   True,  "강제 조정 포함"),
        ("차이 분석 (Variance)",               "재고",   True,  "실제-예상 수량 차이 기록"),
        ("폐기 · 분실 이력 관리",               "재고",   True,  "Scrap / Loss 라우터"),
        # ── 입출고 워크플로 ──────────────────────────────
        ("입출고 단계형 Wizard UX",             "입출고", True,  ""),
        ("입출고 부서별 필터 · 반품 · 자가승인", "입출고", True,  ""),
        ("창고 승인 기반 입출고 흐름",           "입출고", True,  "요청→승인→처리"),
        ("직원별 저장형 입출고 장바구니",        "입출고", True,  "화면 이탈 후 유지"),
        ("생산 / 분해 / 반품 처리 Queue",      "생산",   True,  "2단계 워크플로"),
        ("QR · 바코드 스캔",                  "입출고", True,  "BarcodeDetector + zxing"),
        # ── 출하 ─────────────────────────────────────────
        ("출하 패키지 관리",                   "출하",   True,  "묶음 CRUD + 품목 관리"),
        # ── BOM ──────────────────────────────────────────
        ("BOM 등록 / 조회",                   "BOM",   True,  ""),
        ("BOM 웹 수정 기능",                  "BOM",   True,  ""),
        ("BOM Where-Used 조회",              "BOM",   True,  ""),
        ("BOM 플래너 (위저드 · 일괄 다운로드)", "BOM",   True,  ""),
        ("BOM 관리자 워크벤치",                "BOM",   True,  "검색·자재·소요량 통합 화면"),
        ("BOM 세팅 도구 (오프라인 · 848 품목)", "BOM",   True,  "855건 입력 진행 (출하 PA/PF 추가)"),
        ("BOM 부모 후보 추출 스크립트",         "BOM",   True,  "입출고 관리대장 2024-2026 기반"),
        ("자재 폼 ERP/옵션/모델 슬롯 편집",     "재고",   True,  "ItemFormFields 공통화"),
        ("부서간 재고 조정 4단계 위저드",        "입출고", True,  "입출고 v2에 흡수 정리"),
        ("주간 재고 변화 보고 화면",            "관리",   True,  "공정별 입출고·달력 히스토리"),
        ("모바일 5탭 + 입출고 허브 (사전)",     "UI",    True,  "최종 사용성은 후순위"),
        ("공통 UI 컴포넌트 5종",                "인프라", True,  "FilterChip·SlidePanel·KpiCard 등"),
        # ── 5/9 ~ 5/15 신규 ────────────────────────────
        ("입출고 v2 — 4단계 마법사",            "입출고", True,  "sub_type 분기 + 권한 가드 + 위저드 재디자인"),
        ("입출고 v2 — 중복 제출 방지(멱등성)",   "입출고", True,  "client_request_id + 409/503 UX"),
        ("입출고 내역 화면 전면 재설계",         "입출고", True,  "scope·서버필터·묶음 상세 패널·흐름 라벨"),
        ("주간 재고 변화 보고 화면 시각 재정리",  "관리",   True,  "KPI 상단·헤더 통합·색상 위계·0값 투명"),
        ("UX 일관성 — 모달 Enter 키",            "UI",    True,  "전 화면 모달/팝업 통일"),
        ("부서 결재 + 직원 권한 분리",          "보안",   True,  "부서 결재 API 3건 + 조립 부서 담당 모델"),
        ("재고 대시보드 품목 이미지 컬럼",       "재고",   True,  "Next/Image 최적화"),
        ("데이터 모델 정리 (legacy 제거)",      "인프라", True,  "legacy_model + category + ShipPackage 제거"),
        ("AppSelect 셀렉트 디자인 통일",        "UI",    True,  "공통 드롭다운 컴포넌트"),
        ("로그인 단일 카드 화면",               "보안",   True,  "3단계 → 콤보박스+PIN 1화면"),
        ("관리자 화면 8섹션 + PIN",             "관리",   True,  "BOM 가로분할·KPI 컴팩트화"),
        ("생산 가능 수량 immediate/maximum",   "생산",   True,  "status 4분기 시각화"),
        # ── 5/16 ~ 5/22 신규 ───────────────────────────
        ("운영 인프라 안정화 (WS1~WS10)",       "인프라", True,  "Docker·보안·로깅·예외·헬스/관측"),
        ("입출고 내역 2·3차 정비 (main 머지)",  "입출고", True,  "model·process_step·department 파라미터"),
        ("모바일 UI/UX 전면 개편",              "UI",    True,  "5개 화면 반응형 + 디자이너 PASS"),
        ("불량 처리 흐름 재설계 Phase 1",        "생산",   True,  "라인 결재·격리·처리·테스트 113건"),
        ("BOM 워크벤치 3-column + 완료 워크플로", "BOM",  True,  "BOM 완료 상태 토글 + export"),
        ("입출고 CSV 외부 감시 시스템",          "인프라", True,  "월별 누적 미러"),
        ("ItemCode 도메인 정리",                "인프라", True,  "erp_code 잔재 정리"),
        ("외부 PC 첫 실행 안정화",              "인프라", True,  "gitattributes + start.bat 사전 검사"),
        ("권동환 5월 재고 100% 검증",           "데이터", True,  "엑셀↔DB 동기화·자식 행 반영"),
        # ── 5/23 ~ 5/28 신규 ───────────────────────────
        ("직원별 입출고 권한 토글 신설",          "보안",   True,  "직원 개인 단위 토글, 꺼지면 진입 불가"),
        ("관리자 화면 4종 한꺼번에 보강",         "관리",   True,  "저장 경고·공통 버튼·BOM 전체 필터·대시보드 동선"),
        ("화면 데이터 갱신 방식 정비",            "인프라", True,  "React Query 도입·도메인 10여 개 통일"),
        ("모델 관리 인라인 편집 + 끌어서 정렬",   "관리",   True,  "정렬 순서 DB 컬럼 추가"),
        ("불량 처리 흐름 후속 보완",              "생산",   True,  "5/28 자동결재·재작업 화면·요청자/승인자 표시"),
        ("품목 시리얼 (공정 단위) + 5월 재고 검증","재고",  True,  "공정 단위, 김건호 과장님 검토 반영"),
        ("품목 관리 — 소프트 삭제 보강",          "관리",   True,  "삭제 시 비활성 상태로 목록 아래, 재활성 가능"),
        ("생산 가능 화면 — 모델별 대표 제품 5종", "생산",   True,  "기존 합계 표시 → 대표 완제품 5종으로"),
        ("현장 사용성 피드백 11건 일괄 반영",     "UX",    True,  "안내 메시지·모달·큰 이미지 보기·필터 통일 등"),
        ("백엔드 큰 파일 정리 (종류별 분리)",     "인프라", True,  "models.py·bootstrap_db.py + 로그인 인증 흐름 정리"),
        # ── 5/29 신규 ──────────────────────────────────
        ("운영 로그 시스템 + 사번 기반 추적",      "보안",   True,  "권동환 사원님 요청 3 — 도메인 이벤트 15군데 사번 부착·AdminAuditLog 사번 컬럼"),
        ("부품 코드 자동 갱신 + DB 정합성 정리",   "데이터", True,  "어긋남 34건 → 0건 · item_models 테이블 폐기"),
        ("주간보고 매트릭스 보강 + 백엔드 동결",   "주간보고", True, "거래타입 5종 분류·11종 enum 회귀 테스트"),
        ("다중 사용자 피드백 — PIN·수량·담당자",   "UX",     True, "PIN 4자리·수량 소수점·자동완성 드롭다운"),
        # ── 직원 · 보안 ──────────────────────────────────
        ("직원 명단 관리",                     "직원",   True,  ""),
        ("부서별 담당 색상 배지 표시",           "직원",   True,  ""),
        ("작업자 PIN 로그인 + 감사 이력",       "보안",   True,  "수정 이력 자동 기록"),
        ("로그인 및 권한 관리",                 "보안",   True,  "PIN 기반 — 역할별 접근 제어 예정"),
        # ── UI · 관리 ────────────────────────────────────
        ("PC 화면 (데스크톱)",                 "UI",    True,  ""),
        ("모바일 화면",                        "UI",    True,  ""),
        ("다크 / 라이트 모드",                 "UI",    True,  ""),
        ("관리자 화면",                        "관리",   True,  ""),
        ("부서 관리 (DnD · 컬러 피커 · 검색)", "관리",   True,  ""),
        ("데이터 CSV 내보내기",                "관리",   True,  ""),
        # ── 인프라 ───────────────────────────────────────
        ("원클릭 실행 (start.bat)",            "인프라", True,  ""),
        ("테스트 자동화 (CI)",                 "인프라", True,  "pytest 42건 + vitest 12건"),
        ("WAL-safe 백업 / 복구",              "인프라", True,  "ops 스크립트 자동화"),
        # ── 6/12~6/19 신규 ───────────────────────────────────
        ("입출고 취소 기능",              "입출고", True,  "생산·분해 포함 모든 유형 취소 — 원본 보존·재고 역방향 복원"),
        # ── 6/20~6/26 신규 ───────────────────────────────────
        ("창고 박스 단위 재고 관리 기반", "재고", True, "박스 차감·복원·이동·미배치 알림"),
        ("생산 가능 수량 기준 제품 지정", "생산", True, "모델별 기준 제품 지정·출하 대기 수량 보정"),
        ("승인 전 입출고 요청 수정", "입출고", True, "작성자가 승인 전 요청을 수정 가능"),
        ("알림 삭제 기능", "알림", True, "개별 삭제·읽은 알림 전체 정리"),
        ("모바일 화면 정비 5~8차", "UI", True, "입출고·불량·대시보드·로그인 화면 보강"),
        ("재고 전환 안전 점검", "인프라", True, "위험 취소 차단·감사 준비·백업 확인"),
        ("출고·출하 창고 가져오기 확장", "입출고", True, "분해 작업까지 부족 품목 가져오기 지원"),
        ("창고 지도 권한 보존", "관리", True, "구조 편집 후 권한 설정 유지"),
        # ── 6/27~7/3 신규 ─────────────────────────────────────
        ("로그인 알림 팝업", "알림", True, "로그아웃 중 알림을 로그인 시 1회 팝업 표시 + 개인 켜기/끄기 설정"),
        ("대시보드 부서별 재고 표시", "재고", True, "창고·조립·불량 위치별 수량 알약 표시로 개편"),
        ("창고 지도 전체화면", "관리", True, "PC 창고 지도를 넓게 보는 전체화면 모드"),
        ("입출고·출하 탭 이동 확인", "UX", True, "탭 재클릭 시 첫 화면 복귀 + 작업 중 이탈 확인"),
        # ── 예정 ─────────────────────────────────────────
        ("실 재고 데이터 입력 + 직원 테스트",    "데이터", False, "BOM·입출고 정리 후"),
        ("불량 처리 흐름 Phase 2",              "생산",   False, "부서 결재·집계·자동화"),
        ("역할별 접근 권한 연동",              "보안",  False, "부서·역할 기반 (PIN 로그인 완료)"),
        ("총재고 / 가용재고 / 예약재고 분리",   "재고",  False, "재고 가용성 계산 고도화"),
        ("외부 접근 환경 구성",                "인프라", False, "PC 또는 NAS 서버 + API 인증"),
        ("발주 관리",                         "구매",  False, ""),
        ("생산 실적 / 원가 관리",              "생산",  False, ""),
        ("부서별 생산진행도 대시보드",          "관리",  False, "사장님 요청 — 검토 중"),
        ("출하 거래처 관리",                   "출하",  False, "거래처 마스터 + 출하 이력"),
    ]

    for r, (name, cat, done, note) in enumerate(features, 2):
        no = r - 1
        status = "✅ 완료" if done else "🔲 예정"
        bg = DONE_BG if done else TODO_BG
        apply_data(ws.cell(r, 1), no,     bg, align="center")
        apply_data(ws.cell(r, 2), name,   bg)
        apply_data(ws.cell(r, 3), cat,    bg, align="center")
        apply_data(ws.cell(r, 4), status, bg, bold=True, align="center")
        apply_data(ws.cell(r, 5), note,   bg)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 28

# ── Sheet 3: 커밋 이력 ───────────────────────────────────
_TYPE_MAP = {
    "feat": "feat", "fix": "fix", "refactor": "refactor",
    "docs": "docs", "chore": "chore", "ci": "ci", "test": "test",
    "perf": "perf", "style": "style", "build": "build",
    "desktop": "desktop", "mobile": "mobile", "backend": "backend",
    "frontend": "frontend", "admin": "admin", "scripts": "scripts",
    "devx": "devx", "merge": "merge", "design": "design",
}

def fetch_commits():
    result = subprocess.run(
        ["git", "log", "--format=%ad|%h|%s", "--date=format-local:%Y-%m-%d %H:%M"],
        capture_output=True, text=True, encoding="utf-8"
    )
    commits = []
    for line in result.stdout.strip().splitlines():
        parts = line.split("|", 2)
        if len(parts) != 3:
            continue
        dtstr, hash7, msg = parts
        dt = datetime.strptime(dtstr.strip(), "%Y-%m-%d %H:%M")
        date = dt.strftime("%Y-%m-%d")
        # 날짜 prefix 제거: "2026-04-25 desktop: ..." → "desktop: ..."
        msg_clean = re.sub(r"^\d{4}-\d{2}-\d{2}\s+", "", msg)
        m = re.match(r"^([a-z][a-z+\-]*)(\(.+?\))?[!:]", msg_clean)
        raw = m.group(1) if m else None
        ctype = _TYPE_MAP.get(raw, "기타") if raw else "기타"
        commits.append((date, hash7, msg, ctype, dt))
    return list(reversed(commits))  # 오래된 순으로

def group_commits_by_date(commits):
    grouped = defaultdict(list)
    for date, hash7, msg, ctype, dt in commits:
        grouped[date].append((hash7, msg, ctype, dt))
    return grouped

def count_commits_in_range(commits, start_date, end_date):
    count = 0
    for date, *_ in commits:
        if start_date <= date <= end_date:
            count += 1
    return count

def build_daily(ws, commits):
    ws.title = "일일 개발 현황"
    ws.row_dimensions[1].height = 22

    headers = ["날짜", "커밋 수", "주요 작업", "영역"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    by_date = group_commits_by_date(commits)
    dates_sorted = sorted(by_date.keys())

    r = 2
    for date in dates_sorted:
        msgs = by_date[date]
        count = len(msgs)
        work_summary = "\n".join([msg[:60] for hash7, msg, ctype, dt in msgs[:5]])
        area = "Backend" if any("feat" in msg.lower() for _, msg, _, _ in msgs) else "Frontend"

        # 행 높이는 텍스트 줄 수에 비례 — 사용자가 매번 손으로 조정하지 않도록 자동
        line_count = work_summary.count("\n") + 1
        ws.row_dimensions[r].height = max(27, line_count * 13.5)
        bg = STRIPE_BG if r % 2 == 0 else WHITE_BG
        apply_data(ws.cell(r, 1), date,          bg, align="center")
        apply_data(ws.cell(r, 2), count,         bg, align="center")
        apply_data(ws.cell(r, 3), work_summary,  bg)
        apply_data(ws.cell(r, 4), area,          bg, align="center")
        r += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 14

def build_commits(ws, commits):
    ws.title = "커밋 이력"
    ws.row_dimensions[1].height = 22

    headers = ["날짜", "시각", "해시", "커밋 메시지", "타입", "구분"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(1, col), h)

    for r, (date, hash7, msg, ctype, dt) in enumerate(commits, 2):
        status = classify_work(dt)
        if status == "근무 외":
            bg = OFFHOURS_BG
        elif status == "개인 선행":
            bg = PERSONAL_BG
        else:
            bg = STRIPE_BG if r % 2 == 0 else WHITE_BG
        apply_data(ws.cell(r, 1), date,                 bg, align="center")
        apply_data(ws.cell(r, 2), dt.strftime("%H:%M"), bg, align="center")
        apply_data(ws.cell(r, 3), hash7,                bg, align="center")
        apply_data(ws.cell(r, 4), msg,                  bg)
        apply_data(ws.cell(r, 5), ctype,                bg, align="center")
        apply_data(ws.cell(r, 6), status,               bg,
                   bold=(status == "근무 외"), align="center")

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10

# ── 메인 ─────────────────────────────────────────────────
def main():
    OUTPUT_PATH.parent.mkdir(exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    commits = fetch_commits()
    off_count = sum(1 for c in commits if classify_work(c[4]) == "근무 외")

    build_dashboard(wb.create_sheet(), commits, off_count)
    build_summary(wb.create_sheet(), commits)
    build_features(wb.create_sheet())
    build_daily(wb.create_sheet(), commits)
    build_commits(wb.create_sheet(), commits)

    wb.save(OUTPUT_PATH)
    print(f"저장 완료: {OUTPUT_PATH}  (커밋 {len(commits)}개 · 근무 외 {off_count}건)")

if __name__ == "__main__":
    main()
