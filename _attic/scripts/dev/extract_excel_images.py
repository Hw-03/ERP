"""생산부 자재 xlsx의 이미지 추출 + 품목 매칭 CSV 생성.

지원 시트(시트마다 컬럼 위치가 달라 SHEETS 테이블로 관리):
  - 고압        : 2026.05_생산부 자재_고압,진공,튜닝파트.xlsx
  - 조립 자재    : 2026.05_생산부 자재_조립,출하파트.xlsx
  - 튜브        : 2026.05_생산부 자재_튜브 파트.xlsx

산출:
  data/이미지 추출을 위한 원본/extracted/{시트명}/r{NNN}_{품목명}.{jpg|png}
  data/이미지 추출을 위한 원본/extracted/{시트명}/extracted_index.csv

화질 보존: openpyxl이 보관한 원본 바이트(img._data())를 그대로 write_bytes 한다. 재인코딩 없음.
WMF/EMF/WDP 등은 openpyxl이 지원하지 않아 자동으로 drop 됨 (경고로 표시).
"""
from __future__ import annotations

import csv
import re
import sys
import warnings
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC_DIR = ROOT / "data" / "이미지 추출을 위한 원본"
OUT_BASE = SRC_DIR / "extracted"

# 시트별 설정: (xlsx 파일명, 시트명, 분류 col, 모델 col or None, 품목 col)
SHEETS: list[dict] = [
    {
        "xlsx": "2026.05_생산부 자재_고압,진공,튜닝파트.xlsx",
        "sheet": "고압",
        "col_분류": 3,
        "col_모델": 4,
        "col_품목": 5,
    },
    {
        "xlsx": "2026.05_생산부 자재_조립,출하파트.xlsx",
        "sheet": "조립 자재",
        "col_분류": 2,
        "col_모델": 3,
        "col_품목": 4,
    },
    {
        "xlsx": "2026.05_생산부 자재_튜브 파트.xlsx",
        "sheet": "튜브",
        "col_분류": 2,
        "col_모델": None,
        "col_품목": 3,
    },
]

FORBIDDEN = re.compile(r'[\\/:*?"<>|]')
WHITESPACE = re.compile(r"\s+")


def sanitize(name: str) -> str:
    name = FORBIDDEN.sub(" ", name)
    name = WHITESPACE.sub(" ", name).strip(" .")
    if len(name) > 200:
        name = name[:200].rstrip()
    return name or "unnamed"


def sniff_ext(head: bytes) -> str:
    if head.startswith(b"\x89PNG"):
        return "png"
    if head.startswith(b"\xff\xd8\xff"):
        return "jpg"
    if head.startswith(b"GIF8"):
        return "gif"
    if head.startswith(b"BM"):
        return "bmp"
    return "bin"


def extract(xlsx_path: Path, sheet_name: str, out_dir: Path,
            col_분류: int, col_모델: int | None, col_품목: int) -> tuple[list[tuple], list[str]]:
    """이미지 추출. (records, dropped_warnings) 반환."""
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        dropped = [str(w.message) for w in caught if "not supported" in str(w.message).lower()]

    ws = wb[sheet_name]
    out_dir.mkdir(parents=True, exist_ok=True)

    records: list[tuple] = []
    used_names: set[str] = set()
    if not ws._images:
        return records, dropped

    images = sorted(ws._images, key=lambda i: i.anchor._from.row)

    # 분류/모델 forward-fill: 시트 위에서부터 순회하며 마지막 값 캐시
    max_row_needed = max(img.anchor._from.row + 1 for img in images)
    last_분류: str | None = None
    last_모델: str | None = None
    row_ctx: dict[int, tuple[str | None, str | None]] = {}
    for r in range(1, max_row_needed + 1):
        v_cls = ws.cell(r, col_분류).value
        if v_cls:
            last_분류 = str(v_cls).strip()
        if col_모델 is not None:
            v_mdl = ws.cell(r, col_모델).value
            if v_mdl:
                last_모델 = str(v_mdl).strip()
        row_ctx[r] = (last_분류, last_모델)

    for img in images:
        row = img.anchor._from.row + 1
        품목_raw = ws.cell(row, col_품목).value
        품목 = str(품목_raw).strip() if 품목_raw is not None else ""
        분류, 모델 = row_ctx.get(row, (None, None))

        data = img._data()
        ext = sniff_ext(data[:4])
        if ext == "bin":
            print(f"  ! row {row}: unknown image format, head={data[:8]!r}", file=sys.stderr)

        fname = f"r{row:03d}_{sanitize(품목)}.{ext}"
        # 같은 행에 이미지가 둘 이상 박혀 있는 드문 경우 대비 (이번 실행 내 중복만 처리)
        if fname in used_names:
            base, _, e = fname.rpartition(".")
            n = 2
            while f"{base}-{n}.{e}" in used_names:
                n += 1
            fname = f"{base}-{n}.{e}"
        used_names.add(fname)
        (out_dir / fname).write_bytes(data)
        records.append((row, 분류 or "", 모델 or "", 품목, fname, len(data)))

    return records, dropped


def write_csv(records: list[tuple], csv_path: Path) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["행", "분류", "모델", "품목", "파일명", "바이트수"])
        for rec in records:
            w.writerow(rec)


def run_one(cfg: dict) -> int:
    xlsx = SRC_DIR / cfg["xlsx"]
    sheet = cfg["sheet"]
    out_dir = OUT_BASE / sheet
    if not xlsx.exists():
        print(f"원본 xlsx 없음: {xlsx}", file=sys.stderr)
        return 0

    print("=" * 60)
    print(f"파일: {xlsx.name}")
    print(f"시트: {sheet}")
    print(f"출력: {out_dir}")

    records, dropped = extract(
        xlsx, sheet, out_dir,
        cfg["col_분류"], cfg["col_모델"], cfg["col_품목"],
    )
    csv_path = out_dir / "extracted_index.csv"
    write_csv(records, csv_path)

    exts = Counter(rec[4].rsplit(".", 1)[-1] for rec in records)
    rows = [r[0] for r in records]
    print(f"\n추출: {len(records)}개  / 확장자: {dict(exts)}")
    if rows:
        print(f"행 범위: {min(rows)} ~ {max(rows)}")
    if dropped:
        print(f"[!] openpyxl drop {len(dropped)}건 (WMF/EMF/WDP 미지원)")
    print(f"CSV: {csv_path}\n")
    return len(records)


def main(argv: list[str]) -> int:
    targets = SHEETS
    if len(argv) > 1:
        wanted = set(argv[1:])
        targets = [c for c in SHEETS if c["sheet"] in wanted]
        if not targets:
            print(f"매칭되는 시트 없음. 가능: {[c['sheet'] for c in SHEETS]}", file=sys.stderr)
            return 1
    total = sum(run_one(cfg) for cfg in targets)
    print(f"=== 전체 합계: {total}개 이미지 ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
